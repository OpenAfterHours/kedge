"""The workbook handle: one open ``.xlsx``, two openpyxl views, and the raw zip.

openpyxl cannot give formulas and cached values from a single load, so kedge loads the
workbook twice — once normally and once with ``data_only=True`` — and zips the views
together (PLAN §1.5). Both loads are ``read_only`` because the pitch involves 40MB
workbooks with half a million formula cells and a full object graph for those is not
affordable.

What ``read_only=True`` costs us, and how each is recovered:

===========================  ====================================================
Lost in read-only mode       Recovered by
===========================  ====================================================
``ws.merged_cells``          Byte scan of the sheet part for ``<mergeCell ref=.../>``
``ws.column_dimensions``     Byte scan of the sheet part for ``<col hidden="1"/>``
Sheet-scoped defined names   Parsing ``xl/workbook.xml`` directly
Cell comments                Left to :mod:`kedge.analysis.docs` via :meth:`read_part`
Data validations, styles     Not recovered; the analyser does not use them
===========================  ====================================================

Everything else survives: ``sheet_state``, workbook-scoped defined names, ``calcPr``, and
the document properties all load in read-only mode.

References:
- PLAN.md §1.5 (dual load), §2.3 (bounded reads), §M1.
"""

from __future__ import annotations

import hashlib
import logging
import re
import zipfile
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from types import TracebackType
from typing import TYPE_CHECKING, Any, Literal
from xml.etree import ElementTree

from openpyxl import load_workbook

from kedge.analysis.model import NamedRange, WorkbookIdentity
from kedge.errors import AnalysisError

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import Workbook

logger = logging.getLogger(__name__)

__all__ = [
    "SUPPORTED_FORMATS",
    "SheetStructure",
    "WorkbookHandle",
    "WorkbookOpenError",
    "detect_format",
    "open_workbook",
    "parse_a1_range",
    "read_identity",
]

SUPPORTED_FORMATS = frozenset({"xlsx", "xlsm"})
"""Formats the analyser can actually read. ``.xlsb`` and ``.xls`` are refused (PLAN §6.2)."""

_MAIN_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
_PKG_REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"

_DEFAULT_STRUCTURE_BUDGET = 64 * 1024 * 1024
_HEAD_LIMIT = 1024 * 1024
_TAIL_LIMIT = 1024 * 1024
_CHUNK = 1 << 18

_COL_RE = re.compile(rb"<col\b[^>]*>")
_ATTR_RE = re.compile(rb'([A-Za-z:]+)="([^"]*)"')
_MERGE_RE = re.compile(rb'<mergeCell\b[^>]*ref="([^"]+)"')
_RANGE_RE = re.compile(
    r"^\$?([A-Za-z]{1,3})?\$?([0-9]{1,7})?(?::\$?([A-Za-z]{1,3})?\$?([0-9]{1,7})?)?$"
)

_MAX_READ_ROWS = 500
_MAX_READ_COLS = 128


class WorkbookOpenError(AnalysisError):
    """The file could not be opened as an Excel workbook at all."""


# =============================================================================
# IDENTITY
# =============================================================================


def detect_format(path: Path) -> Literal["xlsx", "xlsm", "xlsb", "xls", "unknown"]:
    """Classify a file by suffix, corrected by its magic bytes.

    A ``.xlsx`` that is really a legacy OLE2 ``.xls`` renamed is a real thing people do, and
    it must be refused with the right message rather than an XML parse error.

    Args:
        path: The file to classify.

    Returns:
        One of the ``WorkbookIdentity.file_format`` literals.
    """
    suffix = path.suffix.lower().lstrip(".")
    try:
        with path.open("rb") as handle:
            magic = handle.read(8)
    except OSError:
        magic = b""

    if magic.startswith(b"\xd0\xcf\x11\xe0"):  # OLE2 compound document
        return "xls"
    if magic.startswith(b"PK") and suffix in ("xlsx", "xlsm", "xlsb"):
        return "xlsb" if suffix == "xlsb" else ("xlsm" if suffix == "xlsm" else "xlsx")
    if suffix in ("xlsx", "xlsm", "xlsb", "xls"):
        return suffix  # type: ignore[return-value]
    return "unknown"


def _sha256(path: Path) -> str:
    """Hash a file in chunks, so a 40MB workbook does not land in memory twice."""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _text_of(data: bytes | None, tag: str) -> str | None:
    """Pull the text of the first element with this local tag name out of an XML part."""
    if not data:
        return None
    try:
        root = ElementTree.fromstring(data)
    except ElementTree.ParseError:
        return None
    for element in root.iter():
        if element.tag.rsplit("}", 1)[-1] == tag and element.text:
            return element.text.strip()
    return None


def read_identity(path: Path, zf: zipfile.ZipFile | None = None) -> WorkbookIdentity:
    """Build the :class:`~kedge.analysis.model.WorkbookIdentity` for a file.

    Works without opening the workbook in openpyxl, so it is safe to call on a format the
    analyser is about to refuse.

    Args:
        path: The workbook path.
        zf: An already-open archive, if the caller has one.

    Returns:
        The identity record. ``has_vba`` and ``created_by`` are left at their defaults when
        the file is not a readable zip.
    """
    stat = path.stat()
    file_format = detect_format(path)
    identity_kwargs: dict[str, Any] = {
        "path": str(path),
        "filename": path.name,
        "sha256": _sha256(path),
        "size_bytes": stat.st_size,
        "file_format": file_format,
        "modified_at": datetime.fromtimestamp(stat.st_mtime, tz=UTC),
    }

    archive = zf
    opened_here = False
    if archive is None and file_format in SUPPORTED_FORMATS:
        try:
            archive = zipfile.ZipFile(path)
            opened_here = True
        except (OSError, zipfile.BadZipFile):
            archive = None

    if archive is not None:
        names = set(archive.namelist())
        identity_kwargs["has_vba"] = "xl/vbaProject.bin" in names
        app = archive.read("docProps/app.xml") if "docProps/app.xml" in names else None
        identity_kwargs["created_by"] = _text_of(app, "Application")
        book = archive.read("xl/workbook.xml") if "xl/workbook.xml" in names else None
        identity_kwargs["iterative_calculation"] = _iterative_calculation(book)
        if opened_here:
            archive.close()

    return WorkbookIdentity(**identity_kwargs)


def _iterative_calculation(workbook_xml: bytes | None) -> bool:
    """Read ``calcPr/@iterate`` out of ``xl/workbook.xml``."""
    if not workbook_xml:
        return False
    try:
        root = ElementTree.fromstring(workbook_xml)
    except ElementTree.ParseError:
        return False
    calc = root.find(f"{_MAIN_NS}calcPr")
    if calc is None:
        return False
    return calc.get("iterate", "0") in ("1", "true")


# =============================================================================
# SHEET STRUCTURE
# =============================================================================


@dataclass(frozen=True, slots=True)
class SheetStructure:
    """The structural facts about a sheet that ``read_only`` mode drops on the floor."""

    name: str
    index: int
    state: str = "visible"
    part: str | None = None
    hidden_columns: tuple[str, ...] = ()
    merged_ranges: tuple[str, ...] = ()
    structure_truncated: bool = False
    """True when the sheet part was too large to scan to the end, so merged ranges may be
    incomplete. Hidden columns are always complete: ``<cols>`` precedes ``<sheetData>``."""

    @property
    def is_hidden(self) -> bool:
        """Whether the sheet is hidden by either mechanism."""
        return self.state in ("hidden", "veryHidden")

    @property
    def is_very_hidden(self) -> bool:
        """Whether the sheet is hidden such that the Excel UI will not reveal it."""
        return self.state == "veryHidden"


def _column_letters(index: int) -> str:
    """1-based column index to letters."""
    letters = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _scan_part(zf: zipfile.ZipFile, part: str, budget: int) -> tuple[bytes, bytes, bool]:
    """Return (head, tail, truncated) for a sheet part without materialising the middle."""
    try:
        info = zf.getinfo(part)
    except KeyError:
        return b"", b"", False

    if info.file_size <= budget:
        data = zf.read(part)
        return data, data, False

    head = b""
    tail = b""
    seen = 0
    buffer = b""
    with zf.open(part) as source:
        while True:
            chunk = source.read(_CHUNK)
            if not chunk:
                break
            seen += len(chunk)
            if not head:
                buffer += chunk
                marker = buffer.find(b"<sheetData")
                if marker != -1:
                    head = buffer[:marker]
                    buffer = b""
                elif len(buffer) > _HEAD_LIMIT:
                    head = buffer
                    buffer = b""
            tail = (tail + chunk)[-_TAIL_LIMIT:]
            if seen > budget:
                return head or buffer, tail, True
    return head or buffer, tail, False


def _hidden_columns(head: bytes) -> tuple[str, ...]:
    """Extract hidden column letters from the ``<cols>`` block."""
    hidden: list[str] = []
    for element in _COL_RE.findall(head):
        attrs = {key.decode(): value.decode() for key, value in _ATTR_RE.findall(element)}
        if attrs.get("hidden") not in ("1", "true"):
            continue
        try:
            first = int(attrs.get("min", "0"))
            last = int(attrs.get("max", "0"))
        except ValueError:
            continue
        if first < 1 or last < first:
            continue
        hidden.extend(_column_letters(index) for index in range(first, min(last, 16_384) + 1))
    return tuple(hidden)


def _merged_ranges(tail: bytes) -> tuple[str, ...]:
    """Extract merged range references from the ``<mergeCells>`` block."""
    return tuple(ref.decode() for ref in _MERGE_RE.findall(tail))


# =============================================================================
# THE HANDLE
# =============================================================================


class WorkbookHandle:
    """One open workbook: the zip, both openpyxl views, and bounded access to cells.

    The handle is the shared object every extractor is given. It owns three file handles
    and must be closed; use it as a context manager.

    Example:
        >>> with open_workbook(Path("process.xlsx")) as handle:  # doctest: +SKIP
        ...     handle.sheet_names
        ...     handle.read_range("Parameters", "A1:B10")
    """

    def __init__(
        self,
        path: Path,
        zf: zipfile.ZipFile,
        formulas: Workbook,
        values: Workbook,
        identity: WorkbookIdentity,
        *,
        structure_budget: int = _DEFAULT_STRUCTURE_BUDGET,
    ) -> None:
        self.path = path
        self.zf = zf
        self.formulas = formulas
        self.values = values
        self.identity = identity
        self._structure_budget = structure_budget
        self._parts = frozenset(zf.namelist())
        self._sheet_parts = _sheet_parts(zf)
        self._structures: dict[str, SheetStructure] = {}
        self._defined_names: list[NamedRange] | None = None
        self._closed = False

    # ── identity and inventory ───────────────────────────────────────────

    @property
    def sheet_names(self) -> list[str]:
        """Worksheet names in workbook order."""
        return list(self.formulas.sheetnames)

    @property
    def part_names(self) -> frozenset[str]:
        """Every part in the archive, for extractors that go straight to the XML."""
        return self._parts

    def has_part(self, name: str) -> bool:
        """Whether the archive contains this part."""
        return name in self._parts

    def read_part(self, name: str) -> bytes | None:
        """Read one part of the archive, or None when it is absent or unreadable.

        Args:
            name: The part path, e.g. ``"xl/connections.xml"``.

        Returns:
            The raw bytes, or None. Never raises for a missing or corrupt member.
        """
        if name not in self._parts:
            return None
        try:
            return self.zf.read(name)
        except (KeyError, OSError, zipfile.BadZipFile) as exc:
            logger.warning("could not read part %s: %s", name, exc)
            return None

    # ── sheets ───────────────────────────────────────────────────────────

    def formula_sheet(self, sheet: str) -> Any:
        """The read-only worksheet from the formula view."""
        return self.formulas[sheet]

    def value_sheet(self, sheet: str) -> Any:
        """The read-only worksheet from the ``data_only`` view (cached values)."""
        return self.values[sheet]

    def structure(self, sheet: str) -> SheetStructure:
        """Merged ranges, hidden columns and visibility for one sheet.

        Cached per sheet; the first call streams the sheet part far enough to find the
        ``<mergeCells>`` block, subject to the handle's structure budget.

        Args:
            sheet: The sheet name.

        Returns:
            A :class:`SheetStructure`. Unknown sheet names yield an empty structure rather
            than raising.
        """
        cached = self._structures.get(sheet)
        if cached is not None:
            return cached

        names = self.sheet_names
        index = names.index(sheet) if sheet in names else -1
        part = self._sheet_parts.get(sheet)
        state = "visible"
        if index >= 0:
            state = str(getattr(self.formulas[sheet], "sheet_state", "visible"))

        hidden: tuple[str, ...] = ()
        merged: tuple[str, ...] = ()
        truncated = False
        if part is not None:
            try:
                head, tail, truncated = _scan_part(self.zf, part, self._structure_budget)
                hidden = _hidden_columns(head)
                merged = _merged_ranges(tail)
            except (OSError, zipfile.BadZipFile) as exc:
                logger.warning("could not scan structure of sheet %r: %s", sheet, exc)
                truncated = True

        structure = SheetStructure(
            name=sheet,
            index=index,
            state=state,
            part=part,
            hidden_columns=hidden,
            merged_ranges=merged,
            structure_truncated=truncated,
        )
        self._structures[sheet] = structure
        return structure

    def table_ranges(self, sheet: str) -> tuple[str, ...]:
        """The ranges of any Excel tables (ListObjects) defined on a sheet.

        Args:
            sheet: The sheet name.

        Returns:
            A1 ranges in part order, or an empty tuple when the sheet has no tables.
        """
        part = self._sheet_parts.get(sheet)
        if part is None:
            return ()
        directory, _, filename = part.rpartition("/")
        rels = self.read_part(f"{directory}/_rels/{filename}.rels")
        if not rels:
            return ()
        try:
            root = ElementTree.fromstring(rels)
        except ElementTree.ParseError:
            return ()

        ranges: list[str] = []
        for relationship in root.iter(f"{_PKG_REL_NS}Relationship"):
            target = relationship.get("Target", "")
            if "tables/" not in target:
                continue
            resolved = (
                target.lstrip("/")
                if target.startswith("/")
                else _resolve_relative(directory, target)
            )
            table = self.read_part(resolved)
            if not table:
                continue
            try:
                reference = ElementTree.fromstring(table).get("ref")
            except ElementTree.ParseError:
                continue
            if reference:
                ranges.append(reference)
        return tuple(ranges)

    # ── defined names ────────────────────────────────────────────────────

    def defined_names(self) -> list[NamedRange]:
        """Every defined name, workbook- and sheet-scoped.

        Parsed from ``xl/workbook.xml`` rather than from openpyxl, because read-only mode
        drops sheet-scoped names.

        Returns:
            The defined names in file order. A name whose target mentions ``#REF!`` or an
            unknown sheet is marked ``is_broken``. Cached after the first call.
        """
        if self._defined_names is not None:
            return list(self._defined_names)

        data = self.read_part("xl/workbook.xml")
        if not data:
            self._defined_names = []
            return []
        try:
            root = ElementTree.fromstring(data)
        except ElementTree.ParseError as exc:
            logger.warning("could not parse xl/workbook.xml: %s", exc)
            self._defined_names = []
            return []

        sheets = self.sheet_names
        known = set(sheets)
        names: list[NamedRange] = []
        for element in root.iter(f"{_MAIN_NS}definedName"):
            name = element.get("name")
            if not name:
                continue
            refers_to = (element.text or "").strip()
            local = element.get("localSheetId")
            scope: str | None = None
            if local is not None and local.isdigit() and int(local) < len(sheets):
                scope = sheets[int(local)]
            names.append(
                NamedRange(
                    name=name,
                    refers_to=refers_to,
                    scope=scope,
                    is_broken=_is_broken_name(refers_to, known),
                    is_hidden=element.get("hidden") in ("1", "true"),
                )
            )
        self._defined_names = names
        return list(names)

    @property
    def defined_name_keys(self) -> frozenset[str]:
        """Upper-cased defined names, for disambiguating name-versus-reference tokens."""
        return frozenset(name.name.upper() for name in self.defined_names())

    # ── bounded reads ────────────────────────────────────────────────────

    def read_range(
        self,
        sheet: str,
        a1: str,
        *,
        view: Literal["values", "formulas"] = "values",
        max_rows: int = _MAX_READ_ROWS,
        max_cols: int = _MAX_READ_COLS,
    ) -> list[list[Any]]:
        """Read a bounded rectangle of cell values.

        The primitive behind the agent's ``read_range`` tool and behind every extractor that
        needs to look at actual data. Always bounded: an unbounded range such as ``"A:D"``
        is clamped rather than refused, and the caller is expected to say so when it
        renders the result (PLAN §2.3).

        Args:
            sheet: Sheet name.
            a1: An A1 range (``"A1:D50"``), a single cell (``"B2"``), or a whole-column or
                whole-row range (``"A:D"``, ``"2:10"``).
            view: ``"values"`` for the cached-value view, ``"formulas"`` for formula text.
            max_rows: Hard row cap.
            max_cols: Hard column cap.

        Returns:
            A list of rows, each a list of values. Empty when the sheet or range is unknown.
        """
        if sheet not in self.formulas.sheetnames:
            logger.debug("read_range: unknown sheet %r", sheet)
            return []
        bounds = parse_a1_range(a1)
        if bounds is None:
            logger.debug("read_range: unparseable range %r", a1)
            return []

        min_row, min_col, max_row, max_col = bounds
        worksheet = self.value_sheet(sheet) if view == "values" else self.formula_sheet(sheet)
        sheet_rows = getattr(worksheet, "max_row", None) or max_row
        sheet_cols = getattr(worksheet, "max_column", None) or max_col
        max_row = min(max_row, sheet_rows, min_row + max_rows - 1)
        max_col = min(max_col, sheet_cols, min_col + max_cols - 1)
        if max_row < min_row or max_col < min_col:
            return []

        rows: list[list[Any]] = []
        for row in worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ):
            rows.append(list(row))
        return rows

    # ── lifecycle ────────────────────────────────────────────────────────

    def close(self) -> None:
        """Release the archive and both openpyxl views. Idempotent."""
        if self._closed:
            return
        self._closed = True
        for closeable in (self.formulas, self.values, self.zf):
            try:
                closeable.close()
            except Exception:
                logger.debug("error closing %s", type(closeable).__name__, exc_info=True)

    def __enter__(self) -> WorkbookHandle:
        return self

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc: BaseException | None,
        traceback: TracebackType | None,
    ) -> None:
        self.close()


def _resolve_relative(directory: str, target: str) -> str:
    """Resolve a relationship target against the part directory that declared it."""
    parts = directory.split("/")
    for segment in target.split("/"):
        if segment == "..":
            if parts:
                parts.pop()
        elif segment not in ("", "."):
            parts.append(segment)
    return "/".join(parts)


def _is_broken_name(refers_to: str, known_sheets: set[str]) -> bool:
    """Whether a defined name's target is unusable."""
    if not refers_to or "#REF" in refers_to.upper():
        return True
    if "!" not in refers_to:
        return False
    sheet = refers_to.split("!", 1)[0].lstrip("=").strip()
    if sheet.startswith("'") and sheet.endswith("'"):
        sheet = sheet[1:-1].replace("''", "'")
    if "[" in sheet:  # external workbook reference; resolvability is checked elsewhere
        return False
    return bool(sheet) and bool(known_sheets) and sheet not in known_sheets


def _sheet_parts(zf: zipfile.ZipFile) -> dict[str, str]:
    """Map sheet name to its part path, via ``xl/workbook.xml`` and its relationships."""
    try:
        book = zf.read("xl/workbook.xml")
        rels = zf.read("xl/_rels/workbook.xml.rels")
    except (KeyError, OSError, zipfile.BadZipFile):
        return {}

    try:
        book_root = ElementTree.fromstring(book)
        rels_root = ElementTree.fromstring(rels)
    except ElementTree.ParseError:
        return {}

    targets: dict[str, str] = {}
    for relationship in rels_root.iter(f"{_PKG_REL_NS}Relationship"):
        rel_id = relationship.get("Id")
        target = relationship.get("Target")
        if not rel_id or not target:
            continue
        targets[rel_id] = target.lstrip("/") if target.startswith("/") else f"xl/{target}"

    parts: dict[str, str] = {}
    for sheet in book_root.iter(f"{_MAIN_NS}sheet"):
        name = sheet.get("name")
        rel_id = sheet.get(f"{_REL_NS}id")
        if name and rel_id and rel_id in targets:
            parts[name] = targets[rel_id]
    return parts


def parse_a1_range(a1: str) -> tuple[int, int, int, int] | None:
    """Parse an A1 range into ``(min_row, min_col, max_row, max_col)``.

    Accepts single cells, rectangles, whole-column ranges (``"A:D"``) and whole-row ranges
    (``"2:10"``). Sheet qualifiers are rejected: the caller names the sheet separately.

    Args:
        a1: The range text.

    Returns:
        The bounds, or None when the text is not a range. Whole-column and whole-row ranges
        come back with the missing dimension spanning the sheet limits.
    """
    text = a1.strip().replace("$", "")
    if not text or "!" in text:
        return None

    left, _, right = text.partition(":")
    first = _parse_cell(left)
    if first is None:
        return None
    if not right:
        row, col = first
        if row is None or col is None:
            return None
        return row, col, row, col

    second = _parse_cell(right)
    if second is None:
        return None
    rows = [value for value in (first[0], second[0]) if value is not None]
    cols = [value for value in (first[1], second[1]) if value is not None]
    min_row, max_row = (min(rows), max(rows)) if len(rows) == 2 else (1, 1_048_576)
    min_col, max_col = (min(cols), max(cols)) if len(cols) == 2 else (1, 16_384)
    return min_row, min_col, max_row, max_col


def _parse_cell(text: str) -> tuple[int | None, int | None] | None:
    """Parse one side of a range into (row, col), either of which may be absent."""
    match = _RANGE_RE.match(text)
    if match is None:
        return None
    letters, digits = match.group(1), match.group(2)
    if not letters and not digits:
        return None
    col = None
    if letters:
        col = 0
        for char in letters.upper():
            col = col * 26 + (ord(char) - 64)
        if not 1 <= col <= 16_384:
            return None
    row = int(digits) if digits else None
    if row is not None and not 1 <= row <= 1_048_576:
        return None
    return row, col


def open_workbook(
    path: Path | str,
    *,
    structure_budget: int = _DEFAULT_STRUCTURE_BUDGET,
) -> WorkbookHandle:
    """Open a workbook for analysis: the archive plus both openpyxl views.

    Args:
        path: Path to an ``.xlsx`` or ``.xlsm`` file.
        structure_budget: How many decompressed bytes of a sheet part to stream when
            hunting for merged ranges. Sheets larger than this report
            ``structure_truncated``.

    Returns:
        An open :class:`WorkbookHandle`. The caller owns it and must close it.

    Raises:
        WorkbookOpenError: The file is missing, is not a supported format, or is not a
            readable workbook. Callers that must not fail should check
            :func:`detect_format` first and record a finding instead.
    """
    resolved = Path(path)
    if not resolved.is_file():
        raise WorkbookOpenError(
            f"cannot analyse {resolved}: the file does not exist. Check the path and try again."
        )

    file_format = detect_format(resolved)
    if file_format not in SUPPORTED_FORMATS:
        raise WorkbookOpenError(
            f"cannot analyse {resolved.name}: format {file_format!r} is not supported. "
            "Re-save the workbook as .xlsx or .xlsm and analyse that."
        )

    try:
        zf = zipfile.ZipFile(resolved)
    except (OSError, zipfile.BadZipFile) as exc:
        raise WorkbookOpenError(
            f"cannot analyse {resolved.name}: it is not a readable Office Open XML archive ({exc})."
        ) from exc

    formulas: Workbook | None = None
    values: Workbook | None = None
    try:
        formulas = load_workbook(resolved, read_only=True, data_only=False, keep_links=True)
        values = load_workbook(resolved, read_only=True, data_only=True, keep_links=True)
        identity = read_identity(resolved, zf)
    except Exception as exc:
        for opened in (formulas, values):
            if opened is not None:
                opened.close()
        zf.close()
        raise WorkbookOpenError(
            f"cannot analyse {resolved.name}: openpyxl could not load it ({exc})."
        ) from exc

    logger.info("opened workbook %s (%d sheets)", resolved.name, len(formulas.sheetnames))
    return WorkbookHandle(
        resolved, zf, formulas, values, identity, structure_budget=structure_budget
    )
