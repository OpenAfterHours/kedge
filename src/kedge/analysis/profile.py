"""Column profiles, bounded sampling and format anomalies (PLAN 2.3).

This is what the model sees about a sheet by default, so it has to be genuinely informative
and genuinely bounded at the same time. Per column: header, inferred dtype, null count,
distinct count, numeric summary where the column is numeric, top values where it is
categorical, the first five rows, the last five rows, and a seeded random sample of five so
two runs over the same workbook produce the same artifact.

**Cost is bounded by construction.** Rows are streamed once through fixed-size accumulators —
a reservoir for the sample, a bounded deque for the tail, a capped counter for cardinality —
so a 200,000-row sheet costs one pass and a few kilobytes, never a materialised list. Where a
cap does bite, the profile says so in ``format_anomalies`` rather than quietly describing a
prefix as though it were the whole column.

**Format anomalies are the point, not decoration.** Numbers stored as text, mixed date
formats, stray whitespace and inconsistent casing in a key column are precisely the silent
join failures from PLAN 2.6 — the ones that produce a wrong number rather than an error. Every
anomaly string is drawn from :data:`FORMAT_ANOMALY_PREFIXES` and carries counts only, never
cell values, which is what makes it safe to keep them on a redacted profile.

References:
- PLAN.md 2.3 (profiles and sampling), 2.6 (Excel/polars semantic gaps), M1.
"""

from __future__ import annotations

import datetime as dt
import logging
import random
import re
from collections import Counter, deque
from collections.abc import Sequence
from dataclasses import dataclass, field
from decimal import Decimal
from typing import TYPE_CHECKING, Any

from openpyxl.utils import get_column_letter

from kedge.analysis.model import ColumnProfile, NumericStats, SheetInfo
from kedge.analysis.redact import redact_profile, should_redact

if TYPE_CHECKING:
    from kedge.analysis.workbook import WorkbookHandle

logger = logging.getLogger(__name__)

__all__ = [
    "FORMAT_ANOMALY_PREFIXES",
    "MAX_COLUMNS",
    "MAX_SCAN_ROWS",
    "classify_value",
    "infer_dtype",
    "profile_sheet",
    "resolve_worksheet",
    "workbook_view",
]

MAX_SCAN_ROWS = 200_000
"""Hard ceiling on rows read per sheet. Hitting it is reported, never silently absorbed."""

MAX_COLUMNS = 256
"""Hard ceiling on columns profiled per sheet. Excel allows 16,384; nobody analyses 16,384."""

MAX_DISTINCT = 10_000
"""Cardinality is counted exactly up to here; beyond it ``distinct_count`` becomes ``None``."""

BLANK_RUN_STOP = 500
"""Consecutive blank rows that end the scan. Excel inflates ``max_row`` with formatting."""

HEAD_ROWS = 5
TAIL_ROWS = 5
SAMPLE_ROWS = 5
TOP_K = 10
SAMPLE_SEED = 1729
"""Fixed seed: the random sample must be reproducible, or two runs diff for no reason."""

_TOP_K_MAX_CARDINALITY = 50
"""Above this many distinct values a continuous column's top-k is noise, so it is omitted."""

FORMAT_ANOMALY_PREFIXES = (
    "mixed types:",
    "numbers stored as text",
    "mixed date formats",
    "text dates alongside real dates",
    "leading or trailing whitespace",
    "non-breaking spaces",
    "empty strings",
    "inconsistent casing",
    "Excel error values",
    "profile truncated",
)
"""Every anomaly a profile can carry. Counts only, never cell values -- which is what lets a
redacted profile keep its anomalies (see :func:`kedge.analysis.redact.redact_profile`)."""

_ERROR_VALUES = frozenset(
    {
        "#N/A",
        "#VALUE!",
        "#REF!",
        "#DIV/0!",
        "#NAME?",
        "#NULL!",
        "#NUM!",
        "#GETTING_DATA",
        "#SPILL!",
        "#CALC!",
        "#FIELD!",
        "#BLOCKED!",
        "#CONNECT!",
        "#BUSY!",
        "#UNKNOWN!",
    }
)

_NUMERIC_TEXT_RE = re.compile(r"^\s*[-+(]?\s*[£$€]?\s*\d{1,3}(?:,\d{3})*(?:\.\d+)?\s*\)?\s*%?\s*$")
_PLAIN_NUMERIC_TEXT_RE = re.compile(
    r"^\s*[-+(]?\s*[£$€]?\s*\d*\.?\d+(?:[eE][-+]?\d+)?\s*\)?\s*%?\s*$"
)

_DATE_TEXT_PATTERNS: tuple[tuple[str, re.Pattern[str]], ...] = (
    ("iso", re.compile(r"^\s*\d{4}-\d{1,2}-\d{1,2}([ T].*)?$")),
    ("slash", re.compile(r"^\s*\d{1,2}/\d{1,2}/\d{2,4}\s*$")),
    ("dotted", re.compile(r"^\s*\d{1,2}\.\d{1,2}\.\d{2,4}\s*$")),
    ("named-month", re.compile(r"(?i)^\s*\d{1,2}[- ][a-z]{3,9}[- ]\d{2,4}\s*$")),
    ("month-first", re.compile(r"(?i)^\s*[a-z]{3,9}\s+\d{1,2},?\s+\d{4}\s*$")),
)

_KEY_HEADER_RE = re.compile(
    r"(?i)\b(id|ids|code|codes|key|keys|ref|reference|no|num|number|account|counterparty|"
    r"customer|client|isin|cusip|sedol|ticker|name|desk|entity|book)\b"
)
_KEY_DISTINCT_RATIO = 0.8


# ── WorkbookHandle access ────────────────────────────────────────────────────────────────────
#
# The two functions below are the only place in this package that reaches into the handle's
# openpyxl views, so `docs.py` borrows them rather than growing its own. Both views are opened
# `read_only=True`, which matters: worksheets stream, cells are `ReadOnlyCell`, and a gap in a
# row is an `EmptyCell` carrying no row or column of its own. Never ask a cell where it is --
# count position from the iteration instead.


def workbook_view(handle: WorkbookHandle, *, data_only: bool) -> Any | None:
    """Return one of the handle's two openpyxl workbooks, or ``None``.

    ``data_only=True`` asks for the cached-value view; ``False`` asks for the formula view.
    openpyxl cannot serve both from one load (PLAN 1.5), which is why the handle holds two.

    Args:
        handle: The open workbook handle.
        data_only: Whether the cached-value view is wanted.

    Returns:
        An ``openpyxl.Workbook``, or ``None`` when the handle does not expose that view.
    """
    book = getattr(handle, "values" if data_only else "formulas", None)
    return book if book is not None and hasattr(book, "sheetnames") else None


def resolve_worksheet(handle: WorkbookHandle, sheet_name: str, *, data_only: bool) -> Any | None:
    """Return one worksheet from the requested view, or ``None`` if it cannot be reached."""
    accessor = getattr(handle, "value_sheet" if data_only else "formula_sheet", None)
    if callable(accessor):
        try:
            return accessor(sheet_name)
        except (KeyError, TypeError, ValueError):
            logger.debug("sheet %r is not in the requested view", sheet_name)
            return None

    book = workbook_view(handle, data_only=data_only)
    if book is None:
        return None
    try:
        return book[sheet_name]
    except (KeyError, TypeError):
        logger.debug("sheet %r is not in the requested view", sheet_name)
        return None


# ── dtype inference ──────────────────────────────────────────────────────────────────────────


def classify_value(value: Any) -> str:
    """Classify one cell value into a base type name.

    ``bool`` is checked before ``int`` deliberately — Python makes the former a subclass of
    the latter, and a TRUE/FALSE flag column profiled as integers would be wrong in a way
    nobody would notice until a join failed.

    Args:
        value: A cell value as openpyxl returns it.

    Returns:
        One of ``empty``, ``boolean``, ``integer``, ``float``, ``datetime``, ``date``,
        ``time``, ``duration``, ``error``, ``string`` or ``other``.
    """
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, (float, Decimal)):
        return "float"
    if isinstance(value, dt.datetime):
        return "datetime"
    if isinstance(value, dt.date):
        return "date"
    if isinstance(value, dt.time):
        return "time"
    if isinstance(value, dt.timedelta):
        return "duration"
    if isinstance(value, str):
        return "error" if value.strip().upper() in _ERROR_VALUES else "string"
    return "other"


def _dtype_from_counts(counts: Counter[str]) -> str:
    """Collapse a tally of base types into the column's dtype.

    Integers and floats are one family, as are the temporal types: an Excel column holding
    ``1`` and ``1.5`` is a number column, not a mixed one. Error values are set aside rather
    than forcing ``mixed``, because ``#N/A`` in a numeric column is a data-quality problem
    reported through ``format_anomalies``, not a change of type. Anything genuinely
    heterogeneous is ``mixed``, and the composition goes into ``format_anomalies`` so that a
    column that is 97% numeric with a few text cells says exactly that.
    """
    kinds = {kind for kind, count in counts.items() if kind != "empty" and count}
    if not kinds:
        return "empty"
    if kinds == {"error"}:
        return "error"

    kinds.discard("error")
    if kinds <= {"integer", "float"}:
        return "float" if counts["float"] else "integer"
    if kinds <= {"date", "datetime", "time"}:
        if kinds == {"time"}:
            return "time"
        return "datetime" if counts["datetime"] else "date"
    if len(kinds) == 1:
        return next(iter(kinds))
    return "mixed"


def infer_dtype(values: Sequence[Any]) -> str:
    """Infer a column dtype from a sequence of cell values.

    Excel columns are heterogeneous and the honest answer is sometimes ``"mixed"``. This is
    the function that decides, and it is unit-tested against that expectation directly rather
    than only through :func:`profile_sheet`.

    Args:
        values: The column's values, nulls included.

    Returns:
        The dtype name. ``"empty"`` when there are no non-null values.
    """
    return _dtype_from_counts(Counter(classify_value(value) for value in values))


# ── per-column accumulation ──────────────────────────────────────────────────────────────────


@dataclass(slots=True)
class _ColumnAggregate:
    """Fixed-size accumulator for one column. Nothing here grows with the row count."""

    kinds: Counter[str] = field(default_factory=Counter)
    distinct: Counter[Any] = field(default_factory=Counter)
    distinct_overflowed: bool = False
    numeric_count: int = 0
    numeric_min: float | None = None
    numeric_max: float | None = None
    numeric_sum: float = 0.0
    zero_count: int = 0
    negative_count: int = 0
    text_numeric_count: int = 0
    whitespace_count: int = 0
    nbsp_count: int = 0
    empty_string_count: int = 0
    date_text_formats: set[str] = field(default_factory=set)

    @property
    def non_null(self) -> int:
        return sum(count for kind, count in self.kinds.items() if kind != "empty")

    def add(self, value: Any) -> None:
        kind = classify_value(value)
        self.kinds[kind] += 1
        if kind == "empty":
            return

        try:
            if len(self.distinct) < MAX_DISTINCT or value in self.distinct:
                self.distinct[value] += 1
            else:
                self.distinct_overflowed = True
        except TypeError:  # an unhashable cell value is possible, if rare
            self.distinct_overflowed = True

        if kind in ("integer", "float"):
            self._add_numeric(float(value))
        elif kind == "string":
            self._add_text(value)

    def _add_numeric(self, number: float) -> None:
        self.numeric_count += 1
        self.numeric_sum += number
        self.numeric_min = number if self.numeric_min is None else min(self.numeric_min, number)
        self.numeric_max = number if self.numeric_max is None else max(self.numeric_max, number)
        if number == 0:
            self.zero_count += 1
        elif number < 0:
            self.negative_count += 1

    def _add_text(self, text: str) -> None:
        if not text.strip():
            self.empty_string_count += 1
            return
        if text != text.strip():
            self.whitespace_count += 1
        if "\xa0" in text:  # non-breaking space: invisible, and it breaks joins
            self.nbsp_count += 1
        if _NUMERIC_TEXT_RE.match(text) or _PLAIN_NUMERIC_TEXT_RE.match(text):
            self.text_numeric_count += 1
            return
        for label, pattern in _DATE_TEXT_PATTERNS:
            if pattern.match(text):
                self.date_text_formats.add(label)
                break


def _is_key_like(header: str | None, distinct_count: int | None, non_null: int) -> bool:
    """Whether a column looks like a join key, and so whether casing drift matters."""
    if header and _KEY_HEADER_RE.search(header):
        return True
    if distinct_count is None or non_null == 0:
        return False
    return distinct_count / non_null >= _KEY_DISTINCT_RATIO


def _composition(aggregate: _ColumnAggregate) -> str:
    """Render the type mix as percentages, most common first. Counts only, no values."""
    non_null = aggregate.non_null or 1
    parts = [
        f"{count * 100 // non_null}% {kind} ({count:,})"
        for kind, count in aggregate.kinds.most_common()
        if kind != "empty" and count
    ]
    return "mixed types: " + ", ".join(parts)


def _anomalies(
    aggregate: _ColumnAggregate,
    *,
    dtype: str,
    header: str | None,
    distinct_count: int | None,
    truncation: str | None,
) -> list[str]:
    """Build the format-anomaly list. Every string starts with a FORMAT_ANOMALY_PREFIXES entry."""
    anomalies: list[str] = []
    non_null = aggregate.non_null

    if dtype == "mixed":
        anomalies.append(_composition(aggregate))

    if aggregate.text_numeric_count:
        share = (
            ""
            if not non_null
            else f", {aggregate.text_numeric_count * 100 // non_null}% of the column"
        )
        anomalies.append(
            f"numbers stored as text ({aggregate.text_numeric_count:,} cells{share}) -- "
            f"a silent join and arithmetic hazard"
        )

    if len(aggregate.date_text_formats) > 1:
        anomalies.append(
            f"mixed date formats ({len(aggregate.date_text_formats)} text date layouts detected)"
        )
    elif aggregate.date_text_formats and (aggregate.kinds["datetime"] or aggregate.kinds["date"]):
        real_dates = aggregate.kinds["datetime"] + aggregate.kinds["date"]
        anomalies.append(
            f"text dates alongside real dates ({len(aggregate.date_text_formats)} text layout(s) "
            f"beside {real_dates:,} real date cells)"
        )

    if aggregate.whitespace_count:
        anomalies.append(f"leading or trailing whitespace ({aggregate.whitespace_count:,} cells)")
    if aggregate.nbsp_count:
        anomalies.append(f"non-breaking spaces ({aggregate.nbsp_count:,} cells)")
    if aggregate.empty_string_count:
        anomalies.append(
            f"empty strings ({aggregate.empty_string_count:,} cells) -- distinct from blank cells"
        )

    if not aggregate.distinct_overflowed and _is_key_like(header, distinct_count, non_null):
        text_values = [value for value in aggregate.distinct if isinstance(value, str)]
        folded = {value.strip().casefold() for value in text_values}
        collisions = len(text_values) - len(folded)
        if collisions > 0:
            anomalies.append(
                f"inconsistent casing ({collisions:,} value(s) differ only by case or padding "
                f"in what looks like a key column)"
            )

    if aggregate.kinds["error"]:
        anomalies.append(f"Excel error values ({aggregate.kinds['error']:,} cells)")

    if truncation:
        anomalies.append(truncation)
    return anomalies


# ── the sheet scan ───────────────────────────────────────────────────────────────────────────


class _SheetScan:
    """One streaming pass over a sheet's data rows.

    Holds the per-column accumulators plus three fixed-size row buffers: the first rows, a
    bounded deque of the last rows, and a seeded reservoir. Sampling at row level rather than
    per column is deliberate — the model gets five coherent rows it could have read off the
    sheet, not five unrelated values per column.
    """

    def __init__(self, columns: int) -> None:
        self.columns = columns
        self.aggregates = [_ColumnAggregate() for _ in range(columns)]
        self.head: list[tuple[Any, ...]] = []
        self.tail: deque[tuple[Any, ...]] = deque(maxlen=TAIL_ROWS)
        self.reservoir: list[tuple[Any, ...]] = []
        self.rows_seen = 0
        self._rng = random.Random(SAMPLE_SEED)

    def consume(self, values: tuple[Any, ...]) -> None:
        self.rows_seen += 1
        if len(self.head) < HEAD_ROWS:
            self.head.append(values)
        self.tail.append(values)

        if len(self.reservoir) < SAMPLE_ROWS:
            self.reservoir.append(values)
        else:
            index = self._rng.randrange(self.rows_seen)
            if index < SAMPLE_ROWS:
                self.reservoir[index] = values

        for aggregate, value in zip(self.aggregates, values, strict=False):
            aggregate.add(value)


def _column_values(rows: Sequence[tuple[Any, ...]], index: int) -> list[Any]:
    return [row[index] for row in rows if index < len(row)]


def _pad(row: tuple[Any, ...], width: int) -> tuple[Any, ...]:
    if len(row) == width:
        return row
    if len(row) > width:
        return row[:width]
    return row + (None,) * (width - len(row))


def _clean_header(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _numeric_stats(aggregate: _ColumnAggregate) -> NumericStats | None:
    """Numeric summary, but only where the column is numeric enough for it to mean anything."""
    non_null = aggregate.non_null
    if not aggregate.numeric_count or non_null == 0:
        return None
    if aggregate.numeric_count * 2 < non_null:
        return None
    return NumericStats(
        min=aggregate.numeric_min,
        max=aggregate.numeric_max,
        mean=aggregate.numeric_sum / aggregate.numeric_count,
        sum=aggregate.numeric_sum,
        zero_count=aggregate.zero_count,
        negative_count=aggregate.negative_count,
    )


def _top_k(aggregate: _ColumnAggregate, dtype: str) -> list[tuple[Any, int]]:
    """Top values with frequencies, for the columns where a frequency table informs anything.

    Categorical columns get one always. A continuous column gets one only while it is low
    cardinality, where it is usually a numeric code rather than a measure; the ten most common
    values of a float measure tell nobody anything.
    """
    if not aggregate.distinct:
        return []
    is_continuous = dtype in ("integer", "float", "date", "datetime", "time", "duration")
    if is_continuous and len(aggregate.distinct) > _TOP_K_MAX_CARDINALITY:
        return []
    return list(aggregate.distinct.most_common(TOP_K))


def _build_profile(
    scan: _SheetScan,
    *,
    sheet_name: str,
    index: int,
    header: str | None,
    truncation: str | None,
) -> ColumnProfile:
    aggregate = scan.aggregates[index]
    dtype = _dtype_from_counts(aggregate.kinds)
    distinct_count = None if aggregate.distinct_overflowed else len(aggregate.distinct)
    return ColumnProfile(
        sheet=sheet_name,
        column=get_column_letter(index + 1),
        index=index + 1,
        header=header,
        dtype=dtype,
        row_count=scan.rows_seen,
        null_count=aggregate.kinds["empty"],
        distinct_count=distinct_count,
        numeric=_numeric_stats(aggregate),
        top_k=_top_k(aggregate, dtype),
        head=_column_values(scan.head, index),
        tail=_column_values(list(scan.tail), index),
        sample=_column_values(scan.reservoir, index),
        format_anomalies=_anomalies(
            aggregate,
            dtype=dtype,
            header=header,
            distinct_count=distinct_count,
            truncation=truncation,
        ),
    )


def _scan_sheet(
    worksheet: Any, sheet: SheetInfo, columns: int
) -> tuple[_SheetScan, list[str | None], str | None]:
    """Stream the sheet once, returning the accumulators, the headers, and any truncation note."""
    max_row = int(sheet.max_row or getattr(worksheet, "max_row", 0) or 0)
    header_row = sheet.header_row
    first_data_row = header_row + 1 if header_row else max(sheet.preamble_rows + 1, 1)
    last_row = min(max_row, first_data_row + MAX_SCAN_ROWS - 1)

    truncation: str | None = None
    if max_row > last_row:
        truncation = (
            f"profile truncated: first {last_row - first_data_row + 1:,} of {max_row - first_data_row + 1:,} "
            f"data rows were read"
        )
        logger.warning(
            "sheet %r has %d rows; profiling the first %d only", sheet.name, max_row, MAX_SCAN_ROWS
        )

    scan = _SheetScan(columns)
    headers: list[str | None] = [None] * columns
    start_row = header_row if header_row and header_row < first_data_row else first_data_row
    blank_run = 0

    rows = worksheet.iter_rows(
        min_row=start_row, max_row=last_row, min_col=1, max_col=columns, values_only=True
    )
    for row_number, row in enumerate(rows, start=start_row):
        values = _pad(tuple(row), columns)
        if header_row is not None and row_number == header_row:
            headers = [_clean_header(value) for value in values]
            continue
        if row_number < first_data_row:
            continue

        if all(value is None for value in values):
            blank_run += 1
            if blank_run >= BLANK_RUN_STOP:
                logger.debug(
                    "stopping the scan of %r at row %d after %d blank rows",
                    sheet.name,
                    row_number,
                    blank_run,
                )
                break
            continue

        # A gap inside the data is real (a blank cell is a null), so it is replayed once the
        # next populated row proves the data had not ended. A run that reaches the end of the
        # sheet never gets replayed, which is what keeps Excel's inflated `max_row` out of the
        # row count. Leading blanks are dropped for the same reason.
        if scan.rows_seen:
            for _ in range(blank_run):
                scan.consume((None,) * columns)
        blank_run = 0
        scan.consume(values)

    return scan, headers, truncation


def profile_sheet(
    handle: WorkbookHandle,
    sheet: SheetInfo,
    *,
    redact_patterns: list[str] | None = None,
) -> list[ColumnProfile]:
    """Profile every column of one sheet in a single bounded pass.

    Reads the cached-value view, so a column of formulas profiles as the numbers Excel last
    calculated rather than as a column of formula strings. Where the workbook carries no
    cached values the column reads as empty, which is the condition ``values.py`` reports as
    :attr:`~kedge.analysis.model.FindingKind.MISSING_CACHED_VALUES`.

    Never raises: a sheet that cannot be read produces an empty list and a warning, because
    one unreadable sheet must not cost the whole analysis (CONVENTIONS non-negotiable 4).

    Args:
        handle: The open workbook handle.
        sheet: The sheet to profile. ``header_row`` and ``preamble_rows`` decide where data
            starts; ``max_row`` and ``max_column`` bound the scan.
        redact_patterns: Case-insensitive regular expressions matched against column headers.
            A matching column keeps its dtype and null count and loses every value
            (PLAN 2.3). ``None`` -- the default -- redacts nothing.

    Returns:
        One :class:`~kedge.analysis.model.ColumnProfile` per non-empty column, in column
        order. Empty when the sheet has no data rows or cannot be read.
    """
    try:
        return _profile_sheet(handle, sheet, redact_patterns)
    except Exception:
        logger.warning("could not profile sheet %r; skipping it", sheet.name, exc_info=True)
        return []


def _profile_sheet(
    handle: WorkbookHandle,
    sheet: SheetInfo,
    redact_patterns: Sequence[str] | None,
) -> list[ColumnProfile]:
    worksheet = resolve_worksheet(handle, sheet.name, data_only=True)
    if worksheet is None:
        worksheet = resolve_worksheet(handle, sheet.name, data_only=False)
        if worksheet is None:
            logger.warning("sheet %r could not be reached through the workbook handle", sheet.name)
            return []
        logger.warning(
            "no cached-value view for sheet %r; profiling the formula view instead", sheet.name
        )

    max_column = int(sheet.max_column or getattr(worksheet, "max_column", 0) or 0)
    if max_column <= 0:
        logger.debug("sheet %r has no columns to profile", sheet.name)
        return []
    columns = min(max_column, MAX_COLUMNS)
    if max_column > columns:
        logger.warning(
            "sheet %r has %d columns; profiling the first %d only", sheet.name, max_column, columns
        )

    scan, headers, truncation = _scan_sheet(worksheet, sheet, columns)
    if scan.rows_seen == 0:
        logger.debug("sheet %r has no data rows below its header", sheet.name)
        return []

    profiles: list[ColumnProfile] = []
    for index in range(columns):
        header = headers[index]
        if header is None and scan.aggregates[index].non_null == 0:
            continue
        profile = _build_profile(
            scan, sheet_name=sheet.name, index=index, header=header, truncation=truncation
        )
        if should_redact(header, redact_patterns):
            profile = redact_profile(profile)
        profiles.append(profile)

    logger.info(
        "profiled %d column(s) of %r over %d row(s)", len(profiles), sheet.name, scan.rows_seen
    )
    return profiles
