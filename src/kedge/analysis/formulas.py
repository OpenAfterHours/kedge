"""Formula tokenisation, reference extraction, and A1 <-> R1C1 normalisation.

R1C1 normalisation is the load-bearing idea of the whole analyser (PLAN §2.1). A formula
filled down a column produces a different A1 string in every row, but exactly one R1C1
string: ``=G2*$B$1`` at ``H2`` and ``=G3*$B$1`` at ``H3`` both normalise to
``=RC[-1]*R1C2``. That is what collapses half a million formula cells into thirty or sixty
logical operations.

openpyxl gives us :class:`~openpyxl.formula.tokenizer.Tokenizer` and nothing else — the
A1 <-> R1C1 conversion is written here on top of it. Working from the token stream rather
than from a regular expression over the raw text is what makes the awkward cases safe: a
string literal that happens to look like a reference (``=IF(A1="A1","A1",B1)``) arrives as
an ``OPERAND/TEXT`` token and is never rewritten.

References:
- PLAN.md §1.5 (Tokenizer is sufficient), §2.1 (R1C1 compression).
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from typing import Any

from openpyxl.formula.tokenizer import Token, Tokenizer

from kedge.analysis.model import Reference

logger = logging.getLogger(__name__)

__all__ = [
    "MAX_COL",
    "MAX_ROW",
    "VOLATILE_FUNCTIONS",
    "ParsedFormula",
    "RefExtent",
    "formula_text",
    "parse_formula",
    "to_a1",
    "to_r1c1",
    "tokenise",
]

MAX_ROW = 1_048_576
MAX_COL = 16_384

VOLATILE_FUNCTIONS: frozenset[str] = frozenset(
    {"NOW", "TODAY", "RAND", "RANDBETWEEN", "RANDARRAY", "OFFSET", "INDIRECT", "INFO", "CELL"}
)
"""Functions that force recalculation on every edit, and that make a translation
non-deterministic or position-dependent (PLAN §M1 findings list)."""

_CELL_RE = re.compile(r"^(\$?)([A-Za-z]{1,3})(\$?)([0-9]{1,7})$")
_COL_RE = re.compile(r"^(\$?)([A-Za-z]{1,3})$")
_ROW_RE = re.compile(r"^(\$?)([0-9]{1,7})$")
_EXTERNAL_RE = re.compile(r"^(?P<prefix>.*)\[(?P<workbook>[^\[\]]+)\](?P<sheet>.*)$")
_R1C1_PART_RE = re.compile(
    r"^(?:R(?P<row>\[-?[0-9]+\]|[0-9]+)?)?(?:C(?P<col>\[-?[0-9]+\]|[0-9]+)?)?$"
)
_STRUCTURED_RE = re.compile(r"^[A-Za-z_\\][A-Za-z0-9_.\\]*\[")


# =============================================================================
# PUBLIC TYPES
# =============================================================================


@dataclass(frozen=True, slots=True)
class RefExtent:
    """The rectangle a reference covers, resolved against the cell that holds it.

    Whole-column and whole-row references expand to the sheet limits, so an extent is
    always a concrete rectangle and overlap tests are a plain interval comparison.
    """

    sheet: str | None
    min_row: int
    min_col: int
    max_row: int
    max_col: int
    is_external: bool = False

    def overlaps(self, min_row: int, min_col: int, max_row: int, max_col: int) -> bool:
        """Whether this extent intersects the given rectangle (sheet is the caller's problem)."""
        return (
            self.min_row <= max_row
            and min_row <= self.max_row
            and self.min_col <= max_col
            and min_col <= self.max_col
        )

    @property
    def cell_count(self) -> int:
        """How many cells the rectangle covers."""
        return (self.max_row - self.min_row + 1) * (self.max_col - self.min_col + 1)


@dataclass(frozen=True, slots=True)
class ParsedFormula:
    """Everything the analyser needs to know about one formula.

    Parsing is expensive relative to the number of *distinct* formulas in a workbook, so
    callers are expected to normalise to R1C1 first (cheap, per cell) and parse once per
    distinct R1C1 string.
    """

    a1: str
    r1c1: str
    functions: tuple[str, ...]
    outermost_function: str | None
    references: tuple[Reference, ...]
    extents: tuple[RefExtent, ...]
    numeric_literals: tuple[float, ...]
    string_literals: tuple[str, ...]
    names: tuple[str, ...]
    volatile_functions: tuple[str, ...]
    has_external_refs: bool = False
    has_ref_error: bool = False

    @property
    def is_volatile(self) -> bool:
        """Whether the formula calls a volatile function."""
        return bool(self.volatile_functions)


# =============================================================================
# TOKENS
# =============================================================================


def formula_text(value: Any) -> str | None:
    """Normalise an openpyxl cell value into a formula string, or None if it is not one.

    openpyxl hands back a plain string for an ordinary formula but an ``ArrayFormula`` or
    ``DataTableFormula`` object for the two special cases, and those carry the text on a
    ``text`` attribute.

    Args:
        value: A cell value from either workbook view.

    Returns:
        The formula including its leading ``=``, or None when the cell is not a formula.
    """
    if isinstance(value, str):
        return value if value.startswith("=") else None
    text = getattr(value, "text", None)
    if isinstance(text, str) and text.startswith("="):
        return text
    return None


def tokenise(formula: str) -> list[Token]:
    """Tokenise a formula, returning an empty list if openpyxl cannot make sense of it.

    Args:
        formula: The formula text, with or without its leading ``=``.

    Returns:
        The token stream, or ``[]`` when tokenisation failed.
    """
    body = formula if formula.startswith("=") else "=" + formula
    try:
        return list(Tokenizer(body).items)
    except Exception:
        logger.debug("could not tokenise formula: %s", formula)
        return []


# =============================================================================
# REFERENCE PARSING
# =============================================================================


@dataclass(frozen=True, slots=True)
class _Body:
    """A reference body (the part after any ``Sheet!`` qualifier), decoded."""

    r1c1: str
    min_row: int
    min_col: int
    max_row: int
    max_col: int
    row_absolute: bool
    col_absolute: bool
    is_range: bool


def _split_sheet(token: str) -> tuple[str, str]:
    """Split a reference token into its sheet prefix (including ``!``) and its body."""
    if token.startswith("'"):
        i = 1
        n = len(token)
        while i < n:
            if token[i] == "'":
                if i + 1 < n and token[i + 1] == "'":
                    i += 2
                    continue
                break
            i += 1
        if i >= n or i + 1 >= n or token[i + 1] != "!":
            return "", token
        return token[: i + 2], token[i + 2 :]
    index = token.find("!")
    if index == -1:
        return "", token
    return token[: index + 1], token[index + 1 :]


def _decode_sheet_prefix(prefix: str) -> tuple[str | None, str | None]:
    """Turn a raw ``'[1]My Sheet'!`` prefix into (sheet name, external workbook)."""
    if not prefix:
        return None, None
    raw = prefix[:-1]
    if raw.startswith("'") and raw.endswith("'") and len(raw) >= 2:
        raw = raw[1:-1].replace("''", "'")
    match = _EXTERNAL_RE.match(raw)
    if match is None:
        return (raw or None), None
    workbook = match.group("workbook")
    path = match.group("prefix")
    sheet = match.group("sheet")
    return (sheet or None), (f"{path}[{workbook}]" if path else workbook)


def _column_index(letters: str) -> int | None:
    """Convert column letters to a 1-based index, or None when out of Excel's range."""
    index = 0
    for char in letters.upper():
        index = index * 26 + (ord(char) - 64)
    return index if 1 <= index <= MAX_COL else None


def _column_letters(index: int) -> str:
    """Convert a 1-based column index to letters."""
    letters = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _part(kind: str, value: int, absolute: bool, current: int) -> str:
    """Render one R1C1 coordinate component: ``R5``, ``R[-3]`` or ``R``."""
    if absolute:
        return f"{kind}{value}"
    offset = value - current
    return kind if offset == 0 else f"{kind}[{offset}]"


def _decode_side(side: str, row: int, col: int) -> tuple[str, int, int, bool, bool, str] | None:
    """Decode one side of a reference body.

    Returns:
        ``(r1c1, row, col, row_absolute, col_absolute, kind)`` where *kind* is one of
        ``"cell"``, ``"col"`` or ``"row"``, or None when the text is not a reference.
    """
    cell = _CELL_RE.match(side)
    if cell is not None:
        col_index = _column_index(cell.group(2))
        row_index = int(cell.group(4))
        if col_index is None or not 1 <= row_index <= MAX_ROW:
            return None
        col_absolute = cell.group(1) == "$"
        row_absolute = cell.group(3) == "$"
        rendered = _part("R", row_index, row_absolute, row) + _part(
            "C", col_index, col_absolute, col
        )
        return rendered, row_index, col_index, row_absolute, col_absolute, "cell"

    column = _COL_RE.match(side)
    if column is not None:
        col_index = _column_index(column.group(2))
        if col_index is None:
            return None
        col_absolute = column.group(1) == "$"
        return _part("C", col_index, col_absolute, col), 0, col_index, False, col_absolute, "col"

    row_only = _ROW_RE.match(side)
    if row_only is not None:
        row_index = int(row_only.group(2))
        if not 1 <= row_index <= MAX_ROW:
            return None
        row_absolute = row_only.group(1) == "$"
        return _part("R", row_index, row_absolute, row), row_index, 0, row_absolute, False, "row"

    return None


def _decode_body(body: str, row: int, col: int) -> _Body | None:
    """Decode a reference body into R1C1 plus its extent, or None if it is not a reference."""
    if ":" in body:
        left, _, right = body.partition(":")
        first = _decode_side(left, row, col)
        second = _decode_side(right, row, col)
        if first is None or second is None or first[5] != second[5]:
            return None
        kind = first[5]
        if kind == "col":
            min_row, max_row = 1, MAX_ROW
        else:
            min_row, max_row = sorted((first[1], second[1]))
        if kind == "row":
            min_col, max_col = 1, MAX_COL
        else:
            min_col, max_col = sorted((first[2], second[2]))
        return _Body(
            r1c1=f"{first[0]}:{second[0]}",
            min_row=min_row,
            min_col=min_col,
            max_row=max_row,
            max_col=max_col,
            row_absolute=first[3] and second[3],
            col_absolute=first[4] and second[4],
            is_range=True,
        )

    single = _decode_side(body, row, col)
    if single is None:
        return None
    kind = single[5]
    min_row, max_row = (1, MAX_ROW) if kind == "col" else (single[1], single[1])
    min_col, max_col = (1, MAX_COL) if kind == "row" else (single[2], single[2])
    return _Body(
        r1c1=single[0],
        min_row=min_row,
        min_col=min_col,
        max_row=max_row,
        max_col=max_col,
        row_absolute=single[3],
        col_absolute=single[4],
        is_range=kind != "cell",
    )


# =============================================================================
# A1 -> R1C1
# =============================================================================


def _convert_token(token: str, row: int, col: int, names: frozenset[str]) -> str | None:
    """Convert one ``OPERAND/RANGE`` token to R1C1, or None to leave it verbatim."""
    prefix, body = _split_sheet(token)
    if not prefix and (token.upper() in names or _STRUCTURED_RE.match(token)):
        return None
    decoded = _decode_body(body, row, col)
    if decoded is None:
        return None
    return prefix + decoded.r1c1


def to_r1c1(formula: str, row: int, col: int, *, names: frozenset[str] = frozenset()) -> str:
    """Normalise an A1 formula to R1C1 relative form as seen from one cell.

    Absolute references keep their coordinates (``$B$1`` -> ``R1C2``); relative ones become
    offsets from the holding cell (``G2`` at ``H2`` -> ``RC[-1]``). Sheet qualifiers,
    external workbook prefixes and quoting are preserved verbatim, so two cells in different
    sheets never collapse into the same operation by accident.

    Args:
        formula: The A1 formula, with or without its leading ``=``.
        row: 1-based row of the cell holding the formula.
        col: 1-based column of the cell holding the formula.
        names: Defined names in scope, upper-cased. A token matching one of these is left
            alone rather than being misread as a cell reference — ``TAX2024`` is a valid
            A1 reference *and* a plausible name, and only the workbook can say which.

    Returns:
        The R1C1 formula including its leading ``=``. A formula that cannot be tokenised is
        returned unchanged, so grouping still works (identical text still groups together).
    """
    tokens = tokenise(formula)
    if not tokens:
        return formula if formula.startswith("=") else "=" + formula

    out: list[str] = ["="]
    for token in tokens:
        if token.type == Token.OPERAND and token.subtype == Token.RANGE:
            converted = _convert_token(token.value, row, col, names)
            out.append(converted if converted is not None else token.value)
        else:
            out.append(token.value)
    return "".join(out)


# =============================================================================
# R1C1 -> A1
# =============================================================================


def _resolve_component(text: str | None, current: int, limit: int) -> int:
    """Resolve one R1C1 coordinate component back to an absolute 1-based index."""
    if text is None:
        return current
    value = current + int(text[1:-1]) if text.startswith("[") else int(text)
    if not 1 <= value <= limit:
        raise ValueError(f"R1C1 component {text!r} resolves to {value}, outside the sheet")
    return value


def _side_to_a1(side: str, row: int, col: int) -> str | None:
    """Convert one R1C1 side back to A1, or None when it is not an R1C1 reference."""
    match = _R1C1_PART_RE.match(side)
    if match is None:
        return None
    has_row = side.startswith("R")
    has_col = "C" in (side[1:] if has_row else side)
    if not has_row and not has_col:
        return None
    row_text = match.group("row")
    col_text = match.group("col")
    row_absolute = has_row and row_text is not None and not row_text.startswith("[")
    col_absolute = has_col and col_text is not None and not col_text.startswith("[")
    row_index = _resolve_component(row_text, row, MAX_ROW) if has_row else 0
    col_index = _resolve_component(col_text, col, MAX_COL) if has_col else 0
    column = f"{'$' if col_absolute else ''}{_column_letters(col_index)}" if has_col else ""
    rows = f"{'$' if row_absolute else ''}{row_index}" if has_row else ""
    return column + rows


def to_a1(formula: str, row: int, col: int) -> str:
    """Convert an R1C1 formula back to A1 as seen from one cell.

    The inverse of :func:`to_r1c1`, used to render a region's formula at a cell other than
    its anchor and as the round-trip check in the tests.

    Args:
        formula: The R1C1 formula, with or without its leading ``=``.
        row: 1-based row to resolve relative offsets against.
        col: 1-based column to resolve relative offsets against.

    Returns:
        The A1 formula including its leading ``=``.
    """
    tokens = tokenise(formula)
    if not tokens:
        return formula if formula.startswith("=") else "=" + formula

    out: list[str] = ["="]
    for token in tokens:
        if token.type != Token.OPERAND or token.subtype != Token.RANGE:
            out.append(token.value)
            continue
        prefix, body = _split_sheet(token.value)
        left, sep, right = body.partition(":")
        first = _side_to_a1(left, row, col)
        second = _side_to_a1(right, row, col) if sep else None
        if first is None or (sep and second is None):
            out.append(token.value)
            continue
        out.append(prefix + first + (f":{second}" if sep else ""))
    return "".join(out)


# =============================================================================
# FULL PARSE
# =============================================================================


def _function_name(token_value: str) -> str:
    """Strip the trailing bracket and any ``_xlfn.`` compatibility prefix."""
    name = token_value[:-1] if token_value.endswith("(") else token_value
    for prefix in ("_xlfn.", "_xlws.", "@"):
        if name.startswith(prefix):
            name = name[len(prefix) :]
    return name.upper()


def _make_reference(
    token: str,
    row: int,
    col: int,
    *,
    sheet: str | None,
    known_sheets: frozenset[str],
) -> tuple[Reference, RefExtent | None] | None:
    """Build the contract-level :class:`Reference` for one range token."""
    prefix, body = _split_sheet(token)
    ref_sheet, external = _decode_sheet_prefix(prefix)
    if "#REF!" in token.upper():
        return (
            Reference(
                raw=token,
                sheet=ref_sheet,
                a1=body,
                is_range=":" in body,
                is_external=external is not None,
                external_workbook=external,
                resolves=False,
            ),
            None,
        )

    decoded = _decode_body(body, row, col)
    if decoded is None:
        if _STRUCTURED_RE.match(body):
            return (
                Reference(
                    raw=token,
                    sheet=ref_sheet,
                    a1=body,
                    is_range=True,
                    is_external=external is not None,
                    external_workbook=external,
                ),
                None,
            )
        return None

    resolves = True
    if (
        external is None
        and ref_sheet is not None
        and known_sheets
        and ref_sheet not in known_sheets
    ):
        resolves = False

    reference = Reference(
        raw=token,
        sheet=ref_sheet,
        a1=body,
        is_range=decoded.is_range,
        absolute_row=decoded.row_absolute,
        absolute_col=decoded.col_absolute,
        is_external=external is not None,
        external_workbook=external,
        resolves=resolves,
    )
    extent = RefExtent(
        sheet=ref_sheet if ref_sheet is not None else sheet,
        min_row=decoded.min_row,
        min_col=decoded.min_col,
        max_row=decoded.max_row,
        max_col=decoded.max_col,
        is_external=external is not None,
    )
    return reference, extent


def parse_formula(
    formula: str,
    row: int,
    col: int,
    *,
    sheet: str | None = None,
    names: frozenset[str] = frozenset(),
    known_sheets: frozenset[str] = frozenset(),
) -> ParsedFormula:
    """Decompose one formula into references, functions, literals and its R1C1 form.

    Args:
        formula: The A1 formula, with or without its leading ``=``.
        row: 1-based row of the holding cell.
        col: 1-based column of the holding cell.
        sheet: Name of the holding sheet, used to qualify same-sheet extents.
        names: Defined names in scope, upper-cased (see :func:`to_r1c1`).
        known_sheets: Sheet names in the workbook. A qualified reference naming a sheet
            outside this set is marked ``resolves=False``.

    Returns:
        A :class:`ParsedFormula`. Tokenisation failure yields an otherwise-empty result
        carrying the original text, never an exception.
    """
    tokens = tokenise(formula)
    a1 = formula if formula.startswith("=") else "=" + formula
    if not tokens:
        return ParsedFormula(
            a1=a1,
            r1c1=a1,
            functions=(),
            outermost_function=None,
            references=(),
            extents=(),
            numeric_literals=(),
            string_literals=(),
            names=(),
            volatile_functions=(),
        )

    out: list[str] = ["="]
    functions: list[str] = []
    references: list[Reference] = []
    extents: list[RefExtent] = []
    numbers: list[float] = []
    strings: list[str] = []
    found_names: list[str] = []
    has_ref_error = False

    for token in tokens:
        if token.type == Token.FUNC and token.subtype == Token.OPEN:
            functions.append(_function_name(token.value))
            out.append(token.value)
            continue
        if token.type == Token.OPERAND and token.subtype == Token.NUMBER:
            try:
                numbers.append(float(token.value))
            except ValueError:
                logger.debug("non-numeric NUMBER token: %s", token.value)
            out.append(token.value)
            continue
        if token.type == Token.OPERAND and token.subtype == Token.TEXT:
            strings.append(token.value.strip('"').replace('""', '"'))
            out.append(token.value)
            continue
        if token.type == Token.OPERAND and token.subtype == Token.ERROR:
            has_ref_error = has_ref_error or token.value.upper() == "#REF!"
            out.append(token.value)
            continue
        if token.type != Token.OPERAND or token.subtype != Token.RANGE:
            out.append(token.value)
            continue

        prefix, _ = _split_sheet(token.value)
        if not prefix and token.value.upper() in names:
            found_names.append(token.value)
            out.append(token.value)
            continue
        made = _make_reference(token.value, row, col, sheet=sheet, known_sheets=known_sheets)
        if made is None:
            found_names.append(token.value)
            out.append(token.value)
            continue
        reference, extent = made
        references.append(reference)
        if extent is not None:
            extents.append(extent)
        if not reference.resolves:
            has_ref_error = has_ref_error or "#REF!" in token.value.upper()
        converted = _convert_token(token.value, row, col, names)
        out.append(converted if converted is not None else token.value)

    return ParsedFormula(
        a1=a1,
        r1c1="".join(out),
        functions=tuple(functions),
        outermost_function=functions[0] if functions else None,
        references=tuple(references),
        extents=tuple(extents),
        numeric_literals=tuple(numbers),
        string_literals=tuple(strings),
        names=tuple(found_names),
        volatile_functions=tuple(sorted({f for f in functions if f in VOLATILE_FUNCTIONS})),
        has_external_refs=any(r.is_external for r in references),
        has_ref_error=has_ref_error,
    )
