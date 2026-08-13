"""The agent loop, driven by a scripted model client rather than by a live endpoint.

Everything here runs with no key, no network and no marimo. That is the point: PLAN M4's loop is
the part of kedge most easily left untested "because it needs a model", and a loop nobody tests is
a loop that silently emits two ``DoneEvent``s on the error path and corrupts a session transcript.

The choreography is asserted against the same shape ``ScriptedAgent`` establishes, because that
fake is what the UI was built and judged against, and a real loop that streams a different shape
would be a regression the UI would show but nothing would catch.
"""

from __future__ import annotations

import asyncio
import json
from functools import cache
from pathlib import Path
from tempfile import TemporaryDirectory
from types import SimpleNamespace
from typing import Any

import httpx
import pytest

from conftest import approved_plan_store
from kedge.agent.context import (
    _EVICTED_SHORT_TAIL,
    _EVICTED_TAIL,
    CARRY_BLOCK_TURNS,
    MAX_EVICTED_SHAPE_CHARS,
    ContextMessage,
    TokenCounter,
)
from kedge.agent.loop import (
    _FINAL_WORD_PROMPT,
    _NO_ACCOUNT,
    AgentError,
    ChatDelta,
    KedgeAgent,
    OpenAIClient,
    Usage,
    _history_before,
    _result_shape,
    _safe_arguments,
)
from kedge.agent.tools import ToolContext, ToolResult
from kedge.analysis.model import WorkbookAnalysis
from kedge.config import ContextConfig
from kedge.errors import KedgeError
from kedge.notebook.model import (
    CellInfo,
    CellRef,
    GraphNode,
    GraphView,
    MutationResult,
    ProbeResult,
)
from kedge.plan.store import PlanStore
from kedge.server.agent_seam import AgentLoop, CancelToken, TurnMessage, TurnRequest
from kedge.server.events import DoneEvent

CLEAN_CELL = "apply_haircuts = load_handin.join(reference_haircuts, on='asset_class', how='left')\n"
PANDAS_CELL = "import pandas as pd\nframe = pd.read_excel('handin.xlsx')\n"

_PLAN_DIRECTORY = TemporaryDirectory(prefix="kedge-plans-")
"""Where `_plans_in_force` writes. Module-level so it outlives the tests and is removed with the
interpreter, rather than a fixture threaded through `build`'s forty call sites."""

_EVICTION_TAILS = (_EVICTED_TAIL, _EVICTED_SHORT_TAIL)
"""The wordings an eviction stub can end in.

Which one it can afford depends on the length of the result it replaces — a stub costs no more
than what it stands in for — so a test asking whether something was evicted asks for either rather
than pinning the rung, and asks for the constants rather than for a phrase copied out of them."""


# ── fakes ────────────────────────────────────────────────────────────────────────────────────


class ScriptedClient:
    """A model that says exactly what the test tells it to, one list of deltas per round trip."""

    def __init__(self, rounds: list[list[ChatDelta]]) -> None:
        self._rounds = list(rounds)
        self.seen: list[list[dict[str, Any]]] = []
        self.tools_offered: list[list[str]] = []
        """What each round was offered, in order. Per round rather than latest-wins because the
        last word a stopped turn gets is offered nothing, and a field that only remembers the most
        recent request cannot tell that apart from a turn that was never asked anything at all."""

        self.models_asked: list[str] = []

    async def stream(self, *, model: str, messages: Any, tools: Any) -> Any:
        self.models_asked.append(model)
        self.seen.append([dict(message) for message in messages])
        self.tools_offered.append([tool["function"]["name"] for tool in tools])
        for delta in self._rounds.pop(0) if self._rounds else []:
            yield delta


class RaisingClient:
    """A model endpoint that fails the way a real one does: mid-stream."""

    def __init__(self, error: BaseException) -> None:
        self._error = error

    async def stream(self, **_kwargs: Any) -> Any:
        yield ChatDelta(text="Let me look at ")
        raise self._error


class CancellingClient:
    """A model that keeps talking while the user presses stop."""

    def __init__(self, cancel: CancelToken) -> None:
        self._cancel = cancel

    async def stream(self, **_kwargs: Any) -> Any:
        yield ChatDelta(text="Working ")
        self._cancel.cancel()
        yield ChatDelta(text="on it ")
        yield ChatDelta(text="still ")


class ThinkingClient:
    """A reasoning model: a long silence on the wire before anything the loop can translate.

    This is the shape that made Stop inert, and it is the ordinary case rather than a contrived
    one. ``responses_delta`` maps every reasoning event to ``None``, so for the whole of a model's
    thinking nothing reaches the loop — and a cancellation check written inside ``async for delta``
    does not run once during exactly the stretch a user reaches for Stop.
    """

    def __init__(self) -> None:
        self.finished = False

    async def stream(self, **_kwargs: Any) -> Any:
        await asyncio.sleep(30)  # thinking, none of which becomes a delta
        self.finished = True  # pragma: no cover - the turn is stopped long before this
        yield ChatDelta(text="...")  # pragma: no cover


class FakeDriver:
    def __init__(self) -> None:
        self.graphs_read = 0
        self.created: list[str] = []

    async def read_graph(self) -> GraphView:
        self.graphs_read += 1
        return GraphView(
            nodes=(
                GraphNode(id="AAaa", name="imports", defs=("pl",)),
                GraphNode(id="MJUe", name="load_handin", defs=("load_handin",), refs=("pl",)),
            )
        )

    async def list_cells(self, *, with_code: bool = True) -> tuple[CellInfo, ...]:
        return (CellInfo(id="AAaa", name="imports", code="import polars as pl"),)

    async def create_cell(self, code: str, *, name: str, **_kw: Any) -> MutationResult:
        self.created.append(name)
        return MutationResult(
            operation="create_cell", cell=CellRef(id="U1", name=name), ran=True, status="idle"
        )

    async def probe(self, code: str) -> ProbeResult:
        return ProbeResult(
            ok=True,
            value_repr="shape: (49999, 3)\n" + "| corp | 1000.0 | 0.02 |\n" * 6,
            value_type="DataFrame",
        )


def call(name: str, arguments: dict[str, Any], *, index: int = 0) -> list[ChatDelta]:
    """The deltas a model emits for one tool call, split the way a real stream splits them."""
    encoded = json.dumps(arguments)
    midpoint = len(encoded) // 2
    return [
        ChatDelta(index=index, call_id=f"call_{index}", name=name),
        ChatDelta(index=index, arguments=encoded[:midpoint]),
        ChatDelta(index=index, arguments=encoded[midpoint:]),
    ]


@cache
def _plans_in_force() -> PlanStore:
    """One store on disk holding an approved plan, shared by every loop built without one.

    `propose_cell` and `edit_cell` are refused until a plan is in force, and most of what is
    asserted here — the event choreography, the validation gate, the retry cap — needs to get past
    that to reach its own subject. Built once and never written to again, so sharing it between
    tests cannot couple them; `_PLAN_DIRECTORY` cleans itself up when the session ends.
    """
    return approved_plan_store(Path(_PLAN_DIRECTORY.name))


def build(
    client: Any, *, driver: Any = None, context: ToolContext | None = None, **kwargs: Any
) -> KedgeAgent:
    return KedgeAgent(
        client=client,
        context=context
        if context is not None
        else ToolContext(driver=driver, plans=_plans_in_force()),
        counter=TokenCounter(allow_download=False),
        system_prompt="SYSTEM PROMPT",
        **kwargs,
    )


async def drive(
    agent: KedgeAgent, message: str = "convert the haircut lookup", **kwargs: Any
) -> list[Any]:
    request = TurnRequest(turn_id="t1", session_id="s1", message=message, **kwargs)
    return [event async for event in agent.run(request, cancel=CancelToken())]


# ── the Protocol ─────────────────────────────────────────────────────────────────────────────


def test_the_loop_satisfies_the_server_s_protocol() -> None:
    assert isinstance(build(ScriptedClient([])), AgentLoop)


async def test_the_tools_offered_are_the_ones_plan_m4_lists() -> None:
    client = ScriptedClient([[ChatDelta(text="Nothing to do.")]])
    await drive(build(client))
    assert "propose_cell" in client.tools_offered[0]
    assert "reconcile" in client.tools_offered[0]
    # PLAN M4's fifteen, plus `propose_plan`: the planning step has to be reachable from the chat
    # too, or an account worked out in conversation stays prose and is compacted away with it.
    assert len(client.tools_offered[0]) == 16


# ── exactly one DoneEvent ────────────────────────────────────────────────────────────────────


async def test_exactly_one_done_event_on_the_success_path() -> None:
    events = await drive(build(ScriptedClient([[ChatDelta(text="Here is what I found.")]])))
    assert [event.type for event in events].count("done") == 1
    assert events[-1].type == "done"
    assert events[-1].turn_id == "t1"


async def test_exactly_one_done_event_on_the_error_path() -> None:
    events = await drive(build(RaisingClient(EndpointRefusedError())))
    types = [event.type for event in events]
    assert types.count("done") == 1
    assert types[-1] == "done"
    errors = [event for event in events if event.type == "error"]
    assert len(errors) == 1
    assert errors[0].recoverable is True


async def test_an_unexpected_exception_is_reported_as_unrecoverable() -> None:
    events = await drive(build(RaisingClient(ValueError("the SDK changed shape"))))
    errors = [event for event in events if event.type == "error"]
    assert len(errors) == 1
    assert errors[0].recoverable is False
    assert "the SDK changed shape" in errors[0].message
    assert [event.type for event in events].count("done") == 1


async def test_exactly_one_done_event_on_the_cancellation_path() -> None:
    token = CancelToken()
    agent = build(CancellingClient(token))
    request = TurnRequest(turn_id="t1", session_id="s1", message="go")
    events = [event async for event in agent.run(request, cancel=token)]
    types = [event.type for event in events]
    assert types.count("done") == 1
    assert types[-1] == "done"
    cancelled = [event for event in events if event.type == "error"]
    assert cancelled[0].message == "Turn cancelled at your request."
    assert cancelled[0].recoverable is True


async def test_a_pre_cancelled_turn_still_finishes_cleanly() -> None:
    token = CancelToken()
    token.cancel()
    agent = build(ScriptedClient([[ChatDelta(text="never seen")]]))
    request = TurnRequest(turn_id="t1", session_id="s1", message="go")
    events = [event async for event in agent.run(request, cancel=token)]
    assert [event.type for event in events].count("done") == 1
    assert not any(event.type == "token" for event in events)


async def test_stop_lands_while_the_model_is_still_thinking() -> None:
    """Cancellation must not wait for the model to say something first.

    The check used to sit inside ``async for delta``, so a reasoning model that spends a minute
    thinking before its first token left Stop doing nothing for that whole minute — and the chat
    went on saying "Thinking" while the user pressed it again. The wait here is 30 seconds of model
    call against a 5 second deadline: it only passes if the turn abandons the call rather than
    outliving it.
    """
    token = CancelToken()
    client = ThinkingClient()
    agent = build(client)
    request = TurnRequest(turn_id="t1", session_id="s1", message="go")

    async def collect() -> list[Any]:
        return [event async for event in agent.run(request, cancel=token)]

    async def press_stop() -> None:
        await asyncio.sleep(0.05)
        token.cancel()

    stopper = asyncio.create_task(press_stop())
    # Timed, for the reason `test_stop_during_the_last_word_is_honoured` sets out at length: `run`
    # turns a cancellation into a tidy finish, so `wait_for` prising the turn off a 30 second call
    # at the deadline looks from here exactly like the turn abandoning it at once. Without the
    # clock this test passes with `_abandon_if_cancelled` taken out -- a false positive sitting
    # directly above the docstring explaining why that must not happen.
    clock = asyncio.get_running_loop().time
    started = clock()
    events = await asyncio.wait_for(collect(), timeout=5)
    elapsed = clock() - started
    await stopper

    assert elapsed < 1, f"the turn waited {elapsed:.1f}s on a model call it should have abandoned"

    types = [event.type for event in events]
    assert types.count("done") == 1
    assert types[-1] == "done"
    stopped = next(event for event in events if event.type == "error")
    assert stopped.message == "Turn cancelled at your request."
    assert client.finished is False, "the model call was abandoned, not waited out"


async def test_a_task_cancellation_that_is_not_the_user_s_is_re_raised() -> None:
    class Torn:
        async def stream(self, **_kwargs: Any) -> Any:
            raise asyncio.CancelledError
            yield  # pragma: no cover

    agent = build(Torn())
    request = TurnRequest(turn_id="t1", session_id="s1", message="go")
    with pytest.raises(asyncio.CancelledError):
        [event async for event in agent.run(request, cancel=CancelToken())]


# ── the choreography ─────────────────────────────────────────────────────────────────────────


async def test_a_realistic_turn_streams_the_same_shape_as_the_scripted_agent() -> None:
    driver = FakeDriver()
    client = ScriptedClient(
        [
            [
                ChatDelta(text="The lookup is an exact VLOOKUP, "),
                ChatDelta(text="which is a left join.\n\n"),
                *call("propose_cell", {"name": "apply_haircuts", "code": CLEAN_CELL}),
            ],
            [ChatDelta(text="Reconciliation is clean.")],
        ]
    )
    events = await drive(build(client, driver=driver))
    types = [event.type for event in events]

    assert types[0] == "status"
    assert "token" in types
    assert types.index("tool_call") < types.index("tool_result")
    assert "validation" in types
    assert "cell_created" in types
    assert "cell_running" in types
    assert "cell_result" in types
    assert types[-1] == "done"

    created = next(event for event in events if event.type == "cell_created")
    assert created.name == "apply_haircuts"
    assert created.preview.startswith("apply_haircuts =")
    assert driver.created == ["apply_haircuts"]


async def test_the_tool_call_event_never_carries_the_cell_body() -> None:
    driver = FakeDriver()
    body = "apply_haircuts = 1  # " + "x" * 500 + "\n"
    client = ScriptedClient(
        [[*call("propose_cell", {"name": "apply_haircuts", "code": body})], [ChatDelta(text="ok")]]
    )
    events = await drive(build(client, driver=driver))
    tool_call = next(event for event in events if event.type == "tool_call")
    assert "xxxx" not in tool_call.args_summary
    assert "chars>" in tool_call.args_summary


async def test_two_tool_calls_in_one_reply_are_both_dispatched() -> None:
    driver = FakeDriver()
    client = ScriptedClient(
        [
            [
                *call("list_cells", {}, index=0),
                *call("probe", {"code": "1 + 1"}, index=1),
            ],
            [ChatDelta(text="done")],
        ]
    )
    events = await drive(build(client, driver=driver))
    names = [event.name for event in events if event.type == "tool_call"]
    assert names == ["list_cells", "probe"]


async def test_the_status_phase_matches_what_the_tool_is_doing() -> None:
    driver = FakeDriver()
    client = ScriptedClient(
        [[*call("propose_cell", {"name": "c", "code": "c = 1\n"})], [ChatDelta(text="ok")]]
    )
    events = await drive(build(client, driver=driver))
    phases = [event.phase for event in events if event.type == "status"]
    assert phases[0] == "analysing"
    assert "editing" in phases


# ── the validation gate in the loop ──────────────────────────────────────────────────────────


async def test_a_rejected_cell_comes_back_to_the_model_as_violations() -> None:
    driver = FakeDriver()
    client = ScriptedClient(
        [
            [*call("propose_cell", {"name": "loader", "code": PANDAS_CELL})],
            [ChatDelta(text="Rewritten with polars.")],
        ]
    )
    events = await drive(build(client, driver=driver))
    validation = next(event for event in events if event.type == "validation")
    assert validation.ok is False
    assert any("polars, never pandas" in violation for violation in validation.violations)
    assert driver.created == []

    # The violations were handed back, so the model's next prompt contains them verbatim.
    second_prompt = json.dumps(client.seen[1])
    assert "polars, never pandas" in second_prompt


async def test_validation_retries_are_capped_at_three() -> None:
    driver = FakeDriver()
    rounds = [[*call("propose_cell", {"name": "loader", "code": PANDAS_CELL})] for _ in range(5)]
    events = await drive(build(client := ScriptedClient(rounds), driver=driver))
    types = [event.type for event in events]

    assert types.count("validation") == 3
    assert types.count("tool_call") == 3
    errors = [event for event in events if event.type == "error"]
    assert len(errors) == 1
    assert "3 times running" in errors[0].message
    assert types.count("done") == 1
    # Three attempts and the last word, rather than the rest of the budget: the fourth request is
    # the one that asks for prose, and it is offered nothing to call.
    assert len(client.seen) == 4
    assert client.tools_offered[-1] == []


# ── the step budget, and carrying a turn across it ───────────────────────────────────────────


async def test_the_step_budget_pauses_a_loop_that_never_answers() -> None:
    driver = FakeDriver()
    rounds = [[*call("list_cells", {}, index=0)] for _ in range(10)]
    events = await drive(build(ScriptedClient(rounds), driver=driver, max_steps=3))
    types = [event.type for event in events]
    assert types.count("tool_call") == 3
    paused = [event for event in events if event.type == "paused"]
    assert len(paused) == 1
    assert paused[0].steps == 3
    assert "3 steps" in paused[0].message
    # It is a question, not a failure: an error here is what makes a user start again.
    assert not any(event.type == "error" for event in events)
    assert types.count("done") == 1


async def test_a_paused_turn_resumes_with_everything_it_had_learnt() -> None:
    driver = FakeDriver()
    rounds = [[*call("list_cells", {}, index=0)] for _ in range(10)]
    agent = build(client := ScriptedClient(rounds), driver=driver, max_steps=2)
    await drive(agent, "convert the haircut lookup")

    # What the server would have persisted of that turn: the question, and no prose to speak of.
    history = (
        TurnMessage(role="user", content="convert the haircut lookup"),
        TurnMessage(role="assistant", content="Looking."),
    )
    await drive(agent, "continue", history=history)

    resumed = client.seen[2]
    assert [message["role"] for message in resumed].count("tool") == 2
    assert any(message.get("tool_calls") for message in resumed)
    assert resumed[-1] == {"role": "user", "content": "continue"}
    # The paused turn is carried whole rather than alongside the flattened record of itself.
    opening = [m for m in resumed if m.get("content") == "convert the haircut lookup"]
    assert len(opening) == 1
    assert not any(message.get("content") == "Looking." for message in resumed)


async def test_a_carried_turn_of_kernel_reads_is_let_go_once_the_model_answers() -> None:
    """Answering is no longer what releases the carry — expiring is, and here everything has.

    The turn read the kernel and nothing else. Once it answers, the conversation has moved past it
    and a listing from before the user's next message describes a notebook they may have edited, so
    every result in the span is evicted and what is left is a shell the server's own history
    already replays.
    """
    driver = FakeDriver()
    agent = build(
        client := ScriptedClient(
            [
                [*call("list_cells", {}, index=0)],
                [*call("list_cells", {}, index=0)],
                [ChatDelta(text="Here is the answer.")],
                [ChatDelta(text="Anything else?")],
            ]
        ),
        driver=driver,
        max_steps=2,
    )
    await drive(agent, "convert the haircut lookup")
    await drive(agent, "continue")
    await drive(agent, "and now the overrides")

    assert not any(message["role"] == "tool" for message in client.seen[-1])


async def test_a_cancelled_turn_keeps_the_tool_results_it_paid_for() -> None:
    """Stop is not "throw it away". The work is on the shelf when the user says what to do next."""

    class StopAfterOneCall:
        """Calls a tool, presses Stop during the next round trip, and answers the turn after."""

        def __init__(self) -> None:
            self.token = CancelToken()
            self.rounds = 0
            self.seen: list[list[dict[str, Any]]] = []

        async def stream(self, *, model: str, messages: Any, tools: Any) -> Any:
            del model, tools
            self.seen.append([dict(message) for message in messages])
            self.rounds += 1
            if self.rounds > 2:
                yield ChatDelta(text="Picking up where I left off.")
                return
            if self.rounds == 2:
                self.token.cancel()
            for delta in call("list_cells", {}, index=0):
                yield delta

    client = StopAfterOneCall()
    agent = build(client, driver=FakeDriver())
    request = TurnRequest(turn_id="t1", session_id="s1", message="convert the haircut lookup")
    events = [event async for event in agent.run(request, cancel=client.token)]
    assert [event.type for event in events].count("done") == 1

    await drive(agent, "use the cached values instead")
    carried = client.seen[-1]
    assert any(message["role"] == "tool" for message in carried)
    # Every call the interrupted turn made is answered, or the endpoint rejects the request whole.
    asked = {
        tool_call["id"] for message in carried for tool_call in message.get("tool_calls") or ()
    }
    answered = {message["tool_call_id"] for message in carried if message["role"] == "tool"}
    assert asked and asked == answered


async def test_a_new_session_forgets_the_turn_that_was_held_for_it() -> None:
    driver = FakeDriver()
    rounds = [[*call("list_cells", {}, index=0)] for _ in range(10)]
    agent = build(client := ScriptedClient(rounds), driver=driver, max_steps=2)
    await drive(agent, "convert the haircut lookup")
    agent.reset_session("s1")
    await drive(agent, "start again")
    assert not any(message["role"] == "tool" for message in client.seen[2])


# ── carrying a turn that answered ────────────────────────────────────────────────────────────


def _tools_in(prompt: list[dict[str, Any]]) -> dict[str, str]:
    """The tool results in one prompt, by the id of the call each answers."""
    return {
        message["tool_call_id"]: message["content"]
        for message in prompt
        if message["role"] == "tool"
    }


async def test_an_answered_turn_hands_its_stable_results_to_the_next_one(
    analysis: WorkbookAnalysis,
) -> None:
    """The bug this exists for: summarise the workbook, then ask for the notebook.

    The first turn read the workbook and answered well. The second re-read all of it, because a
    turn that succeeded had its tool traffic thrown away and the server persists prose alone. The
    expensive turn was being punished for succeeding.
    """
    client = ScriptedClient(
        [
            [*call("inspect_workbook", {"section": "operations"})],
            [ChatDelta(text="Stage 2 is the haircut lookup in Calc!H.")],
            [ChatDelta(text="Starting on it.")],
        ]
    )
    agent = build(client, context=ToolContext(analysis=analysis))
    await drive(agent, "summarise the workbook stage by stage")
    await drive(agent, "now build the notebook")

    produced = _tools_in(client.seen[1])
    assert _tools_in(client.seen[2]) == produced, "the workbook does not change under kedge"
    assert produced


async def test_an_answered_turns_volatile_results_arrive_as_stubs_rather_than_as_holes(
    analysis: WorkbookAnalysis,
) -> None:
    """What is carried is what is still true, and what is not still true is still *there*.

    A kernel listing from before the follow-up describes a notebook the user may have edited in
    the pane next to the chat. Dropping the message is not an option — an assistant message
    carrying ``tool_calls`` without a result against each id is rejected outright — so it is
    evicted instead, which leaves the model the call, its arguments and an invitation to repeat it.
    """
    driver = FakeDriver()
    client = ScriptedClient(
        [
            [
                *call("inspect_workbook", {"section": "operations"}, index=0),
                *call("list_cells", {}, index=1),
            ],
            [ChatDelta(text="Two cells so far, and the lookup is in Calc!H.")],
            [ChatDelta(text="Starting on it.")],
        ]
    )
    agent = build(client, context=ToolContext(analysis=analysis, driver=driver))
    await drive(agent, "summarise the workbook stage by stage")
    await drive(agent, "now build the notebook")

    before, after = _tools_in(client.seen[1]), _tools_in(client.seen[2])
    assert set(after) == set(before), "every call the carried turn made is still answered"
    assert after["call_0"] == before["call_0"], "the workbook cannot have changed"
    assert after["call_1"] != before["call_1"], "the kernel can, and has"
    assert "list_cells" in after["call_1"]
    assert any(tail in after["call_1"] for tail in _EVICTION_TAILS)


async def test_the_carried_turn_still_answers_every_tool_call_it_makes(
    analysis: WorkbookAnalysis,
) -> None:
    """The API-rejection case, asserted over the request as the endpoint would receive it."""
    driver = FakeDriver()
    client = ScriptedClient(
        [
            [
                *call("inspect_workbook", {"section": "operations"}, index=0),
                *call("probe", {"code": "1 + 1"}, index=1),
            ],
            [ChatDelta(text="Both read.")],
            [ChatDelta(text="Right.")],
        ]
    )
    agent = build(client, context=ToolContext(analysis=analysis, driver=driver))
    await drive(agent, "summarise the workbook stage by stage")
    await drive(agent, "now build the notebook")

    prompt = client.seen[2]
    asked = {call["id"] for message in prompt for call in message.get("tool_calls") or ()}
    answered = {message["tool_call_id"] for message in prompt if message["role"] == "tool"}
    assert asked and asked == answered


async def test_a_carried_result_ages_out_after_the_configured_number_of_turns(
    analysis: WorkbookAnalysis,
) -> None:
    """Re-dating is what keeps the carry in order; it must not also make the carry immortal.

    ``resume`` re-dates a carried message into the turn it is resumed as, so a result carried
    every turn would read as one turn old for ever and never reach the horizon
    ``[context] evict_tool_results_after_turns`` sets.

    A span holding nothing but an expired result is not carried at all. The stub that says a call
    has gone earns its place next to results that have not; on its own it is the whole of what the
    server's own history already replays, and a turn more expensive than the record it displaces.
    """
    rounds: list[list[ChatDelta]] = [
        [*call("inspect_workbook", {"section": "operations"})],
        [ChatDelta(text="Stage 2 is the haircut lookup in Calc!H.")],
    ]
    rounds += [[ChatDelta(text="Understood.")] for _ in range(4)]
    agent = build(
        client := ScriptedClient(rounds),
        context=ToolContext(analysis=analysis),
        context_config=ContextConfig(evict_tool_results_after_turns=2),
    )
    await drive(agent, "summarise the workbook stage by stage")
    for _ in range(3):
        await drive(agent, "carry on")

    fetched = _tools_in(client.seen[1])["call_0"]
    assert _tools_in(client.seen[2])["call_0"] == fetched
    assert not _tools_in(client.seen[3])
    assert not _tools_in(client.seen[4])


async def test_an_expired_result_spends_one_turn_as_a_stub_before_it_leaves(
    analysis: WorkbookAnalysis,
) -> None:
    """The extra turn a result spends as a stub, which is signal rather than an off-by-one.

    Content lasts ``evict_tool_results_after_turns`` turns; the stub naming the call lasts one turn
    beyond that, so the model reads "this is gone, ask again" rather than finding an earlier call
    of its own with no answer against it. Each turn here makes a call of its own, which is what
    keeps the span worth carrying — a span holding nothing but the stub is dropped whole.
    """
    rounds: list[list[ChatDelta]] = []
    for index in range(4):
        rounds.append([*call("inspect_workbook", {"section": "operations"}, index=index)])
        rounds.append([ChatDelta(text=f"Read it, pass {index}.")])
    agent = build(
        client := ScriptedClient(rounds),
        context=ToolContext(analysis=analysis),
        context_config=ContextConfig(evict_tool_results_after_turns=2),
    )
    for index in range(4):
        await drive(agent, f"question {index}")

    fetched = _tools_in(client.seen[1])["call_0"]
    assert _tools_in(client.seen[2])["call_0"] == fetched, "content for the configured horizon"
    stub = _tools_in(client.seen[4])["call_0"]
    assert stub != fetched
    assert "inspect_workbook" in stub
    assert any(tail in stub for tail in _EVICTION_TAILS)
    assert "call_0" not in _tools_in(client.seen[6]), "and then it goes"


async def test_the_carry_stays_bounded_over_a_long_session(analysis: WorkbookAnalysis) -> None:
    """Carrying on every turn is what makes this a leak rather than a corner case."""
    turns = 10
    rounds: list[list[ChatDelta]] = []
    for index in range(turns):
        rounds.append([*call("inspect_workbook", {"section": "operations"}, index=index)])
        rounds.append([ChatDelta(text=f"Read it, pass {index}.")])
    agent = build(
        client := ScriptedClient(rounds),
        context=ToolContext(analysis=analysis),
        context_config=ContextConfig(evict_tool_results_after_turns=3),
    )
    for index in range(turns):
        await drive(agent, f"question {index}")

    # The prompt that opens each turn: what the previous ones handed it, before it calls anything.
    opening = [len(_tools_in(client.seen[index * 2])) for index in range(turns)]
    assert opening[-1] == opening[-1 - CARRY_BLOCK_TURNS], "the carry reaches a steady cycle"
    # Bounded by the horizon, and not by being empty: the whole point is that something arrives.
    assert 0 < max(opening) <= 3 - 1 + CARRY_BLOCK_TURNS


async def test_a_one_turn_horizon_carries_nothing_rather_than_a_span_of_stubs(
    analysis: WorkbookAnalysis,
) -> None:
    """``evict_tool_results_after_turns = 1`` is legal, and it is where the mechanism inverts.

    Everything a turn reads is stubbed before the next turn sees it, so there is nothing to carry —
    and holding the span anyway would trim a turn out of history to make room for a shell that
    costs more than the record it displaced.
    """
    client = ScriptedClient(
        [
            [*call("inspect_workbook", {"section": "operations"})],
            [ChatDelta(text="Stage 2 is the haircut lookup in Calc!H.")],
            [ChatDelta(text="Starting on it.")],
        ]
    )
    agent = build(
        client,
        context=ToolContext(analysis=analysis),
        context_config=ContextConfig(evict_tool_results_after_turns=1),
    )
    await drive(agent, "summarise the workbook stage by stage")
    await drive(agent, "now build the notebook")

    assert not _tools_in(client.seen[2])


async def test_consecutive_pauses_expire_what_the_kernel_may_have_changed(
    analysis: WorkbookAnalysis,
) -> None:
    """A paused turn is resumed as itself; the legs behind it are not.

    The user types between the pause and the resume, and edits cells in the pane beside the chat
    while the turn waits — which is why the notebook state is rebuilt every turn. So a ``probe``
    read on the first leg is not still describing the kernel two pauses later, however the turns
    ended, and only the leg that has just run keeps its volatile results.
    """
    client = ScriptedClient(
        [[*call("probe", {"code": "load_handin.height"}, index=index)] for index in range(3)]
    )
    agent = build(
        client,
        context=ToolContext(analysis=analysis, driver=FakeDriver()),
        context_config=ContextConfig(evict_tool_results_after_turns=6),
        max_steps=1,
    )
    await drive(agent, "start reading")
    await drive(agent, "continue")
    await drive(agent, "continue")

    # The leg that just ran keeps its probe; the one behind it has crossed a turn boundary.
    third = _tools_in(client.seen[2])
    assert "49999" in third["call_1"], "the leg the pause is resuming stands"
    assert "49999" not in third["call_0"]
    assert any(tail in third["call_0"] for tail in _EVICTION_TAILS)


async def test_compaction_does_not_take_the_carry_with_it(analysis: WorkbookAnalysis) -> None:
    """Compaction runs at the end of a turn and drops everything older than the one being built.

    That is the whole resumed span. Taken before the carry it collapsed the hand-off to its newest
    leg on every turn a session sat over the compaction threshold — the sessions that most need to
    start warm. The digest is compaction's lasting product and does not care which order it is
    written in.
    """
    client = ScriptedClient(
        [
            [*call("inspect_workbook", {"section": "operations"})],
            [ChatDelta(text="Stage 2 is the haircut lookup in Calc!H.")],
            [ChatDelta(text="Starting on it.")],
        ]
    )
    agent = build(
        client,
        context=ToolContext(analysis=analysis),
        # A budget the pinned blocks alone overrun, so compaction fires on the first turn.
        context_config=ContextConfig(max_context_tokens=5_000, reserve_output_tokens=0),
    )
    await drive(agent, "summarise the workbook stage by stage")
    await drive(agent, "now build the notebook")

    assert _tools_in(client.seen[2]), "the carry survived the compaction that ran beside it"


# ── what the server hands the loop, and what the loop hands back ──────────────────────────────


def _span(*openers: str) -> list[ContextMessage]:
    """A carried span holding one user message per leg, which is what history is counted against."""
    return [
        ContextMessage(role="user", content=opener, kind="user", turn=1, carried_age=age)
        for age, opener in enumerate(reversed(openers))
    ]


def test_history_is_trimmed_by_counting_the_turns_the_span_reinstates() -> None:
    history = [("user", "first"), ("assistant", "one"), ("user", "second"), ("assistant", "two")]
    assert _history_before(history, _span("second")) == history[:2]


def test_a_user_message_repeated_later_does_not_trim_history_at_the_earlier_copy() -> None:
    """The defect this counts its way past, and the phrasing kedge itself suggests.

    ``_pause_message`` asks the user to say "continue", so the colliding text is the likeliest text
    there is, and a two-leg span needs nothing more exotic than two consecutive turns. Searching
    backwards for a message equal to the span's opener stops at the *newer* copy, leaves the turn
    between the two in history, and the model then reads that exchange twice — once flattened out
    of order and once in its place.
    """
    history = [
        ("user", "carry on"),
        ("assistant", "ANSWER ONE"),
        ("user", "carry on"),
        ("assistant", "ANSWER TWO"),
    ]
    assert _history_before(history, _span("carry on", "carry on")) == []


def test_history_is_left_alone_when_it_is_not_what_the_span_was_built_from() -> None:
    """Two copies of a turn is a poor context; a turn silently truncated is a worse one."""
    history = [("user", "something else"), ("assistant", "an answer")]
    assert _history_before(history, _span("convert the haircut lookup")) == history


def test_history_is_left_alone_when_it_holds_fewer_turns_than_the_span() -> None:
    history = [("user", "only one"), ("assistant", "an answer")]
    assert _history_before(history, _span("only one", "and another")) == history


def test_a_span_with_no_user_message_leaves_history_alone() -> None:
    assert _history_before([("user", "a question")], []) == [("user", "a question")]


# ── context ──────────────────────────────────────────────────────────────────────────────────


async def test_notebook_state_is_rebuilt_from_the_kernel_every_turn() -> None:
    driver = FakeDriver()
    agent = build(ScriptedClient([[ChatDelta(text="ok")], [ChatDelta(text="ok")]]), driver=driver)
    await drive(agent)
    assert driver.graphs_read == 1
    await drive(agent)
    assert driver.graphs_read == 2


async def test_the_name_registry_and_notebook_state_are_pinned_into_every_turn() -> None:
    client = ScriptedClient([[ChatDelta(text="ok")]])
    await drive(build(client, driver=FakeDriver()))
    system = client.seen[0][0]
    assert system["role"] == "system"
    assert "SYSTEM PROMPT" in system["content"]
    assert "one owning cell per public name" in system["content"]
    assert "load_handin (MJUe)" in system["content"]
    assert "Cell bodies are not shown" in system["content"]


async def test_history_is_replayed_but_the_notebook_is_not_taken_from_it() -> None:
    driver = FakeDriver()
    client = ScriptedClient([[ChatDelta(text="ok")]])
    await drive(
        build(client, driver=driver),
        history=(
            TurnMessage(role="user", content="what does Calc do?"),
            TurnMessage(role="assistant", content="It applies a haircut."),
        ),
    )
    prompt = client.seen[0]
    assert {"role": "user", "content": "what does Calc do?"} in prompt
    # The notebook facts came from the kernel, not from what the conversation remembered.
    assert driver.graphs_read == 1


async def test_a_workspace_without_a_kernel_still_answers() -> None:
    client = ScriptedClient(
        [[*call("list_cells", {})], [ChatDelta(text="I cannot see the notebook.")]]
    )
    events = await drive(build(client))
    result = next(event for event in events if event.type == "tool_result")
    assert result.ok is False
    assert [event.type for event in events].count("done") == 1


async def test_each_session_gets_its_own_tool_registry() -> None:
    agent = build(ScriptedClient([[ChatDelta(text="a")], [ChatDelta(text="b")]]))
    await drive(agent)
    request = TurnRequest(turn_id="t2", session_id="s2", message="hello")
    [event async for event in agent.run(request, cancel=CancelToken())]
    assert set(agent.registries) == {"s1", "s2"}
    agent.reset_session("s1")
    assert set(agent.registries) == {"s2"}


async def test_tokens_used_is_reported_on_the_done_event() -> None:
    events = await drive(build(ScriptedClient([[ChatDelta(text="a fairly long answer " * 20)]])))
    done = events[-1]
    assert done.type == "done"
    assert done.tokens_used > 0


# ── the server, driven by this loop rather than by the scripted stand-in ──────────────────────


def test_the_server_streams_this_loop_where_it_streamed_the_scripted_agent(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The definition-of-done claim, exercised rather than asserted in prose.

    ``create_app`` is handed a :class:`KedgeAgent` in exactly the place ``create_demo_app`` hands
    it a ``ScriptedAgent``, and the SSE stream that comes back is parsed frame by frame. If the
    loop ever drifts from the seam's Protocol this is the test that notices.
    """
    from fastapi.testclient import TestClient
    from openpyxl import Workbook as OpenpyxlWorkbook

    from kedge.server.app import create_app
    from kedge.server.sessions import SessionStore
    from kedge.workspace import Workspace

    monkeypatch.setenv("KEDGE_HOME", str(tmp_path / "home"))
    workbook = tmp_path / "process.xlsx"
    book = OpenpyxlWorkbook()
    sheet = book.active
    sheet.title = "Calc"
    sheet.append(["haircut"])
    sheet.append([0.02])
    book.save(workbook)

    workspace = Workspace.for_workbook(workbook)
    workspace.ensure_dirs()
    client = ScriptedClient(
        [
            [
                ChatDelta(text="The haircut column is a flat rate.\n"),
                *call("read_range", {"sheet": "Calc", "range": "A1:A2"}),
            ],
            [ChatDelta(text="Two rows, header on row 1.")],
        ]
    )
    agent = KedgeAgent(
        client=client,
        context=ToolContext.for_workspace(workspace),
        counter=TokenCounter(allow_download=False),
    )
    app = create_app(workspace, agent=agent, store=SessionStore(tmp_path / "sessions.sqlite"))

    with TestClient(app) as http:
        session = http.post("/api/sessions", json={"title": "smoke"}).json()["session"]["id"]
        with http.stream(
            "POST", f"/api/sessions/{session}/turns", json={"message": "describe Calc"}
        ) as response:
            assert response.status_code == 200
            body = "".join(response.iter_text())

    frames = [line[len("event: ") :] for line in body.splitlines() if line.startswith("event: ")]
    assert frames[-1] == "done"
    assert frames.count("done") == 1
    assert "tool_call" in frames and "tool_result" in frames

    # The value reached the model, and only the model: the SSE trail carries a summary, never
    # the payload, so a browser tab and the session transcript never hold workbook data.
    assert "0.02" in json.dumps(client.seen[1])
    assert "0.02" not in body

    # And the audit line for that payload landed under the machine directory, values excluded.
    logs = sorted((workspace.logs_dir).glob("outbound-*.jsonl"))
    assert len(logs) == 1
    recorded = logs[0].read_text(encoding="utf-8")
    assert '"tool":"read_range"' in recorded
    assert "0.02" not in recorded


def test_a_paused_turn_resumes_through_the_server_with_what_it_had_learnt(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The pause and the resume as two real requests, against the history the server itself wrote.

    The other tests hand the loop a history they made up. This one lets the server store it, which
    is the version that has to agree: ``_history_before`` trims the server's flattened record of
    the very turn the carried messages are about to put back whole.
    """
    from fastapi.testclient import TestClient
    from openpyxl import Workbook as OpenpyxlWorkbook

    from kedge.server.app import create_app
    from kedge.server.sessions import SessionStore
    from kedge.workspace import Workspace

    monkeypatch.setenv("KEDGE_HOME", str(tmp_path / "home"))
    workbook = tmp_path / "process.xlsx"
    book = OpenpyxlWorkbook()
    sheet = book.active
    sheet.title = "Calc"
    sheet.append(["haircut"])
    sheet.append([0.02])
    book.save(workbook)

    workspace = Workspace.for_workbook(workbook)
    workspace.ensure_dirs()
    reading = [
        ChatDelta(text="Reading it.\n"),
        *call("read_range", {"sheet": "Calc", "range": "A1:A2"}),
    ]
    client = ScriptedClient([reading, reading, [ChatDelta(text="Two rows, header on row 1.")]])
    agent = KedgeAgent(
        client=client,
        context=ToolContext.for_workspace(workspace),
        counter=TokenCounter(allow_download=False),
        max_steps=2,
    )
    app = create_app(workspace, agent=agent, store=SessionStore(tmp_path / "sessions.sqlite"))

    with TestClient(app) as http:
        session = http.post("/api/sessions", json={"title": "smoke"}).json()["session"]["id"]
        turns = f"/api/sessions/{session}/turns"
        with http.stream("POST", turns, json={"message": "describe Calc"}) as response:
            first = "".join(response.iter_text())
        with http.stream("POST", turns, json={"message": "continue"}) as response:
            second = "".join(response.iter_text())

    assert "event: paused" in first
    assert "event: paused" not in second

    resumed = client.seen[2]
    assert [message["role"] for message in resumed].count("tool") == 2
    # Carried whole, not carried alongside the flattened record of the same turn.
    assert [message.get("content") for message in resumed].count("describe Calc") == 1


def test_an_answered_turn_carries_through_the_server_without_being_read_twice(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The answered equivalent, and the case ``_history_before`` has the most to get wrong.

    A paused turn leaves the server a user message and whatever prose had streamed. An answered
    one leaves a user message and the whole answer — and the answer is in the carry too, because
    the assistant message is recorded before the step loop returns. Trimming only the opener would
    hand the model its own answer twice, once flattened out of order and once in its place.
    """
    from fastapi.testclient import TestClient
    from openpyxl import Workbook as OpenpyxlWorkbook

    from kedge.server.app import create_app
    from kedge.server.sessions import SessionStore
    from kedge.workspace import Workspace

    monkeypatch.setenv("KEDGE_HOME", str(tmp_path / "home"))
    workbook = tmp_path / "process.xlsx"
    book = OpenpyxlWorkbook()
    sheet = book.active
    sheet.title = "Calc"
    sheet.append(["haircut"])
    sheet.append([0.02])
    book.save(workbook)

    workspace = Workspace.for_workbook(workbook)
    workspace.ensure_dirs()
    answer = "Two rows, header on row 1."
    client = ScriptedClient(
        [
            [
                ChatDelta(text="Reading it.\n"),
                *call("read_range", {"sheet": "Calc", "range": "A1:A2"}),
            ],
            [ChatDelta(text=answer)],
            [ChatDelta(text="Starting on it.")],
        ]
    )
    agent = KedgeAgent(
        client=client,
        context=ToolContext.for_workspace(workspace),
        counter=TokenCounter(allow_download=False),
    )
    app = create_app(workspace, agent=agent, store=SessionStore(tmp_path / "sessions.sqlite"))

    with TestClient(app) as http:
        session = http.post("/api/sessions", json={"title": "smoke"}).json()["session"]["id"]
        turns = f"/api/sessions/{session}/turns"
        with http.stream("POST", turns, json={"message": "describe Calc"}) as response:
            "".join(response.iter_text())
        with http.stream("POST", turns, json={"message": "now build the notebook"}) as response:
            "".join(response.iter_text())

    second = client.seen[2]
    contents = [message.get("content") for message in second]
    # The workbook read came with it rather than being fetched again.
    assert any("0.02" in (content or "") for content in contents)
    # And the turn appears once: not flattened in history and reinstated in the carry as well.
    assert contents.count("describe Calc") == 1
    assert contents.count(answer) == 1
    assert contents[-1] == "now build the notebook"


def test_a_repeated_message_through_the_server_does_not_read_a_turn_twice(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Two turns saying the same thing, which the carry makes routine rather than exotic.

    Every turn now hands its traffic on, so a span of two legs needs nothing more than two
    consecutive turns — and ``_pause_message`` asks the user to type "continue", which makes the
    colliding text the one kedge itself suggests. Trimming history by searching backwards for the
    span's opening message stops at the *newer* of the two copies and leaves the first turn in the
    prompt twice: once flattened, once in the span.
    """
    from fastapi.testclient import TestClient
    from openpyxl import Workbook as OpenpyxlWorkbook

    from kedge.server.app import create_app
    from kedge.server.sessions import SessionStore
    from kedge.workspace import Workspace

    monkeypatch.setenv("KEDGE_HOME", str(tmp_path / "home"))
    workbook = tmp_path / "process.xlsx"
    book = OpenpyxlWorkbook()
    sheet = book.active
    sheet.title = "Calc"
    sheet.append(["haircut"])
    sheet.append([0.02])
    book.save(workbook)

    workspace = Workspace.for_workbook(workbook)
    workspace.ensure_dirs()
    client = ScriptedClient(
        [
            [*call("read_range", {"sheet": "Calc", "range": "A1:A2"}, index=0)],
            [ChatDelta(text="ANSWER ONE")],
            [*call("read_range", {"sheet": "Calc", "range": "A1:A2"}, index=1)],
            [ChatDelta(text="ANSWER TWO")],
            [ChatDelta(text="Starting on it.")],
        ]
    )
    agent = KedgeAgent(
        client=client,
        context=ToolContext.for_workspace(workspace),
        counter=TokenCounter(allow_download=False),
    )
    app = create_app(workspace, agent=agent, store=SessionStore(tmp_path / "sessions.sqlite"))

    with TestClient(app) as http:
        session = http.post("/api/sessions", json={"title": "smoke"}).json()["session"]["id"]
        turns = f"/api/sessions/{session}/turns"
        for message in ("carry on", "carry on", "and now the notebook"):
            with http.stream("POST", turns, json={"message": message}) as response:
                "".join(response.iter_text())

    contents = [message.get("content") for message in client.seen[4]]
    assert contents.count("ANSWER ONE") == 1
    assert contents.count("ANSWER TWO") == 1
    assert contents.count("carry on") == 2, "one per leg the span reinstates, and no more"
    assert contents[-1] == "and now the notebook"


def test_a_turn_that_read_the_workbook_and_the_kernel_carries_each_as_it_deserves(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The mixed case as the server produces it, rather than through the loop's own front door.

    The workbook is a file kedge never writes, so what was read of it is still true next turn. The
    kernel is a thing the user edits in the pane beside the chat, so what was read of it is not.
    Both messages stay — the endpoint rejects a request whose assistant message carries a call with
    no result against it — and only one of them still says anything.
    """
    from fastapi.testclient import TestClient
    from openpyxl import Workbook as OpenpyxlWorkbook

    from kedge.server.app import create_app
    from kedge.server.sessions import SessionStore
    from kedge.workspace import Workspace

    monkeypatch.setenv("KEDGE_HOME", str(tmp_path / "home"))
    workbook = tmp_path / "process.xlsx"
    book = OpenpyxlWorkbook()
    sheet = book.active
    sheet.title = "Calc"
    sheet.append(["haircut"])
    sheet.append([0.02])
    book.save(workbook)

    workspace = Workspace.for_workbook(workbook)
    workspace.ensure_dirs()
    client = ScriptedClient(
        [
            [
                *call("read_range", {"sheet": "Calc", "range": "A1:A2"}, index=0),
                *call("list_cells", {}, index=1),
            ],
            [ChatDelta(text="Two rows, and two cells so far.")],
            [ChatDelta(text="Starting on it.")],
        ]
    )
    agent = KedgeAgent(
        client=client,
        context=ToolContext.for_workspace(workspace, driver=FakeDriver()),
        counter=TokenCounter(allow_download=False),
    )
    app = create_app(workspace, agent=agent, store=SessionStore(tmp_path / "sessions.sqlite"))

    with TestClient(app) as http:
        session = http.post("/api/sessions", json={"title": "smoke"}).json()["session"]["id"]
        turns = f"/api/sessions/{session}/turns"
        for message in ("describe Calc", "now build the notebook"):
            with http.stream("POST", turns, json={"message": message}) as response:
                "".join(response.iter_text())

    carried = _tools_in(client.seen[2])
    assert set(carried) == {"call_0", "call_1"}, "every call is still answered"
    assert "0.02" in carried["call_0"], "the workbook cannot have changed under kedge"
    assert "import polars as pl" not in carried["call_1"], "the kernel can, and may have"
    assert "list_cells" in carried["call_1"]
    assert any(tail in carried["call_1"] for tail in _EVICTION_TAILS)


class EndpointRefusedError(KedgeError):
    """A kedge-shaped failure from the model endpoint, for the recoverable-error path."""

    def __init__(self) -> None:
        super().__init__("the model endpoint refused the request: no API key is configured")


# ── what a turn cost ─────────────────────────────────────────────────────────────────────────


def _done(events: list[Any]) -> Any:
    return next(event for event in events if isinstance(event, DoneEvent))


async def test_the_turn_total_is_the_whole_turn_rather_than_its_last_step() -> None:
    # Every step re-sends the whole prompt, so a turn is the sum of its steps. Assigning the
    # latest step's figure understated a long turn by as much as that multiple.
    one = await drive(build(ScriptedClient([[ChatDelta(text="done")]]), driver=FakeDriver()))
    three = await drive(
        build(
            ScriptedClient(
                [
                    call("list_cells", {}),
                    call("list_cells", {}, index=1),
                    [ChatDelta(text="done")],
                ]
            ),
            driver=FakeDriver(),
        )
    )
    assert _done(three).tokens_used > 2 * _done(one).tokens_used


async def test_the_endpoints_own_count_is_preferred_to_kedges_estimate() -> None:
    # kedge counts with a fixed encoding that is wrong for most current models, and cannot see
    # the cache at all. Where the endpoint reports, its numbers are the ones that are used.
    reported = Usage(prompt=9_000, completion=40, cached=8_600)
    events = await drive(
        build(ScriptedClient([[ChatDelta(text="done"), ChatDelta(usage=reported)]]))
    )
    assert _done(events).tokens_used == reported.total


async def test_a_step_the_endpoint_did_not_report_still_contributes_its_estimate() -> None:
    # Mixed is possible. The reported step must not also be estimated, and the silent one must
    # not vanish.
    reported = Usage(prompt=5_000, completion=10)
    events = await drive(
        build(
            ScriptedClient(
                [call("list_cells", {}), [ChatDelta(text="done"), ChatDelta(usage=reported)]]
            ),
            driver=FakeDriver(),
        )
    )
    total = _done(events).tokens_used
    assert total > reported.total


# ── which model actually ran ─────────────────────────────────────────────────────────────────


async def test_the_model_recorded_is_the_model_the_turn_ran() -> None:
    """A per-session override runs, so a per-session override is what gets stamped on artifacts.

    `[model] model` is a default: `routes.py` sends `body.model or session.model`, and the user
    can change the session's model while the chat is open. Anything the turn produces and dates —
    a plan, most of all — is read months later by someone asking which model wrote it and whether
    that model is still trusted. A wrong id ends that question instead of answering it, which is
    worse than no id at all.
    """
    client = ScriptedClient([[ChatDelta(text="done")]])
    agent = build(client, driver=FakeDriver(), model="config-default")

    await drive(agent, model="chosen-for-this-session")

    assert client.models_asked == ["chosen-for-this-session"]
    assert agent.registries["s1"].model == "chosen-for-this-session"


async def test_config_is_the_fallback_when_the_turn_names_no_model() -> None:
    client = ScriptedClient([[ChatDelta(text="done")]])
    agent = build(client, driver=FakeDriver(), model="config-default")

    await drive(agent)

    assert client.models_asked == ["config-default"]
    assert agent.registries["s1"].model == "config-default"


# ── a draft the model cannot get right ───────────────────────────────────────────────────────


async def test_a_draft_a_tool_keeps_rejecting_stops_the_turn(
    analysis: WorkbookAnalysis, tmp_path: Path
) -> None:
    """The cell path caps at three, and this is the same failure wearing different clothes.

    Uncapped, a model that cannot produce a valid plan spends the whole step budget on it — up to
    `max_steps` completions, each re-sending the whole tool surface — and the turn ends in a pause
    that reads as though progress were being made. The cap ends it while the useful part is still
    worth having, and says what that is.
    """
    rounds = [call("inspect_workbook", {"section": "operations"})]
    rounds += [call("propose_plan", {"plan": "here you go: {stages: ["}) for _ in range(6)]
    agent = build(
        client := ScriptedClient(rounds),
        context=ToolContext(analysis=analysis, plans=PlanStore(tmp_path / "plans")),
        max_steps=10,
    )

    events = await drive(agent)
    types = [event.type for event in events]

    errors = [event for event in events if event.type == "error"]
    assert len(errors) == 1
    assert "rejected the draft 3 times" in errors[0].message
    assert errors[0].recoverable is True
    # Written for the user, who reads it as a warning chip, and written in the past tense about
    # what happened. The instruction that used to sit here was addressed to the model, which never
    # saw it; the account it asks for is best-effort, so the chip must not promise one.
    assert "The rejections are above" in errors[0].message
    assert "follows" not in errors[0].message
    assert types.count("tool_call") == 4  # the read, then three attempts and no more
    # The fifth request is the last word and calls nothing, so the tool call it comes back with is
    # dropped rather than dispatched -- and the sixth request is never made.
    assert len(client.seen) == 5, "it stopped rather than spending the rest of the budget"
    assert types.count("done") == 1


async def test_a_refusal_the_model_cannot_fix_is_not_counted_as_an_attempt(
    analysis: WorkbookAnalysis, tmp_path: Path
) -> None:
    # The read gate refuses, and no amount of redrafting changes that: the model has to go and
    # read something. Counting it would stop the turn with advice about the wrong problem.
    rounds = [call("propose_plan", {"plan": '{"stages": [], "open_questions": []}'})] * 4
    rounds += [[ChatDelta(text="Let me read the workbook first.")]]
    agent = build(
        ScriptedClient(rounds),
        context=ToolContext(analysis=analysis, plans=PlanStore(tmp_path / "plans")),
        max_steps=10,
    )

    events = await drive(agent)

    assert not any(event.type == "error" for event in events)
    assert [event.type for event in events].count("tool_call") == 4


# ── the last word a stopped turn still gets ──────────────────────────────────────────────────


def _rejected_drafts(count: int) -> list[list[ChatDelta]]:
    """A read, then ``count`` rounds of a plan `propose_plan` cannot parse.

    The read is not decoration. `propose_plan` refuses outright until something has been read this
    session, and that refusal is not counted as an attempt (no redraft fixes it), so a script that
    opens with the draft never reaches the cap at all.
    """
    return [call("inspect_workbook", {"section": "operations"})] + [
        call("propose_plan", {"plan": "here you go: {stages: ["}) for _ in range(count)
    ]


def _stopped_by_a_draft_it_cannot_fix(
    rounds: list[list[ChatDelta]], analysis: WorkbookAnalysis, plans: Path
) -> tuple[KedgeAgent, ScriptedClient]:
    """A loop that reads, hits the draft cap on its next three rounds, then does what ``rounds`` says.

    The last word is therefore the fifth request, and `client.seen[4]` is the one that asks for it.
    """
    client = ScriptedClient(_rejected_drafts(3) + rounds)
    agent = build(
        client,
        context=ToolContext(analysis=analysis, plans=PlanStore(plans)),
        max_steps=10,
    )
    return agent, client


async def test_a_stopped_turn_still_ends_in_prose_the_user_can_read(
    analysis: WorkbookAnalysis, tmp_path: Path
) -> None:
    """The defect: a turn could end having produced no answer at all, and say nothing about it.

    A user spent two full 50-step turns having a workbook read and got back a trail of tool calls,
    a warning chip and not one sentence. The cap that stopped the turn was right to stop it — the
    fourth attempt at a shape the model has got wrong three times is not the one that works — but
    it ended the turn before the model was ever given a completion in which to write down what it
    had worked out, which is the only part of the turn the user can act on.

    The error still comes first. The user has to know why it stopped before they read the account,
    or the account arrives as prose that trails off for no stated reason.
    """
    agent, _ = _stopped_by_a_draft_it_cannot_fix(
        [[ChatDelta(text="Stage 1 loads the hand-in. "), ChatDelta(text="Stage 2 is the lookup.")]],
        analysis,
        tmp_path / "plans",
    )

    events = await drive(agent)
    types = [event.type for event in events]

    assert types.index("error") < types.index("token"), "why it stopped, then what was worked out"
    account = "".join(event.text for event in events if event.type == "token")
    assert account == "Stage 1 loads the hand-in. Stage 2 is the lookup."
    assert types.count("done") == 1


async def test_the_last_word_is_asked_for_with_no_tools_at_all(
    analysis: WorkbookAnalysis, tmp_path: Path
) -> None:
    """Withheld rather than discouraged, because the cap fires where asking nicely has failed.

    Left the tool surface, a model that has just been told to stop drafting has an obvious fourth
    attempt in front of it, and the turn ends silently again. Offered nothing, the only move left
    is to speak — and an empty sequence is sent as no ``tools`` key at all, since an
    OpenAI-compatible endpoint is entitled to refuse ``"tools": []``.
    """
    agent, client = _stopped_by_a_draft_it_cannot_fix(
        [[ChatDelta(text="Here is what the process does.")]], analysis, tmp_path / "plans"
    )

    await drive(agent)

    assert client.tools_offered[0], "the turn itself is worked with the whole surface"
    assert client.tools_offered[-1] == []
    assert len(client.tools_offered) == 5


async def test_the_last_word_is_spoken_to_one_request_rather_than_stored(
    analysis: WorkbookAnalysis, tmp_path: Path
) -> None:
    """The instruction is true of that one request and false of every later one.

    Kept in the window it would be carried into the next turn, where the budget is whole and "there
    is no further step" is a lie. Kept as a *user* message it would be counted by `_history_before`
    as an exchange the span reinstates, and the trim would cut the wrong turn out of history.
    """
    agent, client = _stopped_by_a_draft_it_cannot_fix(
        [[ChatDelta(text="Here is what the process does.")], [ChatDelta(text="Right.")]],
        analysis,
        tmp_path / "plans",
    )

    await drive(agent)
    await drive(agent, "try that again")

    asked_for = {"role": "user", "content": _FINAL_WORD_PROMPT}
    assert client.seen[4][-1] == asked_for, "the request that asks for prose closes with it"
    assert asked_for not in client.seen[5], "and no request after it carries it at all"


async def test_the_last_word_happens_once_and_does_not_re_enter_the_step_loop(
    analysis: WorkbookAnalysis, tmp_path: Path
) -> None:
    """One extra completion, whatever the model does with it.

    A model handed no tools may call one anyway. Dispatching it would put the turn back in the loop
    the cap exists to leave, so the calls are ignored — and dropped from the window as well as from
    the loop, because an assistant message carrying ``tool_calls`` that nothing answers is rejected
    outright by the endpoint, which would break every later turn of the session rather than this one.
    """
    agent, client = _stopped_by_a_draft_it_cannot_fix(
        [
            call("propose_plan", {"plan": "still not json"}, index=7),
            [ChatDelta(text="Right.")],
        ],
        analysis,
        tmp_path / "plans",
    )

    events = await drive(agent)
    types = [event.type for event in events]
    # Tool calls and no prose is one of the ways the account can go missing, so the line that says
    # so stands in for it rather than the turn ending silently after all.
    assert _NO_ACCOUNT in "".join(event.text for event in events if event.type == "token")
    assert types.count("tool_call") == 4, (
        "the read and three attempts; the ignored call ran nothing"
    )
    assert len(client.seen) == 5, "the read, three attempts and one last word"

    await drive(agent, "try that again")
    carried = client.seen[5]
    asked = {
        tool_call["id"] for message in carried for tool_call in message.get("tool_calls") or ()
    }
    answered = {message["tool_call_id"] for message in carried if message["role"] == "tool"}
    assert asked and asked == answered, "the ignored call was never recorded"
    assert "call_7" not in asked


async def test_the_account_is_recorded_so_the_next_turn_still_reads_it(
    analysis: WorkbookAnalysis, tmp_path: Path
) -> None:
    """The one piece of state the last word writes, and nothing else would notice its loss.

    `_history_before` trims the persisted exchange out of history on the understanding that the
    carried span reinstates it. An account the window never recorded is therefore not replaced by
    anything: it appears in the next turn's prompt zero times, and the model loses its own last
    answer — the only message of the turn the user actually read.

    Exactly once, too. Carried *and* left in history, it would be read twice, once flattened and
    once in place.
    """
    account = "Stage 1 loads the hand-in; stage 2 is the haircut lookup in Calc!H."
    agent, client = _stopped_by_a_draft_it_cannot_fix(
        [[ChatDelta(text=account)], [ChatDelta(text="Right.")]], analysis, tmp_path / "plans"
    )

    await drive(agent)
    # What the server persists of a stopped turn: the question, and the prose that reached the user.
    history = (
        TurnMessage(role="user", content="convert the haircut lookup"),
        TurnMessage(role="assistant", content=account),
    )
    await drive(agent, "try that again", history=history)

    carried = [message.get("content") or "" for message in client.seen[5]]
    assert sum(1 for content in carried if account in content) == 1


async def test_a_last_word_that_fails_does_not_brand_the_turn_fatal(
    analysis: WorkbookAnalysis, tmp_path: Path
) -> None:
    """A best-effort extra call must not escalate a stop the turn has already explained.

    Anything raised inside the last word would reach the loop's catch-all, which reports every
    exception as unrecoverable and which `app.js` renders as **Fatal** — a banner saying the turn
    broke, under a chip saying it stopped for a reason it understood, on a turn whose conclusion
    was settled before the call was made. The account failing to arrive is not fatal to anything.
    """

    class FailsWhenTheToolsGoAway:
        """Reads, redrafts until the cap fires, then fails the request that asks for prose."""

        def __init__(self) -> None:
            self.rounds = 0

        async def stream(self, *, model: str, messages: Any, tools: Any) -> Any:
            del model, messages
            self.rounds += 1
            if not tools:
                raise ValueError("the SDK changed shape")
                yield  # pragma: no cover - unreachable, and what makes this a generator
            script = (
                call("inspect_workbook", {"section": "operations"})
                if self.rounds == 1
                else call("propose_plan", {"plan": "here you go: {stages: ["})
            )
            for delta in script:
                yield delta

    agent = build(
        FailsWhenTheToolsGoAway(),
        context=ToolContext(analysis=analysis, plans=PlanStore(tmp_path / "plans")),
        max_steps=10,
    )

    events = await drive(agent)
    errors = [event for event in events if event.type == "error"]

    assert [event.recoverable for event in errors] == [True], "the cap chip, and no Fatal banner"
    assert "rejected the draft" in errors[0].message
    assert _NO_ACCOUNT in "".join(event.text for event in events if event.type == "token")
    assert [event.type for event in events].count("done") == 1


async def test_an_account_cut_short_says_so_on_the_screen_and_in_the_record(
    analysis: WorkbookAnalysis, tmp_path: Path
) -> None:
    """Prose, then a failure — the shape the mid-answer translation exists for, arriving here.

    `reply.content` is truthy, so the empty-account fallback does not fire and the containment
    would otherwise swallow the failure into a WARNING: the account stops mid-word with no marker,
    and the user is left to guess. The record is the worse half. What arrived is recorded, carried
    into the next turn and persisted by the server as the turn's assistant message, so an unmarked
    fragment reads as a finished thought for as long as the transcript survives.
    """
    fragment = "Stage 1 loads the hand-in, stage 2 is the hair"
    died = AgentError(
        "the model endpoint at http://127.0.0.1:1/v1 accepted the request and then failed "
        "part-way through the answer: upstream closed the connection."
    )

    class FailsPartWayThroughTheAccount:
        """Reads, redrafts until the cap fires, then dies half-way through the account."""

        def __init__(self) -> None:
            self.rounds = 0
            self.seen: list[list[dict[str, Any]]] = []

        async def stream(self, *, model: str, messages: Any, tools: Any) -> Any:
            del model
            self.seen.append([dict(message) for message in messages])
            self.rounds += 1
            if not tools:
                yield ChatDelta(text=fragment)
                raise died
            script = (
                call("inspect_workbook", {"section": "operations"})
                if self.rounds == 1
                else call("propose_plan", {"plan": "here you go: {stages: ["})
            )
            for delta in script:
                yield delta

    client = FailsPartWayThroughTheAccount()
    agent = build(
        client,
        context=ToolContext(analysis=analysis, plans=PlanStore(tmp_path / "plans")),
        max_steps=10,
    )

    events = await drive(agent)
    on_screen = "".join(event.text for event in events if event.type == "token")

    assert on_screen.startswith(fragment), "what arrived is kept; it is not thrown away"
    assert "This account stops here" in on_screen
    assert "upstream closed the connection" in on_screen, "and why, in the endpoint's own words"
    assert [event.recoverable for event in events if event.type == "error"] == [True]

    # And the same text is what the next turn reads, rather than a fragment ending mid-word.
    await drive(agent, "try that again")
    recorded = [
        message.get("content") or ""
        for message in client.seen[5]
        if message["role"] == "assistant" and fragment in (message.get("content") or "")
    ]
    assert recorded == [on_screen]


async def test_a_last_word_with_nothing_in_it_says_so_rather_than_going_quiet(
    analysis: WorkbookAnalysis, tmp_path: Path
) -> None:
    """Silence is the defect, so the one path left that could produce it says a line instead.

    An endpoint that answers the last word with no content at all — or with tool calls and no prose,
    which is the same thing once they are dropped — would otherwise leave the user exactly where
    they started: a chip, a trail of tool calls, and nothing written. The line is worth having for
    what it says next, which is that the reading is still on the shelf.
    """
    agent, _ = _stopped_by_a_draft_it_cannot_fix([[]], analysis, tmp_path / "plans")

    events = await drive(agent)

    assert "".join(event.text for event in events if event.type == "token") == _NO_ACCOUNT
    assert [event.type for event in events].count("done") == 1


async def test_the_validation_cap_gets_the_same_last_word() -> None:
    """Both caps leave the user in the same place, so both owe them the same account.

    A cell the gate keeps rejecting is a different failure from a plan a tool keeps refusing, but
    what reaches the chat is identical: a stopped turn. The account of what the cell was for is
    what the user needs in order to write it themselves.
    """
    rounds = [[*call("propose_cell", {"name": "loader", "code": PANDAS_CELL})] for _ in range(3)]
    rounds += [[ChatDelta(text="The loader reads the hand-in; polars wants scan_csv here.")]]
    client = ScriptedClient(rounds)

    events = await drive(build(client, driver=FakeDriver()))
    types = [event.type for event in events]

    assert types.count("validation") == 3
    assert types.index("error") < types.index("token")
    assert "".join(event.text for event in events if event.type == "token").startswith("The loader")
    assert client.tools_offered[-1] == []


async def test_the_last_word_is_counted_like_any_other_step(
    analysis: WorkbookAnalysis, tmp_path: Path
) -> None:
    """It re-sends the whole prompt, so a turn's figure is only truthful if it is metered.

    Every step of a turn re-sends the head, the tool schemas and the conversation, and this one is
    a step like the others. Left out of `meter.record`, the ``DoneEvent`` would understate the turn
    by an entire prompt — and the turns that reach a cap are the long ones.
    """
    step = Usage(prompt=100, completion=10)
    account = Usage(prompt=9_000, completion=40)
    # The read and the three refused drafts each report the same modest figure; the last word
    # reports a whole prompt of its own, which is what it costs.
    rounds = [[*round_, ChatDelta(usage=step)] for round_ in _rejected_drafts(3)]
    rounds += [[ChatDelta(text="Here is the account."), ChatDelta(usage=account)]]
    agent = build(
        ScriptedClient(rounds),
        context=ToolContext(analysis=analysis, plans=PlanStore(tmp_path / "plans")),
        max_steps=10,
    )

    events = await drive(agent)

    assert _done(events).tokens_used == 4 * step.total + account.total


async def test_stop_during_the_last_word_is_honoured(
    analysis: WorkbookAnalysis, tmp_path: Path
) -> None:
    """The last word is one more model call, and Stop means stop during it too.

    It goes through `_complete` rather than opening a stream of its own precisely so that this
    holds: the call is abandoned mid-thought, and the turn leaves through the cancel path with its
    single `DoneEvent` and the tool traffic it paid for still on the shelf.
    """

    class StopWhenTheToolsGoAway:
        """Reads, redrafts until the cap fires, then thinks about the account while Stop is pressed.

        The round with nothing to call is the last word, so the token is set there — and then the
        fake *hangs*, which is the part that makes this a test rather than a formality. A fake that
        cancels and yields in the same breath is stopped by the ordinary check after the await,
        and would pass just as happily with `_abandon_if_cancelled` taken out; the 30 seconds of
        thinking against a 5 second deadline can only be survived by abandoning the call.
        """

        def __init__(self) -> None:
            self.token = CancelToken()
            self.tools_offered: list[list[str]] = []
            self.rounds = 0
            self.finished = False

        async def stream(self, *, model: str, messages: Any, tools: Any) -> Any:
            del model, messages
            self.tools_offered.append([tool["function"]["name"] for tool in tools])
            self.rounds += 1
            if not tools:
                self.token.cancel()
                await asyncio.sleep(30)  # still thinking about the account when Stop lands
                self.finished = True  # pragma: no cover - the turn is abandoned long before
                yield ChatDelta(text="never read")  # pragma: no cover
                return
            script = (
                call("inspect_workbook", {"section": "operations"})
                if self.rounds == 1
                else call("propose_plan", {"plan": "here you go: {stages: ["})
            )
            for delta in script:
                yield delta

    client = StopWhenTheToolsGoAway()
    agent = build(
        client,
        context=ToolContext(analysis=analysis, plans=PlanStore(tmp_path / "plans")),
        max_steps=10,
    )
    request = TurnRequest(turn_id="t1", session_id="s1", message="convert the haircut lookup")

    async def collect() -> list[Any]:
        return [event async for event in agent.run(request, cancel=client.token)]

    # Timed, and the deadline is not the assertion. `run` converts a cancellation into a tidy
    # finish by design, so a turn that sat out the whole 30 seconds -- or that had to be prised off
    # it by `wait_for` -- comes back looking exactly like one that abandoned the call immediately.
    # Only the clock tells them apart, and that difference is the whole subject of this test.
    clock = asyncio.get_running_loop().time
    started = clock()
    events = await asyncio.wait_for(collect(), timeout=5)
    elapsed = clock() - started
    types = [event.type for event in events]

    assert elapsed < 1, f"the turn waited {elapsed:.1f}s on a model call it should have abandoned"
    assert client.tools_offered[-1] == [], "the cap fired and the last word was asked for"
    assert client.finished is False, "the last word was abandoned, not waited out"
    assert not any(event.type == "token" for event in events), "the stop landed before the prose"
    errors = [event for event in events if event.type == "error"]
    assert errors[-1].message == "Turn cancelled at your request."
    assert types.count("done") == 1
    assert types[-1] == "done"


# ── a failure the endpoint reports once it has already started answering ─────────────────────


_CONTEXT_REFUSAL = "This model's maximum context length is 128000 tokens, however you requested 0"


def _bare_api_error() -> BaseException:
    """The exception `openai/_streaming.py` raises for an `error` event delivered after the 200.

    Its MRO is `(APIError, OpenAIError, Exception)`: not an `APIStatusError`, because there is no
    response left to carry a status, and not an `APIConnectionError`, because the connection is
    working — the endpoint is using it to say something. That is precisely why it used to slip past
    every clause in `OpenAIClient.stream` and land in the loop's catch-all.
    """
    from openai import APIError

    return APIError(
        _CONTEXT_REFUSAL,
        httpx.Request("POST", "http://127.0.0.1:1/v1/chat/completions"),
        body={"error": {"message": _CONTEXT_REFUSAL}},
    )


class FailingSDK:
    """Stands in for `AsyncOpenAI`: the stream opens cleanly, then fails while the body is drained.

    The position is the point. A gateway that refuses on context length, or loses its upstream,
    has usually already answered 200 and begun relaying, so the failure arrives among the chunks
    rather than around the request — where the SDK's own translation no longer reaches it.
    """

    def __init__(self, error: BaseException) -> None:
        self._error = error
        self.chat = SimpleNamespace(completions=SimpleNamespace(create=self._create))

    async def _create(self, **_payload: Any) -> Any:
        return self._chunks()

    async def _chunks(self) -> Any:
        delta = SimpleNamespace(content="Reading the reference sheet ", tool_calls=None)
        yield SimpleNamespace(choices=[SimpleNamespace(delta=delta)])
        raise self._error

    async def close(self) -> None:
        return None


def _failing_endpoint() -> OpenAIClient:
    return OpenAIClient(
        base_url="http://127.0.0.1:1/v1",
        api_key="k",
        api="chat_completions",
        client=FailingSDK(_bare_api_error()),
    )


async def test_a_bare_api_error_mid_answer_leaves_the_client_as_an_agent_error() -> None:
    """The invariant CLAUDE.md states: nothing the SDK raises may leave `stream` untranslated.

    `stream` had a clause for a timeout, one for a broken connection and one for a status error,
    and a mid-stream `APIError` is none of the three. The endpoint's own prose is what is quoted,
    not the SDK's wrapper around it.
    """
    with pytest.raises(AgentError) as caught:
        async for _ in _failing_endpoint().stream(
            model="m", messages=[{"role": "user", "content": "convert this"}], tools=[]
        ):
            pass

    assert "maximum context length" in str(caught.value)
    assert "part-way through the answer" in str(caught.value)
    # The endpoint's own prose, not the SDK's wrapper around it (CLAUDE.md). Asserted as the
    # absence of the wrapper rather than the presence of the phrase, because the phrase is inside
    # the wrapper too: interpolating `{exc!r}` passes a test that only looks for it.
    assert "Error code:" not in str(caught.value)
    assert "APIError(" not in str(caught.value)


async def test_a_failure_mid_answer_is_reported_as_recoverable_rather_than_fatal() -> None:
    """Which is the whole reason the translation matters, and it is not a capped-turn problem.

    Everything reaching the loop's catch-all is reported unrecoverable and rendered as **Fatal**,
    so before this an ordinary context-length refusal — mid-answer, on any turn — told the user
    their notebook might be half-written. It is recoverable and it says so, in the endpoint's words.
    """
    events = await drive(build(_failing_endpoint()))
    errors = [event for event in events if event.type == "error"]

    assert [event.recoverable for event in errors] == [True]
    assert "maximum context length" in errors[0].message
    assert [event.type for event in events].count("done") == 1
    # What had streamed before the failure still reached the user; it is all there is of it.
    assert "".join(event.text for event in events if event.type == "token") == (
        "Reading the reference sheet "
    )


# ── arguments kedge cannot decode ────────────────────────────────────────────────────────────


@cache
def nested_past_the_recursion_limit() -> str:
    """JSON nested deeper than the C scanner on *this* platform will follow.

    No single depth is deep enough everywhere, and hard-coding one is how these two tests came to
    pass on Windows and mean nothing on Linux. `Py_C_RECURSION_LIMIT` — what `json`'s C scanner
    spends one of per nesting level — is 3000 on Windows and 10000 elsewhere, so the 3000 this
    used to say was the Windows limit exactly: `RecursionError` there, a cleanly parsed list on
    Ubuntu, and a test asserting the raw string comes back that got the decoded one instead.
    CPython 3.14 drops the constant for a check against the real stack, so the next such number
    would go stale differently. Ask json where the edge is rather than predicting it.
    """
    depth = 1_000
    while depth <= 1_000_000:
        raw = "[" * depth + "1" + "]" * depth
        try:
            json.loads(raw)
        except RecursionError:
            return raw
        depth *= 2
    raise AssertionError("json followed a million levels of nesting; this test needs rethinking")


def test_arguments_nested_past_the_recursion_limit_are_handed_back_rather_than_raised() -> None:
    """`RecursionError` is not a `JSONDecodeError`, and everything reaching the loop's catch-all
    is reported to the user as Fatal — a whole turn ended and the notebook's state put in doubt,
    over one tool call kedge was only trying to summarise for the activity trail."""
    raw = nested_past_the_recursion_limit()

    assert _safe_arguments(raw) == {"arguments": raw}


async def test_a_tool_call_kedge_cannot_decode_does_not_end_the_turn() -> None:
    deep = nested_past_the_recursion_limit()
    rounds = [
        [
            ChatDelta(index=0, call_id="call_0", name="list_cells"),
            ChatDelta(index=0, arguments=deep),
        ],
        [ChatDelta(text="That argument was nonsense; here is what I found instead.")],
    ]
    events = await drive(build(ScriptedClient(rounds), driver=FakeDriver()))

    assert not any(event.type == "error" and event.recoverable is False for event in events), (
        "a malformed tool call is not a fatal turn"
    )
    assert [event.type for event in events].count("done") == 1


async def test_the_pinned_blocks_are_ordered_least_volatile_first() -> None:
    # A prompt cache keys on the prefix: whatever sits ahead of a block that changes stays cached
    # and whatever sits behind it does not. The analysis and the plan hold still for a session;
    # the registry and the notebook state change the moment a cell is created.
    client = ScriptedClient([[ChatDelta(text="done")]])
    await drive(build(client, driver=FakeDriver()))

    head = client.seen[0][0]["content"]
    order = [
        head.index("## Workbook analysis"),
        head.index("## Process plan"),
        head.index("## Name registry"),
        head.index("## Live notebook state"),
    ]
    assert order == sorted(order)
    assert head.index("SYSTEM PROMPT") < order[0]


def test_the_shape_recorded_for_an_eviction_stub_stays_inside_the_window_s_own_cap() -> None:
    """The stub has to cost a fraction of what it replaced, so what describes it must be short.

    It says the same handful of words whether the payload was one number or 32KB — a description
    that grew with the result would defeat the mechanism it belongs to — and the window re-caps it
    at render, so a field added here can crowd out what is already in the shape but can never grow
    the stub.
    """
    shapes = [
        _result_shape(ToolResult(text="x" * 40_000, row_count=49_999, truncated=True)),
        _result_shape(ToolResult(text="6 cells\nAAaa imports", summary="6 cells")),
        _result_shape(ToolResult.note("reconciliation failed on 14 of 400 rows " * 20)),
        _result_shape(ToolResult(text="")),
    ]
    assert all(len(shape) <= MAX_EVICTED_SHAPE_CHARS for shape in shapes), shapes
    assert "49999 rows" in shapes[0] and "truncated" in shapes[0]
    assert "6 cells" in shapes[1]
    # `ToolResult.note` defaults its summary to the payload's first line, so a fragment of the
    # payload reaches the stub. That is deliberate and model-bound: the stub is worth having only
    # if it says what is missing, and the model has already read the whole result.
    assert shapes[2].startswith("reconciliation failed")
