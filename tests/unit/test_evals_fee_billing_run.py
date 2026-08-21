"""The fee-billing-run workbook is still hard, and still hard in the ways it was built to be.

This eval's whole premise is that the workbook is *complex* -- see
``evals/proposals/fee_billing_run.md``. That makes its rot guard different in kind from
``test_evals_adjustment_signoff.py``, which pins figures. Here the thing that can rot silently
is the **structure**, and it can rot in a way no ordinary test would notice.

The mechanism is R1C1 normalisation, and it is the reason this file leads with a structural
band rather than a list of cells. A formula filled down a column and across thirty columns
normalises to one string, so thirty *identical* columns are one logical operation and thirty
*different* ones are thirty. A refactor that tidies ``WORKING_COLUMNS`` into something
uniform -- exactly the sort of change a reviewer would wave through as a simplification --
collapses a 49-operation workbook to single figures while leaving a file that opens fine,
analyses fine and grades fine against every other assertion in the suite. The eval would go on
passing and would have stopped measuring anything.

Both directions matter, hence a band rather than a floor. Too simple and the workbook stresses
nothing the existing corpus does not already stress. Too complex and it triages to ``stop``, at
which point it cannot be converted, so it cannot be graded on the quality of its conversion.
The measured position is recorded in the proposal and reproduced here.

The second half of the file is about the **rubric** rather than the workbook, and it is the
lesson ``adjustment_signoff`` paid for: a rubric quoting figures the generator no longer produces
is worse than no rubric, because it fails a correct conversion and sends somebody hunting a
defect that is not there. Its equivalent test has already caught that happening once. So every
figure in ``expected.yaml``'s ``facts`` is recomputed here from ``build_workbook.compute()`` --
never read back out of the workbook, which would make the assertion tautological -- and the
rubric's ids are checked against the graders in both directions, the way
``test_evals_harness.py`` does for the first eval.

This is not a corpus test: it asserts what the workbook *is*, not what the analyser should find
in general. See ``evals/README.md`` for why those are kept apart.
"""

from __future__ import annotations

import importlib.util
import random
import sys
from pathlib import Path
from types import ModuleType
from typing import Any

import pytest
import yaml

from kedge.analysis.analyse import analyse
from kedge.analysis.model import WorkbookAnalysis
from kedge.plan.propose import build_proposal_context
from kedge.plan.triage import TriageVerdict, complexity, triage

EVAL_ROOT = Path(__file__).resolve().parents[2] / "evals"
EVAL_DIR = EVAL_ROOT / "fee_billing_run"


def _load_generator() -> ModuleType:
    """Import this case's ``build_workbook`` under a name nothing else can claim.

    Every eval directory has a module of that name, and ``test_evals_adjustment_signoff.py``
    puts the *other* one into ``sys.modules`` as the bare name ``build_workbook``. Two modules
    with one name means whichever import ran first wins, so a plain ``import build_workbook``
    here silently binds to whichever test file pytest happened to collect first -- passing
    alone, failing in a full run, and taking the other eval's suite down with it when the order
    goes the other way. The first version of this file did exactly that.

    ``tests/fixtures/generate.py`` was renamed for this reason and the comment recording it
    sits at the top of ``evals/adjustment_signoff/build_workbook.py``. Loading from an explicit
    path under a qualified name is the version that does not depend on anybody else's import
    order, and it leaves the bare name free for the case that already uses it.
    """
    path = EVAL_DIR / "build_workbook.py"
    spec = importlib.util.spec_from_file_location("evals.fee_billing_run.build_workbook", path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _load_case() -> ModuleType:
    """Import ``evals/fee_billing_run/case.py``, the way ``test_evals_harness.py`` does.

    Imported inside the tests that need it rather than at module scope. Everything above the
    rubric section is about the workbook and holds whether or not the graders import cleanly,
    and ``case.py`` reaches for ``harness.model`` and for whatever the harness imports in turn;
    binding the whole file's collection to that would report a broken import as thirty failures
    about the workbook.
    """
    if str(EVAL_ROOT) not in sys.path:
        sys.path.insert(0, str(EVAL_ROOT))
    from fee_billing_run import case

    return case


evalgen = _load_generator()
WORKBOOK = EVAL_DIR / evalgen.WORKBOOK_NAME


@pytest.fixture(scope="module")
def analysis() -> WorkbookAnalysis:
    if not WORKBOOK.is_file():
        pytest.skip(f"{WORKBOOK.name} has not been generated")
    return analyse(WORKBOOK)


@pytest.fixture(scope="module")
def rubric() -> dict[str, Any]:
    return yaml.safe_load((EVAL_DIR / "expected.yaml").read_text(encoding="utf-8"))


@pytest.fixture(scope="module")
def facts(rubric: dict[str, Any]) -> dict[str, Any]:
    return rubric["facts"]


@pytest.fixture(scope="module")
def clients() -> list[Any]:
    """The client list the committed workbook was built from.

    ``build`` seeds from :data:`build_workbook.SEED`, so this is the same run rather than a
    similar one -- and it is generated rather than read back out of the workbook, because a
    figure checked against the file it was derived from asserts nothing at all.
    """
    return evalgen.build_clients(random.Random(evalgen.SEED))


@pytest.fixture(scope="module")
def rows(clients: list[Any]) -> list[dict[str, Any]]:
    return evalgen.compute(clients)


@pytest.fixture(scope="module")
def by_code(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {row["client"]: row for row in rows}


# =============================================================================
# THE STRUCTURAL BAND
# =============================================================================


def test_the_workbook_is_as_complex_as_the_eval_needs(analysis: WorkbookAnalysis) -> None:
    """The measured position, and the thing every other assertion here depends on."""
    low, high = evalgen.TARGET_OPERATIONS
    assert low <= len(analysis.operations) <= high, (
        f"{len(analysis.operations)} logical operations. This is the R1C1 collapse: check "
        f"whether WORKING_COLUMNS has been made uniform, or whether a column's references now "
        f"vary per row and so refuse to compress."
    )

    score = complexity(analysis)
    low, high = evalgen.TARGET_COMPLEXITY
    assert low <= score <= high, f"complexity {score:.3f}, band {evalgen.TARGET_COMPLEXITY}"

    patterns = {op.excel_pattern for op in analysis.operations}
    assert len(patterns) >= evalgen.TARGET_PATTERNS, sorted(p.value for p in patterns)


def test_the_workbook_can_still_be_converted(analysis: WorkbookAnalysis) -> None:
    """A workbook that triages to `stop` cannot be graded on its conversion.

    That is the ceiling on how hard this eval is allowed to get, and it is a real one: the
    corpus's own `hostile.xlsx` is over it.
    """
    result = triage(analysis)
    assert result.verdict is not TriageVerdict.STOP, result.blocker_lines()


def test_it_is_harder_than_anything_the_repository_has_converted(
    analysis: WorkbookAnalysis,
) -> None:
    """The point of the exercise, stated as an assertion.

    ``q2_accrual_adjustment.xlsx`` scores 0.368 and is the most complex workbook kedge has ever
    been asked to convert. If this one ever drops near it, the eval has stopped being a second
    data point and has become a second copy of the first.
    """
    assert complexity(analysis) > 0.55


# =============================================================================
# THE PLANTED DISCRIMINATIONS
# =============================================================================


def test_the_posting_column_is_a_dead_region_among_many(analysis: WorkbookAnalysis) -> None:
    """Discrimination 5, and the reason it is sharper here than in the first eval.

    `Post` must be a real operation with no consumer -- that is what makes dropping it
    plausible -- and it must be one of *many*, so that keeping it is a judgement rather than
    the only available answer to "did you keep the one dead region?".
    """
    carry = [op for op in analysis.operations if op.sheet == "Post"]
    assert len(carry) == 1, "the Post column should compress to exactly one operation"
    assert carry[0].downstream_ref_count == 0, "a manual carry has no consumer, by definition"

    dead = [f for f in analysis.findings if f.kind.value == "dead_region"]
    assert len(dead) >= 10, f"only {len(dead)} dead regions; the needle needs a haystack"


def test_the_posting_column_sorts_into_the_truncated_tail(analysis: WorkbookAnalysis) -> None:
    """Proposal section 2.1, as a regression test rather than an argument.

    `_operation_digest` ranks by fan-out, so every zero-fan-out region sorts last -- and a
    manual carry has zero fan-out *because* it is carried by a person. If this assertion ever
    fails it is good news: it means the ranking has learned to treat a generated-SQL region as
    something other than an abandoned column.
    """
    ranked = sorted(analysis.operations, key=lambda op: (-op.downstream_ref_count, op.id))
    carry = next(op for op in analysis.operations if op.sheet == "Post")
    position = [op.id for op in ranked].index(carry.id)
    assert position >= len(ranked) // 2, (
        "the posting column no longer sorts into the tail of the planner's context"
    )


def test_the_planner_sees_every_operation(analysis: WorkbookAnalysis) -> None:
    """The workbook is hard, not pathological.

    Past ``_MAX_OPERATIONS`` the digest silently drops its tail, and the tail is where the
    posting column lives -- so a workbook over the cap tests truncation rather than conversion
    quality. Those are different questions and the proposal keeps them apart: truncation is
    what a *variant* of this workbook is for.
    """
    context = build_proposal_context(analysis, triage(analysis))
    assert "operations_omitted" not in context, context.get("operations_omitted")


def _posting_statements() -> list[str]:
    """What ``Post!A`` actually holds: Excel's own text, cached alongside the formula."""
    from openpyxl import load_workbook

    sheet = load_workbook(WORKBOOK, data_only=True)["Post"]
    return [
        str(sheet.cell(row=row, column=1).value)
        for row in range(4, 4 + evalgen.CLIENTS)
        if sheet.cell(row=row, column=1).value is not None
    ]


def test_the_workbooks_own_generated_sql_is_broken_for_the_apostrophe_client() -> None:
    """The one place in this eval where reproducing the workbook faithfully is *wrong*.

    ``Post!A`` builds each INSERT with ``&``, which quotes nothing and escapes nothing, and one
    of the eighty-four client names carries an apostrophe -- so that statement is not valid SQL.
    It is asserted to be broken here for the same reason ``adjustment_signoff`` asserts it of
    ``O'Brien & Partners``: a conversion that renders the literal through :mod:`kedge.sql` has
    to be distinguishable from one that copied the workbook's text, and without a literal that
    concatenation gets wrong the two are the same file.
    """
    name = next(iter(evalgen.CLIENT_NAMES.values()))
    statements = _posting_statements()

    assert len(statements) == evalgen.CLIENTS
    assert "'" in name, "the planted name no longer needs escaping"

    broken = [text for text in statements if name in text]
    assert broken, f"no statement carries {name!r}"
    # Six quotes in a valid single-row statement: two each around the code, the name and the
    # month. The unescaped apostrophe makes seven, which is what the client rejects.
    assert broken[0].count("'") % 2 == 1, broken[0]
    assert all(text.count("'") % 2 == 0 for text in statements if name not in text), (
        "only the apostrophe client's statement should be unbalanced"
    )


def test_kedge_sql_renders_that_same_statement_correctly() -> None:
    """The fix the eval is looking for, pinned so its absence is a failing assertion."""
    from kedge.sql import literal, render

    name = next(iter(evalgen.CLIENT_NAMES.values()))
    rendered = render(
        "INSERT INTO fin.fee_invoice (client_code, client_name, period_month, fee_gbp) "
        "VALUES ({client_code}, {client_name}, {period_month}, {fee_gbp})",
        {
            "client_code": "00041",
            "client_name": name,
            "period_month": "2026-11",
            "fee_gbp": 18_500.0,
        },
    )

    assert literal(name) == "'" + name.replace("'", "''") + "'"
    assert literal(name) in rendered
    assert rendered.count("'") % 2 == 0
    assert "'00041'" in rendered, "a client code is text; 41 joins to nothing"


def test_the_posting_column_reads_nothing_that_was_dead(analysis: WorkbookAnalysis) -> None:
    """A guard on where ``Post!A`` gets the client name from, which is not a style question.

    ``Working!B`` holds the same name one lookup away and is far the more natural reference --
    and it is also one of the nineteen ``dead_region`` findings that discrimination 5 hides the
    manual carry among. Pointing the posting column at it takes the haystack to eighteen and the
    complexity to 0.698, so two figures the rubric quotes move because of a formula that reads
    better. The name comes off ``Entity Map`` instead, which ``Working!B`` already reads.
    """
    dead = {
        (finding.sheet, finding.location)
        for finding in analysis.findings
        if finding.kind.value == "dead_region"
    }
    name_column = f"{evalgen.LETTERS['client_name']}{evalgen.WORKING_FIRST_ROW}"

    assert ("Working", name_column) in dead, (
        f"Working!{name_column} is read by something now, so the dead-region count and the "
        f"complexity the rubric quotes have both moved"
    )


def test_the_extract_query_is_also_a_real_connection(analysis: WorkbookAnalysis) -> None:
    """Proposal section 3, which the generator promised and for a year did not keep.

    The extract appears twice: as the SQL somebody left in ``Positions`` rows 3-13, and as the
    query Excel stored when the connection was last refreshed. Two sources for the same claim
    is the point -- it is what lets "the extract is a step to hand over" be checked rather than
    inferred, and without it this workbook was *easier* than ``q2_accrual_adjustment.xlsx`` on
    the one axis the whole case exists to make harder.

    The newlines matter as much as the query. See ``build_workbook.restore_connection``: Excel
    re-encodes them on save into a form the reader does not decode, and a query that reads back
    as one unbroken line is what the planner would be shown.
    """
    assert len(analysis.connections) == 1, "no ODBC connection; see proposal section 3"

    connection = analysis.connections[0]
    assert connection.kind == "odbc"
    assert connection.command_type == "sql"
    assert connection.command is not None
    assert "fin.positions" in connection.command
    assert connection.command.splitlines() == evalgen.EXTRACT_SQL.splitlines(), (
        "the stored query no longer matches the one in the cells, or its newlines did not "
        "survive the round trip"
    )
    assert "DSN=FinanceWarehouse" in (connection.connection_string_redacted or "")
    assert "svc_finread" not in (connection.connection_string_redacted or ""), (
        "the credential should be redacted on the way out"
    )


def test_the_tier_lookup_is_an_approximate_match(analysis: WorkbookAnalysis) -> None:
    """Discrimination 1: a banded rate table, which is `join_asof` and not a join."""
    assert any(op.excel_pattern.value == "vlookup_approx" for op in analysis.operations), (
        "the banded tier lookup has gone; it is the highest-value translation risk here"
    )


def test_the_overrides_are_reachable_and_reasoned(analysis: WorkbookAnalysis) -> None:
    """Discrimination 4: three typed-over fees, each carrying a written reason.

    The reasons are the part that matters. An override with no reason is a number, and a
    runbook cannot re-ask a question nobody wrote down.
    """
    from openpyxl import load_workbook

    sheet = load_workbook(WORKBOOK)["Overrides"]
    reasons = [sheet.cell(row, 4).value for row in (2, 3, 4)]
    assert all(isinstance(r, str) and len(r) > 30 for r in reasons), reasons
    assert len(set(reasons)) == 3, "three overrides, three distinct reasons"


def test_the_subtotals_sit_inside_the_allocation_grid(analysis: WorkbookAnalysis) -> None:
    """Discrimination 6: an embedded subtotal is not data."""
    assert any(
        op.sheet == "Allocation" and op.excel_pattern.value == "subtotal"
        for op in analysis.operations
    )


def test_the_dependency_order_is_not_the_tab_order(analysis: WorkbookAnalysis) -> None:
    """Discrimination 10: `Entity Map` is the eighth tab and the fourth tab depends on it."""
    names = [sheet.name for sheet in analysis.sheets]
    assert names.index("Entity Map") > names.index("Working")


def test_the_process_notes_survive_into_the_analysis(analysis: WorkbookAnalysis) -> None:
    """The briefing has to come from somewhere, and it must cite it."""
    text = " ".join(note.text for note in analysis.notes)
    assert "manual calculation" in text or "Overrides" in text


# =============================================================================
# THE RECONCILIATION BASELINE
# =============================================================================


def test_the_workbook_carries_a_reconciliation_baseline(analysis: WorkbookAnalysis) -> None:
    """Without cached values the whole reconciliation half of the rubric is ungradeable.

    ``infer_regions`` proposes nothing, every region degrades to "not reconciled", and triage
    reports a blocker -- so items about whether a translation reproduces the workbook cannot be
    scored at all, in either direction.
    """
    from kedge.analysis.workbook import open_workbook
    from kedge.reconcile.baseline import infer_regions

    assert analysis.cached_values.cached_present_count > 3_000
    assert all(op.cached_values_present for op in analysis.operations)

    with open_workbook(WORKBOOK) as handle:
        regions = infer_regions(handle, analysis)
    assert len(regions) > 30, f"only {len(regions)} reconcilable regions"


def test_the_allocation_tab_is_stale_on_purpose(analysis: WorkbookAnalysis) -> None:
    """Discrimination 9, and the only place in the eval where the cached values are *wrong*.

    ``Allocation`` is left on manual calculation, so its figures predate the three agreed
    overrides. A conversion that reconciles against them and adjusts itself until they match has
    adopted numbers Excel itself would disown. The correct answer is to declare the region not
    reproduced, with the reason, and report CHECKED WITH EXCEPTIONS.

    The staleness is recorded in the Sign-off prose and nowhere else, because no analyser can
    see it in the cells -- which is exactly why the briefing has to survive the conversion.
    """
    from openpyxl import load_workbook

    values = load_workbook(WORKBOOK, data_only=True)
    working = values["Working"]
    allocation = values["Allocation"]

    agreed = evalgen.LETTERS["agreed_fee"]
    net = evalgen.LETTERS["net_fee"]
    codes = {code for code, _agreed, _reason, _on in evalgen.OVERRIDES}

    live = 0.0
    stale_expected = 0.0
    for offset in range(evalgen.CLIENTS):
        row = evalgen.WORKING_FIRST_ROW + offset
        live += working[f"{agreed}{row}"].value
        stale_expected += working[f"{net}{row}"].value

    # A client row carries a code in column A; a subtotal row does not. Reading the sheet rather
    # than recomputing the layout keeps this a check on the file and not on the generator.
    banked = sum(
        row[4].value
        for row in allocation.iter_rows(min_row=2, max_col=5)
        if row[0].value is not None
    )
    assert abs(banked - stale_expected) < 0.05, "Allocation should hold the pre-override figures"
    assert abs(banked - live) > 1.0, "Allocation must disagree with Working, or nothing is stale"
    assert codes, "the overrides are what make the two differ"


def test_an_empty_string_result_reads_as_an_uncalculated_cell(
    analysis: WorkbookAnalysis,
) -> None:
    """A defect in kedge, pinned here because this workbook is the first thing to reach it.

    ``=IF(agreed=net,"","OVERRIDE")`` returns an empty string on the 81 rows with no override.
    Excel stores that as ``<c t="str"><f>...</f><v/></c>`` -- verified by driving Excel, not
    assumed -- and openpyxl's ``data_only`` view hands back ``None`` for it, which
    ``values.add_formula`` counts as *no cached value*.

    So a workbook that is fully calculated reports partial coverage and a verification blocker,
    and the blocker's remediation is unfollowable: it says "recalculate and re-save in Excel",
    which cannot help, because Excel is what wrote the file. The ``t="str"`` attribute is the
    signal that would tell the two apart, and it is discarded before kedge sees it.

    When that is fixed this assertion inverts, and the eval will have paid for itself once.
    """
    coverage = analysis.cached_values
    missing = coverage.formula_cell_count - coverage.cached_present_count
    empty_string_rows = evalgen.CLIENTS - len(evalgen.OVERRIDES)
    assert missing == empty_string_rows, (
        f"{missing} formula cells report no cached value; expected exactly the "
        f"{empty_string_rows} rows whose override flag is an empty string"
    )


def test_the_tier_column_caches_text_a_conversion_must_still_reconcile_against() -> None:
    """Discrimination 8 where it is hardest: the *baseline* is text, not just the input.

    ``Working!J`` holds whatever ``VLOOKUP`` found, and for the seventeen clients on a
    negotiated rate the schedule cell it found is text -- verified by driving Excel, which is
    how anybody was going to notice. A conversion must type that column or the arithmetic below
    it fails four operations later inside a query plan; and a typed ``20.0`` against a cached
    ``'20.0'`` used to read as a type difference, which made the region unreconcilable *because*
    the conversion did the right thing.

    ``kedge.reconcile.model.as_numeric_pair`` draws the line at information loss rather than at
    type, so this region is now claimed rather than declared. The client codes on the same sheet
    are the other side of that line and must stay text: ``'00417'`` is not the number 417.
    """
    from openpyxl import load_workbook

    from kedge.reconcile.model import as_numeric_pair

    sheet = load_workbook(WORKBOOK, data_only=True)["Working"]
    column = evalgen.LETTERS["tier_bps"]
    cached = [
        sheet[f"{column}{evalgen.WORKING_FIRST_ROW + offset}"].value
        for offset in range(evalgen.CLIENTS)
    ]

    text = [value for value in cached if isinstance(value, str)]
    assert len(text) == 17, f"{len(text)} tier cells cache as text; the paste no longer arrives"
    assert all(as_numeric_pair(value, float(value)) is not None for value in text), (
        "a typed rate against the text Excel cached reads as a type difference again, so "
        "Working!J cannot be reconciled by a conversion that types the column"
    )
    assert as_numeric_pair("00417", 417.0) is None, "a client code is not a number"


def test_the_summary_tab_holds_a_real_pivot(analysis: WorkbookAnalysis) -> None:
    """Discrimination 7, and a rot guard on the one thing plain regeneration destroys.

    openpyxl cannot author a pivot table, so the committed workbook is finished by driving Excel
    (``--with-pivot``). That makes it the only part of the file a pure-Python ``build()`` does
    **not** reproduce -- run the generator without the flag and the pivot silently disappears,
    taking discrimination 7 with it and leaving a workbook that still opens, still analyses and
    still passes every other assertion here.

    So this reads the committed artifact rather than a fresh build, on purpose.
    """
    from openpyxl import load_workbook

    pivots = load_workbook(WORKBOOK)["Summary"]._pivots
    assert pivots, (
        "the Summary pivot has gone -- regenerate with "
        "`uv run --with pywin32 python evals/fee_billing_run/build_workbook.py --with-pivot`"
    )

    pivot = pivots[0]
    source = pivot.cache.cacheSource.worksheetSource
    assert source.sheet == "Allocation", "the pivot must aggregate the Allocation grid"
    names = [field.name for field in pivot.cache.cacheFields]
    assert [names[f.x] for f in pivot.rowFields] == ["legal_entity", "cost_centre"]
    assert [names[f.fld] for f in pivot.dataFields] == ["fee_gbp"]


def test_the_pivot_is_invisible_to_the_analyser(analysis: WorkbookAnalysis) -> None:
    """Proposal section 2.4, pinned on the eval's own workbook.

    A pivot writes its rendered result into the sheet as ordinary cached cells, so the tab does
    not read as empty -- it reads as *data*, which is to say as an input. A plan can therefore
    conclude the aggregated summary is a source table when it is a derived view of the tab beside
    it, and reading a derived aggregate as a source is how a total gets double-counted.

    When the section 7.1 extractor lands this inverts: the sheet becomes ``output``, the pivot
    becomes a ``LogicalOperation`` with ``ExcelPattern.PIVOT``, and it gains a dependency edge to
    ``Allocation``. Asserting the defect rather than skipping past it is what makes that day
    visible.
    """
    summary = analysis.sheet("Summary")
    assert summary is not None
    assert summary.role.value == "data", (
        f"Summary now classifies as {summary.role.value!r} -- if the pivot extractor has landed, "
        f"this assertion should be inverted to expect 'output'"
    )
    assert not [op for op in analysis.operations if op.sheet == "Summary"]


# =============================================================================
# THE RUBRIC QUOTES FIGURES THE GENERATOR PRODUCES
# =============================================================================
#
# `expected.yaml` opens by promising that every figure in `facts` is derived from
# `build_workbook.compute()` and never transcribed. This is where that promise is kept. The
# failure it guards against is asymmetric and nasty: a stale figure does not fail an incorrect
# conversion, it fails a *correct* one, and the person sent to find the defect is looking at code
# that does not have one.
#
# Every assertion recomputes from the generator rather than reading the workbook back. Reading
# the workbook would check that openpyxl can read what openpyxl wrote.

DERIVED_FACTS = frozenset(
    {
        "clients_billed",
        "period_start",
        "period_end",
        "days_in_month",
        "first_client_code",
        "last_client_code",
        "posting_table",
        "posting_statements",
        "fee_bands",
        "negotiated_rate_clients",
        "negotiated_rate_example",
        "gross_fee_total_gbp",
        "prorated_total_gbp",
        "net_fee_total_gbp",
        "agreed_fee_total_gbp",
        "discount_total_gbp",
        "running_total_final_gbp",
        "part_period_clients",
        "minimum_fee_gbp",
        "floored_clients",
        "prior_close_client",
        "prior_close_gbp",
        "overrides",
        "allocation_client_rows",
        "allocation_subtotal_rows",
        "entity_totals_gbp",
        "allocation_stale_total_gbp",
        "allocation_live_total_gbp",
        "allocation_stale_by_gbp",
        "logical_operations",
        "dead_regions",
        "reconcilable_regions",
        "complexity",
    }
)
"""Facts this file recomputes. Every one of them is a measurement of the workbook."""

JUDGED_FACTS = frozenset({"stage_count_band", "minimum_open_questions"})
"""Facts nothing can recompute, because they are judgements about what a good plan looks like.

`expected.yaml` says so of both: the stage band is "a judgement rather than a measurement", and
the question count is a proportion somebody chose. Listing them rather than ignoring them is what
lets :func:`test_every_figure_in_facts_is_checked_or_declared_a_judgement` be an exhaustive
check -- a figure added to the rubric is either recomputed here or explicitly none of this
file's business, and cannot be neither by accident.
"""


def test_every_figure_in_facts_is_checked_or_declared_a_judgement(facts: dict[str, Any]) -> None:
    """A new figure in the rubric is a new thing that can go stale."""
    quoted = set(facts)
    accounted = DERIVED_FACTS | JUDGED_FACTS

    assert quoted - accounted == set(), (
        f"figures in expected.yaml that nothing checks: {sorted(quoted - accounted)}. "
        f"Recompute each from build_workbook, or add it to JUDGED_FACTS with a reason."
    )
    assert accounted - quoted == set(), (
        f"figures checked here that the rubric no longer quotes: {sorted(accounted - quoted)}"
    )


def test_the_rubric_quotes_the_run_the_generator_produces(
    facts: dict[str, Any], rows: list[dict[str, Any]], clients: list[Any]
) -> None:
    """The shape of the run: how many clients, over which month, keyed how."""
    assert facts["clients_billed"] == len(rows) == evalgen.CLIENTS
    assert facts["period_start"] == evalgen.PERIOD_START
    assert facts["period_end"] == evalgen.PERIOD_END
    assert facts["days_in_month"] == evalgen.PERIOD_END.day
    assert {row["days_in_month"] for row in rows} == {facts["days_in_month"]}
    assert facts["first_client_code"] == rows[0]["client"]
    assert facts["last_client_code"] == rows[-1]["client"]

    statements = evalgen.cached_values(clients, rows)["Post"]
    assert facts["posting_statements"] == len(statements)
    posts_to = f"INSERT INTO {facts['posting_table']} "
    assert all(posts_to in text for text in statements.values()), (
        f"the generated statements no longer post to {facts['posting_table']!r}"
    )


def test_the_rubric_quotes_the_band_table_it_grades_against(
    facts: dict[str, Any], rows: list[dict[str, Any]], clients: list[Any], by_code: dict[str, Any]
) -> None:
    """Discrimination 1, and the figures an off-by-one at a band edge would move.

    The client counts are the point. A rubric can quote the rate card from the constants and
    still be useless; what catches a boundary that has moved from ``>=`` to ``>`` is how many of
    the 84 land in each band.
    """
    quoted = [dict(band) for band in facts["fee_bands"]]
    assert [(b["floor"], b["ceiling"], b["bps"]) for b in quoted] == [
        tuple(band) for band in evalgen.BANDS
    ]

    for band in quoted:
        landed = sum(1 for row in rows if row["band_bps"] == band["bps"])
        assert band["clients"] == landed, (
            f"the rubric says {band['clients']} clients at {band['bps']}bps; {landed} land there"
        )
    assert sum(band["clients"] for band in quoted) == len(rows)

    negotiated = [client for client in clients if client.negotiated_bps is not None]
    assert facts["negotiated_rate_clients"] == len(negotiated)

    example = dict(facts["negotiated_rate_example"])
    row = by_code[str(example["client"])]
    # The cell holds text -- discrimination 8 -- and the rubric quotes the number it means.
    assert float(row["tier_bps"]) == example["bps"]
    assert row["band_bps"] == example["band_bps"], (
        "the example client no longer has a negotiated rate that differs from its band"
    )


def test_the_rubric_quotes_the_totals_to_the_penny(
    facts: dict[str, Any], rows: list[dict[str, Any]]
) -> None:
    """Every one of these goes through Excel's 15-significant-digit collapse.

    Summed with :func:`generate.excel_round` rather than :func:`round` for the reason
    ``compute`` gives: a bare rounding is a penny out and the error propagates into the totals a
    conversion is graded on.
    """
    for figure, key in (
        ("gross_fee_total_gbp", "gross_fee"),
        ("prorated_total_gbp", "prorated"),
        ("net_fee_total_gbp", "net_fee"),
        ("agreed_fee_total_gbp", "agreed_fee"),
        ("discount_total_gbp", "discount"),
    ):
        total = evalgen.excel_round(sum(row[key] for row in rows), 2)
        assert facts[figure] == total, f"{figure}: rubric {facts[figure]}, generator {total}"

    assert facts["running_total_final_gbp"] == rows[-1]["running_total"]


def test_the_rubric_quotes_the_rows_the_discriminations_live_on(
    facts: dict[str, Any], rows: list[dict[str, Any]], by_code: dict[str, Any]
) -> None:
    """Pro-rating, the floor and the prior-row reference: discriminations 3, the floor and 2.

    Each is planted on named rows, and each is dead if the rows go: a seed change that gives
    every client a full month makes the pro-rating columns the identity, and the rubric would go
    on quoting three clients that no longer exist.
    """
    quoted = [dict(item) for item in facts["part_period_clients"]]
    part_period = [row for row in rows if row["days_billed"] != row["days_in_month"]]
    assert [item["client"] for item in quoted] == [row["client"] for row in part_period]
    for item, row in zip(quoted, part_period, strict=True):
        assert item["days_billed"] == row["days_billed"]
        assert item["gross_fee_gbp"] == row["gross_fee"]
        assert item["prorated_fee_gbp"] == row["prorated"]

    assert facts["minimum_fee_gbp"] == evalgen.MIN_FEE
    assert facts["floored_clients"] == [
        row["client"] for row in rows if row["floored"] > row["prorated"]
    ], "the minimum fee binds on different clients now, so the floor column has moved"

    prior = by_code[str(facts["prior_close_client"])]
    assert facts["prior_close_gbp"] == prior["prior_close"]
    assert prior["prior_close"] != prior["avg_aum"], (
        "the quoted row no longer shows the prior-row reference doing anything"
    )


def test_the_rubric_quotes_the_overrides_the_billing_manager_typed(
    facts: dict[str, Any], by_code: dict[str, Any]
) -> None:
    """Discrimination 4. Both fees matter, and so does the reason.

    ``computed_fee_gbp`` is what the workbook calculated and ``agreed_fee_gbp`` is what was
    typed instead; a rubric that lost the first could not tell a conversion that surfaced the
    decision from one that quietly adopted the number.
    """
    quoted = [dict(item) for item in facts["overrides"]]
    planted = {code: (agreed, reason) for code, agreed, reason, _on in evalgen.OVERRIDES}
    assert [item["client"] for item in quoted] == list(planted)

    for item in quoted:
        agreed, reason = planted[item["client"]]
        row = by_code[item["client"]]
        assert item["agreed_fee_gbp"] == agreed == row["agreed_fee"]
        assert item["computed_fee_gbp"] == row["net_fee"]
        assert item["reason_fragment"] in reason, (
            f"the rubric quotes {item['reason_fragment']!r}, which is no longer in the reason "
            f"the generator writes: {reason!r}"
        )
        assert row["agreed_fee"] != row["net_fee"], "an override that changes nothing is not one"


def test_the_rubric_quotes_the_allocation_figures_both_ways(
    facts: dict[str, Any], rows: list[dict[str, Any]], clients: list[Any], by_code: dict[str, Any]
) -> None:
    """Discriminations 6 and 9: the embedded subtotals, and the staleness they hide.

    The stale total is the *pre-override* one, because the tab is on manual calculation. Both
    figures and the gap between them are quoted, and a grader that could not tell them apart
    would report the dangerous answer -- a pass against numbers Excel itself would disown -- as
    a correct one.
    """
    layout = evalgen.allocation_rows(clients)
    assert facts["allocation_client_rows"] == sum(1 for _row, c, _e in layout if c is not None)
    assert facts["allocation_subtotal_rows"] == sum(1 for _row, c, _e in layout if c is None)
    assert facts["allocation_subtotal_rows"] == len(evalgen.ENTITIES)

    entity_of = {row["client"]: row["legal_entity"] for row in rows}
    for name, total in facts["entity_totals_gbp"].items():
        live = evalgen.excel_round(
            sum(row["agreed_fee"] for row in rows if entity_of[row["client"]] == name), 2
        )
        assert total == live, f"{name}: rubric {total}, generator {live}"

    stale = evalgen.excel_round(sum(row["net_fee"] for row in rows), 2)
    live = evalgen.excel_round(sum(row["agreed_fee"] for row in rows), 2)
    assert facts["allocation_stale_total_gbp"] == stale
    assert facts["allocation_live_total_gbp"] == live
    assert facts["allocation_stale_by_gbp"] == evalgen.excel_round(stale - live, 2)
    assert stale != live, "nothing is stale, so discrimination 9 has gone"
    assert facts["allocation_stale_by_gbp"] == evalgen.excel_round(
        sum(by_code[code]["net_fee"] - agreed for code, agreed, _r, _o in evalgen.OVERRIDES), 2
    ), "the gap is the three overrides and nothing else"


def test_the_rubric_quotes_the_structure_calibrate_measures(
    facts: dict[str, Any], analysis: WorkbookAnalysis
) -> None:
    """The four figures the rubric takes from ``--calibrate`` rather than from ``compute``.

    They are quoted in the rubric's own prose -- "at least three open questions on a workbook
    scoring 0.699" -- so a drift here makes a grader's failure message say something untrue
    about the file it is grading.
    """
    assert facts["logical_operations"] == len(analysis.operations)
    assert facts["dead_regions"] == len(
        [finding for finding in analysis.findings if finding.kind.value == "dead_region"]
    )
    assert facts["complexity"] == round(complexity(analysis), 3)

    from kedge.analysis.workbook import open_workbook
    from kedge.reconcile.baseline import infer_regions

    with open_workbook(WORKBOOK) as handle:
        regions = infer_regions(handle, analysis)
    assert facts["reconcilable_regions"] == len(regions)


def test_the_judged_figures_are_at_least_self_consistent(facts: dict[str, Any]) -> None:
    """Nothing can measure these, but a band the wrong way round is still a defect."""
    low, high = (int(bound) for bound in facts["stage_count_band"])
    assert 0 < low < high
    assert int(facts["minimum_open_questions"]) > 1, (
        "one open question satisfies `open_questions_warning` on any workbook, so a minimum of "
        "one grades nothing"
    )


# =============================================================================
# THE RUBRIC AND THE GRADERS AGREE
# =============================================================================
#
# The same pair of checks `test_evals_harness.py` runs over `adjustment_signoff`, which had them
# and this case did not. Both directions, because they fail differently. A grader with no rubric
# entry is a check whose reason was never written down. A rubric item with no grader is a
# promise the report never keeps -- and it does not show up as a failure, it shows up as an item
# quietly missing from a denominator nobody counts.


@pytest.mark.parametrize(
    ("tier", "graders"), [("deterministic", "DETERMINISTIC"), ("structural", "STRUCTURAL")]
)
def test_every_grader_has_a_rubric_entry(rubric: dict[str, Any], tier: str, graders: str) -> None:
    case = _load_case()
    declared = {entry["id"] for entry in rubric[tier]}
    implemented = set(getattr(case, graders))

    assert implemented <= declared, f"graders with no rubric entry: {implemented - declared}"


@pytest.mark.parametrize(
    ("tier", "graders"), [("deterministic", "DETERMINISTIC"), ("structural", "STRUCTURAL")]
)
def test_every_rubric_item_has_a_grader(rubric: dict[str, Any], tier: str, graders: str) -> None:
    case = _load_case()
    declared = {entry["id"] for entry in rubric[tier]}
    implemented = set(getattr(case, graders))

    assert declared <= implemented, f"rubric items with no grader: {declared - implemented}"


# =============================================================================
# DETERMINISM
# =============================================================================


def test_generating_twice_produces_identical_bytes(tmp_path: Path) -> None:
    """A workbook nobody can regenerate is a workbook nobody can change."""
    first, second = tmp_path / "a.xlsx", tmp_path / "b.xlsx"
    evalgen.build(first)
    evalgen.build(second)
    assert first.read_bytes() == second.read_bytes()


# =============================================================================
# MEASURING THE WORKBOOK MUST NOT DAMAGE IT
# =============================================================================
#
# `--calibrate` and `--verify-with-excel` both used to call `build()` first, which rewrites the
# file *without* the Summary pivot -- so asking either question destroyed discrimination 7, and
# the answer printed was about a workbook that no longer matched the committed one. Two
# assertions, because the hole has two doors: a measuring run must not build, and a building run
# must not silently drop a pivot it cannot put back.


def test_measuring_the_workbook_never_rebuilds_it(monkeypatch: pytest.MonkeyPatch) -> None:
    """The committed file is the artifact the eval grades; a rebuilt one is a different input."""

    def _refuse(_path: Path) -> None:
        raise AssertionError("a measuring run rebuilt the workbook")

    monkeypatch.setattr(evalgen, "build", _refuse)
    monkeypatch.setattr(evalgen, "calibrate", lambda _path: 0)
    monkeypatch.setattr(evalgen, "verify_with_excel", lambda _path: 0)
    monkeypatch.setattr(sys, "argv", ["build_workbook.py", "--calibrate"])

    assert evalgen.main() == 0

    monkeypatch.setattr(sys, "argv", ["build_workbook.py", "--verify-with-excel"])
    assert evalgen.main() == 0


def test_a_rebuild_refuses_to_drop_the_pivot_it_cannot_author(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """openpyxl reads a pivot table and will not write one, so a plain rebuild deletes it.

    ``--with-pivot`` puts it back and ``--force`` says the loss is intended. Neither given, the
    generator refuses rather than doing it quietly, which is what it did for a year.
    """
    if not WORKBOOK.is_file():
        pytest.skip(f"{WORKBOOK.name} has not been generated")
    assert evalgen.carries_a_pivot(WORKBOOK), (
        "the committed workbook has lost its Summary pivot; regenerate with --with-pivot"
    )

    def _refuse(_path: Path) -> None:
        raise AssertionError("the rebuild went ahead and took the pivot with it")

    monkeypatch.setattr(evalgen, "build", _refuse)
    monkeypatch.setattr(sys, "argv", ["build_workbook.py"])

    assert evalgen.main() == 2
