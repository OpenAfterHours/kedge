"""The adjustment eval's workbook still holds the things it is an eval for.

An eval rots quietly. A generator change, a seed change, a refactor of the sheet layout -- any
of them can remove a discrimination while leaving a workbook that opens fine, analyses fine and
scores well, because the thing it was scoring is no longer there. Every assertion here exists to
make that failure loud.

The other half is the rubric. ``expected.yaml`` quotes figures -- totals to the penny, row
counts, which trade carries the null -- and a rubric quoting numbers the workbook does not
produce is worse than no rubric, because it fails a correct conversion. So the figures are
checked against the generated data rather than trusted.

This is not a corpus test. It asserts nothing about what the analyser finds; see
``tests/corpus`` for that, and ``evals/README.md`` for why the two are kept apart.
"""

from __future__ import annotations

import hashlib
import sys
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from typing import Any

import pytest
import yaml
from openpyxl import load_workbook

from kedge.sql import literal, render

EVAL_DIR = Path(__file__).resolve().parents[2] / "evals" / "adjustment_signoff"
FIXTURE_DIR = Path(__file__).resolve().parents[1] / "fixtures"
for directory in (EVAL_DIR, FIXTURE_DIR):
    if str(directory) not in sys.path:
        sys.path.insert(0, str(directory))

import build_workbook as evalgen  # noqa: E402
import generate as fixtures  # noqa: E402

WORKBOOK = EVAL_DIR / evalgen.WORKBOOK_NAME


@pytest.fixture(scope="module")
def trades() -> list[Any]:
    return evalgen.build_trades()


@pytest.fixture(scope="module")
def adjusted(trades: list[Any]) -> list[Any]:
    return [trade for trade in trades if trade.adjusted]


@pytest.fixture(scope="module")
def rubric() -> dict[str, Any]:
    return yaml.safe_load((EVAL_DIR / "expected.yaml").read_text(encoding="utf-8"))


@pytest.fixture(scope="module")
def cached() -> Any:
    """The workbook as Excel would see it: formulas replaced by their cached results."""
    return load_workbook(WORKBOOK, data_only=True)


# ── the file on disk ─────────────────────────────────────────────────────────


def test_the_committed_workbook_matches_a_fresh_generation(tmp_path: Path) -> None:
    """Byte-identical, so the committed file is provably what the generator produces."""
    fresh = tmp_path / evalgen.WORKBOOK_NAME
    evalgen.build(fresh)

    assert hashlib.sha256(fresh.read_bytes()).hexdigest() == (
        hashlib.sha256(WORKBOOK.read_bytes()).hexdigest()
    )


def test_generating_twice_produces_identical_bytes(tmp_path: Path) -> None:
    first, second = tmp_path / "a.xlsx", tmp_path / "b.xlsx"
    evalgen.build(first)
    evalgen.build(second)

    assert first.read_bytes() == second.read_bytes()


def test_the_workbook_has_cached_values(cached: Any) -> None:
    """Without these the eval's reconciliation tier is dead on arrival (PLAN 6.2)."""
    sheet = cached["Adjustment"]
    computed = [
        sheet.cell(row=row, column=6).value
        for row in range(evalgen.ADJ_HEADER_ROW + 1, evalgen.ADJ_HEADER_ROW + 20)
    ]

    assert all(isinstance(value, int | float) for value in computed)


# ── the discriminations ──────────────────────────────────────────────────────


def test_the_rounding_trap_is_present(adjusted: list[Any]) -> None:
    """At least three uplifts need Excel's 15-significant-digit collapse to round correctly.

    The generator refuses to build without this, so the test is a second lock on the same door
    -- worth having, because the guard lives in the code somebody would be editing.
    """
    traps = 0
    for trade in adjusted:
        if trade.accrual is None:
            continue
        raw = trade.accrual * (1 + evalgen.UPLIFT_RATE)
        naive = float(Decimal(raw).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
        if naive != fixtures.excel_round(raw, 2):
            traps += 1

    assert traps >= 3


def test_the_null_row_and_the_apostrophe_row_are_different_rows(adjusted: list[Any]) -> None:
    """Two findings need two rows, or fixing either one looks like fixing both."""
    null_trades = [trade for trade in adjusted if trade.accrual is None]
    apostrophe = [trade for trade in adjusted if "'" in trade.entity_name]

    assert len(null_trades) == 1
    assert apostrophe
    assert null_trades[0].trade_id not in {trade.trade_id for trade in apostrophe}


def test_excel_predicts_zero_where_the_warehouse_keeps_null(adjusted: list[Any]) -> None:
    """The divergence the whole eval turns on, asserted on both sides."""
    blank = next(trade for trade in adjusted if trade.accrual is None)

    assert blank.excel_after() == 0.0
    assert blank.warehouse_after() is None


def test_the_post_adjustment_tab_leaves_that_row_empty(cached: Any, adjusted: list[Any]) -> None:
    blank = next(trade for trade in adjusted if trade.accrual is None)
    sheet = cached["Post-Adjustment"]
    values = {
        sheet.cell(row=row, column=1).value: sheet.cell(row=row, column=5).value
        for row in range(evalgen.PRE_HEADER_ROW + 1, evalgen.PRE_HEADER_ROW + 121)
    }

    assert values[blank.trade_id] is None


def test_the_workbooks_own_generated_sql_is_broken_for_the_apostrophe_entity(
    cached: Any,
) -> None:
    """The workbook builds its UPDATE by concatenation, and for one entity that is invalid SQL.

    This is the one discrimination where reproducing the workbook faithfully is the *wrong*
    answer. The statement is asserted to be broken so that a conversion which fixes it can be
    told apart from one that never noticed.
    """
    sheet = cached["Adjustment"]
    statements = [
        sheet.cell(row=row, column=7).value
        for row in range(evalgen.ADJ_HEADER_ROW + 1, evalgen.ADJ_HEADER_ROW + 77)
    ]
    broken = [text for text in statements if "O'Brien" in str(text)]

    assert broken, "no statement carries the apostrophe entity"
    # Four quotes in a valid single-row statement: two around the note, two around the trade id.
    # The unescaped apostrophe makes five, which is what a client rejects.
    assert broken[0].count("'") % 2 == 1


def test_kedge_sql_renders_that_same_row_correctly() -> None:
    """The fix the eval is looking for, pinned so its absence is a failing assertion."""
    rendered = render(
        "UPDATE fin.accruals SET adjustment_note = {note} WHERE trade_id = {trade_id}",
        {"note": "Q2 uplift for O'Brien & Partners", "trade_id": "ACC-00041"},
    )

    assert "'Q2 uplift for O''Brien & Partners'" in rendered
    assert rendered.count("'") % 2 == 0
    assert literal("O'Brien & Partners") == "'O''Brien & Partners'"


def test_cancelled_rows_sit_inside_adjusted_entities(trades: list[Any]) -> None:
    """Otherwise the WHERE clause's `status <> 'CANCELLED'` discriminates nothing."""
    cancelled_in_scope = [
        trade for trade in trades if trade.in_scope and trade.status == "CANCELLED"
    ]

    assert cancelled_in_scope
    assert all(not trade.adjusted for trade in cancelled_in_scope)


def test_the_memo_disagrees_with_the_scope(cached: Any) -> None:
    """Says three entities; the scope row and the statement both name four."""
    sheet = cached["Sign-off"]
    prose = " ".join(str(sheet.cell(row=row, column=1).value or "") for row in range(1, 20))

    assert evalgen.SIGNOFF_ENTITY_COUNT_CLAIM in prose
    assert len(evalgen.IN_SCOPE) == 4
    assert all(entity in prose for entity in evalgen.IN_SCOPE)


def test_the_memos_impact_figures_are_stale(cached: Any, adjusted: list[Any]) -> None:
    """Close enough to survive a glance, wrong by an amount only recomputation finds."""
    signoff = cached["Sign-off"]
    quoted_movement = next(
        signoff.cell(row=row, column=4).value
        for row in range(1, 40)
        if isinstance(signoff.cell(row=row, column=4).value, int | float)
    )
    true_movement = fixtures.excel_round(sum(trade.excel_uplift() for trade in adjusted), 2)

    assert quoted_movement != true_movement
    assert 0 < (true_movement - quoted_movement) < true_movement * 0.05


# ── the rubric agrees with the workbook ──────────────────────────────────────


def test_the_rubric_quotes_the_figures_the_workbook_actually_produces(
    rubric: dict[str, Any], trades: list[Any], adjusted: list[Any]
) -> None:
    """A rubric quoting numbers the generator no longer produces fails a correct conversion."""
    facts = rubric["facts"]

    assert facts["trades_extracted"] == len(trades)
    assert facts["rows_the_update_reaches"] == len(adjusted)
    assert facts["entities_in_scope"] == len(evalgen.IN_SCOPE)
    assert facts["uplift_rate"] == evalgen.UPLIFT_RATE
    assert facts["accrual_total_before"] == fixtures.excel_round(
        sum(trade.accrual or 0.0 for trade in adjusted), 2
    )
    assert facts["uplift_total"] == fixtures.excel_round(
        sum(trade.excel_uplift() for trade in adjusted), 2
    )
    assert facts["accrual_total_after"] == fixtures.excel_round(
        sum(trade.excel_after() for trade in adjusted), 2
    )


def test_every_scored_item_declares_a_weight(rubric: dict[str, Any]) -> None:
    for tier in ("deterministic", "structural"):
        for item in rubric[tier]:
            assert isinstance(item.get("weight"), int), f"{tier}/{item['id']} has no weight"
            assert item.get("check"), f"{tier}/{item['id']} says nothing to check"


def test_the_weights_still_lean_on_the_deterministic_tier(rubric: dict[str, Any]) -> None:
    """A plan that reads well and produces the wrong pennies has failed, and the weights say so.

    The structural tier has grown twice -- once when it stopped grading the presence of a field
    and started grading the shape the scaffolder consumes, and again for the item that grades
    where a hand-in is emitted. Both were good reasons to move the balance and neither is a reason
    to lose track of it.

    Asserted as a **band**, because the bound this replaced (``deterministic > structural * 1.5``)
    permitted anything from 60% to 100% and the README quotes one figure out of that range. A
    range nobody can drift out of unnoticed is the point; the width is there so that adding a
    two-point item does not fail a test for no reason.
    """
    weights = {
        tier: sum(int(item["weight"]) for item in rubric[tier])
        for tier in ("deterministic", "structural")
    }
    share = weights["deterministic"] / sum(weights.values())

    assert 0.60 <= share <= 0.66, (
        f"the deterministic tier is {share:.1%} of the declared points ({weights}). The README "
        f"quotes 62.5%; move both together or neither."
    )


def test_the_workbook_still_explains_itself_where_the_briefing_item_looks(cached: Any) -> None:
    """The rot guard for ``the_briefing_survives_the_workbook``.

    That item is only a discrimination while the Sign-off tab actually carries prose somebody
    wrote for the next reader. Strip the headings out of the generator and a plan with
    ``briefing: null`` becomes the *honest* answer, the item becomes a trap, and nothing else in
    the suite would notice: the workbook would still open, still analyse and still reconcile.
    """
    sheet = cached["Sign-off"]
    headings = {str(sheet.cell(row=row, column=1).value or "").strip() for row in range(1, 20)}
    prose = " ".join(str(sheet.cell(row=row, column=1).value or "") for row in range(1, 20))

    assert {"Purpose", "Background", "Scope", "Known issues"} <= headings
    assert "reforecast" in prose, "the background a conversion cannot reconstruct has gone"
    assert str(evalgen.UPLIFT_RATE * 100) in prose or "4.5%" in prose


def test_the_committed_plan_keeps_that_prose_and_says_where_it_came_from() -> None:
    """The other half: the reference plan must pass the item it is the positive control for.

    Both halves are asserted because either alone is satisfiable by accident. A plan citing cells
    that hold nothing is as useless as prose citing nothing at all, so every source names a sheet
    the workbook has.
    """
    from kedge.plan.store import plan_from_yaml

    plan = plan_from_yaml((EVAL_DIR / "plan.yaml").read_text(encoding="utf-8"))
    briefing = plan.briefing

    assert briefing is not None and not briefing.is_empty
    assert briefing.purpose and briefing.background
    assert briefing.sources
    sheets = {source.split("!")[0] for source in briefing.sources if "!" in source}
    assert "Sign-off" in sheets, briefing.sources


# ── the reference plan says what the statement does ──────────────────────────


def test_the_committed_plan_declares_the_update_as_a_statement_that_writes() -> None:
    """``mutates`` is what puts a confirmation step between the UPDATE and the re-extract.

    This plan carried ``mutates: false`` over an ``UPDATE fin.accruals``, and the notebook it
    scaffolded therefore had no confirmation cell, no token for the re-extract selector to read,
    and so a re-extract box on screen from the moment the notebook opened -- inviting a grid
    taken *before* the update ran, which looks exactly like one taken after and cannot be told
    apart afterwards. Asserted on the committed file rather than on the generator, because the
    committed file is what every other test loads.
    """
    from kedge.plan.store import plan_from_yaml
    from kedge.sql import changes_data

    plan = plan_from_yaml((EVAL_DIR / "plan.yaml").read_text(encoding="utf-8"))
    handoffs = [stage for stage in plan.stages if stage.is_handoff]
    assert handoffs, "the reference plan has stopped handing anything over"

    writing = [
        stage
        for stage in handoffs
        if changes_data(stage.effective_handoff().template or stage.effective_handoff().statement)
    ]
    assert [stage.id for stage in writing] == ["update_statement"]
    for stage in writing:
        handoff = stage.effective_handoff()
        assert handoff.mutates, f"{stage.id} writes but the plan declares it read-only"
        assert handoff.needs_confirmation
        assert not handoff.contradicts_its_own_statement
