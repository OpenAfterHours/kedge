"""The scaffolder: an approved plan in, the notebook the user actually gets out.

This module is the last thing between a reviewed plan and forty cells somebody has to live
with, so the assertions here are mostly about properties of the *emitted text* rather than
about return values:

- every cell body parses, because a cell that does not is a notebook that will not open;
- no public name is defined by two cells, because marimo rejects the batch if one is;
- the head calls `kedge.ingest` and `kedge.contracts` and the tail calls `kedge.reconcile`,
  rather than re-implementing any of them;
- and the reconciliation cell says "not reconciled" wherever it cannot say anything else.

That last one is not a style point. A reconciliation that reports a pass it did not earn is
the single most dangerous thing kedge could emit (PLAN 6.2), so the tail cell is executed
here -- against a missing workbook and against a real one with no cached values -- and the
resulting panel is asserted to be falsy and to say so in words.
"""

from __future__ import annotations

import ast
import inspect
import re
from pathlib import Path
from types import SimpleNamespace
from typing import TYPE_CHECKING, Any

import polars as pl
import pytest
from openpyxl import Workbook

import kedge
import kedge.contracts
import kedge.ingest
import kedge.ingest.drift
import kedge.reconcile
import kedge.reconcile.acceptance
import kedge.runs
from conftest import make_draft, make_plan
from kedge.notebook import scaffold
from kedge.notebook.driver import CellNameError, StaleCellError
from kedge.notebook.scaffold import (
    HEAD_CELL_NAMES,
    PlanNotApprovedError,
    ScaffoldCell,
    ScaffoldError,
    build_cells,
    cell_name_for,
    sync_notebook,
)
from kedge.plan.model import (
    Approval,
    ApprovalState,
    Checkpoint,
    Handoff,
    ProcessPlan,
    SourceOrigin,
    Stage,
    StageKind,
    StageSource,
)
from kedge.reconcile.model import ReconciliationStatus

if TYPE_CHECKING:
    from collections.abc import Iterable


# ── helpers ──────────────────────────────────────────────────────────────────


def approved(**overrides: Any) -> ProcessPlan:
    """The shared fixture plan, with the review gate satisfied."""
    plan = make_plan(**overrides)
    return plan.model_copy(update={"approval": Approval(state=ApprovalState.APPROVED)})


def cells_for(plan: ProcessPlan | None = None, **kwargs: Any) -> list[ScaffoldCell]:
    return build_cells(plan if plan is not None else approved(), **kwargs)


def named(cells: Iterable[ScaffoldCell], name: str) -> ScaffoldCell:
    """The one cell with this name; a KeyError here is itself a useful failure."""
    for cell in cells:
        if cell.name == name:
            return cell
    raise KeyError(name)


def public_names(code: str) -> set[str]:
    """Every module-level name a cell body binds, excluding marimo's `_` escape hatch.

    Deliberately generous -- it walks the whole tree rather than only the top level -- because
    a name bound inside a `try` or an `if` is just as much a definition as far as marimo's
    static analysis is concerned.
    """
    tree = ast.parse(code)
    found: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Name) and isinstance(node.ctx, ast.Store):
            found.add(node.id)
        elif isinstance(node, (ast.Import, ast.ImportFrom)):
            for alias in node.names:
                found.add(alias.asname or alias.name.split(".")[0])
        elif isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)):
            found.add(node.name)
    return {name for name in found if not name.startswith("_")}


def plain_workbook(path: Path) -> Path:
    """A workbook of literal values and not one formula: the no-baseline case."""
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["id", "amount"])
    for row in range(1, 6):
        sheet.append([row, row * 1.5])
    book.save(path)
    return path


class StoppedError(Exception):
    """Stands in for marimo's `MarimoStopError`, which needs a kernel to raise for real."""


class FakeMarimo:
    """Just enough of `mo` for the head cells to run outside a notebook.

    `mo.stop` halting the cell *and its descendants* is the mechanism the contract gate and the
    empty-selection guard both rely on, so it is modelled as an exception rather than stubbed
    out: a test that could not tell a block from a pass would be worse than no test.

    Everything passed to `mo.md` is kept, so a test can assert on the panel the user reads and
    not merely on the values the cell happened to bind.
    """

    class ui:  # noqa: N801 - mirrors marimo's own lowercase namespace
        pass

    def __init__(self) -> None:
        self.rendered: list[str] = []

    def md(self, text: str) -> str:
        self.rendered.append(text)
        return text

    def vstack(self, items: list[Any]) -> list[Any]:
        return items

    @staticmethod
    def stop(predicate: object, output: Any = None) -> None:
        if predicate:
            raise StoppedError(output)

    def panel(self, heading: str) -> str:
        """The one rendered panel starting with this markdown heading."""
        for text in self.rendered:
            if text.startswith(heading):
                return text
        raise AssertionError(f"no panel headed {heading!r} in {self.rendered}")


class Selection:
    """A stand-in for a marimo file input: the `.value` is all the cells touch."""

    def __init__(self, value: tuple[Any, ...] = ()) -> None:
        self.value = value


def run_head(
    cells: list[ScaffoldCell],
    *,
    store: Path,
    picked: Path | None = None,
    contract: Path | None = None,
    pasted: str = "",
    run_mode: str = "resume the run in progress",
) -> dict[str, Any]:
    """Execute the five head cells below the selector against the real ingest machinery.

    The setup and selector cells are supplied rather than run: one imports marimo for real and
    the other builds UI elements that need a kernel. Everything below them is ordinary Python
    and is exactly what this is here to exercise.

    ``pasted`` stands in for the third entry point. Its value is a string rather than a tuple
    because that is what ``mo.ui.text_area`` yields, and the head cell's precedence rule depends
    on the difference.
    """
    import datetime as datetime_module
    import pathlib

    namespace: dict[str, Any] = {
        "mo": FakeMarimo(),
        "pl": pl,
        "kedge": kedge,
        "datetime": datetime_module,
        "pathlib": pathlib,
        "HANDIN_DIR": store,
        "CONTRACT_PATH": contract if contract is not None else store / "absent.yaml",
        "WORKBOOK": store / "absent.xlsx",
        "RUNS_DIR": store / "runs",
        "handin_drop": Selection(),
        "handin_pick": Selection((picked,) if picked is not None else ()),
        "handin_paste": Selection(pasted),
        "kedge_run_mode": Selection(run_mode),
    }
    # Three cells are supplied rather than run: `kedge_setup` imports marimo for real, and the
    # two selector cells build UI elements that need a kernel. Everything else is ordinary
    # Python -- including the resume logic, which is exercised here rather than mocked, because
    # "does reopening pick the same run up" is the whole point of it.
    supplied = {"kedge_setup", "kedge_briefing", "kedge_run_mode", "handin_source"}
    for name in HEAD_CELL_NAMES:
        if name in supplied:
            continue
        exec(compile(named(cells, name).code, f"<{name}>", "exec"), namespace)
    return namespace


def run_reconciliation(cells: Iterable[ScaffoldCell], workbook: Path, **frames: Any) -> Any:
    """Execute the reconciliation cells in order and hand back the panel they produced.

    Every ``reconcile`` cell, not one: the region map and the panel are separate cells because
    they are owned by different people -- the map is a translation judgement carrying a
    ``TODO(kedge)``, the panel is machinery that stays kedge's -- and the panel reads what the map
    defines. Running only one of them would test a notebook nobody will ever have.

    They read `WORKBOOK`, `kedge` and the stage cells' frames out of the notebook's globals, so
    supplying those is the whole of the harness. `mo` is deliberately absent: the tail must not
    need marimo to state that nothing was checked.
    """
    namespace: dict[str, Any] = {
        "WORKBOOK": workbook,
        "ACCEPTANCE_PATH": workbook.parent / "reconciliation.json",
        "handin": SimpleNamespace(sha256="a" * 64),
        "kedge": kedge,
        "__file__": str(workbook.parent / "notebook.py"),
        **frames,
    }
    for cell in cells:
        if cell.role != "reconcile":
            continue
        exec(compile(cell.code, f"<{cell.name}>", "exec"), namespace)
    return namespace["reconciliation"]


# ── the review gate ──────────────────────────────────────────────────────────


def test_build_cells_refuses_an_unapproved_plan() -> None:
    """The structural half of the review gate (PLAN 2.2, M2 step 3)."""
    with pytest.raises(PlanNotApprovedError) as caught:
        build_cells(make_plan())
    assert "'draft', not 'approved'" in str(caught.value)


def test_the_refusal_names_the_blockers_a_reviewer_has_to_clear() -> None:
    with pytest.raises(PlanNotApprovedError) as caught:
        build_cells(make_plan())

    assert "Calc!AK:AP" in str(caught.value)
    assert "has not been acknowledged" in str(caught.value)


def test_an_unapproved_plan_can_be_previewed_but_not_written() -> None:
    """`build_cells` has an escape hatch because it writes nothing; scaffolding has none."""
    preview = build_cells(make_plan(), allow_unapproved=True)
    assert [cell.name for cell in preview[: len(HEAD_CELL_NAMES)]] == list(HEAD_CELL_NAMES)


# ── writing an approved plan into the notebook ───────────────────────────────
#
# `sync_notebook` is the only way in. The scaffolder it replaced created every cell
# unconditionally, which collided on the second call and is why scaffolding used to happen once,
# inside the open sequence. The things worth asserting hardest are what this one does *not* do:
# it never overwrites a cell it cannot prove kedge wrote, and it never deletes.


class SyncingDriver:
    """A `CellSyncer` over a dict. No kernel, no marimo, and no file either."""

    def __init__(self, cells: dict[str, str] | None = None, *, refuse: str = "") -> None:
        self.cells: dict[str, str] = dict(cells or {})
        self.created: list[tuple[str, str | None]] = []
        self.hidden: list[bool] = []
        self.edited: list[str] = []
        self.deleted: list[str] = []
        self._refuse = refuse

    async def list_cells(self, *, with_code: bool = True) -> tuple[SimpleNamespace, ...]:
        return tuple(
            SimpleNamespace(name=name, code=code if with_code else None)
            for name, code in self.cells.items()
        )

    async def create_cell(
        self, code: str, *, name: str, after: str | None = None, hide_code: bool = False
    ) -> str:
        # The collision the old scaffolder hit on every reopen. Raised rather than tolerated so a
        # sync that ever calls this for a name already in the notebook fails loudly here.
        if name in self.cells:
            raise CellNameError(f"the name {name!r} is already taken")
        self.created.append((name, after))
        self.hidden.append(hide_code)
        self.cells[name] = code
        return f"id-{name}"

    async def edit_cell(self, target: str, code: str, *, run: bool = True) -> str:
        if target == self._refuse:
            raise StaleCellError(f"{target} changed since kedge last read it")
        self.edited.append(target)
        self.cells[target] = code
        return f"id-{target}"


def synced_names(plan: ProcessPlan | None = None) -> list[str]:
    return [cell.name for cell in cells_for(plan)]


@pytest.mark.parametrize(
    "state",
    [ApprovalState.DRAFT, ApprovalState.CHANGES_REQUESTED, ApprovalState.REJECTED],
)
async def test_sync_notebook_writes_nothing_without_approval(state: ApprovalState) -> None:
    """Approval is state on the plan, and this is the one place that consumes it."""
    plan = make_plan().model_copy(update={"approval": Approval(state=state)})
    driver = SyncingDriver()

    with pytest.raises(PlanNotApprovedError):
        await sync_notebook(plan, driver)

    assert driver.created == []
    assert "allow_unapproved" not in inspect.signature(sync_notebook).parameters


async def test_sync_notebook_scaffolds_an_empty_notebook_whole() -> None:
    """The approve-a-proposal case: nothing there, so everything the plan calls for is written."""
    driver = SyncingDriver()

    result = await sync_notebook(approved(), driver)

    assert [name for name, _ in driver.created] == synced_names()
    assert list(result.named("created")) == synced_names()
    assert result.wrote_anything


async def test_sync_notebook_creates_every_cell_visible() -> None:
    """`create_cell` hides code by default, and kedge always overrides that (PLAN 1.1)."""
    driver = SyncingDriver()

    await sync_notebook(approved(), driver)

    assert driver.hidden == [False] * len(synced_names())


async def test_sync_notebook_passes_the_paths_through_to_the_setup_cell(tmp_path: Path) -> None:
    driver = SyncingDriver()

    await sync_notebook(
        approved(),
        driver,
        handins_dir=tmp_path / "handins",
        workbook_path=tmp_path / "source.xlsx",
        contract_path=tmp_path / "agreed.yaml",
    )

    setup = driver.cells["kedge_setup"]
    assert repr(str(tmp_path / "source.xlsx")) in setup
    assert repr(str(tmp_path / "agreed.yaml")) in setup


async def test_sync_notebook_places_a_new_cell_behind_the_one_the_plan_emits_before_it() -> None:
    """Order is presentation, but a stage under the reconciliation tail reads as a mistake."""
    driver = SyncingDriver()

    await sync_notebook(approved(), driver)

    names = synced_names()
    assert driver.created[0] == (names[0], None)  # nothing to sit behind yet
    assert [after for _, after in driver.created[1:]] == names[:-1]


async def test_sync_notebook_is_idempotent_and_writes_nothing_the_second_time() -> None:
    """The reopen case. This used to die on CellNameError at the first cell."""
    driver = SyncingDriver()
    await sync_notebook(approved(), driver)
    driver.created.clear()

    result = await sync_notebook(approved(), driver)

    assert driver.created == []
    assert driver.edited == []
    assert list(result.named("unchanged")) == synced_names()
    assert not result.wrote_anything


async def test_sync_notebook_leaves_a_cell_somebody_has_worked_on_alone() -> None:
    """The rule the whole function exists for.

    By the time a plan is amended the stage cells have usually been translated: the scaffolded
    body is a documented passthrough with a TODO in it, and the body beside it a week later is
    the user's and the agent's work. Overwriting that is the one unrecoverable thing here.
    """
    plan = approved()
    stage = synced_names(plan)[len(HEAD_CELL_NAMES)]
    translated = f"{stage} = handin_frame.filter(pl.col('amount') > 0)  # mine, not kedge's"
    driver = SyncingDriver({stage: translated})

    result = await sync_notebook(plan, driver)

    assert driver.cells[stage] == translated
    assert driver.edited == []
    assert list(result.named("diverged")) == [stage]
    assert stage in result.summary(plan.version)


async def test_sync_notebook_updates_a_cell_it_can_prove_it_wrote_itself() -> None:
    """The amendment case: a note attached to a stage reaches the untouched cell for that stage."""
    before = approved()
    stage_id = before.stages[0].id
    after = before.model_copy(
        update={
            "version": before.version + 1,
            "stages": [
                before.stages[0].model_copy(update={"notes": "FX rates come from Treasury."}),
                *before.stages[1:],
            ],
        }
    )
    driver = SyncingDriver({cell.name: cell.code for cell in cells_for(before)})
    name = next(cell.name for cell in cells_for(after) if cell.stage_id == stage_id)

    result = await sync_notebook(after, driver, previous=before)

    assert name in driver.edited
    assert "FX rates come from Treasury." in driver.cells[name]
    assert name in result.named("updated")
    # The setup cell names the version it was generated from, so a bump rewrites its header too.
    # Both are cells kedge wrote and nobody has touched, which is the whole test.
    assert set(driver.edited) == {name, "kedge_setup", "kedge_briefing"}
    assert not result.named("diverged")


async def test_sync_notebook_without_a_previous_plan_reports_rather_than_overwrites() -> None:
    """No evidence that kedge wrote the cell is the same answer as evidence that it did not.

    Safe in the only direction that matters: the sync declines and says so, rather than
    overwriting on the assumption that a differing cell must be stale scaffolding.
    """
    before = approved()
    after = before.model_copy(
        update={
            "stages": [
                before.stages[0].model_copy(update={"notes": "an amendment"}),
                *before.stages[1:],
            ]
        }
    )
    driver = SyncingDriver({cell.name: cell.code for cell in cells_for(before)})

    result = await sync_notebook(after, driver)  # no `previous`

    assert driver.edited == []
    assert result.named("diverged")


async def test_sync_notebook_never_deletes_a_cell_the_plan_has_stopped_mentioning() -> None:
    """A plan edit is not consent to lose the code that implemented the old decomposition."""
    before = approved()
    after = before.model_copy(update={"stages": before.stages[:1]})
    driver = SyncingDriver({cell.name: cell.code for cell in cells_for(before)})
    dropped = set(synced_names(before)) - set(synced_names(after))

    result = await sync_notebook(after, driver, previous=before)

    assert dropped
    assert dropped <= set(driver.cells)
    assert set(result.obsolete) == dropped
    assert not hasattr(driver, "delete_cell") or driver.deleted == []
    assert "nothing was deleted" in result.summary(after.version)


async def test_a_cell_the_kernel_refuses_does_not_stop_the_rest_being_written() -> None:
    """Nine tenths of a notebook in line with the plan beats giving up at the first refusal."""
    before = approved()
    after = before.model_copy(
        update={
            "stages": [
                before.stages[0].model_copy(update={"notes": "an amendment"}),
                *before.stages[1:],
            ]
        }
    )
    stale = next(cell.name for cell in cells_for(after) if cell.stage_id == before.stages[0].id)
    existing = {cell.name: cell.code for cell in cells_for(before)}
    del existing[synced_names(after)[-1]]  # one still to create, after the refusal
    driver = SyncingDriver(existing, refuse=stale)

    result = await sync_notebook(after, driver, previous=before)

    assert list(result.named("refused")) == [stale]
    assert list(result.named("created")) == [synced_names(after)[-1]]
    assert "The kernel refused 1" in result.summary(after.version)


# ── every cell is real Python ────────────────────────────────────────────────


@pytest.mark.parametrize("cell", cells_for(), ids=lambda cell: cell.name)
def test_every_generated_cell_body_parses(cell: ScaffoldCell) -> None:
    ast.parse(cell.code)


@pytest.mark.parametrize(
    "plan",
    [
        approved(),
        approved(draft=make_draft(dropped=[])),
        approved(
            draft=make_draft(
                stages=[Stage(id="only", intent="One stage and nothing else")],
                open_questions=[],
                dropped=[],
            )
        ),
        approved(
            draft=make_draft(
                stages=[
                    Stage(
                        id="sign_off",
                        intent="Someone has to look at this",
                        kind=StageKind.CHECKPOINT,
                        checkpoint=Checkpoint(question="Fine?", require_note=False),
                    )
                ],
                open_questions=[],
                dropped=[],
            )
        ),
    ],
    ids=["standard", "no-drops", "single-stage", "checkpoint-only"],
)
def test_every_cell_parses_for_every_plan_shape(plan: ProcessPlan) -> None:
    for cell in build_cells(plan):
        ast.parse(cell.code)


def test_a_cell_that_would_not_parse_is_refused_rather_than_written(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The scaffolder marks its own homework; a broken cell must not reach the kernel."""
    monkeypatch.setattr(scaffold, "_SOURCE_CELL", "handin_source = mo.ui.tabs(")

    with pytest.raises(ScaffoldError, match="would not parse"):
        build_cells(approved())


def test_a_cell_that_would_import_pandas_is_refused(monkeypatch: pytest.MonkeyPatch) -> None:
    """Non-negotiable 1, checked on the emitted text rather than trusted (PLAN 2.5)."""
    plan = approved(
        draft=make_draft(
            stages=[Stage(id="load", intent="Read it", notes="import pandas as pd first")],
            open_questions=[],
            dropped=[],
        )
    )

    with pytest.raises(ScaffoldError, match="pandas"):
        build_cells(plan)


def test_no_cell_mentions_pandas_or_the_private_marimo_api() -> None:
    """Non-negotiables 1 and 2, on the text a user will run."""
    for cell in cells_for():
        assert "pandas" not in cell.code
        assert "_code_mode" not in cell.code


# ── names, ordering and the single-definition rule ───────────────────────────


def test_the_head_comes_first_and_the_stages_follow_in_dependency_order() -> None:
    cells = cells_for()
    names = [cell.name for cell in cells]

    assert names[: len(HEAD_CELL_NAMES)] == list(HEAD_CELL_NAMES)
    # Reconciliation sits after `apply_haircuts`, the last stage that names any analysis
    # operations -- so the evidence that the arithmetic matches the workbook is in front of the
    # user *before* the checkpoint asking them to approve it, rather than in a footnote below.
    # Two cells, not one: the region map is a translation judgement and carries a TODO, the panel
    # is machinery and stays kedge's. They must stay adjacent and in this order, because the
    # panel reads what the map defines.
    assert [name for name in names if name not in HEAD_CELL_NAMES] == [
        "load_handin",
        "apply_haircuts",
        "reconciliation_values",
        "reconciliation",
        "manual_overrides_ui",
        "manual_overrides",
        "write_output",
    ]


def test_reconciliation_lands_after_the_last_stage_it_can_report_on() -> None:
    """Not at the end of the file, which is where it used to be and where it read as a footnote.

    The panel depends only on the computing stages, so marimo runs it as soon as those are done.
    Left at the bottom that meant a user blocked at an approval three steps earlier got a wall
    of reconciliation output directly beneath the sentence telling them what to type -- and it
    put the evidence *after* the decision it is evidence for. Reconciliation against the
    workbook's own cached values is exactly what somebody wants in front of them when they
    approve a change to production.
    """
    plan = make_plan()
    last_computing = [
        stage.id
        for stage in plan.ordered_stages()
        if stage.operations and not stage.generates_no_code
    ][-1]
    names = [cell.name for cell in cells_for(approved())]

    # The map, then the panel, immediately after the last stage either could report on. The panel
    # reads what the map defines, so a gap between them would be a name defined below its reader.
    assert names.index("reconciliation_values") == names.index(last_computing) + 1
    assert names.index("reconciliation") == names.index(last_computing) + 2


def test_a_plan_whose_stages_name_no_operations_keeps_reconciliation_at_the_end() -> None:
    """There is nothing to place it relative to, so it stays where it was."""
    draft = make_draft()
    bare = draft.model_copy(
        update={"stages": [stage.model_copy(update={"operations": []}) for stage in draft.stages]}
    )
    names = [cell.name for cell in cells_for(approved(draft=bare))]

    assert names[-1] == "reconciliation"


def test_the_head_reports_drift_before_the_contract_check() -> None:
    """PLAN 2.8: a renamed column is a better message than a schema-validation traceback."""
    names = list(HEAD_CELL_NAMES)
    assert names.index("handin_drift") < names.index("handin_check")
    assert names.index("handin_check") < names.index("handin_frame")


def test_the_contract_is_loaded_before_anything_that_has_to_read_the_same_rows() -> None:
    """Loading is not checking: the sheet is needed by the profile, the check and the frame."""
    names = list(HEAD_CELL_NAMES)
    assert names.index("handin") < names.index("handin_contract")
    assert names.index("handin_contract") < names.index("handin_drift")

    loader = named(cells_for(), "handin_contract")
    assert "kedge.contracts.load(CONTRACT_PATH)" in loader.code
    assert "kedge.contracts.validate" not in loader.code  # loading is not checking


def test_stage_cells_carry_their_stage_id_and_the_fixed_cells_do_not() -> None:
    cells = cells_for()
    assert named(cells, "apply_haircuts").stage_id == "apply_haircuts"
    assert named(cells, "manual_overrides_ui").stage_id == "manual_overrides"
    assert named(cells, "kedge_setup").stage_id is None
    assert named(cells, "reconciliation").stage_id is None


def test_every_cell_has_a_role_from_the_vocabulary() -> None:
    roles = {cell.role for cell in cells_for()}
    assert roles == {"setup", "handin", "stage", "checkpoint", "reconcile"}


def test_no_public_name_is_defined_by_two_cells() -> None:
    """marimo's single-definition rule; `_`-prefixed names are the escape hatch."""
    owner: dict[str, str] = {}
    for cell in cells_for():
        for name in public_names(cell.code):
            assert name not in owner, f"{name!r} defined by {owner.get(name)!r} and {cell.name!r}"
            owner[name] = cell.name


def test_a_checkpoints_ui_names_cannot_collide_with_another_stage() -> None:
    """A checkpoint owns three more public names than its own, and all four must be unique."""
    plan = approved(
        draft=make_draft(
            stages=[
                Stage(id="review_decision", intent="A stage that wants the checkpoint's name"),
                Stage(
                    id="review",
                    intent="The checkpoint itself",
                    kind=StageKind.CHECKPOINT,
                    checkpoint=Checkpoint(question="Approved?"),
                ),
            ],
            open_questions=[],
            dropped=[],
        )
    )

    owner: dict[str, str] = {}
    for cell in build_cells(plan):
        for name in public_names(cell.code):
            assert name not in owner, f"{name!r} defined by {owner.get(name)!r} and {cell.name!r}"
            owner[name] = cell.name


@pytest.mark.parametrize(
    ("stage_id", "expected"),
    [
        ("apply_haircuts", "apply_haircuts"),
        ("Reconcile-EOD", "reconcile_eod"),
        ("step 1", "step_1"),
        ("1_load", "stage_1_load"),
        ("class", "class_stage"),
        ("match", "match_stage"),
        ("!!!x!!!", "x"),
    ],
)
def test_cell_name_for_produces_a_usable_identifier(stage_id: str, expected: str) -> None:
    name = cell_name_for(stage_id)
    assert name == expected
    assert name.isidentifier()


@pytest.mark.parametrize("reserved", ["handin", "handin_frame", "reconciliation", "mo", "kedge"])
def test_cell_name_for_never_collides_with_a_fixed_cell_or_an_import(reserved: str) -> None:
    assert cell_name_for(reserved) != reserved


def test_cell_name_for_suffixes_rather_than_overwriting() -> None:
    assert cell_name_for("load", taken=["load"]) == "load_2"
    assert cell_name_for("load", taken=["load", "load_2"]) == "load_3"


def test_cell_name_for_gives_up_loudly_rather_than_returning_a_duplicate() -> None:
    """A silent duplicate would be a notebook marimo refuses to open."""
    taken = ["load", *(f"load_{index}" for index in range(2, 1000))]

    with pytest.raises(ScaffoldError, match="unique cell name"):
        cell_name_for("load", taken=taken)


# ── the fixed head calls the real machinery ──────────────────────────────────


def test_the_head_carries_no_unfinished_markers() -> None:
    """The plumbing carries no TODOs. Stage cells do, and so does the region map.

    ``reconciliation_values`` is the exception, and a deliberate one: which column of which frame
    reproduces a given workbook region is a translation judgement the scaffolder cannot make, so
    it is a hole like any other and is found the same way -- by the marker. It used to carry none,
    which meant nothing that fills the scaffolder's holes was ever asked to finish it. The panel
    beside it stays free of markers, because that one is machinery and stays kedge's.
    """
    fixed = [
        cell
        for cell in cells_for()
        if cell.role in ("setup", "handin", "reconcile") and cell.name != "reconciliation_values"
    ]
    for cell in fixed:
        assert "TODO" not in cell.code, cell.name

    assert "TODO(kedge)" in named(cells_for(), "reconciliation_values").code


def test_the_handin_cell_calls_ingest_rather_than_hashing_by_hand() -> None:
    code = named(cells_for(), "handin").code
    assert "kedge.ingest.receive(" in code
    assert "store_dir=HANDIN_DIR" in code
    assert "hashlib" not in code


def test_nothing_treats_the_handin_as_a_dict() -> None:
    """`receive` returns a HandIn record; subscripting it would be an AttributeError."""
    for cell in cells_for():
        assert 'handin["' not in cell.code
        assert "handin['" not in cell.code


def test_the_handin_cell_waits_rather_than_erroring_on_an_empty_selection() -> None:
    code = named(cells_for(), "handin").code
    stop = code.index("mo.stop(")
    assert stop < code.index("kedge.ingest.receive(")
    # It names its step, like every other stop in the scaffold. This was the one that did not --
    # it said "Waiting for a hand-in" and stopped there -- because the head hand-in belongs to
    # the notebook rather than to a stage. In app mode a stopped cell is the whole user
    # interface, and "waiting" without a position is indistinguishable from a page that died.
    assert re.search(r"\*\*Step \d+ of \d+: .+\.\*\*", code), code
    assert "Drop the file above" in code


def test_the_drift_cell_calls_check_drift_and_degrades_to_a_message() -> None:
    code = named(cells_for(), "handin_drift").code
    assert "kedge.ingest.check_drift(" in code
    assert "except kedge.IngestError" in code
    assert "Could not profile the hand-in" in code


def test_the_check_cell_validates_against_a_contract_when_there_is_one() -> None:
    code = named(cells_for(), "handin_check").code
    assert "kedge.contracts.validate(handin, handin_contract)" in code


def test_the_contract_cell_scaffolds_and_runs_with_no_contract_at_all() -> None:
    """A new notebook has no contract, and that is a state, not an error."""
    code = named(cells_for(), "handin_contract").code
    assert "CONTRACT_PATH.is_file()" in code
    assert "nothing below is enforced" in code
    assert "kedge contract infer" in code


def test_a_failed_contract_stops_the_frame_and_everything_below_it() -> None:
    """mo.stop halts the cell and its descendants, so the check is a gate (PLAN 2.8)."""
    code = named(cells_for(), "handin_frame").code
    stop = code.index("mo.stop(")
    assert stop < code.index("handin_frame = ")
    assert "handin_check is not None and not handin_check.ok" in code


def test_the_frame_goes_through_the_same_reader_the_contract_check_does() -> None:
    """Not an equivalent-looking read: the same function, with the contract's own arguments.

    `kedge.contracts.validate` does `read_data(handin.path, sheet=contract.sheet,
    header_row=contract.header_row)`. Anything else here and the frame the notebook computes on
    is not the frame that was validated. The behaviour is pinned by the executed tests below;
    this pins the mechanism, because a mirrored four-line equivalent is one refactor from
    diverging again.
    """
    code = named(cells_for(), "handin_frame").code
    assert "kedge.ingest.read_data(" in code
    assert "sheet=handin_contract.sheet if handin_contract is not None else None" in code
    assert "header_row=handin_contract.header_row if handin_contract is not None else None" in code
    assert "handin_frame = _data.lazy()" in code
    assert "detect_layout" not in code  # the reader does that, and does it once


def test_the_setup_cell_owns_the_three_paths_and_nothing_computes_them() -> None:
    code = named(cells_for(), "kedge_setup").code
    assert "HANDIN_DIR = pathlib.Path(" in code
    assert "WORKBOOK = pathlib.Path(" in code
    assert "CONTRACT_PATH = pathlib.Path(" in code


def test_the_paths_default_to_the_standard_workspace_layout() -> None:
    """`<workbook>.kedge/handins` beside the workbook is where kedge puts things."""
    code = named(cells_for(handins_dir=Path("C:/p/rwa.kedge/handins")), "kedge_setup").code
    assert repr(str(Path("C:/p/rwa_monthly_v14.xlsx"))) in code
    assert repr(str(Path("C:/p/rwa.kedge/contract.yaml"))) in code


def test_an_explicit_workbook_path_beats_the_derived_one(tmp_path: Path) -> None:
    elsewhere = tmp_path / "archive" / "2026-06.xlsx"
    code = named(cells_for(handins_dir=tmp_path / "h", workbook_path=elsewhere), "kedge_setup").code
    assert repr(str(elsewhere)) in code


# ── the head, executed against the real ingest and contracts packages ────────


def handin_file(directory: Path) -> Path:
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / "exposures.csv"
    path.write_text("id,amount\n1,10.5\n2,20.25\n3,30.0\n", encoding="utf-8")
    return path


def test_the_head_runs_end_to_end_with_no_contract(tmp_path: Path) -> None:
    """The state a notebook is in on its very first run: a file, no contract, nothing enforced."""
    picked = handin_file(tmp_path / "incoming")

    namespace = run_head(cells_for(), store=tmp_path / "store", picked=picked)

    handin = namespace["handin"]
    assert handin.original_name == "exposures.csv"
    assert handin.path.parent.is_relative_to(tmp_path / "store")  # the managed copy, not the drop
    assert handin.sha256
    assert namespace["handin_drift"].is_first_hand_in
    assert namespace["handin_contract"] is None
    assert namespace["handin_check"] is None
    assert namespace["handin_frame"].collect().height == 3


def test_the_head_takes_a_pasted_grid_and_stores_it_as_a_managed_csv(tmp_path: Path) -> None:
    """The third entry point, end to end through the real ingest machinery.

    The step before this one is usually "run this query", and what comes back is a grid on the
    clipboard rather than a file on disk. What lands in the store has to be readable by the same
    reader as a file, or the notebook is green through ingestion and wrong at the first
    calculation.
    """
    namespace = run_head(
        cells_for(),
        store=tmp_path / "store",
        pasted="id\tead\nA\t100.0\nB\t200.0\n(2 rows affected)",
    )

    handin = namespace["handin"]
    assert handin.source == "pasted"
    assert handin.path.suffix == ".csv"
    assert handin.path.parent.is_relative_to(tmp_path / "store")
    frame = namespace["handin_frame"].collect()
    assert frame.columns == ["id", "ead"]
    assert frame.height == 2  # the SSMS trailer is not a row


def test_a_selected_file_wins_over_a_paste_left_in_the_box(tmp_path: Path) -> None:
    """Precedence is by reproducibility. A stale paste must not override a chosen file."""
    picked = handin_file(tmp_path / "incoming")

    namespace = run_head(
        cells_for(), store=tmp_path / "store", picked=picked, pasted="id,ead\nZ,1.0\n"
    )

    assert namespace["handin"].source == "selected"
    assert namespace["handin"].original_name == "exposures.csv"


def test_the_head_waits_rather_than_erroring_when_nothing_is_selected(tmp_path: Path) -> None:
    with pytest.raises(StoppedError, match=r"Step \d+ of \d+"):
        run_head(cells_for(), store=tmp_path / "store")


def test_a_hand_in_that_fails_its_contract_never_reaches_the_frame(tmp_path: Path) -> None:
    """The contract gate, executed: mo.stop fires before handin_frame is defined (PLAN 2.8)."""
    from kedge.contracts import ColumnContract, Contract, save

    picked = handin_file(tmp_path / "incoming")
    contract = tmp_path / "contract.yaml"
    save(
        Contract(
            name="exposures",
            columns=[ColumnContract(name="counterparty_id", dtype="Int64")],
        ),
        contract,
    )

    with pytest.raises(StoppedError, match="does not satisfy its contract"):
        run_head(cells_for(), store=tmp_path / "store", picked=picked, contract=contract)


def test_a_hand_in_that_satisfies_its_contract_passes_through(tmp_path: Path) -> None:
    from kedge.contracts import ColumnContract, Contract, save

    picked = handin_file(tmp_path / "incoming")
    contract = tmp_path / "contract.yaml"
    save(
        Contract(
            name="exposures",
            columns=[
                ColumnContract(name="id", dtype="Int64"),
                ColumnContract(name="amount", dtype="Float64"),
            ],
        ),
        contract,
    )

    namespace = run_head(cells_for(), store=tmp_path / "store", picked=picked, contract=contract)

    assert namespace["handin_check"].ok
    assert namespace["handin_frame"].collect().height == 3


def test_a_contract_that_will_not_load_says_so_without_blocking(tmp_path: Path) -> None:
    """A YAML typo must not put the data out of reach, but it must not look controlled either."""
    picked = handin_file(tmp_path / "incoming")
    contract = tmp_path / "contract.yaml"
    contract.write_text("name: [unclosed\n", encoding="utf-8")

    namespace = run_head(cells_for(), store=tmp_path / "store", picked=picked, contract=contract)

    assert namespace["handin_check"] is None
    assert namespace["handin_contract"] is None
    assert namespace["handin_frame"].collect().height == 3


def test_the_frame_skips_a_preamble_the_way_the_contract_check_does(tmp_path: Path) -> None:
    """Reading it any other way means the notebook computes on a different frame than was
    validated, and the difference is a row."""
    incoming = tmp_path / "incoming"
    incoming.mkdir(parents=True)
    picked = incoming / "exposures.csv"
    picked.write_text("Monthly extract,,\nid,amount\n1,10.5\n2,20.25\n", encoding="utf-8")

    namespace = run_head(cells_for(), store=tmp_path / "store", picked=picked)

    frame = namespace["handin_frame"].collect()
    assert frame.columns[:2] == ["id", "amount"]
    assert frame.height == 2
    # And it says what it skipped, so a row count that does not match the file is explicable.
    assert "_Layout: skipped 1 preamble row(s) above the header._" in namespace["mo"].rendered


# ── the frame is the frame that was validated ────────────────────────────────
#
# The dangerous shape of this bug is not a crash. It is a green contract panel over rows
# nobody computed on: `validate` reads `contract.sheet` and `contract.header_row`, so anything
# in the frame cell that does not is a check of one thing and a calculation on another.


def two_sheet_workbook(directory: Path) -> Path:
    """A cover page first, the data on a named sheet second. The layout that exposed this."""
    directory.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    cover = book.active
    cover.title = "Cover"
    cover.append(["report", "generated"])
    cover.append(["exposures", "2026-06-30"])
    data = book.create_sheet("Data")
    data.append(["id", "amount"])
    for row in range(1, 6):
        data.append([row, row * 1.5])
    path = directory / "exposures.xlsx"
    book.save(path)
    return path


def contract_at(path: Path, **fields: Any) -> Path:
    from kedge.contracts import ColumnContract, Contract, save

    save(
        Contract(
            name="exposures",
            columns=[
                ColumnContract(name="id", dtype="Int64"),
                ColumnContract(name="amount", dtype="Float64"),
            ],
            **fields,
        ),
        path,
    )
    return path


def test_a_contract_naming_a_sheet_is_checked_and_computed_on_the_same_sheet(
    tmp_path: Path,
) -> None:
    """The reviewed defect: the check passed on 'Data' while the frame held the cover page."""
    picked = two_sheet_workbook(tmp_path / "incoming")
    contract = contract_at(tmp_path / "contract.yaml", sheet="Data")

    namespace = run_head(cells_for(), store=tmp_path / "store", picked=picked, contract=contract)

    validated, _layout = kedge.ingest.drift.read_data(
        namespace["handin"].path, sheet="Data", header_row=None
    )
    frame = namespace["handin_frame"].collect()
    assert namespace["handin_check"].ok
    assert frame.columns == validated.columns == ["id", "amount"]
    assert frame.height == validated.height == 5


def test_a_contract_pinning_a_header_row_is_checked_and_computed_on_the_same_rows(
    tmp_path: Path,
) -> None:
    """A header the detector would not have chosen, so honouring the contract is observable."""
    incoming = tmp_path / "incoming"
    incoming.mkdir(parents=True)
    picked = incoming / "exposures.csv"
    picked.write_text("region,currency\nid,amount\n1,10.5\n2,20.25\n", encoding="utf-8")
    contract = contract_at(tmp_path / "contract.yaml", header_row=1)

    # What the frame cell would produce if it let kedge detect the header, as it used to.
    detected, _layout = kedge.ingest.drift.read_data(picked)
    assert detected.columns == ["region", "currency"]

    namespace = run_head(cells_for(), store=tmp_path / "store", picked=picked, contract=contract)

    frame = namespace["handin_frame"].collect()
    assert namespace["handin_check"].ok
    assert frame.columns == ["id", "amount"]
    assert frame.height == 2


def test_the_shape_profile_is_taken_from_the_sheet_the_contract_names(tmp_path: Path) -> None:
    """Drift off the wrong sheet reports every column as renamed, twice a year, wrongly."""
    picked = two_sheet_workbook(tmp_path / "incoming")
    contract = contract_at(tmp_path / "contract.yaml", sheet="Data")

    namespace = run_head(cells_for(), store=tmp_path / "store", picked=picked, contract=contract)

    profile = namespace["handin_profile"]
    assert profile.sheet == "Data"
    assert profile.column_names == ("id", "amount")
    assert profile.row_count == 5


def test_the_shape_panel_says_what_actually_arrived(tmp_path: Path) -> None:
    """PLAN 2.8: the user sees what arrived before any processing happens."""
    picked = handin_file(tmp_path / "incoming")

    namespace = run_head(cells_for(), store=tmp_path / "store", picked=picked)

    panel = namespace["mo"].panel("### Hand-in shape")
    assert "3 rows, 2 columns" in panel
    assert "`id` Int64" in panel
    assert "`amount` Float64" in panel


def test_a_very_wide_hand_in_is_summarised_rather_than_rendered_whole(tmp_path: Path) -> None:
    """A 200-column extract as one markdown line is a wall, not information."""
    incoming = tmp_path / "incoming"
    incoming.mkdir(parents=True)
    picked = incoming / "wide.csv"
    headers = [f"c{index}" for index in range(30)]
    picked.write_text(
        ",".join(headers) + "\n" + ",".join(str(index) for index in range(30)) + "\n",
        encoding="utf-8",
    )

    namespace = run_head(cells_for(), store=tmp_path / "store", picked=picked)

    panel = namespace["mo"].panel("### Hand-in shape")
    assert "`c23` Int64" in panel
    assert "`c24`" not in panel
    assert "and 6 more -- see `handin_profile`" in panel


def test_the_contract_panel_names_the_rows_everything_below_will_read(tmp_path: Path) -> None:
    picked = two_sheet_workbook(tmp_path / "incoming")
    contract = contract_at(tmp_path / "contract.yaml", sheet="Data", header_row=0)

    namespace = run_head(cells_for(), store=tmp_path / "store", picked=picked, contract=contract)

    panel = namespace["mo"].panel("### Contract")
    assert "sheet `Data`" in panel
    assert "header row 0" in panel


def test_the_contract_panel_says_so_when_there_is_no_contract(tmp_path: Path) -> None:
    picked = handin_file(tmp_path / "incoming")

    namespace = run_head(cells_for(), store=tmp_path / "store", picked=picked)

    assert "nothing below is enforced" in namespace["mo"].panel("### Contract")
    assert "nothing was checked" in namespace["mo"].panel("### Contract check")


# ── checkpoints ──────────────────────────────────────────────────────────────


def test_a_checkpoint_scaffolds_a_ui_cell_and_a_blocking_gate() -> None:
    cells = cells_for()
    ui = named(cells, "manual_overrides_ui")
    gate = named(cells, "manual_overrides")

    assert ui.role == "checkpoint" and gate.role == "checkpoint"
    assert "mo.ui.dropdown(" in ui.code
    assert "mo.ui.text_area(" in ui.code
    assert "mo.stop(" not in ui.code
    assert gate.code.count("mo.stop(") == 2


def test_the_gate_blocks_on_anything_but_the_first_option() -> None:
    """The first option is the one that unblocks; everything else stops the graph.

    The comparison is against `_decision` rather than the widget, because a decision already
    recorded against this run stands: marimo's widget state dies with the kernel, so a reopened
    notebook must not re-ask for something somebody already signed.
    """
    gate = named(cells_for(), "manual_overrides").code
    assert "manual_overrides_decision.value or (_recorded.decision if _recorded else None)" in gate
    assert "_decision != 'approve'" in gate


def test_a_checkpoint_that_does_not_require_a_note_gets_one_gate_not_two() -> None:
    plan = approved(
        draft=make_draft(
            stages=[
                Stage(
                    id="sign_off",
                    intent="Someone has to look at this",
                    kind=StageKind.CHECKPOINT,
                    checkpoint=Checkpoint(question="Fine?", require_note=False),
                )
            ],
            open_questions=[],
            dropped=[],
        )
    )

    gate = named(build_cells(plan), "sign_off").code
    assert gate.count("mo.stop(") == 1
    assert "A reason is required" not in gate


def test_a_downstream_stage_reads_the_checkpoint_so_the_block_reaches_it() -> None:
    """A gated stage has to reference the checkpoint, or mo.stop never touches it."""
    code = named(cells_for(), "write_output").code
    assert '_gate_write_output = manual_overrides["decision"]' in code


def test_a_checkpoint_renders_the_guidance_a_reviewer_needs() -> None:
    plan = approved(
        draft=make_draft(
            stages=[
                Stage(
                    id="sign_off",
                    intent="Someone has to look at this",
                    kind=StageKind.CHECKPOINT,
                    checkpoint=Checkpoint(
                        question="Fine?",
                        guidance="Compare the override sheet against last month's sign-off.",
                    ),
                )
            ],
            open_questions=[],
            dropped=[],
        )
    )

    ui = named(build_cells(plan), "sign_off_ui").code
    assert "# Guidance: Compare the override sheet" in ui


def test_a_checkpoint_never_becomes_the_frame_a_stage_builds_on() -> None:
    """Its output is a decision record, not a frame."""
    code = named(cells_for(), "write_output").code
    assert "write_output = apply_haircuts.collect()" in code


def test_a_stage_listing_its_checkpoint_first_still_builds_on_a_frame() -> None:
    """The checkpoint gates the stage; the frame it transforms is the next dependency along."""
    plan = approved(
        draft=make_draft(
            stages=[
                Stage(id="load_it", intent="Read the hand-in", kind=StageKind.LOAD),
                Stage(
                    id="sign_off",
                    intent="Someone has to look at this",
                    kind=StageKind.CHECKPOINT,
                    depends_on=["load_it"],
                    checkpoint=Checkpoint(question="Fine?"),
                ),
                Stage(
                    id="summarise",
                    intent="Total it up",
                    depends_on=["sign_off", "load_it"],
                ),
            ],
            open_questions=[],
            dropped=[],
        )
    )

    code = named(build_cells(plan), "summarise").code
    assert "summarise = load_it" in code
    assert '_gate_summarise = sign_off["decision"]' in code


# ── a checkpoint that reads a file of its own ────────────────────────────────


def runbook_plan(**checkpoint_overrides: Any) -> ProcessPlan:
    """`load -> verify`, with the re-extract declared on the checkpoint that asks about it.

    Where a model puts it, and where the process itself puts it: "does the re-extract agree with
    what we predicted?" is one step, and the file it needs arrives at the moment it is asked for.
    """
    return approved(
        draft=make_draft(
            stages=[
                Stage(
                    id="load_it",
                    intent="Read the opening extract",
                    kind=StageKind.LOAD,
                    sources=[StageSource(origin=SourceOrigin.HANDIN, ref="opening extract")],
                ),
                Stage(
                    id="verify",
                    intent="Check the re-extract against what was predicted",
                    kind=StageKind.CHECKPOINT,
                    sources=[
                        StageSource(origin=SourceOrigin.HANDIN, ref="post-adjustment extract")
                    ],
                    depends_on=["load_it"],
                    checkpoint=Checkpoint(question="Does the re-extract agree?"),
                    **checkpoint_overrides,
                ),
            ],
            open_questions=[],
            dropped=[],
        )
    )


def test_a_checkpoint_that_reads_a_file_of_its_own_gets_the_cells_to_receive_it() -> None:
    """The runbook whose re-extract had nowhere to arrive.

    `build_cells` used to `continue` past a checkpoint before it looked for a hand-in source, so
    a `{origin: handin, ref: ...}` declared there scaffolded the approval card and nothing else:
    the file was read out of the plan, validated, rendered on the card the user approved, and
    then dropped. A real hub conversion stopped dead at that step, waiting for a grid the
    notebook never asked for -- and the shape that needed a second hand-in most was the one
    shape that could not have one, because a re-extract arrives at the checkpoint by definition.
    """
    cells = build_cells(runbook_plan())
    receivers = [cell.name for cell in cells if cell.role == "handin" and cell.stage_id == "verify"]

    assert receivers == ["verify_input", "verify_handin", "verify_frame"]
    assert "kedge.ingest.receive(" in named(cells, "verify_handin").code
    assert "kedge.ingest.read_data(verify_handin.path)" in named(cells, "verify_frame").code


def test_the_file_is_asked_for_before_the_decision_it_is_evidence_for() -> None:
    """Supply the re-extract, then record what you make of it -- the order a user works in.

    Emitted the other way round the page reads as an approval with its evidence filed underneath
    it, which is the same mistake `_with_reconciliation` exists to undo one cell further down.
    """
    names = [cell.name for cell in build_cells(runbook_plan())]

    assert names.index("verify_input") < names.index("verify_ui")
    assert names.index("verify_frame") < names.index("verify_ui")
    assert names.index("verify_ui") < names.index("verify")


def test_the_approval_card_stays_dark_until_that_file_has_arrived() -> None:
    """Position on the page hides nothing; only a dataflow edge does.

    A cell that just builds `mo.ui` elements reads no upstream name, so marimo has nothing to
    gate it on and it renders from the moment the notebook opens -- offering a sign-off on data
    nobody has supplied yet, above the selector asking for it. Reading the frame is the edge.
    """
    ui = named(build_cells(runbook_plan()), "verify_ui").code

    assert "_after_verify = verify_frame" in ui
    assert "mo.ui.dropdown(" in ui


def test_a_checkpoint_with_a_file_of_its_own_still_reserves_its_decision_names() -> None:
    """Both sets of derived names, not whichever the name map matched last.

    The name map reassigned one satellite function over another, so a stage that was a checkpoint
    *and* carried a hand-in reserved the receiver names and quietly stopped reserving
    `<name>_decision` and `<name>_note`. A plan with a checkpoint `review` beside a stage
    `review_decision` would then scaffold two cells defining one name, which marimo rejects as
    multiply defined and the user meets as a notebook that will not open.
    """
    plan = approved(
        draft=make_draft(
            stages=[
                Stage(id="review_decision", intent="A stage that wants the checkpoint's name"),
                Stage(
                    id="review",
                    intent="The checkpoint itself",
                    kind=StageKind.CHECKPOINT,
                    sources=[StageSource(origin=SourceOrigin.HANDIN, ref="the re-extract")],
                    checkpoint=Checkpoint(question="Approved?"),
                ),
            ],
            open_questions=[],
            dropped=[],
        )
    )

    owner: dict[str, str] = {}
    for cell in build_cells(plan):
        for name in public_names(cell.code):
            assert name not in owner, f"{name!r} defined by {owner.get(name)!r} and {cell.name!r}"
            owner[name] = cell.name


def test_a_checkpoints_own_file_does_not_summon_the_head_hand_in() -> None:
    """The head hand-in blocks the whole notebook, so it is emitted only where something reads it.

    A checkpoint reads no frame at all, and where it declares a hand-in it reads *that* one, under
    its own name. Neither is `handin_frame`, so a checkpoint's file never summons the head on its
    own account -- and here nothing else does either, because `load_it` reads a file of its own.
    """
    names = {cell.name for cell in build_cells(runbook_plan())}

    assert "handin_source" not in names, "the head hand-in was emitted for nobody to read"
    assert "handin_frame" not in names
    assert {"verify_input", "load_it_input"} <= names


def _checkpoint_only_plan(consumer: Stage) -> ProcessPlan:
    """A checkpoint carrying the plan's only named hand-in, and one stage built on it.

    The shape `_stranded_handin_warnings` tells a plan's author to take apart, scaffolded anyway,
    because a warning is advice and the notebook still has to be built.
    """
    return approved(
        draft=make_draft(
            stages=[
                Stage(
                    id="verify",
                    intent="Sign off the re-extract",
                    kind=StageKind.CHECKPOINT,
                    sources=[
                        StageSource(origin=SourceOrigin.HANDIN, ref="post-adjustment extract")
                    ],
                    checkpoint=Checkpoint(question="Does the re-extract agree?"),
                ),
                consumer,
            ],
            open_questions=[],
            dropped=[],
        )
    )


def test_a_stage_built_on_a_checkpoint_falls_through_and_the_head_is_emitted_for_it() -> None:
    """Correcting a docstring that promised the opposite, rather than the behaviour.

    `head_handin_reader` claimed a checkpoint's own hand-in "must therefore not put six cells and
    a blocking `mo.stop` at the top of the notebook for a file no step of the process names", and
    for a plan of checkpoint-plus-output it did exactly that: two file boxes for one declared
    file, and the blocking one is the box nothing named.

    It is not the walk's doing. `_upstream_name` never treats a checkpoint as a frame, so a stage
    depending only on one falls through to `handin_frame` -- something genuinely does read the
    head hand-in, and emitting it is right. What is wrong is the plan, and the approval card says
    so in `_stranded_handin_warnings`; the docstring now says which of the two it is.
    """
    consumer = Stage(
        id="report",
        intent="The impact statement",
        kind=StageKind.OUTPUT,
        depends_on=["verify"],
    )
    cells = build_cells(_checkpoint_only_plan(consumer))
    names = [cell.name for cell in cells]

    assert "handin_frame" in names
    assert "report = handin_frame.collect()" in named(cells, "report").code
    assert [name for name in names if name.endswith("_input")] == ["verify_input"]


def test_the_acceptance_is_keyed_to_the_file_the_arithmetic_actually_ran_on() -> None:
    """PLAN 6.2 and CLAUDE.md non-negotiable 6: a wrong baseline is worse than a loud break.

    A checkpoint's hand-in is an ancestor of anything downstream of the checkpoint, and
    `_baseline_handin` walked ancestors without asking whether the frame was *read*. So the panel
    cited `verify_handin.sha256` while `compare` was built on `handin_frame` -- the head hand-in,
    reached by fall-through -- and `check_translation` decides by digest whether a later run may
    re-compare itself against the workbook. Keyed to a file that takes no part in the computation,
    that decision means nothing, which is what `_baseline_handin`'s own docstring forbids.

    Before a checkpoint's hand-in scaffolded any cells this was a `NameError`: loud, and broken.
    Quiet and wrong is the worse of the two.
    """
    consumer = Stage(
        id="compare",
        intent="Compare the re-extract against the prediction",
        depends_on=["verify"],
        operations=["calc_h2_h500"],
    )
    cells = build_cells(_checkpoint_only_plan(consumer))

    computed = named(cells, "compare").code
    cited = [
        line.strip()
        for line in named(cells, "reconciliation").code.splitlines()
        if "handin_sha256" in line
    ]

    assert "compare = handin_frame" in computed
    assert cited == ["handin_sha256=handin.sha256,"]


def test_a_gated_checkpoint_with_a_file_of_its_own_reads_both_and_assigns_once() -> None:
    """Two `_after_<name> = ...` lines assigning one name made the first a dead store.

    Harmless to marimo, which took the edges either way, and not harmless in a cell a reviewer is
    meant to read: the line that looks like it holds the frame holds the checkpoint above it.
    """
    plan = approved(
        draft=make_draft(
            stages=[
                Stage(id="approve", intent="Approve first", kind=StageKind.CHECKPOINT),
                Stage(
                    id="verify",
                    intent="Sign off the re-extract",
                    kind=StageKind.CHECKPOINT,
                    sources=[
                        StageSource(origin=SourceOrigin.HANDIN, ref="post-adjustment extract")
                    ],
                    depends_on=["approve"],
                    checkpoint=Checkpoint(question="Does the re-extract agree?"),
                ),
            ],
            open_questions=[],
            dropped=[],
        )
    )

    ui = named(build_cells(plan), "verify_ui").code
    reads = [line for line in ui.splitlines() if line.startswith("_after_verify")]

    assert reads == ["_after_verify = (verify_frame, approve)"]


# ── a hand-off, and the two halves of one step ───────────────────────────────
#
# "Run this query" and "give me what it returned" are one step of a runbook, and they were
# coming out in the wrong order. The box asking for a file builds `mo.ui` elements and reads
# nothing, so nothing could hide it; the query above it was behind a `mo.stop` on parameters
# nobody had been shown a reason to fill in. A fresh open therefore offered a drop zone for an
# extract and nowhere the extract's query -- which a real user met as "where is the sql to run
# to get the starting data?", then "i can't see it".


def handoff_plan() -> ProcessPlan:
    """The shape every runbook has: extract, load, update, re-extract, load.

    Both hand-offs take a period end, because that is the parameter almost every periodic
    process has and it is the one that used to hide the step asking for it.
    """
    return approved(
        draft=make_draft(
            stages=[
                Stage(
                    id="extract",
                    intent="Hand over the opening extract",
                    kind=StageKind.HANDOFF,
                    handoff=Handoff(
                        instruction="Run this against the warehouse and bring the grid back.",
                        statement="SELECT * FROM t WHERE period_end = {period_end}",
                        connection="Warehouse",
                        parameters=["period_end"],
                    ),
                ),
                Stage(
                    id="load_extract",
                    intent="Read the opening extract",
                    kind=StageKind.LOAD,
                    sources=[StageSource(origin=SourceOrigin.HANDIN, ref="opening extract")],
                    depends_on=["extract"],
                ),
                Stage(
                    id="apply",
                    intent="Hand over the update",
                    kind=StageKind.HANDOFF,
                    depends_on=["load_extract"],
                    handoff=Handoff(
                        instruction="Run this in one transaction.",
                        statement="UPDATE t SET a = 1 WHERE period_end = {period_end}",
                        parameters=["period_end"],
                        mutates=True,
                    ),
                ),
                Stage(
                    id="re_extract",
                    intent="Hand over the same extract again",
                    kind=StageKind.HANDOFF,
                    depends_on=["apply"],
                    handoff=Handoff(
                        instruction="Run the extract again now the update has been applied.",
                        statement="SELECT * FROM t WHERE period_end = {period_end}",
                        parameters=["period_end"],
                    ),
                ),
                Stage(
                    id="reload",
                    intent="Read the re-extract",
                    kind=StageKind.LOAD,
                    sources=[StageSource(origin=SourceOrigin.HANDIN, ref="post-update extract")],
                    depends_on=["re_extract", "apply"],
                ),
            ],
            open_questions=[],
            dropped=[],
        )
    )


def run_handoff(cells: list[ScaffoldCell], name: str, **values: Any) -> dict[str, Any]:
    """Execute one hand-off cell with its inputs supplied, or deliberately not supplied.

    Executed rather than read, because the claim is about what a user is shown. A test that
    matched on the emitted text would have passed just as happily against the `mo.stop` this
    replaced: the step's heading was in that cell's source the whole time, and never once on
    screen.
    """
    recorded: list[dict[str, Any]] = []
    namespace: dict[str, Any] = {
        "mo": FakeMarimo(),
        "kedge": SimpleNamespace(
            sql=kedge.sql,
            runs=SimpleNamespace(
                record_parameters=lambda _runs, _id, **kw: recorded.append(kw),
            ),
        ),
        "KEDGE_RUNS": Path("runs"),
        "KEDGE_RUN_ID": "20260827T000000Z",
        "recorded": recorded,
        **{f"{name}_{key}": Selection(value) for key, value in values.items()},
    }
    exec(compile(named(cells, name).code, f"<{name}>", "exec"), namespace)
    return namespace


def test_a_handoff_waiting_on_its_parameters_still_shows_the_user_the_step() -> None:
    """The first thing a runbook asks of anybody must be on screen when it opens.

    A `mo.stop` here took the heading and the instruction down with the statement, and in app
    mode everything below a stopped cell goes too -- so the page a user met was a date box and
    a drop zone for a file, with nothing saying what either was for.
    """
    namespace = run_handoff(build_cells(handoff_plan()), "extract", period_end=None)

    rendered = namespace["mo"].rendered
    assert "### Step 1 of 5 -- run this: extract" in rendered
    assert "Run this against the warehouse and bring the grid back." in rendered
    assert namespace["extract"] is None, "an unfilled parameter must not build a statement"
    assert namespace["recorded"] == [], "nor record itself against the run"
    waiting = [text for text in rendered if "Still needed" in text]
    assert waiting and "period end" in waiting[0], (
        f"the step does not say which input it is waiting for: {rendered}"
    )
    assert not any("SELECT" in text for text in rendered), (
        "an unscoped statement is worse than none: it is copyable"
    )


def test_a_handoff_with_its_parameters_filled_in_renders_the_statement() -> None:
    """The other half of the same rule -- withholding the statement, never the step."""
    import datetime as dt

    cells = build_cells(handoff_plan())
    namespace = run_handoff(cells, "extract", period_end=dt.date(2026, 6, 30))

    assert namespace["extract"] == "SELECT * FROM t WHERE period_end = DATE '2026-06-30'"
    assert namespace["recorded"] == [{"period_end": dt.date(2026, 6, 30)}]
    assert any("```sql" in text for text in namespace["mo"].rendered)


def test_a_handin_selector_cannot_render_above_the_query_that_produces_the_file() -> None:
    """The ordering rule, and the only mechanism marimo offers for it.

    A read-only hand-off scaffolds no confirmation, so before this the plan's own
    `depends_on: [extract]` bought the selector nothing at all and the box rendered from the
    moment the notebook opened. Reading the statement is a weaker claim than reading a
    confirmation -- it says the step above is on screen, not that anybody has done it -- and
    ordering the two halves of one step is exactly what it is for.
    """
    selector = named(build_cells(handoff_plan()), "load_extract_input").code
    reads = [line for line in selector.splitlines() if line.startswith("_after_load_extract")]

    assert reads == ["_after_load_extract = extract"]


def test_a_re_extract_still_waits_for_the_confirmation_and_not_merely_the_statement() -> None:
    """Ordering must not be bought at the price of the gate that matters.

    A mutating hand-off's token stays its confirmation, so the box for the re-extract appears
    only once somebody has said the UPDATE was run. A statement token here would put the box on
    screen as soon as the UPDATE was *rendered*, which is the defect
    `MUST_BE_HIDDEN_AT_THE_START` exists for: a grid pasted before the update ran looks exactly
    like one pasted after, and nothing afterwards can tell them apart.
    """
    cells = build_cells(handoff_plan())
    reload_reads = [
        line for line in named(cells, "reload_input").code.splitlines() if line.startswith("_after")
    ]
    inputs_reads = [
        line
        for line in named(cells, "re_extract_inputs").code.splitlines()
        if line.startswith("_after")
    ]

    assert reload_reads == ["_after_reload = (re_extract, apply_confirmed)"]
    assert inputs_reads == ["_after_re_extract_inputs = apply_confirmed"], (
        "the boxes asking which period to re-extract for render before the update is confirmed"
    )


def test_a_confirmation_is_not_offered_for_a_statement_that_was_never_built() -> None:
    """ "I have run the statement above" over an unbuilt statement is a way to record a lie.

    Reading the name is enough of an edge where the statement is fixed or generated. Where it
    waits on a parameter the name is bound either way -- to `None` -- and an edge cannot see
    that, so the check has to be on the value.
    """
    ui = named(build_cells(handoff_plan()), "apply_ran").code

    assert "mo.stop(" in ui
    assert "apply is None" in ui
    assert "Step 3 of 5" in ui


# ── what a stage cell tells the person who has to finish it ──────────────────


def test_a_stage_cell_says_where_each_of_its_inputs_comes_from() -> None:
    """The comment is what somebody reads before writing the translation. `manual` and `query`
    are the two it most needs to say: one is a value nobody can reproduce, the other is a table
    that arrives from outside the workbook entirely."""
    plan = approved(
        draft=make_draft(
            stages=[
                Stage(
                    id="load_exposures",
                    intent="Pull exposures and this month's adjustments",
                    kind=StageKind.LOAD,
                    sources=[
                        StageSource(origin=SourceOrigin.QUERY, ref="MonthlyExposures"),
                        StageSource(origin=SourceOrigin.HANDIN),
                        StageSource(origin=SourceOrigin.MANUAL, ref="Adjustments!B2:B15"),
                    ],
                )
            ],
            open_questions=[],
            dropped=[],
        )
    )

    code = named(build_cells(plan), "load_exposures").code

    assert "# Sources: query MonthlyExposures, handin, manual Adjustments!B2:B15" in code


def test_a_long_sources_comment_never_splits_an_origin_from_its_ref() -> None:
    """`_comment` wraps prose at any space, and since 1.1 a rendered source has one in it.

    Wrapped, a long list could end a comment line at `power_query` and put the ref on the next,
    which reads to the person finishing the cell as a table the plan never identified.
    """
    plan = approved(
        draft=make_draft(
            stages=[
                Stage(
                    id="load_everything",
                    intent="Pull every input this month needs",
                    kind=StageKind.LOAD,
                    sources=[
                        StageSource(
                            origin=SourceOrigin.POWER_QUERY, ref=f"CounterpartyRatings{index}"
                        )
                        for index in range(6)
                    ],
                )
            ],
            open_questions=[],
            dropped=[],
        )
    )

    lines = named(build_cells(plan), "load_everything").code.splitlines()

    for line in lines:
        assert not line.rstrip().endswith("power_query")
    for index in range(6):
        assert any(f"power_query CounterpartyRatings{index}" in line for line in lines)


# ── the reconciliation tail ──────────────────────────────────────────────────


def test_the_tail_calls_reconcile_rather_than_asserting_an_outcome() -> None:
    code = named(cells_for(), "reconciliation").code
    assert "kedge.reconcile.check_translation(" in code


def test_the_tail_maps_analysis_operation_ids_to_the_cells_that_reproduce_them() -> None:
    """`Stage.operations` is the link back to the facts, and this is what consumes it."""
    code = named(cells_for(), "reconciliation_values").code
    assert "'calc_h2_h500': apply_haircuts," in code


def test_an_operation_claimed_by_two_stages_is_mapped_once() -> None:
    """A duplicate key in the literal would silently keep whichever stage came last."""
    plan = approved(
        draft=make_draft(
            stages=[
                Stage(id="first", intent="Half of it", operations=["calc_h2_h500"]),
                Stage(
                    id="second",
                    intent="The other half",
                    depends_on=["first"],
                    operations=["calc_h2_h500"],
                ),
            ],
            open_questions=[],
            dropped=[],
        )
    )

    code = named(build_cells(plan), "reconciliation_values").code
    assert code.count("'calc_h2_h500':") == 1
    assert "'calc_h2_h500': first," in code


def test_a_stage_naming_no_operation_contributes_no_region() -> None:
    """There is no honest guess about which workbook range an unlinked stage reproduces."""
    plan = approved(
        draft=make_draft(
            stages=[Stage(id="only", intent="No operations named")],
            open_questions=[],
            dropped=[],
        )
    )

    code = named(build_cells(plan), "reconciliation_values").code
    assert "reconciliation_values = {}" in code
    assert "nothing to map" in code


def test_the_tail_never_uses_the_word_passed_as_a_verdict() -> None:
    """The only route to a pass is through kedge.reconcile, which refuses an unearned one."""
    code = named(cells_for(), "reconciliation").code
    assert "PASSED" not in code
    assert 'reconciliation = "' not in code


def test_a_missing_workbook_reconciles_to_not_reconciled(tmp_path: Path) -> None:
    """Non-negotiable 6, executed rather than asserted on the text."""
    panel = run_reconciliation(
        cells_for(),
        tmp_path / "never-existed.xlsx",
        apply_haircuts=pl.DataFrame({"haircut": [0.1, 0.2]}),
    )

    assert panel.status is ReconciliationStatus.NOT_RECONCILED
    assert not panel
    assert "NOT RECONCILED" in panel.report.headline()
    assert "not a pass" in " ".join(panel.report.notes)


def test_a_workbook_with_no_cached_values_reconciles_to_not_reconciled(tmp_path: Path) -> None:
    """The dangerous case: a real workbook, a real run, and nothing to compare against."""
    workbook = plain_workbook(tmp_path / "no-formulas.xlsx")
    panel = run_reconciliation(
        cells_for(), workbook, apply_haircuts=pl.DataFrame({"haircut": [0.1, 0.2]})
    )

    assert panel.status is ReconciliationStatus.NOT_RECONCILED
    assert not panel
    assert "PASSED" not in kedge.reconcile.panel_html(panel.report)
    assert "not a pass" in " ".join(panel.report.notes).lower()


def test_the_tail_survives_a_reconciliation_that_raises(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A failure to run is still not a pass, and it must not reach the user as a traceback."""

    def explode(*_args: Any, **_kwargs: Any) -> Any:
        raise kedge.ReconciliationError("the workbook is a zip bomb")

    # Patched on the module the acceptance layer calls through, not on the package: a converted
    # notebook now goes via `check_translation`, and what has to survive is the whole path.
    monkeypatch.setattr(kedge.reconcile.acceptance, "reconcile_workbook", explode)
    workbook = plain_workbook(tmp_path / "unreadable.xlsx")

    panel = run_reconciliation(
        cells_for(),
        workbook,
        apply_haircuts=pl.DataFrame({"haircut": [0.1]}),
    )

    assert panel.status is ReconciliationStatus.NOT_RECONCILED
    assert "zip bomb" in " ".join(panel.report.notes)


def test_the_tail_scaffolds_for_a_plan_that_is_nothing_but_checkpoints() -> None:
    plan = approved(
        draft=make_draft(
            stages=[
                Stage(
                    id="sign_off",
                    intent="Someone has to look at this",
                    kind=StageKind.CHECKPOINT,
                    checkpoint=Checkpoint(question="Fine?"),
                )
            ],
            open_questions=[],
            dropped=[],
        )
    )

    code = named(build_cells(plan), "reconciliation_values").code
    ast.parse(code)
    assert "reconciliation_values = {}" in code
