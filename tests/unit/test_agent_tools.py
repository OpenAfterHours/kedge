"""The tool surface: the caps, the choke point, and what every tool does when its collaborator
is missing.

The cap tests are the important ones. PLAN 2.3's rule is not "each tool should remember to
truncate" — it is that a payload cannot reach the model uncapped — so the tests here go at the
choke point rather than at the individual tools: a deliberately leaky handler is registered and
:meth:`ToolRegistry.dispatch` is asserted to cap it anyway. That is the property that survives
somebody adding a sixteenth tool.
"""

from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook as OpenpyxlWorkbook
from openpyxl import load_workbook

from conftest import approved_plan_store, make_analysis, make_plan
from kedge.agent import tools as tools_module
from kedge.agent.audit import OutboundLog
from kedge.agent.context import CellFacts, NotebookState
from kedge.agent.tools import (
    MAX_CACHED_RESULTS,
    MAX_PAYLOAD_BYTES,
    MAX_ROWS,
    Caps,
    ToolContext,
    ToolRegistry,
    ToolResult,
    ToolSpec,
    Volatility,
    tool_names,
    tool_schemas,
    volatility_of,
)
from kedge.analysis.model import CachedValueCoverage, WorkbookAnalysis
from kedge.notebook.driver import MultiplyDefinedError, StaleCellError
from kedge.notebook.model import (
    CellInfo,
    CellRef,
    GraphNode,
    GraphView,
    MutationResult,
    ProbeResult,
)
from kedge.plan.model import ApprovalState
from kedge.plan.propose import plan_json_schema
from kedge.plan.review import acknowledge_all_drops, approve
from kedge.plan.store import PlanStore
from kedge.plan.triage import TriageVerdict, triage
from kedge.workspace import Workspace

SENTINEL = "ZEPHYR-QX-9931-COUNTERPARTY"
RESAVED = "MERIDIAN-VT-4417-COUNTERPARTY"
"""The value the workbook carries after the user has saved it from Excel mid-session.

Distinct from :data:`SENTINEL` so a stale read is a failed assertion rather than a coincidence:
the tests that exercise the reopen assert both that the new value came back *and* that the old
one did not.
"""


# ── fakes ────────────────────────────────────────────────────────────────────────────────────


class FakeDriver:
    """Enough of :class:`~kedge.notebook.model.NotebookBridge` to exercise the tools."""

    def __init__(self, *, fail: Exception | None = None) -> None:
        self.created: list[tuple[str, str]] = []
        self.edited: list[tuple[str, str]] = []
        self.probes: list[str] = []
        self.deleted: list[str] = []
        self.probe_result: ProbeResult | None = None
        self._fail = fail

    async def read_graph(self) -> GraphView:
        return GraphView(
            nodes=(
                GraphNode(id="AAaa", name="imports", defs=("pl",)),
                GraphNode(id="MJUe", name="load_handin", defs=("load_handin",), refs=("pl",)),
            )
        )

    async def list_cells(self, *, with_code: bool = True) -> tuple[CellInfo, ...]:
        return (
            CellInfo(id="AAaa", name="imports", code="import polars as pl" if with_code else None),
            CellInfo(
                id="MJUe",
                name="load_handin",
                code='load_handin = pl.scan_parquet("h.parquet")' if with_code else None,
            ),
        )

    async def get_cell(self, target: str | int) -> CellInfo:
        return CellInfo(id="MJUe", name="load_handin", code="load_handin = 1", status="idle")

    async def create_cell(self, code: str, *, name: str, **_kw: Any) -> MutationResult:
        if self._fail is not None:
            raise self._fail
        self.created.append((name, code))
        return MutationResult(
            operation="create_cell", cell=CellRef(id="U1", name=name), ran=True, status="idle"
        )

    async def edit_cell(
        self, target: str | int, code: str | None = None, **_kw: Any
    ) -> MutationResult:
        if self._fail is not None:
            raise self._fail
        self.edited.append((str(target), code or ""))
        return MutationResult(
            operation="edit_cell",
            cell=CellRef(id="MJUe", name=str(target)),
            ran=True,
            status="idle",
        )

    async def run_cell(self, target: str | int) -> MutationResult:
        return MutationResult(
            operation="run_cell", cell=CellRef(id="MJUe", name=str(target)), ran=True, status="idle"
        )

    async def delete_cell(self, target: str | int) -> MutationResult:
        self.deleted.append(str(target))
        return MutationResult(operation="delete_cell", cell=CellRef(id="MJUe", name=str(target)))

    async def probe(self, code: str) -> ProbeResult:
        self.probes.append(code)
        if self.probe_result is not None:
            return self.probe_result
        return ProbeResult(ok=True, value_repr=f"['{SENTINEL}']", value_type="list")


@pytest.fixture
def state() -> NotebookState:
    return NotebookState(
        cells=(
            CellFacts(id="AAaa", name="imports", defs=("pl",)),
            CellFacts(id="MJUe", name="load_handin", defs=("load_handin",), refs=("pl",)),
        )
    )


@pytest.fixture
def registry(state: NotebookState, tmp_path: Path) -> ToolRegistry:
    return _notebook_tools(state, tmp_path)


def _notebook_tools(
    state: NotebookState, plans: Path, *, driver: Any = None, **context: Any
) -> ToolRegistry:
    """A registry that may write cells: a driver, and a plan the user has approved.

    The approved plan is a default rather than something each test arranges, because
    `propose_cell` and `edit_cell` are refused without one and almost nothing here is *about*
    that. A test whose subject is the gate itself builds its own store and says so.
    """
    tools = ToolRegistry(
        ToolContext(driver=driver or FakeDriver(), plans=approved_plan_store(plans), **context)
    )
    tools.refresh(state)
    return tools


# ── schemas ──────────────────────────────────────────────────────────────────────────────────


def test_every_tool_plan_m4_lists_is_offered() -> None:
    assert set(tool_names()) == {
        "list_cells",
        "propose_cell",
        "edit_cell",
        "run_cell",
        "delete_cell",
        "inspect_workbook",
        "sample_data",
        "profile_column",
        "read_range",
        "probe",
        "get_plan",
        "propose_plan",
        "amend_plan",
        "reconcile",
        "list_utils",
        "get_knowledge",
    }


def test_schemas_are_valid_openai_function_definitions() -> None:
    for schema in tool_schemas():
        assert schema["type"] == "function"
        function = schema["function"]
        assert function["name"] and function["description"]
        parameters = function["parameters"]
        assert parameters["type"] == "object"
        assert parameters["additionalProperties"] is False
        assert set(parameters["required"]) <= set(parameters["properties"])


def test_the_defaults_are_the_ones_plan_2_3_states() -> None:
    assert MAX_ROWS == 100
    assert MAX_PAYLOAD_BYTES == 32_768


def test_caps_come_from_config() -> None:
    workspace = Workspace.for_workbook("book.xlsx")
    caps = Caps.from_config(workspace.config)
    assert caps.max_rows == workspace.config.sampling.max_rows
    assert caps.max_payload_bytes == workspace.config.sampling.max_payload_bytes


def test_the_policy_allowlists_reach_the_gate_from_config(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """This is the only production construction of `Policy`, so an allowlist that does not
    arrive here is an allowlist no user can reach."""
    monkeypatch.setenv("KEDGE_HOME", str(tmp_path / "home"))
    workbook = tmp_path / "rwa_monthly.xlsx"
    _write_workbook(workbook)
    (tmp_path / "kedge.toml").write_text(
        "[policy]\n"
        'network_allowlist = ["rates.internal.bank"]\n'
        'database_allowlist = ["RiskWarehouse"]\n',
        encoding="utf-8",
    )
    workspace = Workspace.for_workbook(workbook)
    workspace.ensure_dirs()

    context = ToolContext.for_workspace(workspace)

    assert context.policy.allows_host("rates.internal.bank")
    assert context.policy.permits_database("riskwarehouse")
    assert not context.policy.permits_database("warehouse.internal")
    assert context.policy.working_dir == workspace.project_dir


def test_with_no_policy_section_the_gate_permits_neither_network_nor_database(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setenv("KEDGE_HOME", str(tmp_path / "home"))
    workbook = tmp_path / "rwa_monthly.xlsx"
    _write_workbook(workbook)
    workspace = Workspace.for_workbook(workbook)
    workspace.ensure_dirs()

    context = ToolContext.for_workspace(workspace)

    assert context.policy.network_allowlist == frozenset()
    assert context.policy.database_allowlist == frozenset()


# ── volatility ───────────────────────────────────────────────────────────────────────────────


def test_only_the_tools_reading_files_kedge_never_writes_are_session_stable() -> None:
    stable = {name for name in tool_names() if volatility_of(name) is Volatility.SESSION_STABLE}
    assert stable == {"inspect_workbook", "sample_data", "profile_column", "read_range"}


def test_the_tools_reading_kedges_own_artifacts_are_tied_to_them() -> None:
    tied = {name for name in tool_names() if volatility_of(name) is Volatility.ARTIFACT_TIED}
    assert tied == {"get_plan", "list_utils", "get_knowledge"}


def test_everything_touching_the_kernel_or_the_notebook_is_volatile() -> None:
    volatile = {name for name in tool_names() if volatility_of(name) is Volatility.VOLATILE}
    assert volatile == {
        "list_cells",
        "propose_cell",
        "edit_cell",
        "run_cell",
        "delete_cell",
        "probe",
        "reconcile",
        # Both plan tools put a decision to the user, and the answer to "may I propose a plan?"
        # changes the moment one is approved. Neither may ever be answered from a cache.
        "propose_plan",
        "amend_plan",
    }


def test_a_spec_that_says_nothing_about_volatility_is_volatile() -> None:
    # Opting in is the whole point: a sixteenth tool must not inherit a cache by omission.
    spec = ToolSpec(name="sixteenth", description="added later", properties={})
    assert spec.volatility is Volatility.VOLATILE


def test_an_unknown_tool_name_is_treated_as_volatile() -> None:
    assert volatility_of("invented_by_the_model") is Volatility.VOLATILE


def test_volatility_never_reaches_the_schemas_the_model_is_sent() -> None:
    # The schemas are the head of the system prompt, which is the prompt cache's prefix. A key
    # added there for kedge's own use is a silent cost regression on every turn of every session.
    for schema in tool_schemas():
        assert set(schema) == {"type", "function"}
        assert set(schema["function"]) == {"name", "description", "parameters"}
        assert set(schema["function"]["parameters"]) == {
            "type",
            "properties",
            "required",
            "additionalProperties",
        }
    assert "volatility" not in json.dumps(tool_schemas())


# ── the caps ─────────────────────────────────────────────────────────────────────────────────


def test_the_row_cap_truncates_with_the_marker_plan_m4_requires() -> None:
    rows = [[index, index * 2] for index in range(500)]
    result = ToolResult.from_rows(rows, columns=["a", "b"], caps=Caps(max_rows=100))
    assert result.row_count == 100
    assert result.omitted_rows == 400
    assert result.truncated
    assert result.text.endswith("[… 400 more rows omitted]")


def test_the_marker_counts_rows_that_were_never_read_at_all() -> None:
    rows = [[index] for index in range(10)]
    result = ToolResult.from_rows(rows, columns=["a"], caps=Caps(), total=49_999)
    assert result.omitted_rows == 49_989
    assert "[… 49989 more rows omitted]" in result.text


def test_the_payload_cap_sheds_rows_rather_than_cutting_mid_value() -> None:
    rows = [[f"counterparty-{index:05d}", index * 1.5] for index in range(400)]
    result = ToolResult.from_rows(
        rows, columns=["name", "exposure"], caps=Caps(max_rows=400, max_payload_bytes=2_000)
    )
    assert result.byte_count <= 2_000
    assert result.row_count < 400
    assert result.truncated
    assert f"[… {result.omitted_rows} more rows omitted]" in result.text
    # Every rendered row is still complete JSON, which is the point of shedding rather than cutting.
    for line in result.text.splitlines():
        if line.startswith("[") and "omitted" not in line:
            json.loads(line)


def test_an_untruncated_result_carries_no_marker() -> None:
    result = ToolResult.from_rows([[1], [2]], columns=["a"], caps=Caps())
    assert not result.truncated
    assert "omitted" not in result.text


class _LeakyRegistry(ToolRegistry):
    """A handler that ignores the caps entirely, to prove the choke point does not."""

    async def _tool_get_knowledge(self, args: Any) -> ToolResult:
        return ToolResult(text="x" * 200_000, row_count=200_000)


async def test_dispatch_caps_a_handler_that_did_not_cap_itself() -> None:
    tools = _LeakyRegistry(ToolContext(caps=Caps(max_payload_bytes=4_096)))
    result = await tools.dispatch("get_knowledge", {})
    assert result.byte_count <= 4_096
    assert result.truncated
    assert "truncated at the payload cap" in result.text


async def test_dispatch_audits_the_capped_payload_not_the_original(tmp_path: Path) -> None:
    log = OutboundLog(tmp_path / "outbound-s1.jsonl", session="s1")
    tools = _LeakyRegistry(ToolContext(caps=Caps(max_payload_bytes=4_096)), log=log)
    result = await tools.dispatch("get_knowledge", {}, turn_id="t1")
    entries = list(log.entries())
    assert len(entries) == 1
    assert entries[0]["byte_count"] == result.byte_count <= 4_096
    assert entries[0]["truncated"] is True
    assert entries[0]["turn_id"] == "t1"


# ── dispatch ─────────────────────────────────────────────────────────────────────────────────


async def test_an_unknown_tool_is_a_result_rather_than_an_exception(registry: ToolRegistry) -> None:
    result = await registry.dispatch("delete_everything", {})
    assert not result.ok
    assert "no tool called" in result.text
    assert "propose_cell" in result.text


async def test_malformed_arguments_are_a_result_rather_than_an_exception(
    registry: ToolRegistry,
) -> None:
    result = await registry.dispatch("read_range", "{not json")
    assert not result.ok
    assert "not valid JSON" in result.text


async def test_missing_required_arguments_are_reported(registry: ToolRegistry) -> None:
    result = await registry.dispatch("read_range", {"sheet": "Calc"})
    assert not result.ok
    assert "range" in result.text


async def test_a_cell_body_with_raw_newlines_is_still_accepted(registry: ToolRegistry) -> None:
    raw = '{"name": "apply_haircuts", "code": "apply_haircuts = 1\n"}'
    result = await registry.dispatch("propose_cell", raw)
    assert result.ok, result.text


# ── notebook tools ───────────────────────────────────────────────────────────────────────────


async def test_propose_cell_goes_through_the_validation_gate(
    state: NotebookState, tmp_path: Path
) -> None:
    driver = FakeDriver()
    tools = _notebook_tools(state, tmp_path, driver=driver)
    result = await tools.dispatch(
        "propose_cell",
        {"name": "loader", "code": "import pandas as pd\nframe = pd.DataFrame()\n"},
    )
    assert not result.ok
    assert result.validated is False
    assert any("polars, never pandas" in violation for violation in result.violations)
    assert driver.created == []  # nothing reached the kernel


async def test_propose_cell_reports_a_kernel_side_rejection_as_violations(
    state: NotebookState, tmp_path: Path
) -> None:
    message = "Multiply-defined names:\n  - 'pl' is already defined in cell 'AAaa' (imports)\n"
    driver = FakeDriver(fail=MultiplyDefinedError(message))
    tools = _notebook_tools(state, tmp_path, driver=driver)
    result = await tools.dispatch("propose_cell", {"name": "second", "code": "value = 1\n"})
    assert not result.ok
    assert result.validated is False
    assert "'pl'" in result.violations[0]


async def test_edit_cell_surfaces_staleness_as_a_retryable_result(
    state: NotebookState, tmp_path: Path
) -> None:
    driver = FakeDriver(fail=StaleCellError("cell 'MJUe' changed since it was read"))
    tools = _notebook_tools(state, tmp_path, driver=driver)
    result = await tools.dispatch("edit_cell", {"cell": "load_handin", "code": "value = 1\n"})
    assert not result.ok
    assert "list_cells" in result.text
    assert "resubmitting the same body" in result.text


# ── the review gate on the writing tools ─────────────────────────────────────────────────────
#
# `sync_notebook` refuses an unapproved plan structurally and with no override, and its
# docstring calls itself "the one place that consumes" approval. It was not: `propose_cell` and
# `edit_cell` reach the same kernel and never looked at the plan at all, so the only thing between
# a plan the user had declined and the cells implementing it was a sentence in the system prompt.
# These are the tests for the gate that replaced the sentence.


def _unapproved(directory: Path) -> PlanStore:
    """A store holding a plan the model proposed and the user has not approved."""
    store = PlanStore(directory)
    store.save(make_plan())
    return store


@pytest.mark.parametrize(
    ("tool", "args"),
    [
        ("propose_cell", {"name": "apply_haircuts", "code": "apply_haircuts = 1\n"}),
        ("edit_cell", {"cell": "load_handin", "code": "load_handin = 1\n"}),
    ],
)
async def test_writing_a_cell_is_refused_until_the_user_has_approved_a_plan(
    state: NotebookState, tmp_path: Path, tool: str, args: dict[str, str]
) -> None:
    driver = FakeDriver()
    tools = ToolRegistry(
        ToolContext(driver=driver, plans=_unapproved(tmp_path / "plans")),
    )
    tools.refresh(state)

    result = await tools.dispatch(tool, args)

    assert not result.ok
    assert "propose_plan" in result.text, "the refusal has to say what to do instead"
    assert driver.created == [] and driver.edited == [], "nothing reached the kernel"


@pytest.mark.parametrize(
    ("tool", "args"),
    [
        ("propose_cell", {"name": "apply_haircuts", "code": "apply_haircuts = 1\n"}),
        ("edit_cell", {"cell": "load_handin", "code": "load_handin = 1\n"}),
    ],
)
async def test_an_approved_plan_lets_the_writing_tools_through(
    state: NotebookState, tmp_path: Path, tool: str, args: dict[str, str]
) -> None:
    tools = _notebook_tools(state, tmp_path / "plans")

    result = await tools.dispatch(tool, args)

    assert result.ok, result.text
    driver = tools.context.driver
    assert driver.created or driver.edited


async def test_the_writing_tools_are_refused_when_there_is_no_plan_store_at_all(
    state: NotebookState,
) -> None:
    """No store is not a reason to proceed. Nothing can have been approved."""
    driver = FakeDriver()
    tools = ToolRegistry(ToolContext(driver=driver))
    tools.refresh(state)

    result = await tools.dispatch("propose_cell", {"name": "c", "code": "c = 1\n"})

    assert not result.ok
    assert "no plan store" in result.text
    assert "propose_plan" in result.text
    assert driver.created == []


async def test_the_plan_refusal_does_not_read_as_something_to_retry(
    state: NotebookState, tmp_path: Path
) -> None:
    """A decision the user has yet to make is not a transient failure.

    The loop counts rejected drafts and re-prompts; a refusal a model reads as flaky spends the
    rest of the turn resending the same cell rather than telling the user what it is waiting for.
    """
    tools = ToolRegistry(ToolContext(driver=FakeDriver(), plans=_unapproved(tmp_path / "plans")))
    tools.refresh(state)

    result = await tools.dispatch("propose_cell", {"name": "c", "code": "c = 1\n"})

    assert "not a failure to retry" in result.text
    assert result.summary == "refused: no approved plan is in force"


async def test_the_plan_gate_comes_before_the_validation_gate(
    state: NotebookState, tmp_path: Path
) -> None:
    """The most fundamental reason first, and the one the model cannot fix by rewriting the cell.

    Reported the other way round, a model whose plan was declined is told its polars is wrong,
    fixes it, and is refused for a reason it was never given — three times, into the retry cap.
    """
    tools = ToolRegistry(ToolContext(driver=FakeDriver(), plans=_unapproved(tmp_path / "plans")))
    tools.refresh(state)

    result = await tools.dispatch(
        "propose_cell", {"name": "loader", "code": "import pandas as pd\nframe = pd.DataFrame()\n"}
    )

    assert not result.ok
    assert "propose_plan" in result.text
    assert result.violations == (), "the body was never validated, so there is nothing to fix"


async def test_running_and_deleting_a_cell_are_not_gated_on_the_plan(
    state: NotebookState,
) -> None:
    """The gate is on writing logic. Re-running existing code writes none, and a deletion already
    stops and asks the user, which is a stronger control than this one."""
    tools = ToolRegistry(ToolContext(driver=FakeDriver()))
    tools.refresh(state)

    ran = await tools.dispatch("run_cell", {"cell": "load_handin"})
    deleted = await tools.dispatch("delete_cell", {"cell": "load_handin", "reason": "superseded"})

    assert ran.ok, ran.text
    assert "propose_plan" not in deleted.text
    assert tools.pending_deletions[0].cell == "load_handin"


async def test_a_later_unapproved_draft_does_not_retire_the_approval_in_force(
    state: NotebookState, tmp_path: Path
) -> None:
    """The gate reads `latest_approved()`, which is what "in force" means everywhere else.

    Read as `latest()` instead, proposing an amendment would stop the model working — the newest
    plan on disk would be an unapproved one, and the approved plan it is still meant to be
    implementing would stop counting.
    """
    store = approved_plan_store(tmp_path / "plans")
    store.save_next(make_plan())
    tools = ToolRegistry(ToolContext(driver=FakeDriver(), plans=store))
    tools.refresh(state)

    result = await tools.dispatch("propose_cell", {"name": "c", "code": "c = 1\n"})

    assert store.latest() is not None and not store.latest().approval.approved
    assert result.ok, result.text


async def test_delete_cell_never_deletes(registry: ToolRegistry) -> None:
    driver = registry.context.driver
    result = await registry.dispatch(
        "delete_cell", {"cell": "load_handin", "reason": "superseded by apply_haircuts"}
    )
    assert not result.ok
    assert "NOT happened" in result.text
    assert driver.deleted == []
    assert registry.pending_deletions[0].cell == "load_handin"
    assert registry.pending_deletions[0].reason == "superseded by apply_haircuts"


async def test_delete_cell_names_the_cells_that_would_break(registry: ToolRegistry) -> None:
    result = await registry.dispatch("delete_cell", {"cell": "imports", "reason": "tidy up"})
    assert "load_handin" in result.text


async def test_list_cells_returns_the_bodies_and_the_graph(registry: ToolRegistry) -> None:
    result = await registry.dispatch("list_cells", {})
    assert result.ok
    assert "import polars as pl" in result.text
    assert "defines: pl" in result.text


async def test_probe_returns_the_kernel_value(registry: ToolRegistry) -> None:
    result = await registry.dispatch("probe", {"code": "load_handin.height"})
    assert result.ok
    assert SENTINEL in result.text
    assert registry.context.driver.probes == ["load_handin.height"]


async def test_an_over_long_probe_is_refused(registry: ToolRegistry) -> None:
    result = await registry.dispatch("probe", {"code": "x" * 5_000})
    assert not result.ok
    assert "wants to be a cell" in result.text


async def test_a_probe_with_no_value_says_how_to_ask_again(registry: ToolRegistry) -> None:
    # "nothing" is not an answer the model can act on; it has to be told what to change.
    registry.context.driver.probe_result = ProbeResult(ok=True)
    result = await registry.dispatch("probe", {"code": "import polars as pl"})
    assert not result.ok
    assert "bare expression" in result.text
    assert result.summary == "probe returned nothing"


async def test_a_printing_probe_leads_with_what_it_printed(registry: ToolRegistry) -> None:
    # print(...) binds None. The printed text is the answer, so it must not sit under a header
    # that reads as an empty result.
    registry.context.driver.probe_result = ProbeResult(
        ok=True, value_repr="None", value_type="NoneType", stdout=f"shape: (1, 1)\n{SENTINEL}\n"
    )
    result = await registry.dispatch("probe", {"code": "print(load_handin)"})
    assert result.ok
    assert result.text.startswith("stdout:")
    assert SENTINEL in result.text
    assert result.summary == "probe printed output"


# ── degradation ──────────────────────────────────────────────────────────────────────────────


async def test_notebook_tools_say_so_when_no_kernel_is_attached() -> None:
    tools = ToolRegistry(ToolContext())
    result = await tools.dispatch("list_cells", {})
    assert not result.ok
    assert "no live marimo notebook" in result.text


async def test_inspect_workbook_says_so_when_no_analysis_is_loaded() -> None:
    tools = ToolRegistry(ToolContext())
    result = await tools.dispatch("inspect_workbook", {"section": "sheets"})
    assert not result.ok
    assert "kedge inspect" in result.text


async def test_get_plan_says_so_when_no_plan_exists(tmp_path: Path) -> None:
    from kedge.plan.store import PlanStore

    tools = ToolRegistry(ToolContext(plans=PlanStore(tmp_path / "plans")))
    result = await tools.dispatch("get_plan", {})
    assert not result.ok
    # `kedge plan propose` exists and is the batch route, but the model is mid-conversation and
    # its route is the tool. Sending the user to a shell abandons the turn to do what the model
    # could have done and put in front of them for approval.
    assert "propose_plan" in result.text
    assert "kedge plan" not in result.text


async def test_list_utils_degrades_when_the_package_is_absent() -> None:
    tools = ToolRegistry(ToolContext(utils_dir=None))
    result = await tools.dispatch("list_utils", {})
    assert result.ok
    assert "no utils package" in result.text


async def test_list_utils_degrades_when_the_package_is_empty(tmp_path: Path) -> None:
    empty = tmp_path / "utils"
    empty.mkdir()
    tools = ToolRegistry(ToolContext(utils_dir=empty))
    result = await tools.dispatch("list_utils", {})
    assert result.ok
    assert "empty" in result.text


async def test_get_knowledge_degrades_when_there_is_no_pack() -> None:
    tools = ToolRegistry(ToolContext(knowledge_dir=None))
    result = await tools.dispatch("get_knowledge", {})
    assert result.ok
    assert "no knowledge pack" in result.text


async def test_reconcile_without_a_workspace_degrades_to_not_reconciled() -> None:
    tools = ToolRegistry(ToolContext(driver=FakeDriver()))
    result = await tools.dispatch(
        "reconcile", {"variable": "apply_haircuts", "reference": "Calc!H2:H50000"}
    )
    assert not result.ok
    assert "NOT RECONCILED" in result.text
    assert "has not passed" in result.text


async def test_reconcile_refuses_an_unqualified_reference() -> None:
    tools = ToolRegistry(ToolContext(driver=FakeDriver()))
    result = await tools.dispatch("reconcile", {"variable": "x", "reference": "H2:H500"})
    assert not result.ok
    assert "sheet-qualified" in result.text


async def test_amend_plan_records_rather_than_applies() -> None:
    tools = ToolRegistry(ToolContext())
    result = await tools.dispatch(
        "amend_plan",
        {
            "rationale": "Calc!AF is a second haircut the plan does not mention",
            "change": "add a stage between apply_haircuts and reconcile",
            "stage": "apply_haircuts",
        },
    )
    assert result.ok
    assert "NOT in force" in result.text
    assert tools.pending_amendments[0].stage == "apply_haircuts"


# ── propose_plan ─────────────────────────────────────────────────────────────────────────────
#
# The tool exists because the account of a workbook the model works out in conversation *is* a
# plan, and said only in the chat it is compacted away as the conversation grows. It records and
# refuses, exactly as `amend_plan` does, and its three refusals are the tests that matter. It is
# the one tool that turns a conversation into a large, structured, confident-looking artifact, and
# it is offered on every step of every turn, so what must not happen is refused rather than
# discouraged: no replacing an approved plan (that is the amendment gate walked around rather than
# through), no plan from a session that has read nothing, and no plan larger than a notebook
# anyone will review.


def _payload(**overrides: Any) -> dict[str, Any]:
    """The plan a model would send: stages and open questions, and no assessment."""
    payload: dict[str, Any] = {
        "summary": "A monthly regulatory calculation with one lookup and a manual override step.",
        "stages": [
            {
                "id": "load_handin",
                "intent": "Read counterparty exposures from the hand-in",
                "kind": "load",
                "sources": ["handin"],
                "confidence": "high",
            },
            {
                "id": "apply_haircuts",
                "intent": "Collateral haircut lookup by asset class",
                "sources": ["Calc!H2:H500"],
                "depends_on": ["load_handin"],
                "confidence": "high",
            },
        ],
        "open_questions": ["Column AF is computed but never referenced. Dead, or read manually?"],
    }
    payload.update(overrides)
    return payload


def _proposed(**overrides: Any) -> str:
    """That plan as the JSON string the schema asks for."""
    return json.dumps(_payload(**overrides))


async def _plan_tools(
    directory: Path,
    *,
    read: bool = True,
    analysis: WorkbookAnalysis | None = None,
    **context: Any,
) -> tuple[ToolRegistry, PlanStore]:
    """A registry with a plan store and, unless told otherwise, a workbook it has read.

    `propose_plan` refuses a session that has read nothing, so every test about what it does with
    a *draft* has to get past that gate first. It is got past with a real `inspect_workbook` call
    rather than by reaching into the registry, so that nothing here depends on how the reading is
    recorded — which is the half of the gate most likely to be quietly rewired.
    """
    store = PlanStore(directory)
    tools = ToolRegistry(ToolContext(analysis=analysis or make_analysis(), plans=store, **context))
    if read:
        assert (await tools.dispatch("inspect_workbook", {"section": "operations"})).ok
    return tools, store


async def test_propose_plan_records_the_proposal_and_writes_nothing(tmp_path: Path) -> None:
    tools, store = await _plan_tools(tmp_path / "plans")

    result = await tools.dispatch("propose_plan", {"plan": _proposed()})

    assert result.ok
    assert "NOT in force" in result.text
    assert len(tools.pending_proposals) == 1
    assert tools.pending_proposals[0].plan.stage_ids == ["load_handin", "apply_haircuts"]
    assert store.versions() == [], "a proposal is not a plan until the user says it is"


async def test_propose_plan_refuses_once_a_plan_has_been_approved(tmp_path: Path) -> None:
    """The one rule that matters: this must not become a way round amendment review."""
    tools, store = await _plan_tools(tmp_path / "plans")
    store.save(approve(acknowledge_all_drops(make_plan(), note="reviewed"), by="tests"))
    before = store.versions()

    result = await tools.dispatch("propose_plan", {"plan": _proposed()})

    assert not result.ok
    assert "amend_plan" in result.text
    assert tools.pending_proposals == []
    assert store.versions() == before, "the approved plan is untouched"


async def test_propose_plan_supersedes_a_draft_nobody_has_approved(tmp_path: Path) -> None:
    """A draft is not in force, so there is no gate for a replacement to get past."""
    tools, store = await _plan_tools(tmp_path / "plans")
    store.save(make_plan())

    result = await tools.dispatch("propose_plan", {"plan": _proposed()})

    assert result.ok
    proposed = tools.pending_proposals[0].plan
    assert proposed.version == 2
    assert proposed.based_on_version == 1, "the history chain names the draft it supersedes"


async def test_malformed_json_comes_back_as_a_tool_result_not_an_exception(tmp_path: Path) -> None:
    tools, store = await _plan_tools(tmp_path / "plans")

    result = await tools.dispatch("propose_plan", {"plan": "here you go: {stages: ["})

    assert not result.ok
    assert "nothing was recorded" in result.text
    assert tools.pending_proposals == []
    assert store.versions() == []


async def test_a_plan_that_does_not_validate_comes_back_as_the_fields_to_fix(
    tmp_path: Path,
) -> None:
    tools, _store = await _plan_tools(tmp_path / "plans")

    result = await tools.dispatch(
        "propose_plan", {"plan": _proposed(stages=[{"id": "load_handin"}])}
    )

    assert not result.ok
    assert "stages.0.intent" in result.text, "the model is told which field, not that it failed"
    assert tools.pending_proposals == []


async def test_convertibility_is_kedges_triage_and_not_the_models_own_score(
    tmp_path: Path,
) -> None:
    """A model scoring its own decomposition has nothing to score it against (PLAN 2.2)."""
    tools, _store = await _plan_tools(tmp_path / "plans")
    invented = {"convertible": 0.05, "blockers": ["I am not sure about any of this"]}

    result = await tools.dispatch("propose_plan", {"plan": _proposed(assessment=invented)})

    assert result.ok
    assessment = tools.pending_proposals[0].plan.assessment
    assert assessment.convertible == triage(make_analysis()).convertible
    assert assessment.blockers == []
    assert "triage" in (assessment.rationale or "")


async def test_a_proposed_plan_records_the_model_that_actually_wrote_it(
    workspace: Workspace, state: NotebookState
) -> None:
    """The model driving the turn, which is not necessarily the one in config.

    `[model] model` is a default. The turn runs `request.model or self._model`, and
    `request.model` comes from the session, which the user can change while the chat is open. The
    id is stamped by `refresh` from that same expression, so an override is recorded as itself —
    a plan is read months later by someone asking which model wrote it and whether that model is
    still trusted, and a confidently wrong id ends that question rather than answering it.
    """
    tools, _store = await _plan_tools(workspace.plans_dir, workspace=workspace)
    tools.refresh(state, model="an-experimental-model-this-session")

    await tools.dispatch("propose_plan", {"plan": _proposed()})

    plan = tools.pending_proposals[0].plan
    assert plan.generated_by == "llm"
    assert plan.llm_model == "an-experimental-model-this-session"
    assert plan.llm_model != workspace.config.model.model, "config is the fallback, not the answer"
    assert plan.approval.state is ApprovalState.DRAFT, "authoring is not approving"


async def test_a_plan_proposed_by_an_unidentified_model_says_so_rather_than_guessing(
    workspace: Workspace,
) -> None:
    # Nothing stamped the model, so nothing is claimed. An absent id sends a reviewer looking; a
    # wrong one stops them looking, which is the more expensive failure of the two.
    tools, _store = await _plan_tools(workspace.plans_dir, workspace=workspace)

    await tools.dispatch("propose_plan", {"plan": _proposed()})

    assert tools.pending_proposals[0].plan.llm_model is None


# ── propose_plan: the triage verdict ─────────────────────────────────────────────────────────


def _declined() -> WorkbookAnalysis:
    """An analysis triage refuses outright: kedge cannot read a `.xlsb` at all."""
    return make_analysis(workbook_fields={"file_format": "xlsb"})


async def test_a_workbook_triage_declined_is_not_planned_through_the_chat(tmp_path: Path) -> None:
    """`propose.propose_plan` refuses a STOP rather than spend a model call on it. So does this.

    Without it, a workbook kedge has decided against can be planned in the chat, approved through
    the web UI, and the word "stop" never reaches the user.
    """
    tools, store = await _plan_tools(tmp_path / "plans", analysis=_declined())

    result = await tools.dispatch("propose_plan", {"plan": _proposed()})

    assert not result.ok
    assert "recommends against converting" in result.text
    assert "re-save as .xlsx" in result.text, "the refusal carries triage's own reasons"
    assert "not a candidate" in result.text, "it says what to do, not only what was wrong"
    assert tools.pending_proposals == []
    assert store.versions() == []


async def test_the_stop_refusal_is_not_a_draft_worth_resending(tmp_path: Path) -> None:
    # No corrected draft can clear this, so the loop must not count it against its draft cap and
    # then advise about the wrong problem when it reaches one.
    tools, _store = await _plan_tools(tmp_path / "plans", analysis=_declined())

    result = await tools.dispatch("propose_plan", {"plan": _proposed()})

    assert not result.draft_rejected


async def test_the_stop_refusal_comes_before_the_read_gate(tmp_path: Path) -> None:
    """The most fundamental reason first: this one is about the workbook, the other two are about
    the session. Reading cannot change the verdict, so sending the model off through four tools
    and refusing it when it came back would spend the turn and imply the answer might differ."""
    tools, _store = await _plan_tools(tmp_path / "plans", analysis=_declined(), read=False)

    result = await tools.dispatch("propose_plan", {"plan": _proposed()})

    assert not result.ok
    assert "recommends against converting" in result.text
    assert "read nothing about this workbook" not in result.text


async def test_a_blocker_short_of_a_stop_is_still_planned(tmp_path: Path) -> None:
    """The gate is the verdict, not the presence of blockers: `proceed_with_care` proceeds."""
    with_vba = make_analysis(workbook_fields={"has_vba": True})
    assert triage(with_vba).verdict is TriageVerdict.PROCEED_WITH_CARE
    tools, _store = await _plan_tools(tmp_path / "plans", analysis=with_vba)

    result = await tools.dispatch("propose_plan", {"plan": _proposed()})

    assert result.ok, result.text
    proposed = tools.pending_proposals[0].plan
    assert proposed.assessment.convertible == triage(with_vba).convertible


# ── propose_plan: the read gate ──────────────────────────────────────────────────────────────


async def test_a_plan_cannot_be_written_by_a_session_that_has_read_nothing(
    tmp_path: Path,
) -> None:
    """The first tool call of a session must not be able to be a whole process plan."""
    tools, store = await _plan_tools(tmp_path / "plans", read=False)

    result = await tools.dispatch("propose_plan", {"plan": _proposed()})

    assert not result.ok
    assert "read nothing about this workbook" in result.text
    assert "guess with a confident tone" in result.text
    assert tools.pending_proposals == []
    assert store.versions() == []


@pytest.mark.parametrize(
    "tool", ["inspect_workbook", "sample_data", "profile_column", "read_range"]
)
async def test_any_workbook_read_opens_the_gate(
    tool: str, workspace: Workspace, tmp_path: Path
) -> None:
    """Four tools count, and each of them alone is enough. The floor is "you looked"."""
    store = PlanStore(tmp_path / "plans")
    tools = ToolRegistry(ToolContext(workspace=workspace, analysis=make_analysis(), plans=store))
    arguments = {
        "inspect_workbook": {"section": "operations"},
        "sample_data": {"sheet": "Ref", "rows": 3},
        # The one column the fixture analysis actually profiled.
        "profile_column": {"sheet": "Calc", "column": "H"},
        "read_range": {"sheet": "Ref", "range": "A1:C3"},
    }[tool]
    try:
        assert (await tools.dispatch(tool, arguments)).ok
        assert (await tools.dispatch("propose_plan", {"plan": _proposed()})).ok
    finally:
        await tools.aclose()


async def test_a_probe_is_not_a_reading_of_the_workbook(
    tmp_path: Path, state: NotebookState
) -> None:
    """`probe` describes the kernel and the hand-in loaded in it, not the process the workbook runs.

    role.md says exactly that, and a plan is a decomposition of the process. It could not be a
    route round the gate in any case: probing needs cells, and cells need an approved plan.
    """
    store = PlanStore(tmp_path / "plans")
    tools = ToolRegistry(
        ToolContext(analysis=make_analysis(), plans=store, driver=FakeDriver()),
    )
    tools.refresh(state)

    assert (await tools.dispatch("probe", {"code": "load_handin.height"})).ok
    result = await tools.dispatch("propose_plan", {"plan": _proposed()})

    assert not result.ok
    assert "read nothing about this workbook" in result.text


async def test_a_read_that_failed_has_read_nothing(tmp_path: Path) -> None:
    # A call that came back with no payload is not looking at the workbook, however it is logged.
    tools, _store = await _plan_tools(tmp_path / "plans", read=False)

    assert not (await tools.dispatch("inspect_workbook", {"section": "invented"})).ok
    result = await tools.dispatch("propose_plan", {"plan": _proposed()})

    assert not result.ok
    assert "read nothing about this workbook" in result.text


async def test_the_read_gate_is_not_reached_before_there_is_anything_to_read(
    tmp_path: Path,
) -> None:
    # With no analysis, "read the workbook first" is a dead end: kedge has not read it either.
    tools = ToolRegistry(ToolContext(plans=PlanStore(tmp_path / "plans")))

    result = await tools.dispatch("propose_plan", {"plan": _proposed()})

    assert not result.ok
    assert "kedge inspect" in result.text


# ── propose_plan: the shape of what arrives ──────────────────────────────────────────────────


async def test_a_plan_sent_as_an_object_is_read_as_one(tmp_path: Path) -> None:
    """Models emit an object where a description says "object", whatever the schema declares.

    Both readings of that instruction are defensible, and `str()` on the decoded object would
    produce a Python repr — so the model would be told its JSON was invalid at the first property
    name, which is false and leaves it nothing to fix.
    """
    tools, _store = await _plan_tools(tmp_path / "plans")

    result = await tools.dispatch("propose_plan", {"plan": _payload()})

    assert result.ok, result.text
    assert tools.pending_proposals[0].plan.stage_ids == ["load_handin", "apply_haircuts"]


async def test_a_plan_with_more_stages_than_a_notebook_can_carry_is_refused(
    tmp_path: Path,
) -> None:
    """One stage becomes one cell, so this is a ceiling on the notebook the plan describes."""
    tools, store = await _plan_tools(tmp_path / "plans")
    stages = [
        {"id": f"stage_{index}", "intent": f"step {index}"}
        for index in range(tools_module.MAX_PROPOSED_STAGES + 1)
    ]

    result = await tools.dispatch("propose_plan", {"plan": _proposed(stages=stages)})

    assert not result.ok
    assert str(tools_module.MAX_PROPOSED_STAGES) in result.text
    assert "two notebooks" in result.text, "the refusal says what to do, not only what was wrong"
    assert tools.pending_proposals == []
    assert store.versions() == []


async def test_a_plan_at_the_ceiling_is_still_a_plan(tmp_path: Path) -> None:
    tools, _store = await _plan_tools(tmp_path / "plans")
    stages = [
        {"id": f"stage_{index}", "intent": f"step {index}"}
        for index in range(tools_module.MAX_PROPOSED_STAGES)
    ]

    result = await tools.dispatch("propose_plan", {"plan": _proposed(stages=stages)})

    assert result.ok, result.text


@pytest.mark.parametrize(
    ("plan", "expected"),
    [
        ("here you go: {stages: [", "could not be read"),
        (json.dumps({"stages": [{"id": "load_handin"}], "open_questions": []}), "did not validate"),
    ],
)
async def test_a_rejected_draft_is_marked_for_the_loop_to_count(
    plan: str, expected: str, tmp_path: Path
) -> None:
    """The loop caps these at three. Without the flag it cannot tell them from a refusal.

    A model that cannot produce a valid plan otherwise loops to `max_steps` — up to fifty
    completions, each re-sending the whole tool surface — and ends in a pause that reads as
    though progress were being made.
    """
    tools, _store = await _plan_tools(tmp_path / "plans")

    result = await tools.dispatch("propose_plan", {"plan": plan})

    assert not result.ok
    assert expected in result.text
    assert result.draft_rejected


@pytest.mark.parametrize("read", [True, False])
async def test_a_refusal_no_retry_can_change_is_not_counted_against_the_model(
    read: bool, tmp_path: Path
) -> None:
    # Capping these would count the wrong thing, and the message at the cap would be advice about
    # a problem the model does not have.
    tools, store = await _plan_tools(tmp_path / "plans", read=read)
    if read:
        store.save(approve(acknowledge_all_drops(make_plan(), note="reviewed"), by="tests"))

    result = await tools.dispatch("propose_plan", {"plan": _proposed()})

    assert not result.ok
    assert not result.draft_rejected


async def test_propose_plan_without_an_analysis_says_what_is_missing(tmp_path: Path) -> None:
    tools = ToolRegistry(ToolContext(plans=PlanStore(tmp_path / "plans")))

    result = await tools.dispatch("propose_plan", {"plan": _proposed()})

    assert not result.ok
    assert "kedge inspect" in result.text


async def test_propose_plan_without_a_plan_store_degrades_to_a_sentence() -> None:
    tools = ToolRegistry(ToolContext(analysis=make_analysis()))

    result = await tools.dispatch("propose_plan", {"plan": _proposed()})

    assert not result.ok
    assert "no plan store" in result.text


def test_the_propose_plan_schema_costs_far_less_than_the_json_schema_would() -> None:
    """PLAN M4: the tool schemas are the prompt cache's prefix, re-sent on every completion.

    A turn is up to `max_steps` of them, so inlining `plan_json_schema()` into this tool's
    parameters would be paid for once per step for the whole session. One JSON string argument
    and a prose description of the shape costs a fraction of it, and a response that does not fit
    comes back as a tool result the model can read and correct — which is how every other refusal
    on this surface already works.
    """
    spec = next(item for item in tools_module.TOOL_SPECS if item.name == "propose_plan")
    inline = len(json.dumps(plan_json_schema(), separators=(",", ":")))
    assert len(json.dumps(spec.schema(), separators=(",", ":"))) < inline // 4


# The budget, in characters of serialised tool schema. Roughly a token per four characters, so
# TOOL_BLOCK_BUDGET is about 3,000 tokens — sent ahead of the conversation on every one of up to
# `max_steps` completions in every turn of every session, whether or not the turn uses a tool.
#
# The numbers are deliberately fixed rather than derived. Nothing computed from the specs can
# fail, and the failure is the point: this is the only place growth in the surface is visible at
# all, since it is paid for in the endpoint's bill rather than in anything a test would notice.
# Both sit a little under a tenth above what the surface costs today, which is enough headroom to
# reword a description and not enough to add a tool or double one without saying so out loud.
#
# That claim is only true while it is checked. `propose_plan` grew to 1,975 characters against
# LARGEST_SCHEMA_BUDGET -- 1.3% of headroom, on a budget whose whole point is that there is enough
# room to reword something. Raising the ceiling would have been the wrong half of the trade: the
# number is the only place growth in the surface is visible, so moving it to fit the schema makes
# it decorative. The schema was trimmed back to ~1,820 instead, which is the tenth the comment
# above claims. If it creeps up again, cut words before moving the number.
TOOL_BLOCK_BUDGET = 12_000
LARGEST_SCHEMA_BUDGET = 2_000


def test_the_whole_tool_surface_stays_inside_its_budget() -> None:
    """Adding to this surface should cost a decision, not happen quietly.

    `propose_plan` took the block from 9,287 characters to 11,028, +19%, and nothing failed. The
    cost is real and it is per completion: this is the prompt cache's prefix, so it is also what
    every later block is cached behind. If this fails, the question is not "raise it?" — it is
    which description is carrying words the model does not act on.
    """
    block = json.dumps(tool_schemas(), separators=(",", ":"))
    assert len(block) <= TOOL_BLOCK_BUDGET, (
        f"the tool block is {len(block)} characters against a budget of {TOOL_BLOCK_BUDGET}"
    )


def test_no_single_tool_dominates_the_surface() -> None:
    """One tool at a third of the block is a tool whose description has become documentation.

    `propose_plan` is already the largest by 88%, which is defensible — it is describing a nested
    structure to a model that cannot see the schema for it — and is exactly why it wants a ceiling
    rather than a comparison against its neighbours, which would ratchet.
    """
    sizes = {
        spec.name: len(json.dumps(spec.schema(), separators=(",", ":")))
        for spec in tools_module.TOOL_SPECS
    }
    largest = max(sizes, key=lambda name: sizes[name])
    assert sizes[largest] <= LARGEST_SCHEMA_BUDGET, (
        f"{largest} is {sizes[largest]} characters against a budget of {LARGEST_SCHEMA_BUDGET}"
    )


# ── reading a real workbook ──────────────────────────────────────────────────────────────────


def _write_workbook(path: Path) -> None:
    book = OpenpyxlWorkbook()
    sheet = book.active
    sheet.title = "Ref"
    sheet.append(["asset_class", "counterparty", "haircut"])
    sheet.append(["gilts", SENTINEL, 0.02])
    for index in range(2, 60):
        sheet.append([f"class-{index}", f"cp-{index}", index / 100])
    book.save(path)


@pytest.fixture
def workspace(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Workspace:
    monkeypatch.setenv("KEDGE_HOME", str(tmp_path / "home"))
    workbook = tmp_path / "rwa_monthly.xlsx"
    _write_workbook(workbook)
    built = Workspace.for_workbook(workbook)
    built.ensure_dirs()
    return built


async def test_read_range_returns_verbatim_values(workspace: Workspace) -> None:
    tools = ToolRegistry(ToolContext(workspace=workspace))
    try:
        result = await tools.dispatch("read_range", {"sheet": "Ref", "range": "A1:C3"})
        assert result.ok
        assert SENTINEL in result.text
        assert result.sheet == "Ref"
        assert result.columns == ("A", "B", "C")
    finally:
        await tools.aclose()


async def test_read_range_reports_what_it_could_not_return(workspace: Workspace) -> None:
    tools = ToolRegistry(ToolContext(workspace=workspace, caps=Caps(max_rows=5)))
    try:
        result = await tools.dispatch("read_range", {"sheet": "Ref", "range": "A1:C60"})
        assert result.row_count == 5
        assert "[… 55 more rows omitted]" in result.text
    finally:
        await tools.aclose()


async def test_sample_data_resolves_the_header_row(workspace: Workspace) -> None:
    tools = ToolRegistry(ToolContext(workspace=workspace))
    try:
        result = await tools.dispatch("sample_data", {"sheet": "Ref", "rows": 3})
        assert result.ok
        assert "asset_class" in result.text
        assert "haircut" in result.text
        assert SENTINEL in result.text
    finally:
        await tools.aclose()


async def test_sample_data_hashes_a_redacted_column(workspace: Workspace) -> None:
    tools = ToolRegistry(ToolContext(workspace=workspace, redaction_patterns=("counterparty",)))
    try:
        result = await tools.dispatch("sample_data", {"sheet": "Ref", "rows": 3})
        assert SENTINEL not in result.text
        assert "sha256:" in result.text
        assert result.redacted_columns == 1
    finally:
        await tools.aclose()


async def test_an_unknown_sheet_is_reported_rather_than_raised(workspace: Workspace) -> None:
    tools = ToolRegistry(ToolContext(workspace=workspace))
    try:
        result = await tools.dispatch("read_range", {"sheet": "Nope", "range": "A1:B2"})
        assert "does not exist" in result.text
    finally:
        await tools.aclose()


# ── the session cache ────────────────────────────────────────────────────────────────────────


class _CountingRegistry(ToolRegistry):
    """Counts handler runs, so a cache hit shows up as a handler that did not run.

    The run number is in the payload as well, which is how the byte-identical property is
    asserted: a second call that returns "run 1" cannot have re-run anything.
    """

    def __init__(self, *args: Any, **kwargs: Any) -> None:
        super().__init__(*args, **kwargs)
        self.runs = 0
        self.succeed = True
        self.filler = ""

    async def _tool_inspect_workbook(self, args: Any) -> ToolResult:
        self.runs += 1
        section = args.get("section")
        sheet = args.get("sheet") or "-"
        return ToolResult.note(f"run {self.runs}: {section}/{sheet} {self.filler}", ok=self.succeed)


def _touch_from_excel(path: Path) -> None:
    """A save that changed nothing kedge can see in the bytes: only the mtime moves.

    Excel rewrites the whole file on save, and the result can come out the same length. The
    fingerprint has to move on the timestamp alone, which is what this isolates.
    """
    stamp = path.stat().st_mtime_ns + 60_000_000_000
    os.utime(path, ns=(stamp, stamp))


def _save_from_excel(path: Path, *, counterparty: str) -> None:
    """The user opening the workbook in Excel, changing a value, and saving it.

    Written in place rather than through a temp file and :func:`os.replace`, because on Windows
    the rename is refused outright while kedge holds the file open — which is itself half the
    reason a handle must not be held across a save.

    ``os.utime`` afterwards rather than trusting the write to move the clock: the fingerprint is
    ``st_mtime_ns`` and ``st_size``, and a rewrite that lands inside the filesystem's timestamp
    resolution and happens to be the same length would move neither.
    """
    book = load_workbook(path)
    book["Ref"]["B2"] = counterparty
    book.save(path)
    _touch_from_excel(path)


async def test_a_repeated_session_stable_call_is_answered_from_the_cache(
    workspace: Workspace,
) -> None:
    tools = _CountingRegistry(ToolContext(workspace=workspace))
    first = await tools.dispatch("inspect_workbook", {"section": "sheets", "sheet": "Ref"})
    # The same question, with the keys in the order some other completion happened to emit them.
    second = await tools.dispatch("inspect_workbook", {"sheet": "Ref", "section": "sheets"})
    assert tools.runs == 1
    assert second.text == first.text
    assert second.ok


async def test_a_different_question_is_a_different_cache_entry(workspace: Workspace) -> None:
    tools = _CountingRegistry(ToolContext(workspace=workspace))
    await tools.dispatch("inspect_workbook", {"section": "sheets"})
    await tools.dispatch("inspect_workbook", {"section": "findings"})
    assert tools.runs == 2


async def test_a_workbook_saved_mid_session_invalidates_what_was_cached(
    workspace: Workspace,
) -> None:
    """The user is free to edit the workbook mid-chat, and must not be told what it used to say.

    This asserted only that the handler ran again, and it passed while the model was still being
    handed the old rows: the cache missed on the new fingerprint, the handler re-read the
    workbook handle opened before the save, and the identical bytes came back at the cost of the
    recompute. So it asserts the content now, which is the assertion the bug could not satisfy.
    """
    tools = ToolRegistry(ToolContext(workspace=workspace))
    try:
        first = await tools.dispatch("read_range", {"sheet": "Ref", "range": "A1:C2"})
        assert SENTINEL in first.text

        _save_from_excel(workspace.workbook_path, counterparty=RESAVED)

        second = await tools.dispatch("read_range", {"sheet": "Ref", "range": "A1:C2"})
        assert RESAVED in second.text
        assert SENTINEL not in second.text
    finally:
        await tools.aclose()


async def test_a_re_save_that_changed_no_bytes_still_invalidates_what_was_cached(
    workspace: Workspace,
) -> None:
    # The other half: a save whose output happens to be the same length must still invalidate,
    # so the fingerprint cannot be allowed to rest on the size alone.
    tools = _CountingRegistry(ToolContext(workspace=workspace))
    await tools.dispatch("inspect_workbook", {"section": "sheets"})
    _touch_from_excel(workspace.workbook_path)
    result = await tools.dispatch("inspect_workbook", {"section": "sheets"})
    assert tools.runs == 2
    assert "run 2" in result.text


async def test_nothing_is_cached_when_no_fingerprint_can_be_taken() -> None:
    # No workspace and no analysis: nothing can be stat'd, so nothing can be shown to be current.
    # Degrading to a second read is the safe half of that trade.
    tools = _CountingRegistry(ToolContext())
    await tools.dispatch("inspect_workbook", {"section": "sheets"})
    await tools.dispatch("inspect_workbook", {"section": "sheets"})
    assert tools.runs == 2


async def test_a_failed_result_is_never_cached(workspace: Workspace) -> None:
    # A transient failure that pinned itself for the session would be a tool that stays broken
    # until the user restarts kedge.
    tools = _CountingRegistry(ToolContext(workspace=workspace))
    tools.succeed = False
    await tools.dispatch("inspect_workbook", {"section": "sheets"})
    await tools.dispatch("inspect_workbook", {"section": "sheets"})
    assert tools.runs == 2


async def test_a_volatile_tool_is_asked_again_every_time(
    workspace: Workspace, state: NotebookState
) -> None:
    tools = ToolRegistry(ToolContext(workspace=workspace, driver=FakeDriver()))
    tools.refresh(state)
    await tools.dispatch("probe", {"code": "load_handin.height"})
    await tools.dispatch("probe", {"code": "load_handin.height"})
    assert tools.context.driver.probes == ["load_handin.height"] * 2


async def test_the_cache_drops_the_oldest_entry_once_it_is_full(workspace: Workspace) -> None:
    tools = _CountingRegistry(ToolContext(workspace=workspace))
    for index in range(MAX_CACHED_RESULTS + 1):
        await tools.dispatch("inspect_workbook", {"section": "sheets", "sheet": f"s{index}"})
    assert tools.runs == MAX_CACHED_RESULTS + 1

    newest = f"s{MAX_CACHED_RESULTS}"
    await tools.dispatch("inspect_workbook", {"section": "sheets", "sheet": newest})
    assert tools.runs == MAX_CACHED_RESULTS + 1  # still held
    await tools.dispatch("inspect_workbook", {"section": "sheets", "sheet": "s0"})
    assert tools.runs == MAX_CACHED_RESULTS + 2  # the oldest made room for it


async def test_a_redaction_pattern_added_mid_session_invalidates_what_was_cached(
    workspace: Workspace,
) -> None:
    # The cached payload was rendered before the pattern existed. Reissuing it would be kedge
    # quietly unredacting a column the user has since asked it to hash.
    tools = _CountingRegistry(ToolContext(workspace=workspace))
    await tools.dispatch("inspect_workbook", {"section": "sheets"})
    tools.set_context(ToolContext(workspace=workspace, redaction_patterns=("counterparty",)))
    await tools.dispatch("inspect_workbook", {"section": "sheets"})
    assert tools.runs == 2


async def test_the_cache_holds_a_bounded_number_of_bytes(
    workspace: Workspace, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr(tools_module, "MAX_CACHED_BYTES", 64)
    tools = _CountingRegistry(ToolContext(workspace=workspace))
    tools.filler = "x" * 500
    await tools.dispatch("inspect_workbook", {"section": "sheets"})
    await tools.dispatch("inspect_workbook", {"section": "sheets"})
    assert tools.runs == 2  # a payload larger than the whole budget is read again, not held


async def test_a_cache_hit_still_writes_its_audit_line(
    workspace: Workspace, tmp_path: Path
) -> None:
    # Every payload handed to the model is one line in the outbound log. A log with holes in it
    # wherever a result came from memory would be a defect in the audit trail (SECURITY.md).
    log = OutboundLog(tmp_path / "outbound-cache.jsonl", session="cache")
    tools = _CountingRegistry(ToolContext(workspace=workspace), log=log)
    first = await tools.dispatch("inspect_workbook", {"section": "sheets"}, turn_id="t1")
    second = await tools.dispatch("inspect_workbook", {"section": "sheets"}, turn_id="t2")
    entries = list(log.entries())
    assert tools.runs == 1
    assert [entry["turn_id"] for entry in entries] == ["t1", "t2"]
    assert entries[0]["byte_count"] == entries[1]["byte_count"] == first.byte_count
    assert second.byte_count == first.byte_count


# ── the open workbook handle ─────────────────────────────────────────────────────────────────
#
# The cache invalidating over a handle that did not was the whole defect: the fingerprint moved,
# the cache missed, the handler re-ran, and it re-read a zip archive and two openpyxl views
# loaded before the save. The model paid for a recompute and got the same answer. So these tests
# assert on the *value that came back*, never on the handler having run — that was true
# throughout, and it is precisely what made the bug invisible.


async def test_sample_data_re_reads_a_workbook_saved_mid_session(workspace: Workspace) -> None:
    # The scenario kedge itself creates: "open the workbook in Excel, allow it to calculate, and
    # save it" is the documented remedy, so a user who does it and says "try again" must not be
    # told the same thing about the file they have just fixed.
    tools = ToolRegistry(ToolContext(workspace=workspace))
    try:
        first = await tools.dispatch("sample_data", {"sheet": "Ref", "rows": 3})
        assert SENTINEL in first.text

        _save_from_excel(workspace.workbook_path, counterparty=RESAVED)

        second = await tools.dispatch("sample_data", {"sheet": "Ref", "rows": 3})
        assert RESAVED in second.text
        assert SENTINEL not in second.text
    finally:
        await tools.aclose()


async def test_an_untouched_workbook_is_not_reopened(workspace: Workspace) -> None:
    # Reopening per call would be a fresh zip and two openpyxl loads on every read, which is the
    # cost the session-long handle exists to avoid. Asserted on the object rather than on timing.
    tools = ToolRegistry(ToolContext(workspace=workspace))
    try:
        await tools.dispatch("read_range", {"sheet": "Ref", "range": "A1:C2"})
        opened = tools._handle
        assert opened is not None
        await tools.dispatch("read_range", {"sheet": "Ref", "range": "A1:C3"})
        assert tools._handle is opened
    finally:
        await tools.aclose()


async def test_the_handle_a_reopen_replaces_is_closed_rather_than_dropped(
    workspace: Workspace,
) -> None:
    # `WorkbookHandle.close` releases three OS handles. Left to the collector they sit on a file
    # the user is editing in Excel, which on Windows is worse than a leak.
    tools = ToolRegistry(ToolContext(workspace=workspace))
    try:
        await tools.dispatch("read_range", {"sheet": "Ref", "range": "A1:C2"})
        stale = tools._handle

        _save_from_excel(workspace.workbook_path, counterparty=RESAVED)
        await tools.dispatch("read_range", {"sheet": "Ref", "range": "A1:C2"})

        assert tools._handle is not stale
        assert stale._closed
    finally:
        await tools.aclose()


async def test_a_workbook_that_will_not_reopen_is_reported_rather_than_answered_stale(
    workspace: Workspace,
) -> None:
    """The failure that matters: the file has moved and the new one cannot be read.

    Serving the old copy quietly would be the worst of the options — the model would answer
    confidently about a version of the workbook that no longer exists — so the handle is
    released either way and the model is told, in a result it can read, that its earlier reads
    describe the previous version.
    """
    tools = ToolRegistry(ToolContext(workspace=workspace))
    try:
        first = await tools.dispatch("read_range", {"sheet": "Ref", "range": "A1:C2"})
        assert SENTINEL in first.text

        # A save caught mid-write, or one that landed as something kedge cannot read.
        workspace.workbook_path.write_bytes(b"PK\x03\x04 half of a save")
        _touch_from_excel(workspace.workbook_path)

        result = await tools.dispatch("read_range", {"sheet": "Ref", "range": "A1:C2"})
        assert not result.ok
        assert SENTINEL not in result.text
        assert "changed on disk" in result.text
        assert "must not be quoted as current" in result.text
        assert tools._handle is None
    finally:
        await tools.aclose()


async def test_a_reopen_failure_is_not_pinned_for_the_session(workspace: Workspace) -> None:
    # A workbook caught mid-save is not a workbook that cannot be opened, and pinning the
    # session on it would strand the user at the moment they had done what kedge asked.
    tools = ToolRegistry(ToolContext(workspace=workspace))
    try:
        await tools.dispatch("read_range", {"sheet": "Ref", "range": "A1:C2"})

        workspace.workbook_path.write_bytes(b"PK\x03\x04 half of a save")
        _touch_from_excel(workspace.workbook_path)
        failed = await tools.dispatch("read_range", {"sheet": "Ref", "range": "A1:C2"})
        assert not failed.ok

        # The save finishes, and the next call goes through the ordinary open path.
        _write_workbook(workspace.workbook_path)
        _touch_from_excel(workspace.workbook_path)
        recovered = await tools.dispatch("read_range", {"sheet": "Ref", "range": "A1:C2"})
        assert recovered.ok, recovered.text
        assert SENTINEL in recovered.text
    finally:
        await tools.aclose()


async def test_a_handle_is_kept_when_no_fingerprint_can_be_taken(workspace: Workspace) -> None:
    """No fingerprint is not evidence of a change, so it is not a reason to reopen.

    With nothing left to stat the read still answers from the handle already in hand, which is
    the behaviour that was there before any of this. Treating silence as "changed" would reopen
    on every call — and here there would be nothing to reopen from, so the read would fail
    outright rather than degrade.
    """
    tools = ToolRegistry(ToolContext(workspace=workspace))
    try:
        await tools.dispatch("read_range", {"sheet": "Ref", "range": "A1:C2"})
        opened = tools._handle

        tools.set_context(ToolContext())  # no workspace and no analysis: nothing to stat

        result = await tools.dispatch("read_range", {"sheet": "Ref", "range": "A1:C3"})
        assert result.ok
        assert SENTINEL in result.text
        assert tools._handle is opened
    finally:
        await tools.aclose()


# ── reconciliation against a real workbook ───────────────────────────────────────────────────


def _write_calc_workbook(path: Path, values: list[float]) -> None:
    """Column H as Excel left it, column K as a tool left it.

    K holds formulas nothing has ever calculated, so it reads back with no cached value at all —
    the state the whole file is in when it was written by a tool rather than saved by Excel, one
    column wide. Having both in one workbook is what makes "no baseline" a question about a range
    rather than about the file.
    """
    book = OpenpyxlWorkbook()
    sheet = book.active
    sheet.title = "Calc"
    sheet["H1"] = "haircut_exposure"
    sheet["K1"] = "doubled_exposure"
    for offset, value in enumerate(values, start=2):
        sheet[f"H{offset}"] = value
        sheet[f"K{offset}"] = f"=H{offset}*2"
    book.save(path)


def _write_uncalculated_workbook(path: Path) -> None:
    """A workbook no Excel has ever opened: formulas, and nothing cached behind any of them."""
    book = OpenpyxlWorkbook()
    sheet = book.active
    sheet.title = "Calc"
    sheet["H1"] = "haircut_exposure"
    for row in range(2, 5):
        sheet[f"H{row}"] = f"=ROW()*{row}"
    book.save(path)


def _write_notebook(workspace: Workspace, produced: list[float]) -> None:
    workspace.notebook_path.parent.mkdir(parents=True, exist_ok=True)
    workspace.notebook_path.write_text(f"apply_haircuts = {produced!r}\n", encoding="utf-8")


@pytest.fixture
def reconcilable(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Workspace:
    monkeypatch.setenv("KEDGE_HOME", str(tmp_path / "home"))
    workbook = tmp_path / "rwa_monthly.xlsx"
    _write_calc_workbook(workbook, [1.0, 2.0, 3.0])
    built = Workspace.for_workbook(workbook)
    built.ensure_dirs()
    return built


@pytest.fixture
def uncalculated(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Workspace:
    """A workspace whose workbook carries no cached values anywhere."""
    monkeypatch.setenv("KEDGE_HOME", str(tmp_path / "home"))
    workbook = tmp_path / "rwa_monthly.xlsx"
    _write_uncalculated_workbook(workbook)
    built = Workspace.for_workbook(workbook)
    built.ensure_dirs()
    return built


def _no_coverage() -> WorkbookAnalysis:
    """An analysis that says the workbook had nothing cached when it was last analysed."""
    return make_analysis(
        cached_values=CachedValueCoverage(
            formula_cell_count=3, cached_present_count=0, coverage=0.0, status="absent"
        )
    )


async def test_reconcile_passes_when_the_notebook_reproduces_the_workbook(
    reconcilable: Workspace,
) -> None:
    _write_notebook(reconcilable, [1.0, 2.0, 3.0])
    tools = ToolRegistry(ToolContext(workspace=reconcilable))
    result = await tools.dispatch(
        "reconcile",
        {"variable": "apply_haircuts", "reference": "Calc!H2:H4", "region_id": "haircuts"},
    )
    assert result.ok, result.text
    assert "PASSED" in result.text
    assert "3" in result.text


async def test_reconcile_reports_the_mismatching_rows_side_by_side(
    reconcilable: Workspace,
) -> None:
    _write_notebook(reconcilable, [1.0, 2.0, 3.5])
    tools = ToolRegistry(ToolContext(workspace=reconcilable))
    result = await tools.dispatch(
        "reconcile", {"variable": "apply_haircuts", "reference": "Calc!H2:H4"}
    )
    assert not result.ok
    assert "FAILED" in result.text
    assert "3.5" in result.text
    assert "expected" in result.text and "actual" in result.text


async def test_reconcile_says_so_when_the_notebook_does_not_define_the_variable(
    reconcilable: Workspace,
) -> None:
    _write_notebook(reconcilable, [1.0])
    tools = ToolRegistry(ToolContext(workspace=reconcilable))
    result = await tools.dispatch(
        "reconcile", {"variable": "manual_overrides", "reference": "Calc!H2:H4"}
    )
    assert not result.ok
    assert "does not define 'manual_overrides'" in result.text
    assert "apply_haircuts" in result.text


async def test_reconcile_says_so_when_the_notebook_will_not_run(
    reconcilable: Workspace,
) -> None:
    reconcilable.notebook_path.parent.mkdir(parents=True, exist_ok=True)
    reconcilable.notebook_path.write_text("raise RuntimeError('boom')\n", encoding="utf-8")
    tools = ToolRegistry(ToolContext(workspace=reconcilable))
    result = await tools.dispatch(
        "reconcile", {"variable": "apply_haircuts", "reference": "Calc!H2:H4"}
    )
    assert not result.ok
    assert "NOT RECONCILED" in result.text
    assert "could not be run" in result.text


# ── reconciliation and the analysis in memory ────────────────────────────────────────────────
#
# `context.analysis` is loaded once when the loop is built and never regenerated. A pre-check here
# used to refuse the whole tool when it reported no coverage, which made the sequence kedge itself
# prints — reconcile, be told there is no baseline, open the workbook in Excel, let it calculate,
# save it, ask again — return the identical refusal for the problem the user had just fixed. The
# baseline is read from the file as it is now, so the question belongs to the reconciler, which
# asks it per range and has always answered it correctly.


async def test_a_workbook_with_no_cached_values_is_never_reported_as_passed(
    uncalculated: Workspace,
) -> None:
    """The invariant, whatever the route to it (CLAUDE.md non-negotiable 6).

    The notebook's numbers here are not wrong. There is simply nothing in the file to check them
    against, and "not checked" is the only honest report of that.
    """
    _write_notebook(uncalculated, [4.0, 9.0, 16.0])
    tools = ToolRegistry(ToolContext(workspace=uncalculated))

    result = await tools.dispatch(
        "reconcile", {"variable": "apply_haircuts", "reference": "Calc!H2:H4"}
    )

    assert not result.ok
    assert "NOT RECONCILED" in result.text
    assert "PASSED" not in result.text
    assert "no_cached_values" in result.text


async def test_the_no_baseline_report_says_what_the_removed_pre_check_said(
    uncalculated: Workspace,
) -> None:
    """The remedy the user acts on, and the diagnosis that makes sense of it.

    The pre-check's one piece of content the reason's own explanation lacked was *why* a workbook
    might have nothing cached. That moved into the explanation rather than justifying a stale
    check that kept it.
    """
    _write_notebook(uncalculated, [4.0, 9.0, 16.0])
    tools = ToolRegistry(ToolContext(workspace=uncalculated))

    result = await tools.dispatch(
        "reconcile", {"variable": "apply_haircuts", "reference": "Calc!H2:H4"}
    )

    assert "written by a tool rather than saved by Excel" in result.text
    assert "Open the workbook in Excel, let it recalculate, save it" in result.text


async def test_reconcile_reads_the_workbook_rather_than_an_analysis_that_predates_the_save(
    reconcilable: Workspace,
) -> None:
    """The bug this replaced: the refusal survived the fix it asked for.

    The analysis in memory still says the workbook has nothing cached, because it was taken before
    the user opened it in Excel and saved it. The file disagrees, and the file is what is being
    reconciled against.
    """
    _write_notebook(reconcilable, [1.0, 2.0, 3.0])
    tools = ToolRegistry(ToolContext(workspace=reconcilable, analysis=_no_coverage()))

    result = await tools.dispatch(
        "reconcile", {"variable": "apply_haircuts", "reference": "Calc!H2:H4"}
    )

    assert result.ok, result.text
    assert "PASSED" in result.text


async def test_a_range_without_a_baseline_does_not_condemn_the_ranges_that_have_one(
    reconcilable: Workspace,
) -> None:
    """The question is per range. One workbook, two answers, and neither is the other's."""
    _write_notebook(reconcilable, [1.0, 2.0, 3.0])
    tools = ToolRegistry(ToolContext(workspace=reconcilable))

    cached = await tools.dispatch(
        "reconcile", {"variable": "apply_haircuts", "reference": "Calc!H2:H4"}
    )
    uncached = await tools.dispatch(
        "reconcile", {"variable": "apply_haircuts", "reference": "Calc!K2:K4"}
    )

    assert cached.ok, cached.text
    assert not uncached.ok
    assert "no_cached_values" in uncached.text


async def test_a_baseline_with_holes_in_it_is_not_signed_off_either(
    reconcilable: Workspace,
) -> None:
    """Partial coverage is handled where whole-workbook coverage never was.

    Every row that had a baseline matched, and the region is still NOT RECONCILED: a region cannot
    be signed off on a baseline with holes in it. The removed pre-check only ever fired on a
    workbook with nothing cached at all, so this case has always been the reconciler's, decided
    per range and per row.
    """
    _write_notebook(reconcilable, [1.0])
    tools = ToolRegistry(ToolContext(workspace=reconcilable))

    result = await tools.dispatch(
        "reconcile", {"variable": "apply_haircuts", "reference": "Calc!H2:K2"}
    )

    assert not result.ok
    assert "partial_cached_values" in result.text
    assert "PASSED" not in result.text


async def test_a_stale_analysis_cannot_manufacture_a_pass_either(
    uncalculated: Workspace,
) -> None:
    """The staleness runs both ways, and only one direction would be dangerous.

    Here the analysis says the workbook is fully cached and the file says otherwise. The reconciler
    reads the file, so this is NOT RECONCILED — a report that believed the analysis would be a
    signed-off process with nothing behind it.
    """
    _write_notebook(uncalculated, [4.0, 9.0, 16.0])
    tools = ToolRegistry(ToolContext(workspace=uncalculated, analysis=make_analysis()))

    result = await tools.dispatch(
        "reconcile", {"variable": "apply_haircuts", "reference": "Calc!H2:H4"}
    )

    assert not result.ok
    assert "NOT RECONCILED" in result.text
