"""The fixture generator should be deterministic and agree with what is committed.

Three things are held together here:

* ``generate.py`` reproduces the committed ``.xlsx`` files byte for byte.
* The committed files really do have the properties the corpus depends on —
  cached values present in one fixture and absent in another, a decodable
  DataMashup, parseable connections, and the hostile traps at their stated spots.
* ``manifest.py`` does not drift away from the files. Only the directly
  checkable claims are asserted here; sheet roles, region counts and findings
  need the analyser and belong to ``tests/corpus/``.
"""

from __future__ import annotations

import base64
import io
import re
import sys
import zipfile
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

FIXTURE_DIR = Path(__file__).resolve().parents[1] / "fixtures"
if str(FIXTURE_DIR) not in sys.path:  # pragma: no cover - import plumbing
    sys.path.insert(0, str(FIXTURE_DIR))

import generate  # noqa: E402
import manifest  # noqa: E402

COMPANIONS = ("documented_procedure.docx", "procedure_legacy.doc")


@pytest.fixture(scope="session")
def regenerated(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """Generate the whole corpus once into a temporary directory."""
    target = tmp_path_factory.mktemp("corpus")
    generate.generate(target)
    return target


def _formula_cells(path: Path) -> dict[str, list[str]]:
    """Map sheet name to its formula cell coordinates."""
    workbook = load_workbook(path)
    found: dict[str, list[str]] = {}
    for name in workbook.sheetnames:
        sheet = workbook[name]
        coordinates = [
            cell.coordinate
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
        if coordinates:
            found[name] = coordinates
    workbook.close()
    return found


# =============================================================================
# Determinism
# =============================================================================


@pytest.mark.parametrize("filename", sorted(manifest.FIXTURES))
def test_regenerating_reproduces_the_committed_workbook(regenerated: Path, filename: str) -> None:
    """`generate.py` should reproduce each committed workbook byte for byte."""
    committed = (FIXTURE_DIR / filename).read_bytes()
    fresh = (regenerated / filename).read_bytes()
    if committed == fresh:
        return

    # Byte equality is the claim, but a part-level diff is a far more useful
    # failure message than "1_048_576 bytes differ".
    left = generate.read_parts(FIXTURE_DIR / filename)
    right = generate.read_parts(regenerated / filename)
    differing = sorted(name for name in set(left) | set(right) if left.get(name) != right.get(name))
    pytest.fail(
        f"{filename} differs from the committed copy.\n"
        f"  differing parts: {differing or '(none — container-level difference)'}\n"
        f"  committed {len(committed)} bytes, regenerated {len(fresh)} bytes\n"
        f"  re-run: uv run python tests/fixtures/generate.py"
    )


@pytest.mark.parametrize("filename", COMPANIONS)
def test_regenerating_reproduces_the_committed_companion(regenerated: Path, filename: str) -> None:
    """`generate.py` should reproduce the Word companions byte for byte."""
    assert (FIXTURE_DIR / filename).read_bytes() == (regenerated / filename).read_bytes()


def test_generating_twice_produces_identical_bytes(tmp_path: Path) -> None:
    """Two runs of the generator should agree, with no wall-clock leakage."""
    first, second = tmp_path / "first", tmp_path / "second"
    generate.generate(first)
    generate.generate(second)
    for filename in (*sorted(manifest.FIXTURES), *COMPANIONS):
        assert (first / filename).read_bytes() == (second / filename).read_bytes(), filename


@pytest.mark.parametrize("filename", sorted(manifest.FIXTURES))
def test_document_timestamps_are_pinned(filename: str) -> None:
    """Each workbook should carry the pinned timestamp, not a build time."""
    core = generate.read_parts(FIXTURE_DIR / filename)["docProps/core.xml"].decode("utf-8")
    stamps = re.findall(r"<dcterms:(?:created|modified)[^>]*>([^<]*)</dcterms:", core)
    assert stamps, "no dcterms timestamps found"
    assert set(stamps) == {"2026-01-01T00:00:00Z"}, stamps


# =============================================================================
# Every fixture loads
# =============================================================================


@pytest.mark.parametrize("filename", sorted(manifest.FIXTURES))
def test_workbook_opens_with_formulas(filename: str) -> None:
    """Each fixture should open under `load_workbook(path)` without raising."""
    workbook = load_workbook(FIXTURE_DIR / filename)
    assert workbook.sheetnames
    workbook.close()


@pytest.mark.parametrize("filename", sorted(manifest.FIXTURES))
def test_workbook_opens_with_data_only(filename: str) -> None:
    """Each fixture should open under `data_only=True` without raising."""
    workbook = load_workbook(FIXTURE_DIR / filename, data_only=True)
    assert workbook.sheetnames
    workbook.close()


# =============================================================================
# Cached values — the positive and negative reconciliation paths
# =============================================================================


def test_clean_pipeline_returns_real_cached_values() -> None:
    """`clean_pipeline.xlsx` should yield real numbers under `data_only=True`."""
    workbook = load_workbook(manifest.CLEAN_PIPELINE.path, data_only=True)
    calc = workbook["Calc"]

    rwa = [calc[f"G{row}"].value for row in range(2, 502)]
    assert all(isinstance(value, (int, float)) for value in rwa)
    assert len([v for v in rwa if v]) == 500
    assert sum(rwa) > 0

    # Text results must survive as text, not as a number or None.
    assert calc["A2"].value == "TRD-00001"
    assert isinstance(calc["B2"].value, str)

    totals = [workbook["Output"][f"B{row}"].value for row in range(2, 8)]
    assert all(value is not None for value in totals)
    assert workbook["Output"]["B5"].value == 500  # trade_count

    workbook.close()


def test_clean_pipeline_keeps_formulas_alongside_cached_values() -> None:
    """Injecting cached values should not disturb the formulas."""
    workbook = load_workbook(manifest.CLEAN_PIPELINE.path)
    calc = workbook["Calc"]
    assert calc["D2"].value == "=VLOOKUP(B2,Ref!$A$2:$C$7,2,FALSE)"
    assert calc["G2"].value == "=ROUND(E2*F2,2)"
    assert calc["H501"].value == "=ROUND(G501*capital_ratio,2)"
    workbook.close()


def test_clean_pipeline_cached_rounding_matches_excel_semantics() -> None:
    """Cached ROUND results should be half-away-from-zero at 15 significant digits."""
    workbook = load_workbook(manifest.CLEAN_PIPELINE.path, data_only=True)
    calc = workbook["Calc"]
    for row in range(2, 502):
        adjusted = calc[f"E{row}"].value
        weight = calc[f"F{row}"].value
        rwa = calc[f"G{row}"].value
        assert rwa == generate.excel_round(adjusted * weight, 2), f"row {row}"
    workbook.close()


def test_no_cached_values_fixture_yields_none_for_every_formula() -> None:
    """`no_cached_values.xlsx` should return None for every formula cell."""
    path = manifest.NO_CACHED_VALUES.path
    cached = load_workbook(path, data_only=True)

    checked = 0
    for sheet_name, coordinates in _formula_cells(path).items():
        for coordinate in coordinates:
            assert cached[sheet_name][coordinate].value is None, f"{sheet_name}!{coordinate}"
            checked += 1

    assert checked == manifest.NO_CACHED_VALUES.formula_cell_count
    cached.close()


def test_only_the_declared_fixtures_carry_cached_values() -> None:
    """Cached values should be present in exactly the fixtures that claim them."""
    for expectation in manifest.iter_fixtures():
        formula_cells = _formula_cells(expectation.path)
        if not formula_cells:
            continue
        cached = load_workbook(expectation.path, data_only=True)
        populated = any(
            cached[sheet][coordinate].value is not None
            for sheet, coordinates in formula_cells.items()
            for coordinate in coordinates
        )
        cached.close()
        assert populated == expectation.cached_values_present, expectation.filename


# =============================================================================
# Hand-built XML parts
# =============================================================================


def test_power_query_mashup_is_not_at_item1() -> None:
    """The DataMashup should sit behind decoys, forcing a schema match."""
    parts = generate.read_parts(manifest.POWERQUERY.path)
    items = sorted(name for name in parts if re.fullmatch(r"customXml/item\d+\.xml", name))
    assert items == ["customXml/item1.xml", "customXml/item2.xml", "customXml/item3.xml"]

    carriers = [name for name in items if generate.DATAMASHUP_NS.encode() in parts[name]]
    assert carriers == ["customXml/item3.xml"]


def test_power_query_section_m_round_trips() -> None:
    """Decoding the DataMashup should give back the exact M source."""
    parts = generate.read_parts(manifest.POWERQUERY.path)
    match = re.search(
        rb"<DataMashup[^>]*>(.*?)</DataMashup>", parts["customXml/item3.xml"], re.DOTALL
    )
    assert match is not None

    envelope = base64.b64decode(match.group(1))
    blocks = generate.parse_datamashup(envelope)

    with zipfile.ZipFile(io.BytesIO(blocks["package"])) as inner:
        assert "Formulas/Section1.m" in inner.namelist()
        recovered = inner.read("Formulas/Section1.m").decode("utf-8")

    assert recovered == generate.SECTION_1_M
    for query in manifest.POWERQUERY.power_query_names:
        assert f"shared {query} =" in recovered
    assert 'Excel.CurrentWorkbook(){[Name="tbl_Exposures"]}[Content]' in recovered


def test_power_query_envelope_carries_the_qdeff_header() -> None:
    """The payload should be a real MS-QDEFF envelope, not a bare zip."""
    parts = generate.read_parts(manifest.POWERQUERY.path)
    match = re.search(
        rb"<DataMashup[^>]*>(.*?)</DataMashup>", parts["customXml/item3.xml"], re.DOTALL
    )
    assert match is not None
    envelope = base64.b64decode(match.group(1))

    # Version 0, then the package length, then the package itself at offset 8.
    assert envelope[:4] == b"\x00\x00\x00\x00"
    assert envelope[8:12] == b"PK\x03\x04"

    blocks = generate.parse_datamashup(envelope)
    # The envelope is materially longer than the package it contains: the
    # permissions and metadata blocks trail it.
    trailing = len(envelope) - 8 - len(blocks["package"])
    assert trailing > 0
    assert len(blocks["permissions"]) > 0
    assert len(blocks["metadata"]) > 0


def test_naive_unzip_of_the_envelope_still_works_in_python() -> None:
    """Python's zipfile should tolerate the QDEFF framing, prefix and trailer alike.

    Recorded deliberately rather than assumed. Python locates the central
    directory by scanning backwards from the end, so it opens the envelope in
    spite of the eight leading bytes and the permissions and metadata blocks
    trailing the package. An extractor that base64-decodes and hands the result
    straight to zipfile therefore happens to work here.

    That is worth knowing in both directions: this fixture does NOT punish a
    naive implementation, so it cannot be cited as proof that a DataMashup
    reader handles MS-QDEFF properly. It also means the shortcut silently
    discards the permissions and metadata blocks. See README.md.
    """
    parts = generate.read_parts(manifest.POWERQUERY.path)
    match = re.search(
        rb"<DataMashup[^>]*>(.*?)</DataMashup>", parts["customXml/item3.xml"], re.DOTALL
    )
    assert match is not None
    envelope = base64.b64decode(match.group(1))

    with zipfile.ZipFile(io.BytesIO(envelope)) as naive:
        assert naive.testzip() is None
        assert "Formulas/Section1.m" in naive.namelist()
        assert naive.read("Formulas/Section1.m").decode("utf-8") == generate.SECTION_1_M


def test_power_query_custom_xml_parts_are_wired_up() -> None:
    """Each customXml part should have props, rels and a content-type override."""
    parts = generate.read_parts(manifest.POWERQUERY.path)
    content_types = parts["[Content_Types].xml"]
    workbook_rels = parts["xl/_rels/workbook.xml.rels"]
    for index in (1, 2, 3):
        assert f"customXml/itemProps{index}.xml" in parts
        assert f"customXml/_rels/item{index}.xml.rels" in parts
        assert f"/customXml/itemProps{index}.xml".encode() in content_types
        assert f"../customXml/item{index}.xml".encode() in workbook_rels


def test_legacy_sql_connections_round_trip() -> None:
    """`connections.xml` should parse back with both statements intact."""
    import xml.etree.ElementTree as ElementTree

    parts = generate.read_parts(manifest.LEGACY_SQL.path)
    assert "xl/connections.xml" in parts
    assert b"/xl/connections.xml" in parts["[Content_Types].xml"]
    assert b'Target="connections.xml"' in parts["xl/_rels/workbook.xml.rels"]

    root = ElementTree.fromstring(parts["xl/connections.xml"].decode("utf-8"))
    connections = root.findall(f"{{{generate.SHEET_MAIN_NS}}}connection")
    assert len(connections) == manifest.LEGACY_SQL.connection_count

    by_name: dict[str, Any] = {c.get("name"): c for c in connections}
    assert set(by_name) == {"RiskWarehouse", "FinanceCube"}
    assert by_name["RiskWarehouse"].get("type") == "1"  # ODBC
    assert by_name["FinanceCube"].get("type") == "5"  # OLEDB

    odbc = by_name["RiskWarehouse"].find(f"{{{generate.SHEET_MAIN_NS}}}dbPr")
    assert odbc.get("command") == generate.ODBC_SQL
    assert odbc.get("connection") == generate.ODBC_CONNECTION_STRING

    oledb = by_name["FinanceCube"].find(f"{{{generate.SHEET_MAIN_NS}}}dbPr")
    assert oledb.get("command") == generate.OLEDB_SQL


def test_legacy_sql_statement_keeps_its_line_breaks() -> None:
    """Multi-line SQL should survive XML attribute-value normalisation."""
    import xml.etree.ElementTree as ElementTree

    parts = generate.read_parts(manifest.LEGACY_SQL.path)
    # The newlines must be stored as character references, or a conforming
    # parser would turn them into spaces on the way back in.
    assert b"&#10;" in parts["xl/connections.xml"]

    root = ElementTree.fromstring(parts["xl/connections.xml"].decode("utf-8"))
    dbpr = root.find(f"{{{generate.SHEET_MAIN_NS}}}connection/{{{generate.SHEET_MAIN_NS}}}dbPr")
    command = dbpr.get("command", "")
    assert command.count("\n") == 17
    assert command.startswith("SELECT\n")
    assert "FROM risk.trades AS t" in command


# =============================================================================
# The hostile traps
# =============================================================================


def test_hostile_region_is_broken_at_row_47() -> None:
    """Row 47 should break an otherwise uniform formula region."""
    workbook = load_workbook(manifest.HOSTILE.path)
    messy = workbook["Messy"]
    assert messy["D46"].value == "=B46*C46"
    assert messy["D47"].value == "=B47*C47*1.1"
    assert messy["D48"].value == "=B48*C48"
    workbook.close()


def test_hostile_contains_a_genuine_circular_reference() -> None:
    """The Circular sheet should hold a real cycle and a self-reference."""
    workbook = load_workbook(manifest.HOSTILE.path)
    circular = workbook["Circular"]
    assert circular["C2"].value == "=C4*2"
    assert circular["C3"].value == "=C2+1"
    assert circular["C4"].value == "=C3-5"
    assert circular["E3"].value == "=E3+1"
    workbook.close()


def test_hostile_declares_an_unresolvable_external_link() -> None:
    """The external workbook link should exist as a real part and not resolve."""
    parts = generate.read_parts(manifest.HOSTILE.path)
    assert "xl/externalLinks/externalLink1.xml" in parts
    assert b"externalReferences" in parts["xl/workbook.xml"]

    rels = parts["xl/externalLinks/_rels/externalLink1.xml.rels"].decode("utf-8")
    for target in manifest.HOSTILE.external_links:
        assert target in rels
    assert 'TargetMode="External"' in rels

    workbook = load_workbook(manifest.HOSTILE.path)
    assert workbook["Broken"]["A2"].value == "=[1]Rates!$B$2"
    workbook.close()


def test_hostile_hides_a_sheet_and_two_columns() -> None:
    """The hidden sheet and hidden columns should be present and hidden."""
    workbook = load_workbook(manifest.HOSTILE.path)
    assert workbook["_Archive"].sheet_state == "hidden"
    messy = workbook["Messy"]
    assert messy.column_dimensions["H"].hidden
    assert messy.column_dimensions["I"].hidden
    workbook.close()


def test_hostile_carries_an_error_cell_and_a_broken_defined_name() -> None:
    """A hard #REF! cell and a defined name over a deleted range should exist."""
    workbook = load_workbook(manifest.HOSTILE.path)
    assert workbook["Messy"]["K9"].data_type == "e"
    assert workbook["Messy"]["K9"].value == "#REF!"
    assert "obsolete_rate_table" in workbook.defined_names
    assert "#REF!" in workbook.defined_names["obsolete_rate_table"].value
    workbook.close()


def test_hostile_header_row_repeats_a_column_name() -> None:
    """The header row should contain a duplicate column name."""
    workbook = load_workbook(manifest.HOSTILE.path)
    messy = workbook["Messy"]
    headers = [messy.cell(row=4, column=c).value for c in range(1, 10)]
    assert headers[1] == headers[3] == "Amount"
    workbook.close()


def test_hostile_mixes_text_numbers_and_text_dates() -> None:
    """Columns E and G should hold text where numbers and dates belong."""
    workbook = load_workbook(manifest.HOSTILE.path)
    messy = workbook["Messy"]
    assert isinstance(messy["E5"].value, str)
    assert float(messy["E5"].value) > 0  # it parses, it is just the wrong type

    text_dates = {messy[f"G{row}"].value for row in range(5, 15)}
    assert all(isinstance(value, str) for value in text_dates)
    assert len(text_dates) >= 4, "expected several different date formats"
    workbook.close()


def test_hostile_has_totals_inside_and_below_the_data() -> None:
    """A totals row should sit mid-data as well as at the bottom."""
    workbook = load_workbook(manifest.HOSTILE.path)
    messy = workbook["Messy"]
    assert messy["A30"].value == "Subtotal to date"
    assert str(messy["B30"].value).startswith("=SUM(")
    assert messy["A82"].value == "GRAND TOTAL"
    workbook.close()


# =============================================================================
# Companions
# =============================================================================


def test_word_procedure_is_readable() -> None:
    """`documented_procedure.docx` should open under python-docx."""
    docx = pytest.importorskip("docx")
    document = docx.Document(str(FIXTURE_DIR / "documented_procedure.docx"))
    paragraphs = [p.text for p in document.paragraphs if p.text.strip()]
    assert len(paragraphs) >= 8
    assert any("Operating Procedure" in text for text in paragraphs)


def test_legacy_doc_stub_is_not_a_real_doc() -> None:
    """`procedure_legacy.doc` should not be an OLE2 file, so parsing fails clearly."""
    head = (FIXTURE_DIR / "procedure_legacy.doc").read_bytes()[:8]
    assert head != b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


# =============================================================================
# The manifest agrees with the files
# =============================================================================


def test_manifest_covers_every_committed_workbook() -> None:
    """Every committed .xlsx should have a manifest entry, and vice versa."""
    committed = {path.name for path in FIXTURE_DIR.glob("*.xlsx")}
    assert committed == set(manifest.FIXTURES)


def test_manifest_lists_every_builder() -> None:
    """The manifest and the generator should describe the same corpus."""
    assert set(manifest.FIXTURES) == set(generate.BUILDERS)


@pytest.mark.parametrize("expectation", manifest.iter_fixtures(), ids=lambda e: e.filename)
def test_manifest_sheet_names_match_the_workbook(
    expectation: manifest.FixtureExpectation,
) -> None:
    """Declared sheet names should match the workbook, in order."""
    workbook = load_workbook(expectation.path)
    assert tuple(workbook.sheetnames) == expectation.sheet_names
    workbook.close()


@pytest.mark.parametrize("expectation", manifest.iter_fixtures(), ids=lambda e: e.filename)
def test_manifest_formula_counts_match_the_workbook(
    expectation: manifest.FixtureExpectation,
) -> None:
    """The declared formula cell count should match the workbook."""
    found = sum(len(coords) for coords in _formula_cells(expectation.path).values())
    assert found == expectation.formula_cell_count


@pytest.mark.parametrize("expectation", manifest.iter_fixtures(), ids=lambda e: e.filename)
def test_manifest_sheet_formula_claims_match_the_workbook(
    expectation: manifest.FixtureExpectation,
) -> None:
    """Sheets flagged as carrying formulas should carry them, and only those."""
    with_formulas = set(_formula_cells(expectation.path))
    declared = {sheet.name for sheet in expectation.sheets if sheet.has_formulas}
    assert with_formulas == declared


@pytest.mark.parametrize("expectation", manifest.iter_fixtures(), ids=lambda e: e.filename)
def test_manifest_hidden_sheet_claims_match_the_workbook(
    expectation: manifest.FixtureExpectation,
) -> None:
    """Sheets flagged hidden should be hidden, and only those."""
    workbook = load_workbook(expectation.path)
    hidden = {name for name in workbook.sheetnames if workbook[name].sheet_state != "visible"}
    workbook.close()
    assert hidden == {sheet.name for sheet in expectation.sheets if sheet.hidden}


@pytest.mark.parametrize("expectation", manifest.iter_fixtures(), ids=lambda e: e.filename)
def test_manifest_named_ranges_match_the_workbook(
    expectation: manifest.FixtureExpectation,
) -> None:
    """Declared named ranges should match the workbook's defined names."""
    workbook = load_workbook(expectation.path)
    defined = set(workbook.defined_names)
    workbook.close()
    assert defined == set(expectation.named_ranges)


@pytest.mark.parametrize("expectation", manifest.iter_fixtures(), ids=lambda e: e.filename)
def test_manifest_connection_count_matches_the_workbook(
    expectation: manifest.FixtureExpectation,
) -> None:
    """Only fixtures declaring connections should have a connections part."""
    parts = generate.read_parts(expectation.path)
    if expectation.connection_count == 0:
        assert "xl/connections.xml" not in parts
        return
    found = parts["xl/connections.xml"].count(b"<connection ")
    assert found == expectation.connection_count


@pytest.mark.parametrize("expectation", manifest.iter_fixtures(), ids=lambda e: e.filename)
def test_manifest_power_query_claims_match_the_workbook(
    expectation: manifest.FixtureExpectation,
) -> None:
    """Only fixtures declaring queries should carry a DataMashup part."""
    parts = generate.read_parts(expectation.path)
    carriers = [
        name
        for name in parts
        if re.fullmatch(r"customXml/item\d+\.xml", name)
        and generate.DATAMASHUP_NS.encode() in parts[name]
    ]
    if not expectation.power_query_names:
        assert carriers == []
        return

    assert len(carriers) == 1
    match = re.search(rb"<DataMashup[^>]*>(.*?)</DataMashup>", parts[carriers[0]], re.DOTALL)
    assert match is not None
    blocks = generate.parse_datamashup(base64.b64decode(match.group(1)))
    with zipfile.ZipFile(io.BytesIO(blocks["package"])) as inner:
        source = inner.read("Formulas/Section1.m").decode("utf-8")
    found = tuple(re.findall(r"^shared (\w+) =", source, re.MULTILINE))
    assert found == expectation.power_query_names


@pytest.mark.parametrize("expectation", manifest.iter_fixtures(), ids=lambda e: e.filename)
def test_manifest_companion_files_exist(
    expectation: manifest.FixtureExpectation,
) -> None:
    """Declared companion files should be committed alongside the workbook."""
    for filename in expectation.companion_files:
        assert (FIXTURE_DIR / filename).is_file(), filename


def test_manifest_finding_kinds_come_from_the_declared_vocabulary() -> None:
    """Every expected finding should use a kind listed in FINDING_KINDS."""
    for expectation in manifest.iter_fixtures():
        for finding in expectation.findings:
            assert finding.kind in manifest.FINDING_KINDS, (
                f"{expectation.filename}: unknown finding kind {finding.kind!r}"
            )


def test_manifest_sheet_roles_come_from_the_declared_vocabulary() -> None:
    """Every expected sheet role should be one of the five roles in PLAN 2.4."""
    for expectation in manifest.iter_fixtures():
        for sheet in expectation.sheets:
            roles = (sheet.role, *sheet.role_alternatives)
            for role in roles:
                assert role in manifest.SHEET_ROLES, (
                    f"{expectation.filename}/{sheet.name}: unknown role {role!r}"
                )


def test_exactly_one_fixture_offers_a_reconciliation_baseline() -> None:
    """The corpus should have one positive path for reconciliation, and a contrast."""
    with_cache = manifest.fixtures_with_cached_values()
    assert [f.filename for f in with_cache] == ["clean_pipeline.xlsx"]
    assert manifest.NO_CACHED_VALUES.cached_values_present is False
