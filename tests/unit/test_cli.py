"""Tests for the command line: argument surface, output, and the milestone shims."""

from __future__ import annotations

import importlib
import json
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest
from typer.testing import CliRunner

from conftest import make_draft
from kedge import cli
from kedge import config as config_module
from kedge.cli import app
from kedge.notebook import BridgeReport
from kedge.reconcile import ReconciliationStatus

SECRET = "sk-must-never-be-printed-0123456789"

runner = CliRunner()

_ANSI = re.compile(r"\x1b\[[0-9;]*m")

WATCH_MODULE = importlib.import_module("kedge.ingest.watch")
"""The watched-folder module, fetched the long way round.

``kedge.ingest`` re-exports the ``watch`` *function*, which shadows the submodule of the same
name -- so ``import kedge.ingest.watch as m``, and monkeypatch's dotted string form, both bind
the function instead of the module. Patching a name inside the module needs the module itself.
"""

ANALYSE_MODULE = importlib.import_module("kedge.analysis.analyse")
"""The analyser module, fetched the same way round and for the same reason as ``WATCH_MODULE``.

``kedge.analysis`` re-exports ``analyse``, so the dotted string ``kedge.analysis.analyse``
resolves to the function. `kedge.plan.load_analysis` imports the name off the module, so that is
what a test wanting to prove the analyser did *not* run has to patch.
"""


def unstyled(output: str) -> str:
    """Rendered output with the escape sequences taken out.

    Anything asserting on a *token* in rich-rendered output has to go through this. rich splits
    a styled run wherever its highlighter says to, and for an option it emits
    ``\\x1b[1;36m-\\x1b[0m\\x1b[1;36m--out\\x1b[0m`` -- so ``"--out" in result.output`` is false
    while the text on screen plainly reads ``--out``. Whether that happens depends on whether
    rich thinks colour is available, which is why this passes on a developer's machine and fails
    under CI, where ``FORCE_COLOR`` is set.
    """
    return _ANSI.sub("", output)


def flattened(output: str) -> str:
    """Unstyled output with every run of whitespace collapsed to one space.

    For asserting on a *phrase*. rich word-wraps at the console width, so a sentence that reads
    ``the watched folder ... does not exist`` on screen may carry a newline in the middle of it,
    and the length of a `tmp_path` decides where. Collapsing the whitespace makes the assertion
    about the message rather than about the terminal.
    """
    return " ".join(unstyled(output).split())


@pytest.fixture(autouse=True)
def isolated_home(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    """Every CLI test runs against a temporary ~/.kedge and a temporary working directory."""
    home = tmp_path / "kedge-home"
    home.mkdir()
    monkeypatch.setenv("KEDGE_HOME", str(home))
    working = tmp_path / "cwd"
    working.mkdir()
    monkeypatch.chdir(working)
    return home


@pytest.fixture(autouse=True)
def no_real_keyring(monkeypatch: pytest.MonkeyPatch) -> None:
    """The real Windows Credential Manager is never consulted by a test."""
    monkeypatch.setattr(config_module.keyring, "get_password", lambda service, username: None)


@pytest.fixture
def workbook(tmp_path: Path) -> Path:
    path = tmp_path / "cwd" / "process.xlsx"
    path.write_bytes(b"")
    return path


@pytest.fixture
def inbox(tmp_path: Path) -> Path:
    """The folder another team drops hand-ins into."""
    directory = tmp_path / "inbox"
    directory.mkdir()
    return directory


def moved_bridge() -> BridgeReport:
    """A report from a marimo whose private API has shifted under kedge."""
    return BridgeReport(
        version="0.24.0",
        pinned=cli.MARIMO_PIN,
        problems=("AsyncCodeModeContext.edit_cell() no longer accepts hide_code",),
    )


def pretend_bridge(monkeypatch: pytest.MonkeyPatch, report: BridgeReport) -> None:
    """Point every route to the preflight at one fixed report.

    ``kedge.notebook`` re-exports ``check_bridge``, and ``verify_bridge`` calls the copy in
    ``kedge.notebook.driver``. Patching one and not the other proves nothing: `doctor` would
    read the fake report and `open` the real one.
    """
    monkeypatch.setattr("kedge.notebook.check_bridge", lambda: report)
    monkeypatch.setattr("kedge.notebook.driver.check_bridge", lambda: report)


# ── help and version ─────────────────────────────────────────────────────────────────────────


def test_help_lists_every_command() -> None:
    result = runner.invoke(app, ["--help"])

    assert result.exit_code == 0
    for command in (
        "open",
        "hub",
        "inspect",
        "plan",
        "reconcile",
        "watch",
        "contract",
        "config",
        "doctor",
    ):
        assert command in result.output


def test_no_arguments_shows_help_rather_than_failing_silently() -> None:
    result = runner.invoke(app, [])

    assert "Usage" in result.output


def test_version_reports_the_package_and_the_marimo_pin() -> None:
    result = runner.invoke(app, ["--version"])

    assert result.exit_code == 0
    assert "kedge" in result.output
    assert "0.23.15" in result.output


@pytest.mark.parametrize(
    "command", ["open", "hub", "inspect", "plan", "reconcile", "watch", "config", "doctor"]
)
def test_each_command_has_help(command: str) -> None:
    result = runner.invoke(app, [command, "--help"])

    assert result.exit_code == 0


def test_contract_infer_is_reachable() -> None:
    result = runner.invoke(app, ["contract", "infer", "--help"])

    assert result.exit_code == 0
    assert "--out" in unstyled(result.output)


# ── config ───────────────────────────────────────────────────────────────────────────────────


def test_config_shows_defaults_when_no_files_exist() -> None:
    result = runner.invoke(app, ["config"])

    assert result.exit_code == 0
    assert "sampling.max_rows" in result.output
    assert "reconciliation.absolute_tolerance" in result.output


def test_config_names_the_file_each_value_came_from(isolated_home: Path, tmp_path: Path) -> None:
    (isolated_home / "config.toml").write_text("[sampling]\ntop_k = 3\n", encoding="utf-8")
    (tmp_path / "cwd" / "kedge.toml").write_text("[sampling]\nmax_rows = 25\n", encoding="utf-8")

    result = runner.invoke(app, ["config", "--json"])

    assert result.exit_code == 0
    payload = json.loads(result.output)
    assert payload["values"]["sampling.max_rows"]["value"] == 25
    assert payload["values"]["sampling.max_rows"]["source"].endswith("kedge.toml")
    assert payload["values"]["sampling.top_k"]["source"].endswith("config.toml")
    assert payload["values"]["sampling.head_rows"]["source"] == "default"


def test_config_never_prints_the_api_key(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(config_module.keyring, "get_password", lambda service, username: SECRET)

    plain = runner.invoke(app, ["config"])
    machine = runner.invoke(app, ["config", "--json"])

    # Through unstyled(): a secret broken across two style spans is still a leaked secret, and
    # asserting on the raw output would let exactly that through.
    assert SECRET not in unstyled(plain.output)
    assert SECRET not in unstyled(machine.output)
    assert json.loads(machine.output)["api_key"]["status"] == "present"


def test_config_tells_you_how_to_set_a_missing_key() -> None:
    result = runner.invoke(app, ["config"])

    assert "uv run keyring set kedge" in result.output


def test_config_reports_a_broken_config_file_clearly(tmp_path: Path) -> None:
    (tmp_path / "cwd" / "kedge.toml").write_text(
        "[sampling]\nmax_rows = 'lots'\n", encoding="utf-8"
    )

    result = runner.invoke(app, ["config"])

    assert result.exit_code == 1
    assert "sampling.max_rows" in result.output
    assert "kedge.toml" in result.output


# ── doctor ───────────────────────────────────────────────────────────────────────────────────


def test_doctor_runs_and_reports_each_check() -> None:
    result = runner.invoke(app, ["doctor", "--no-network"])

    assert result.exit_code == 0
    for check in ("python", "marimo", "marimo bridge", "config", "keyring", "marker files"):
        assert check in result.output


def test_doctor_json_is_machine_readable() -> None:
    result = runner.invoke(app, ["doctor", "--no-network", "--json"])

    assert result.exit_code == 0
    payload = json.loads(result.output)
    assert payload["ok"] is True
    names = {check["check"] for check in payload["checks"]}
    assert {
        "python",
        "marimo",
        "marimo bridge",
        "config",
        "keyring",
        "user directory",
        "marker files",
    } <= names


def test_doctor_warns_about_a_missing_keyring_entry() -> None:
    result = runner.invoke(app, ["doctor", "--no-network", "--json"])

    keyring_check = next(c for c in json.loads(result.output)["checks"] if c["check"] == "keyring")
    assert keyring_check["status"] == "warn"
    assert "uv run keyring set kedge" in keyring_check["detail"]


def test_doctor_notices_a_stale_marker(isolated_home: Path, workbook: Path) -> None:
    """A marker whose port answers nothing is a leftover from a crashed run."""
    from kedge.config import Config, LoadedConfig
    from kedge.workspace import Workspace

    workspace = Workspace.for_workbook(workbook, loaded_config=LoadedConfig(config=Config()))
    workspace.ensure_dirs()
    workspace.attach_marimo(host="127.0.0.1", port=1, token="t", pid=999_999)
    workspace.write_marker(kedge_version="0.1.0")

    result = runner.invoke(app, ["doctor", "--no-network", "--json"])

    markers = next(c for c in json.loads(result.output)["checks"] if c["check"] == "marker files")
    assert markers["status"] == "warn"
    assert "stale" in markers["detail"]


def test_doctor_fails_when_the_marimo_pin_does_not_match(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(cli, "installed_marimo_version", lambda: "0.23.9")

    result = runner.invoke(app, ["doctor", "--no-network", "--json"])

    assert result.exit_code == 1
    payload = json.loads(result.output)
    assert payload["ok"] is False
    marimo_check = next(c for c in payload["checks"] if c["check"] == "marimo")
    assert marimo_check["status"] == "fail"
    assert "0.23.15" in marimo_check["detail"]


def test_doctor_introspects_the_private_marimo_api_the_bridge_depends_on() -> None:
    """PLAN 6.1 mitigation 5, run rather than merely built.

    The `marimo` check above compares two version strings. This one imports ``_code_mode`` and
    looks at it, which is the only thing that catches a surface that moved without the version
    kedge pins moving with it.
    """
    result = runner.invoke(app, ["doctor", "--no-network", "--json"])

    bridge = next(c for c in json.loads(result.output)["checks"] if c["check"] == "marimo bridge")
    assert bridge["status"] == "ok"
    assert cli.MARIMO_PIN in bridge["detail"]


def test_doctor_fails_and_names_the_marimo_version_when_the_bridge_has_moved(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """ "marimo changed" is not actionable; "edit_cell() lost hide_code in 0.24.0" is."""
    pretend_bridge(monkeypatch, moved_bridge())

    result = runner.invoke(app, ["doctor", "--no-network", "--json"])

    assert result.exit_code == 1
    payload = json.loads(result.output)
    assert payload["ok"] is False
    bridge = next(c for c in payload["checks"] if c["check"] == "marimo bridge")
    assert bridge["status"] == "fail"
    assert "0.24.0" in bridge["detail"]
    assert "edit_cell" in bridge["detail"]


def test_doctor_only_warns_when_an_unpinned_marimo_still_has_the_shape_kedge_drives(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A version kedge has not verified against is worth saying; it is not a broken install."""
    pretend_bridge(monkeypatch, BridgeReport(version="0.24.0", pinned=cli.MARIMO_PIN))

    result = runner.invoke(app, ["doctor", "--no-network", "--json"])

    assert result.exit_code == 0
    bridge = next(c for c in json.loads(result.output)["checks"] if c["check"] == "marimo bridge")
    assert bridge["status"] == "warn"
    assert "0.24.0" in bridge["detail"]


def test_doctor_reports_a_workspace_for_a_named_workbook(workbook: Path) -> None:
    result = runner.invoke(app, ["doctor", "--no-network", "--json", "--workbook", str(workbook)])

    workspace_check = next(
        c for c in json.loads(result.output)["checks"] if c["check"] == "workspace"
    )
    assert workspace_check["status"] == "ok"
    assert "process.kedge" in workspace_check["detail"]
    assert re.search(r"session id s_[\da-z]{6}", workspace_check["detail"])


# ── argument validation ──────────────────────────────────────────────────────────────────────


@pytest.mark.parametrize(
    "argv",
    [
        ["open", "absent.xlsx"],
        ["inspect", "absent.xlsx"],
        ["watch", "absent.xlsx"],
        ["contract", "infer", "absent.csv"],
        ["plan", "propose", "absent.xlsx"],
        ["plan", "show", "absent.xlsx"],
        ["plan", "approve", "absent.xlsx"],
        ["plan", "history", "absent.xlsx"],
    ],
)
def test_a_missing_input_file_is_reported_before_anything_else_happens(argv: list[str]) -> None:
    result = runner.invoke(app, argv)

    assert result.exit_code == 1
    assert "no such" in result.output


def test_reconcile_checks_both_of_its_inputs(workbook: Path) -> None:
    result = runner.invoke(app, ["reconcile", "absent.py", str(workbook)])

    assert result.exit_code == 1
    assert "no such file" in result.output


# ── milestone shims ──────────────────────────────────────────────────────────────────────────


def test_inspect_actually_analyses_now_that_m1_has_landed(workbook: Path, tmp_path: Path) -> None:
    """`inspect` is wired all the way through: no shim, a real analysis on disk."""
    destination = tmp_path / "analysis.json"
    result = runner.invoke(app, ["inspect", str(workbook), "--out", str(destination)])

    assert result.exception is None, result.output
    assert destination.is_file()
    analysis = json.loads(destination.read_text(encoding="utf-8"))
    assert analysis["workbook"]["filename"] == workbook.name
    assert analysis["schema_version"]


def test_inspect_writes_a_self_contained_report(workbook: Path, tmp_path: Path) -> None:
    report = tmp_path / "report.html"
    result = runner.invoke(app, ["inspect", str(workbook), "--report", str(report)])

    assert result.exception is None, result.output
    html = report.read_text(encoding="utf-8")
    # A report on someone's workbook must not fetch anything when opened.
    assert "https://" not in html
    assert "<title>" in html


def test_a_command_whose_milestone_has_not_landed_names_the_owing_module() -> None:
    """Nothing silently no-ops; an unimplemented command says whose it is.

    Asserted against `_resolve` itself rather than against whichever command still has a shim.
    Every command this test used to reach through has since been implemented -- which is the
    point of the shim, and also how the test kept being rewritten. The mechanism is the durable
    subject: it must name the module, the milestone and the file that owes the work.
    """
    with pytest.raises(NotImplementedError) as caught:
        cli._resolve("kedge.contracts.infer", "infer_the_whole_thing", "M5 (contracts)")

    message = str(caught.value)
    assert "kedge.contracts.infer.infer_the_whole_thing" in message
    assert "M5 (contracts)" in message
    assert "src/kedge/contracts/infer.py" in message


def test_a_shim_for_a_module_that_does_not_exist_says_so_the_same_way() -> None:
    """The other half: an absent module, not merely an absent attribute in a present one."""
    with pytest.raises(NotImplementedError, match="M9 \\(not a milestone\\)"):
        cli._resolve("kedge.does_not_exist", "anything", "M9 (not a milestone)")


def test_reconcile_never_reports_passed_when_it_could_not_run(
    workbook: Path, tmp_path: Path
) -> None:
    """The most dangerous failure mode in the project, pinned at the CLI boundary.

    An empty notebook produces no values to compare. That is "not reconciled", and it must exit
    non-zero so the command is usable as a gate -- a vacuous pass would be worse than a crash
    (PLAN 4.5).
    """
    notebook = tmp_path / "cwd" / "process.py"
    notebook.parent.mkdir(parents=True, exist_ok=True)
    notebook.write_text("", encoding="utf-8")

    result = runner.invoke(app, ["reconcile", str(notebook), str(workbook), "--json"])

    assert result.exit_code != 0, "a reconciliation that could not run must not exit 0"
    report = json.loads(result.stdout)
    statuses = {region["status"] for region in report["regions"]}
    assert "passed" not in statuses


def test_contract_infer_writes_a_commented_draft_that_validates_its_own_source(
    tmp_path: Path,
) -> None:
    """The whole point of a draft: it passes against the file it was drafted from.

    Regression, and a sharp one. The command resolved `kedge.contracts.infer.infer_contract`,
    which has never existed -- the module exports `infer`, `infer_with_notes` and `write_yaml`
    -- so every invocation died with `NotImplementedError`. The scaffolded notebook tells the
    user to run this exact command when it finds no contract, so a broken command made that
    guidance a dead end.
    """
    handin = tmp_path / "cwd" / "handin.csv"
    handin.parent.mkdir(parents=True, exist_ok=True)
    handin.write_text("id,amount\n1,10.5\n2,20.25\n", encoding="utf-8")
    destination = tmp_path / "contract.yaml"

    result = runner.invoke(app, ["contract", "infer", str(handin), "--out", str(destination)])

    assert result.exit_code == 0, result.output
    assert destination.is_file()
    body = destination.read_text(encoding="utf-8")
    assert "# Contract:" in body  # the commentary, not just the model
    assert "id" in body and "amount" in body

    from kedge.contracts import load
    from kedge.contracts.validate import validate_frame
    from kedge.ingest import read_data

    contract = load(destination)
    frame, layout = read_data(handin, sheet=contract.sheet, header_row=contract.header_row)
    report = validate_frame(frame, contract, handin_name=handin.name, layout=layout)
    assert report.ok, report.summary_line()


def test_contract_infer_defaults_its_output_beside_the_hand_in(tmp_path: Path) -> None:
    handin = tmp_path / "cwd" / "handin.csv"
    handin.parent.mkdir(parents=True, exist_ok=True)
    handin.write_text("id,amount\n1,10.5\n", encoding="utf-8")

    result = runner.invoke(app, ["contract", "infer", str(handin)])

    assert result.exit_code == 0, result.output
    assert (tmp_path / "cwd" / "handin.contract.yaml").is_file()


def test_contract_infer_reports_a_file_it_cannot_describe_rather_than_tracebacking(
    tmp_path: Path,
) -> None:
    """A file with no columns is a `_fail`, not a stack trace: it names what went wrong."""
    handin = tmp_path / "cwd" / "empty.csv"
    handin.parent.mkdir(parents=True, exist_ok=True)
    handin.write_text("", encoding="utf-8")

    result = runner.invoke(app, ["contract", "infer", str(handin)])

    assert result.exit_code == 1
    assert "error" in flattened(result.output)
    assert not (tmp_path / "cwd" / "empty.contract.yaml").exists()


def _pasted_workbook(tmp_path: Path) -> Path:
    """The committed fixture, copied where a command may write a project directory beside it."""
    fixture = Path(__file__).resolve().parents[1] / "fixtures" / "legacy_sql.xlsx"
    if not fixture.is_file():  # pragma: no cover - the fixture is committed
        pytest.skip("tests/fixtures/legacy_sql.xlsx has not landed yet")
    destination = tmp_path / "cwd" / "legacy_sql.xlsx"
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_bytes(fixture.read_bytes())
    return destination


def test_contract_sketch_drafts_from_the_workbooks_own_pasted_sheet(tmp_path: Path) -> None:
    """The bootstrap: a contract before the first export, drawn off last month's paste."""
    workbook = _pasted_workbook(tmp_path)
    destination = tmp_path / "contract.yaml"

    result = runner.invoke(app, ["contract", "sketch", str(workbook), "--out", str(destination)])

    assert result.exit_code == 0, result.output
    from kedge.contracts import load

    contract = load(destination)
    assert contract.column_names[:2] == ["trade_id", "counterparty_name"]
    assert "legacy_sql.xlsx" in (contract.generated_from or "")
    assert "sketch" in flattened(result.output)


def test_contract_sketch_defaults_to_the_contract_the_notebook_reads(tmp_path: Path) -> None:
    """Writing it anywhere else would leave the scaffolded check cell still enforcing nothing."""
    workbook = _pasted_workbook(tmp_path)

    result = runner.invoke(app, ["contract", "sketch", str(workbook)])

    assert result.exit_code == 0, result.output
    from kedge.workspace import Workspace

    assert Workspace.for_workbook(workbook).contract_path.is_file()


def test_contract_sketch_will_not_overwrite_a_contract_without_being_told(
    tmp_path: Path,
) -> None:
    workbook = _pasted_workbook(tmp_path)
    destination = tmp_path / "contract.yaml"
    destination.write_text("name: tightened-by-hand\n", encoding="utf-8")

    result = runner.invoke(app, ["contract", "sketch", str(workbook), "--out", str(destination)])

    assert result.exit_code == 1
    assert "already a contract" in flattened(result.output)
    assert destination.read_text(encoding="utf-8") == "name: tightened-by-hand\n"

    forced = runner.invoke(
        app, ["contract", "sketch", str(workbook), "--out", str(destination), "--force"]
    )
    assert forced.exit_code == 0, forced.output
    assert "tightened-by-hand" not in destination.read_text(encoding="utf-8")


def test_contract_sketch_reports_a_workbook_it_cannot_speak_for(tmp_path: Path) -> None:
    """A named sheet that is not there is a message naming the ones that are."""
    workbook = _pasted_workbook(tmp_path)

    result = runner.invoke(
        app,
        [
            "contract",
            "sketch",
            str(workbook),
            "--sheet",
            "Trades",
            "--out",
            str(tmp_path / "c.yaml"),
        ],
    )

    assert result.exit_code == 1
    assert "has no sheet called 'Trades'" in flattened(result.output)
    assert not (tmp_path / "c.yaml").exists()


def test_contract_sketch_says_which_sheet_it_chose_and_why(tmp_path: Path) -> None:
    workbook = _pasted_workbook(tmp_path)

    result = runner.invoke(
        app, ["contract", "sketch", str(workbook), "--out", str(tmp_path / "c.yaml")]
    )

    output = flattened(result.output)
    assert "'Extract' sheet" in output
    assert "only sheet" in output
    assert "not a checked contract" in output


def test_open_runs_the_same_sequence_the_hub_runs(
    workbook: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """`open` and the hub must not keep separate copies of the open sequence.

    Ordering in that sequence is not cosmetic: scaffolding needs a live driver, so it cannot
    precede the marimo launch. `open` used to have its own copy with the steps in the wrong order
    and raised a TypeError before it ever launched marimo. It now delegates.
    """
    called: dict[str, object] = {}

    async def _capture(state, path, **kwargs):
        called["path"] = Path(path)
        called["reattach"] = kwargs.get("reattach")
        called["workspace"] = kwargs.get("workspace")
        raise SystemExit(0)

    monkeypatch.setattr("kedge.server.hub.open_workbook", _capture)
    runner.invoke(app, ["open", str(workbook), "--no-browser"])

    assert called["path"] == workbook
    assert called["reattach"] is True, "a crashed previous run is the common case on the CLI"


def test_the_port_option_actually_pins_the_marimo_port(
    workbook: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """An accepted flag that is only printed is worse than no flag at all."""
    seen: dict[str, int] = {}

    async def _capture(_state, _path, **kwargs):
        workspace = kwargs["workspace"]
        seen["port"] = workspace.config.marimo.port
        raise SystemExit(0)

    monkeypatch.setattr("kedge.server.hub.open_workbook", _capture)

    runner.invoke(app, ["open", str(workbook), "--port", "2718", "--no-browser"])

    assert seen["port"] == 2718, "the pinned port must reach the workspace marimo is launched from"


# ── hub ──────────────────────────────────────────────────────────────────────────────────────


def test_hub_starts_the_server_with_no_workbook_and_opens_the_browser_at_it(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """`kedge hub` differs from `kedge open` in exactly one thing: nothing is attached."""
    served: dict[str, object] = {}

    def _capture(app_: object, **kwargs: object) -> None:
        served["state"] = app_.state.kedge
        served["port"] = kwargs.get("port")

    monkeypatch.setattr("kedge.server.app.run_server", _capture)
    monkeypatch.setattr("webbrowser.open", lambda url: served.setdefault("url", url))

    result = runner.invoke(app, ["hub", "--port", "8731"])

    assert result.exit_code == 0
    assert served["port"] == 8731
    assert served["url"] == "http://127.0.0.1:8731"
    assert served["state"].workspace is None
    assert served["state"].agent is None


def test_hub_can_be_told_not_to_open_a_browser(monkeypatch: pytest.MonkeyPatch) -> None:
    opened: list[str] = []
    monkeypatch.setattr("kedge.server.app.run_server", lambda *_a, **_k: None)
    monkeypatch.setattr("webbrowser.open", lambda url: opened.append(url))

    result = runner.invoke(app, ["hub", "--no-browser"])

    assert result.exit_code == 0
    assert opened == []


def test_hub_reports_how_many_workbooks_it_knows_about(
    isolated_home: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr("kedge.server.app.run_server", lambda *_a, **_k: None)

    result = runner.invoke(app, ["hub", "--no-browser"])

    assert "no workbooks registered yet" in result.output


def test_opening_a_workbook_registers_it_so_the_hub_and_the_terminal_agree(
    isolated_home: Path, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import zipfile

    from kedge.registry import WorkbookRegistry

    path = tmp_path / "cwd" / "real.xlsx"
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr("[Content_Types].xml", "<Types/>")

    async def _stop(*_args, **_kwargs):
        raise SystemExit(0)

    monkeypatch.setattr("kedge.server.hub._run_open", _stop)
    runner.invoke(app, ["open", str(path), "--no-browser"])

    assert [entry.name for entry in WorkbookRegistry.for_user().entries()] == ["real.xlsx"]


def test_a_workbook_that_will_not_register_does_not_stop_open(
    isolated_home: Path, workbook: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The `workbook` fixture is an empty file, which is not a real workbook. `open` still runs:
    the user has already said which file they mean."""
    from kedge.registry import WorkbookRegistry

    reached: list[str] = []

    async def _capture(*_args, **_kwargs):
        reached.append("sequence started")
        raise SystemExit(0)

    monkeypatch.setattr("kedge.server.hub._run_open", _capture)
    runner.invoke(app, ["open", str(workbook), "--no-browser"])

    assert reached == ["sequence started"], "registration failing must not abort the open"
    assert WorkbookRegistry.for_user().entries() == []


def test_a_project_config_beside_the_workbook_is_picked_up(workbook: Path) -> None:
    """kedge.toml is hand-written, so it lives beside the workbook, not in the generated dir."""
    (workbook.parent / "kedge.toml").write_text("[sampling]\nmax_rows = 7\n", encoding="utf-8")

    result = runner.invoke(app, ["config", "--json", "--workbook", str(workbook)])

    payload = json.loads(result.output)
    assert payload["values"]["sampling.max_rows"]["value"] == 7


# ── plan ─────────────────────────────────────────────────────────────────────────────────────
#
# `propose` is exercised offline throughout, against `ScriptedCompleter`. That fake lives in the
# library rather than in this suite precisely so the whole path -- context, prompt, parse,
# validate, assemble, save -- runs with no endpoint; a CLI test that stubbed `run_plan` itself
# would prove nothing about the command's own wiring, which is the part that was missing.


def plans_under(workbook: Path) -> list[Path]:
    """Every saved plan version beside a workbook, in version order."""
    return sorted(workbook.parent.rglob("plan-v*.yaml"))


def latest_plan_text(workbook: Path) -> str:
    return plans_under(workbook)[-1].read_text(encoding="utf-8")


def write_workbook(path: Path, *, rate: float = 1.5) -> Path:
    """A real workbook, with enough in it that triage says proceed.

    `rate` varies the data so two workbooks written this way differ in content, and therefore in
    digest -- which is what a check on workbook identity has to be tested against. Two files with
    the same bytes are the same workbook under two names, and that is a different question.
    """
    from openpyxl import Workbook

    book = Workbook()
    data = book.active
    data.title = "Data"
    data["A1"], data["B1"] = "id", "amount"
    for row in range(2, 12):
        data.cell(row=row, column=1, value=row)
        data.cell(row=row, column=2, value=row * rate)
    calculation = book.create_sheet("Calc")
    calculation["A1"] = "doubled"
    for row in range(2, 12):
        calculation.cell(row=row, column=1, value=f"=Data!B{row}*2")
    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    return path


def change_the_workbook(path: Path) -> None:
    """Edit a workbook in place, so it no longer hashes to what an artifact recorded for it."""
    from openpyxl import load_workbook

    book = load_workbook(path)
    book["Data"]["A20"] = "a column somebody added after the plan was written"
    book.save(path)


@pytest.fixture
def planned_workbook(tmp_path: Path) -> Path:
    """The workbook every plan test converts.

    The `workbook` fixture is an empty file, which triage refuses outright -- correct, and
    useless for reaching the path that actually produces a plan.
    """
    return write_workbook(tmp_path / "cwd" / "process.xlsx")


@pytest.fixture
def other_workbook(tmp_path: Path) -> Path:
    """A second, different workbook, in a directory of its own.

    Kept away from `planned_workbook`'s directory on purpose: `plans_under` globs a workbook's
    parent, so a second project directory beside the first would make "nothing was written"
    unassertable.
    """
    return write_workbook(tmp_path / "elsewhere" / "beta.xlsx", rate=2.25)


@pytest.fixture
def offline_model(monkeypatch: pytest.MonkeyPatch) -> None:
    """Answer `kedge plan propose` from a scripted response instead of an endpoint."""
    from kedge.plan.propose import scripted_from_plan

    monkeypatch.setattr(
        "kedge.plan.completer_from_config", lambda _config: scripted_from_plan(make_draft())
    )


def propose(workbook: Path) -> None:
    """Put one draft plan on disk, failing the test loudly if that did not work."""
    result = runner.invoke(app, ["plan", "propose", str(workbook)])
    assert result.exit_code == 0, result.output


@pytest.mark.parametrize(
    "verb", ["propose", "show", "approve", "acknowledge", "reject", "request-changes", "history"]
)
def test_each_plan_verb_has_help(verb: str) -> None:
    result = runner.invoke(app, ["plan", verb, "--help"])

    assert result.exit_code == 0


def test_propose_has_no_way_of_approving_in_the_same_breath() -> None:
    """The gate is the product, so the absence of the flag that would defeat it is a test.

    Proposing and approving in one command would put a decomposition into force without anyone
    having read it, which is the single thing the plan artifact exists to prevent (PLAN 2.2).
    """
    result = runner.invoke(app, ["plan", "propose", "--help"])

    output = unstyled(result.output)
    assert "--approve" not in output
    assert "--dry-run" in output


def test_propose_writes_a_draft_and_nothing_more(
    planned_workbook: Path, offline_model: None
) -> None:
    result = runner.invoke(app, ["plan", "propose", str(planned_workbook)])

    assert result.exit_code == 0, result.output
    saved = plans_under(planned_workbook)
    assert [path.name for path in saved] == ["plan-v001.yaml"]
    assert "state: draft" in saved[0].read_text(encoding="utf-8")
    assert "kedge plan approve" in flattened(result.output)


def test_propose_dry_run_writes_nothing_at_all(
    planned_workbook: Path, offline_model: None, isolated_home: Path
) -> None:
    """PLAN 7 step 4 judges `propose` across the corpus with this, so it is load-bearing.

    Not merely "no plan file": no project directory beside the workbook, and nothing under
    `~/.kedge` either. A sweep over somebody's workbooks that leaves a `.kedge` folder beside each
    one is not a read-only sweep, and neither is one that leaves a trail in the machine-wide
    directory instead.
    """
    result = runner.invoke(app, ["plan", "propose", str(planned_workbook), "--dry-run"])

    assert result.exit_code == 0, result.output
    assert "STAGES" in result.output, "the plan is still rendered; only the write is skipped"
    assert plans_under(planned_workbook) == []
    assert list(planned_workbook.parent.glob("*.kedge")) == []
    assert list(isolated_home.rglob("*")) == [], "KEDGE_HOME is untouched too"


def test_propose_json_carries_the_triage_the_warnings_and_the_plan(
    planned_workbook: Path, offline_model: None
) -> None:
    result = runner.invoke(app, ["plan", "propose", str(planned_workbook), "--dry-run", "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.stdout)
    assert payload["stopped"] is False
    assert payload["saved_to"] is None, "--dry-run saved nothing, and must say so"
    assert payload["triage"]["verdict"]
    assert payload["plan"]["approval"]["state"] == "draft"
    assert [stage["id"] for stage in payload["plan"]["stages"]]


def test_propose_analyses_the_workbook_when_none_has_been_saved(
    planned_workbook: Path, offline_model: None
) -> None:
    """`load_analysis`'s fallback branch, on the command line rather than in a unit test.

    That branch shipped a live `TypeError` for a while -- `from kedge.analysis import analyse`
    bound the *module*, not the function -- and survived because nothing ran it
    (`docs/ty-diagnostics.md` 5). `kedge inspect` does not go through it; this does.
    """
    assert not list(planned_workbook.parent.rglob("analysis.json"))

    result = runner.invoke(app, ["plan", "propose", str(planned_workbook), "--dry-run"])

    assert result.exit_code == 0, result.output
    assert "STAGES" in result.output


def test_propose_uses_the_analysis_it_is_handed_rather_than_analysing_again(
    planned_workbook: Path, tmp_path: Path, offline_model: None, monkeypatch: pytest.MonkeyPatch
) -> None:
    """An accepted `--analysis` that is quietly ignored would be worse than no flag at all."""
    saved = tmp_path / "analysis.json"
    assert (
        runner.invoke(app, ["inspect", str(planned_workbook), "--out", str(saved)]).exit_code == 0
    )

    def _refuse(*_args: object, **_kwargs: object) -> None:
        raise AssertionError("the saved analysis must be preferred over re-analysing")

    # The module, not the re-exported function of the same name: `kedge.analysis` exports
    # `analyse`, so monkeypatch's dotted string form resolves `kedge.analysis.analyse` to the
    # function and never reaches the module the planner imports from.
    monkeypatch.setattr(ANALYSE_MODULE, "analyse", _refuse)

    result = runner.invoke(
        app, ["plan", "propose", str(planned_workbook), "--analysis", str(saved), "--dry-run"]
    )

    assert result.exit_code == 0, result.output


def test_propose_names_an_analysis_file_that_is_not_there(planned_workbook: Path) -> None:
    result = runner.invoke(
        app, ["plan", "propose", str(planned_workbook), "--analysis", "absent.json"]
    )

    assert result.exit_code == 1
    assert "no such analysis" in flattened(result.output)


def test_propose_stops_when_triage_says_stop_and_does_not_exit_zero(
    workbook: Path, offline_model: None
) -> None:
    """An honest refusal is a legitimate result -- and must not read as success to a script."""
    result = runner.invoke(app, ["plan", "propose", str(workbook)])

    assert result.exit_code == cli.TRIAGE_REFUSED
    assert plans_under(workbook) == []
    assert "--force" in flattened(result.output), "the way past a refusal has to be named"


def test_a_triage_refusal_does_not_look_like_a_broken_invocation(
    workbook: Path, offline_model: None
) -> None:
    """ "This workbook should not be converted" and "no such workbook" are different answers.

    Both were exit 1, so a script sweeping a folder could not tell an editorial judgement from a
    typo in a path or a missing API key. `PlanRun.stopped` models a refusal as a result; the exit
    code has to as well.
    """
    refused = runner.invoke(app, ["plan", "propose", str(workbook)])
    broken = runner.invoke(app, ["plan", "propose", "absent.xlsx"])

    assert cli.TRIAGE_REFUSED == 2
    assert refused.exit_code == 2
    assert broken.exit_code == 1


def test_forcing_past_a_stop_verdict_still_only_produces_a_draft(
    workbook: Path, offline_model: None
) -> None:
    result = runner.invoke(app, ["plan", "propose", str(workbook), "--force"])

    assert result.exit_code == 0, result.output
    assert "state: draft" in latest_plan_text(workbook)


def test_show_renders_the_plan_and_what_stands_between_it_and_approval(
    planned_workbook: Path, offline_model: None
) -> None:
    propose(planned_workbook)

    result = runner.invoke(app, ["plan", "show", str(planned_workbook)])

    assert result.exit_code == 0, result.output
    output = flattened(result.output)
    assert "STAGES" in output
    assert "cannot be approved yet" in output
    assert "Calc!AK:AP" in output, "the unacknowledged drop is the blocker; it has to be named"


def test_show_can_render_a_superseded_version(planned_workbook: Path, offline_model: None) -> None:
    """History is retained rather than overwritten, so it has to be readable."""
    propose(planned_workbook)
    runner.invoke(app, ["plan", "acknowledge", str(planned_workbook), "--all"])

    result = runner.invoke(app, ["plan", "show", str(planned_workbook), "--version", "1"])

    assert result.exit_code == 0, result.output
    assert "Process plan v1" in flattened(result.output)


def test_show_names_a_version_that_is_not_there(
    planned_workbook: Path, offline_model: None
) -> None:
    propose(planned_workbook)

    result = runner.invoke(app, ["plan", "show", str(planned_workbook), "--version", "9"])

    assert result.exit_code == 1
    assert "no plan version 9" in flattened(result.output)


def test_approve_refuses_while_a_drop_is_unacknowledged_and_says_what_would_clear_it(
    planned_workbook: Path, offline_model: None
) -> None:
    """The blocker that matters most: a drop nobody signed off (PLAN 2.2).

    Every blocker is listed, not just the first -- a user told only the first will fix it, retry,
    and be told the second -- and each one comes with the command that clears it.
    """
    propose(planned_workbook)

    result = runner.invoke(app, ["plan", "approve", str(planned_workbook), "--by", "phil"])

    assert result.exit_code == 1
    output = flattened(result.output)
    assert "cannot be approved" in output
    assert "Calc!AK:AP" in output
    assert "kedge plan acknowledge" in output
    assert "state: draft" in latest_plan_text(planned_workbook), "nothing was approved"


def test_acknowledging_the_drops_is_what_lets_an_approval_through(
    planned_workbook: Path, offline_model: None
) -> None:
    propose(planned_workbook)

    acknowledged = runner.invoke(
        app, ["plan", "acknowledge", str(planned_workbook), "--all", "--note", "read and agreed"]
    )
    approved = runner.invoke(app, ["plan", "approve", str(planned_workbook), "--by", "phil"])

    assert acknowledged.exit_code == 0, acknowledged.output
    assert approved.exit_code == 0, approved.output
    saved = plans_under(planned_workbook)
    assert [path.name for path in saved] == [
        "plan-v001.yaml",
        "plan-v002.yaml",
        "plan-v003.yaml",
    ], "an edit is a new version, a decision is a new version, and neither replaces what it names"
    assert "state: draft" in saved[0].read_text(encoding="utf-8")
    approved_text = saved[2].read_text(encoding="utf-8")
    assert "state: approved" in approved_text
    assert "by: phil" in approved_text
    assert "based_on_version: 2" in approved_text, "the decision says which version it was about"


def test_approving_the_plan_already_in_force_records_nothing(
    planned_workbook: Path, offline_model: None
) -> None:
    """Idempotent, and quiet about it.

    Every other decision is a new version because it is a change of mind worth reading later.
    Approving what is already approved is not one: the only difference between the version it
    would write and the version before it is a timestamp, and a history of those says nothing.
    """
    propose(planned_workbook)
    runner.invoke(app, ["plan", "acknowledge", str(planned_workbook), "--all"])
    runner.invoke(app, ["plan", "approve", str(planned_workbook), "--by", "phil"])
    before = [path.name for path in plans_under(planned_workbook)]

    again = runner.invoke(app, ["plan", "approve", str(planned_workbook), "--by", "mallory"])

    assert again.exit_code == 0, again.output
    assert "already approved" in flattened(again.output)
    assert [path.name for path in plans_under(planned_workbook)] == before
    latest = plans_under(planned_workbook)[-1].read_text(encoding="utf-8")
    assert "by: phil" in latest, "the reviewer of record is unchanged"
    assert "mallory" not in latest, "a second name is not recorded by approving twice"


def test_an_approval_records_who_gave_it_without_having_to_be_told(
    planned_workbook: Path, offline_model: None, monkeypatch: pytest.MonkeyPatch
) -> None:
    """kedge is single-user and local, so `--by` defaults rather than nagging.

    An approval with no name against it would make the audit trail worse than the spreadsheet
    it replaced, so the default is the operating system's user. Pinned against a known name
    rather than against `cli._reviewer(None)`: checking the function against itself would pass
    just as happily if it returned a constant.
    """
    monkeypatch.setattr(cli.getpass, "getuser", lambda: "philm")
    propose(planned_workbook)
    runner.invoke(app, ["plan", "acknowledge", str(planned_workbook), "--all"])

    result = runner.invoke(app, ["plan", "approve", str(planned_workbook)])

    assert result.exit_code == 0, result.output
    assert "by: philm" in latest_plan_text(planned_workbook)
    assert cli._reviewer("  phil  ") == "phil", "an explicit --by wins, trimmed"


def test_a_name_nobody_typed_is_not_recorded_as_though_somebody_had(
    planned_workbook: Path, offline_model: None, monkeypatch: pytest.MonkeyPatch
) -> None:
    """`getpass.getuser()` reads the environment before it asks the OS, so it is a weak claim.

    `USER=mallory kedge plan approve` records mallory with no privilege at all; under CI the same
    default records `runner`, and in a container `root`. Making `--by` mandatory would turn the
    audit trail into paperwork somebody types `x` into, so the default stays and the *record*
    carries the difference instead: months later, "philm" and "philm, because the shell said so"
    are not the same claim.
    """
    for variable in ("LOGNAME", "USER", "LNAME", "USERNAME"):
        monkeypatch.setenv(variable, "mallory")  # the four `getpass` consults, in its own order
    propose(planned_workbook)
    runner.invoke(app, ["plan", "acknowledge", str(planned_workbook), "--all"])

    inferred = runner.invoke(app, ["plan", "approve", str(planned_workbook)])
    body = latest_plan_text(planned_workbook)

    assert inferred.exit_code == 0, inferred.output
    assert "by: mallory (inferred from the OS user)" in body
    assert body.count("inferred from the OS user") == 1, (
        "the marker belongs to the claim about who reviewed it, not to the reviewer's own note"
    )


def test_a_reviewer_who_names_themselves_is_recorded_verbatim(
    planned_workbook: Path, offline_model: None
) -> None:
    """The other half of the same claim: a name somebody typed carries no hedge."""
    propose(planned_workbook)
    runner.invoke(app, ["plan", "acknowledge", str(planned_workbook), "--all"])

    result = runner.invoke(app, ["plan", "approve", str(planned_workbook), "--by", "phil"])

    assert result.exit_code == 0, result.output
    body = latest_plan_text(planned_workbook)
    assert "by: phil\n" in body
    assert "inferred" not in body


def test_refusing_a_drop_keeps_the_range_and_keeps_approval_blocked(
    planned_workbook: Path, offline_model: None
) -> None:
    """Confirming and refusing are different outcomes, and both must be reachable.

    Refusing a drop means the range has to be kept, which leaves no stage consuming it -- so
    approval stays blocked until somebody says which stage should. Treating a refusal as an
    acknowledgement would be exactly the quiet hole the gate exists to prevent.
    """
    propose(planned_workbook)

    refused = runner.invoke(
        app,
        [
            "plan",
            "acknowledge",
            str(planned_workbook),
            "--range",
            "Calc!AK:AP",
            "--reject",
            "--note",
            "the desk still reads it",
        ],
    )
    blocked = runner.invoke(app, ["plan", "approve", str(planned_workbook)])

    assert refused.exit_code == 0, refused.output
    assert "blocker(s) still stand" in flattened(refused.output)
    assert blocked.exit_code == 1
    assert "must be kept" in flattened(blocked.output)


def test_acknowledge_names_the_drops_the_plan_actually_proposes(
    planned_workbook: Path, offline_model: None
) -> None:
    propose(planned_workbook)

    result = runner.invoke(
        app, ["plan", "acknowledge", str(planned_workbook), "--range", "Sheet1!ZZ:ZZ"]
    )

    assert result.exit_code == 1
    output = flattened(result.output)
    assert "no dropped range" in output
    assert "Calc!AK:AP" in output, "saying which ranges it does propose is the useful half"


@pytest.mark.parametrize("extra", [[], ["--range", "Calc!AK:AP", "--all"]])
def test_acknowledge_wants_exactly_one_of_range_and_all(
    planned_workbook: Path, offline_model: None, extra: list[str]
) -> None:
    propose(planned_workbook)

    result = runner.invoke(app, ["plan", "acknowledge", str(planned_workbook), *extra])

    assert result.exit_code == 1
    assert "exactly one" in flattened(result.output)


def test_acknowledge_will_not_refuse_every_drop_in_one_go(
    planned_workbook: Path, offline_model: None
) -> None:
    """`--all` confirms; refusing is per-range because each refusal raises its own question."""
    propose(planned_workbook)

    result = runner.invoke(app, ["plan", "acknowledge", str(planned_workbook), "--all", "--reject"])

    assert result.exit_code == 1
    assert "--reject" in flattened(result.output)


def test_reject_records_who_turned_it_down_and_why(
    planned_workbook: Path, offline_model: None
) -> None:
    propose(planned_workbook)

    result = runner.invoke(
        app,
        [
            "plan",
            "reject",
            str(planned_workbook),
            "--by",
            "phil",
            "--reason",
            "the override step is a judgement call, not a stage",
        ],
    )

    assert result.exit_code == 0, result.output
    body = latest_plan_text(planned_workbook)
    assert "state: rejected" in body
    assert "by: phil" in body
    assert "judgement call" in body, "the reason is the point of the verb"


def test_a_rejected_plan_can_never_then_be_approved(
    planned_workbook: Path, offline_model: None
) -> None:
    """REJECTED is the one terminal state. The way forward is a new plan, not an approval."""
    propose(planned_workbook)
    runner.invoke(app, ["plan", "reject", str(planned_workbook), "--reason", "wrong shape"])

    result = runner.invoke(app, ["plan", "approve", str(planned_workbook)])

    assert result.exit_code == 1
    output = flattened(result.output)
    assert "rejected" in output
    assert "kedge plan propose" in output


def test_reject_will_not_run_without_a_reason(planned_workbook: Path) -> None:
    result = runner.invoke(app, ["plan", "reject", str(planned_workbook)])

    assert result.exit_code != 0
    assert "--reason" in unstyled(result.output)


def test_request_changes_records_the_note_without_closing_the_plan_off(
    planned_workbook: Path, offline_model: None
) -> None:
    propose(planned_workbook)

    result = runner.invoke(
        app,
        [
            "plan",
            "request-changes",
            str(planned_workbook),
            "--by",
            "phil",
            "--note",
            "split the haircut stage in two",
        ],
    )

    assert result.exit_code == 0, result.output
    body = latest_plan_text(planned_workbook)
    assert "state: changes_requested" in body
    assert "by: phil" in body
    assert "split the haircut stage in two" in body


def test_history_lists_every_version_with_its_approval_state(
    planned_workbook: Path, offline_model: None
) -> None:
    """`PlanStore.history()` is the change record, and until now it had no surface at all."""
    propose(planned_workbook)
    runner.invoke(app, ["plan", "acknowledge", str(planned_workbook), "--all"])
    runner.invoke(app, ["plan", "approve", str(planned_workbook), "--by", "phil"])

    result = runner.invoke(app, ["plan", "history", str(planned_workbook)])

    assert result.exit_code == 0, result.output
    output = flattened(result.output)
    assert "draft" in output and "approved" in output
    assert output.index("v1") < output.index("v2"), "oldest first, so the newest is where you land"
    assert "phil" in output


def test_history_does_not_credit_a_human_edit_to_the_model(
    planned_workbook: Path, offline_model: None
) -> None:
    """`human (gpt-4o)` under a column headed *author* reads as "a human wrote this, with GPT-4o".

    Every review edit goes through `kedge.plan.review`, which sets `generated_by="human"` and
    leaves `llm_model` exactly where the model left it. The version the model did write says so on
    its own row.
    """
    propose(planned_workbook)
    runner.invoke(app, ["plan", "acknowledge", str(planned_workbook), "--all"])

    result = runner.invoke(app, ["plan", "history", str(planned_workbook)])

    assert result.exit_code == 0, result.output
    output = flattened(result.output)
    assert "human (" not in output
    assert "llm (" in output, "the version a model did write still names it"


def test_a_versions_author_names_the_model_only_where_one_wrote_it() -> None:
    from conftest import make_plan

    written = make_plan(generated_by="llm", llm_model="gpt-4o")
    edited = make_plan(generated_by="human", llm_model="gpt-4o")

    assert cli._plan_author(written) == "llm (gpt-4o)"
    assert cli._plan_author(edited) == "human"


def test_history_says_which_zone_it_is_printing_times_in(
    planned_workbook: Path, offline_model: None
) -> None:
    """The store writes UTC. Unlabelled, every reader takes the column for local time -- and a
    plan approved at 23:40 UTC then belongs to the wrong day for most of the world."""
    propose(planned_workbook)

    result = runner.invoke(app, ["plan", "history", str(planned_workbook)])

    assert result.exit_code == 0, result.output
    assert "created (UTC)" in flattened(result.output)


def test_a_timestamp_in_another_zone_is_still_printed_in_utc() -> None:
    tokyo = datetime(2026, 7, 24, 8, 40, tzinfo=timezone(timedelta(hours=9)))

    assert cli._stamp(tokyo) == "2026-07-23 23:40"


def test_an_approval_state_this_cli_has_no_colour_for_still_renders() -> None:
    """A fifth `ApprovalState` must not turn `kedge plan history` into a KeyError over a colour."""
    from kedge.plan.model import ApprovalState

    assert {state.value for state in ApprovalState} <= set(cli._APPROVAL_STYLE), (
        "every state kedge has today should be styled; the fallback is for the one it does not"
    )
    assert cli._approval_cell("quarantined") == "quarantined"


# ── plan: the record of a decision ───────────────────────────────────────────────────────────


def test_every_approval_decision_on_a_version_is_kept(
    planned_workbook: Path, offline_model: None
) -> None:
    """Three reviewers, three decisions, three files. Not one surviving line.

    Recording a decision over the version it names destroys every earlier decision on it. The
    dangerous step is the middle one: a plan a notebook may already have been scaffolded from
    would be un-approved with nothing left on disk saying it ever had been.
    """
    propose(planned_workbook)
    runner.invoke(app, ["plan", "acknowledge", str(planned_workbook), "--all"])

    alice = runner.invoke(
        app,
        [
            "plan",
            "approve",
            str(planned_workbook),
            "--by",
            "alice",
            "--note",
            "checked the haircut lookup",
        ],
    )
    bob = runner.invoke(
        app,
        [
            "plan",
            "request-changes",
            str(planned_workbook),
            "--by",
            "bob",
            "--note",
            "actually no",
            "--withdraw-approval",
        ],
    )
    carol = runner.invoke(app, ["plan", "approve", str(planned_workbook), "--by", "carol", "--yes"])

    assert [alice.exit_code, bob.exit_code, carol.exit_code] == [0, 0, 0], [
        result.output for result in (alice, bob, carol)
    ]
    saved = [path.read_text(encoding="utf-8") for path in plans_under(planned_workbook)]
    assert len(saved) == 5, "propose, acknowledge, and one file per decision"
    assert "state: approved" in saved[2] and "by: alice" in saved[2]
    assert "checked the haircut lookup" in saved[2], "alice's note is still there to read"
    assert "state: changes_requested" in saved[3] and "by: bob" in saved[3]
    assert "state: approved" in saved[4] and "by: carol" in saved[4]


def test_an_approved_plan_cannot_be_sent_back_by_accident(
    planned_workbook: Path, offline_model: None
) -> None:
    """The library refuses it; the command line has to say so as a message, not a traceback."""
    propose(planned_workbook)
    runner.invoke(app, ["plan", "acknowledge", str(planned_workbook), "--all"])
    runner.invoke(app, ["plan", "approve", str(planned_workbook), "--by", "alice"])

    result = runner.invoke(
        app, ["plan", "request-changes", str(planned_workbook), "--note", "second thoughts"]
    )

    assert result.exit_code == 1
    assert isinstance(result.exception, SystemExit)
    output = flattened(result.output)
    assert "--withdraw-approval" in output, "the flag that would allow it has to be named"
    assert "state: approved" in latest_plan_text(planned_workbook), "nothing was withdrawn"


def test_a_rejected_plan_cannot_be_moved_back_into_review(
    planned_workbook: Path, offline_model: None
) -> None:
    """REJECTED is terminal, and the refusal has to arrive as a message rather than a traceback.

    The route out of a rejection was the laundering hole: `changes_requested` approves cleanly, so
    reject, then request-changes, then approve put the whole of a turned-down decomposition into
    force. `--withdraw-approval` is deliberately *not* offered here -- it does not apply to a
    rejection, and naming it would send the user round a loop ending in this same message.
    """
    propose(planned_workbook)
    runner.invoke(app, ["plan", "reject", str(planned_workbook), "--reason", "wrong shape"])

    result = runner.invoke(
        app, ["plan", "request-changes", str(planned_workbook), "--note", "one more look"]
    )

    assert result.exit_code == 1
    assert isinstance(result.exception, SystemExit), "a refusal is a message, not a traceback"
    output = flattened(result.output)
    assert "rejected" in output
    assert "--withdraw-approval" not in output, "the flag is not the way out of a rejection"
    assert "state: rejected" in latest_plan_text(planned_workbook)


# ── plan: proposing over a plan already in force ─────────────────────────────────────────────


def a_different_draft() -> object:
    """The same plan with one stage's intent changed, so a diff of it has something to say."""
    draft = make_draft()
    return make_draft(
        stages=[
            stage.model_copy(update={"intent": "Read exposures from the hand-in, netted"})
            if stage.id == "load_handin"
            else stage
            for stage in draft.stages
        ]
    )


def reproposes(monkeypatch: pytest.MonkeyPatch, draft: object) -> None:
    """Point the next `plan propose` at a different decomposition."""
    from kedge.plan.propose import scripted_from_plan

    monkeypatch.setattr(
        "kedge.plan.completer_from_config", lambda _config: scripted_from_plan(draft)
    )


def approved_plan(workbook: Path, *, by: str = "alice") -> None:
    """Get one approved plan onto disk: propose, acknowledge the drop, approve."""
    propose(workbook)
    runner.invoke(app, ["plan", "acknowledge", str(workbook), "--all"])
    result = runner.invoke(app, ["plan", "approve", str(workbook), "--by", by, "--yes"])
    assert result.exit_code == 0, result.output


def test_proposing_over_an_approved_plan_shows_what_would_change(
    planned_workbook: Path, offline_model: None, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A second proposal is a whole replacement decomposition, and it must not arrive unseen.

    Refusing it is the chat tool's answer, and it works there because `amend_plan` is the thing to
    redirect to. There is no `amend` verb here, so refusing would leave hand-editing YAML as the
    only way forward, and would break the batch route PLAN 7 step 4 judges the corpus with. The
    diff is the answer instead -- `diff_plans` and `render_diff` were fully written and called
    from nowhere in `src/`.
    """
    approved_plan(planned_workbook)
    reproposes(monkeypatch, a_different_draft())

    result = runner.invoke(app, ["plan", "propose", str(planned_workbook)])

    assert result.exit_code == 0, result.output
    output = flattened(result.output)
    assert "plan v3 is approved and in force" in output
    assert "by alice" in output
    assert "Plan v3 -> v4" in output, "the diff names the version this would replace"
    assert "load_handin" in output and "intent" in output
    assert [path.name for path in plans_under(planned_workbook)][-1] == "plan-v004.yaml"


def test_the_json_form_of_a_proposal_says_what_it_would_replace(
    planned_workbook: Path, offline_model: None, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A script reading JSON gets the same warning the terminal does, or it gets none at all."""
    approved_plan(planned_workbook)
    reproposes(monkeypatch, a_different_draft())

    result = runner.invoke(app, ["plan", "propose", str(planned_workbook), "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.stdout)
    assert payload["replaces_approved_version"] == 3
    assert "load_handin" in payload["diff_from_approved"]


def test_a_dry_run_over_an_approved_plan_replaces_nothing_and_says_so(
    planned_workbook: Path, offline_model: None, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Nothing was written, so nothing is being replaced and there is no diff to draw."""
    approved_plan(planned_workbook)
    reproposes(monkeypatch, a_different_draft())

    result = runner.invoke(app, ["plan", "propose", str(planned_workbook), "--dry-run", "--json"])

    assert result.exit_code == 0, result.output
    payload = json.loads(result.stdout)
    assert payload["saved_to"] is None
    assert payload["replaces_approved_version"] is None
    assert payload["diff_from_approved"] is None


def test_approving_over_an_approved_plan_shows_the_diff_and_asks_first(
    planned_workbook: Path, offline_model: None, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The last cheap moment to notice a replaced decomposition is while somebody is standing there.

    Answering no records nothing at all: the plan in force stays the one that was reviewed.
    """
    approved_plan(planned_workbook)
    reproposes(monkeypatch, a_different_draft())
    propose(planned_workbook)
    runner.invoke(app, ["plan", "acknowledge", str(planned_workbook), "--all"])

    declined = runner.invoke(
        app, ["plan", "approve", str(planned_workbook), "--by", "mallory"], input="n\n"
    )

    assert declined.exit_code == 1
    output = flattened(declined.output)
    assert "plan v3 is already approved" in output
    assert "Plan v3 -> v5" in output
    assert "load_handin" in output, "what differs is shown, not just that something does"
    assert [path.name for path in plans_under(planned_workbook)] == [
        f"plan-v00{version}.yaml" for version in (1, 2, 3, 4, 5)
    ], "nothing was recorded"
    assert "state: draft" in latest_plan_text(planned_workbook)


def test_a_confirmed_replacement_is_approved_as_asked(
    planned_workbook: Path, offline_model: None, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Confirming is not a refusal in disguise: the newer decomposition does go into force."""
    approved_plan(planned_workbook)
    reproposes(monkeypatch, a_different_draft())
    propose(planned_workbook)
    runner.invoke(app, ["plan", "acknowledge", str(planned_workbook), "--all"])

    accepted = runner.invoke(
        app, ["plan", "approve", str(planned_workbook), "--by", "mallory"], input="y\n"
    )

    assert accepted.exit_code == 0, accepted.output
    body = latest_plan_text(planned_workbook)
    assert "state: approved" in body
    assert "by: mallory" in body


def test_a_script_can_approve_a_replacement_without_a_terminal(
    planned_workbook: Path, offline_model: None, monkeypatch: pytest.MonkeyPatch
) -> None:
    """`--yes` exists because a prompt with nothing on stdin is an abort, not a default."""
    approved_plan(planned_workbook)
    reproposes(monkeypatch, a_different_draft())
    propose(planned_workbook)
    runner.invoke(app, ["plan", "acknowledge", str(planned_workbook), "--all"])

    result = runner.invoke(app, ["plan", "approve", str(planned_workbook), "--by", "ci", "--yes"])

    assert result.exit_code == 0, result.output
    assert "state: approved" in latest_plan_text(planned_workbook)


def test_the_first_approval_of_all_is_not_interrupted_by_a_question(
    planned_workbook: Path, offline_model: None
) -> None:
    """Nothing is being replaced when nothing is approved yet, so nothing is asked.

    Without stdin a prompt would abort, so an unwanted one is not a papercut -- it is a command
    that stops working.
    """
    propose(planned_workbook)
    runner.invoke(app, ["plan", "acknowledge", str(planned_workbook), "--all"])

    result = runner.invoke(app, ["plan", "approve", str(planned_workbook), "--by", "phil"])

    assert result.exit_code == 0, result.output
    assert "already approved" not in flattened(result.output)


# ── plan: whose workbook is this ─────────────────────────────────────────────────────────────


def test_propose_refuses_an_analysis_taken_from_a_different_workbook(
    planned_workbook: Path, other_workbook: Path, tmp_path: Path, offline_model: None
) -> None:
    """A plan takes its whole identity from its analysis, and nothing downstream re-checks it.

    Accepted, this files a plan under alpha's project directory saying `workbook: beta.xlsx` with
    beta's digest -- so `kedge plan show alpha.xlsx` prints a plan for beta while the approval
    prints alpha, and `workbook_sha256`, documented as tying a plan to the exact file it was
    written for, is read nowhere in `src/`.
    """
    elsewhere = tmp_path / "beta-analysis.json"
    inspected = runner.invoke(app, ["inspect", str(other_workbook), "--out", str(elsewhere)])
    assert inspected.exit_code == 0, inspected.output

    result = runner.invoke(
        app, ["plan", "propose", str(planned_workbook), "--analysis", str(elsewhere)]
    )

    assert result.exit_code == 1
    assert isinstance(result.exception, SystemExit)
    output = flattened(result.output)
    assert "beta.xlsx" in output and "process.xlsx" in output
    assert "kedge inspect" in output, "the way out is to write the right analysis"
    assert plans_under(planned_workbook) == [], "nothing claiming to be beta was written"


def test_propose_warns_but_proceeds_on_an_analysis_of_a_workbook_that_has_since_changed(
    planned_workbook: Path, tmp_path: Path, offline_model: None
) -> None:
    """Same file, older facts. Re-planning after a workbook changes is a legitimate thing to do."""
    saved = tmp_path / "analysis.json"
    assert (
        runner.invoke(app, ["inspect", str(planned_workbook), "--out", str(saved)]).exit_code == 0
    )
    change_the_workbook(planned_workbook)

    result = runner.invoke(
        app, ["plan", "propose", str(planned_workbook), "--analysis", str(saved), "--dry-run"]
    )

    assert result.exit_code == 0, result.output
    output = flattened(result.output)
    assert "out of date" in output
    assert "STAGES" in output, "warned, not refused"


def test_approve_warns_when_the_workbook_has_moved_on_since_the_plan(
    planned_workbook: Path, offline_model: None
) -> None:
    """Approving in silence against a file that has changed is the part that is not legitimate."""
    propose(planned_workbook)
    runner.invoke(app, ["plan", "acknowledge", str(planned_workbook), "--all"])
    change_the_workbook(planned_workbook)

    result = runner.invoke(app, ["plan", "approve", str(planned_workbook), "--by", "phil"])

    assert result.exit_code == 0, result.output
    assert "has changed since plan v2" in flattened(result.output)
    assert "state: approved" in latest_plan_text(planned_workbook), "a warning, not a refusal"


def test_an_error_that_names_an_excel_range_prints_all_of_it(
    planned_workbook: Path, offline_model: None
) -> None:
    """rich reads square brackets as markup, and an external reference is exactly that shape.

    `[budget.xlsx]nope!A1` renders as `nope!A1` unescaped -- so the user is told to name a range
    whose name they have just been shown a truncated version of.
    """
    propose(planned_workbook)

    result = runner.invoke(
        app, ["plan", "acknowledge", str(planned_workbook), "--range", "[budget.xlsx]nope!A1"]
    )

    assert result.exit_code == 1
    assert "[budget.xlsx]nope!A1" in flattened(result.output)


@pytest.mark.parametrize(
    "argv",
    [
        ["show"],
        ["approve"],
        ["acknowledge", "--all"],
        ["reject", "--reason", "no"],
        ["request-changes", "--note", "no"],
        ["history"],
    ],
)
def test_a_review_verb_with_no_plan_saved_says_how_to_make_one(
    planned_workbook: Path, argv: list[str]
) -> None:
    result = runner.invoke(app, ["plan", argv[0], str(planned_workbook), *argv[1:]])

    assert result.exit_code == 1
    assert isinstance(result.exception, SystemExit), "a missing plan is a message, not a traceback"
    output = flattened(result.output)
    assert "no process plan saved" in output
    assert "kedge plan propose" in output


def test_a_hand_edited_plan_that_will_not_parse_names_the_file(
    planned_workbook: Path, offline_model: None
) -> None:
    """The store's YAML is explicitly a file users may edit, so a typo in it must be a message."""
    propose(planned_workbook)
    plans_under(planned_workbook)[0].write_text("stages: [\n", encoding="utf-8")

    result = runner.invoke(app, ["plan", "show", str(planned_workbook)])

    assert result.exit_code == 1
    assert isinstance(result.exception, SystemExit)
    assert "plan-v001.yaml" in flattened(result.output)


def test_every_review_verb_works_with_no_model_endpoint_at_all(
    planned_workbook: Path, offline_model: None, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Only `propose` needs an LLM (PLAN M2). Somebody approving a plan on a train must not.

    Proved rather than asserted. No API key exists in any of these tests (`no_real_keyring`), and
    the model seam is additionally replaced with something that raises, so a review verb that
    reached for a completer would fail here rather than quietly working on a developer's machine.
    """
    propose(planned_workbook)

    def _explode(*_args: object, **_kwargs: object) -> None:
        raise AssertionError("a review verb must never reach for a model endpoint")

    monkeypatch.setattr("kedge.plan.completer_from_config", _explode)
    monkeypatch.setattr("kedge.plan.propose.completer_from_config", _explode)
    monkeypatch.setattr("kedge.plan.propose.OpenAICompleter", _explode)
    monkeypatch.setattr("kedge.config.get_api_key", _explode)

    verbs = [
        ["show"],
        ["history"],
        ["acknowledge", "--all"],
        ["approve", "--by", "phil"],
        # Each of these runs against the plan the one before it approved, and taking an approval
        # back is a deliberate act in `kedge.plan.review` rather than something a verb does on the
        # way past.
        ["request-changes", "--note", "one more look", "--withdraw-approval"],
        ["reject", "--reason", "second thoughts"],
    ]
    results = [
        runner.invoke(app, ["plan", argv[0], str(planned_workbook), *argv[1:]]) for argv in verbs
    ]

    assert [result.exit_code for result in results] == [0] * len(verbs), [
        result.output for result in results
    ]


def test_propose_is_the_one_verb_that_does_need_the_endpoint(planned_workbook: Path) -> None:
    """No `offline_model` here. Without it the claim above would be vacuous.

    With no key in the keyring the proposal stops before it spends anything, and says which
    keyring entry to fill in rather than raising a `MissingApiKeyError` at the user.
    """
    result = runner.invoke(app, ["plan", "propose", str(planned_workbook)])

    assert result.exit_code == 1
    assert isinstance(result.exception, SystemExit)
    assert "keyring" in flattened(result.output)
    assert plans_under(planned_workbook) == []


# ── the bridge preflight ─────────────────────────────────────────────────────────────────────


def test_open_refuses_before_it_spawns_anything_when_the_bridge_has_moved(
    workbook: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """PLAN 6.1 mitigation 5, at the entry point it was written for.

    Without this the same mismatch surfaces as a TypeError from inside a tool call, after the
    analysis, the scaffold, a spawned marimo and several minutes of conversation.
    """
    pretend_bridge(monkeypatch, moved_bridge())
    reached: list[str] = []

    async def _capture(*_args, **_kwargs):
        reached.append("open sequence started")
        raise SystemExit(0)

    monkeypatch.setattr("kedge.server.hub.open_workbook", _capture)

    result = runner.invoke(app, ["open", str(workbook), "--no-browser"])

    assert result.exit_code == 1
    assert reached == [], "nothing may be spawned once the preflight has failed"
    output = flattened(result.output)
    assert "0.24.0" in output, "the error has to name the marimo that is actually installed"
    assert cli.MARIMO_PIN in output


def test_hub_refuses_before_it_starts_the_server_when_the_bridge_has_moved(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The hub spawns marimo too -- from the page rather than the terminal, but just as far in."""
    pretend_bridge(monkeypatch, moved_bridge())
    served: list[str] = []
    monkeypatch.setattr("kedge.server.app.run_server", lambda *_a, **_k: served.append("served"))

    result = runner.invoke(app, ["hub", "--no-browser"])

    assert result.exit_code == 1
    assert served == []
    assert "0.24.0" in flattened(result.output)


# ── watch ────────────────────────────────────────────────────────────────────────────────────


class _StubFolder:
    """A started watcher that never sees a file, and optionally takes a Ctrl-C while waiting."""

    def __init__(self, *, interrupt: bool = False) -> None:
        self._interrupt = interrupt
        self.stopped = False

    def wait(self, *, timeout_seconds: float | None = None) -> bool:
        del timeout_seconds
        if self._interrupt:
            raise KeyboardInterrupt
        return True

    def stop(self) -> None:
        self.stopped = True


def test_watch_says_which_setting_to_fill_in_when_no_folder_is_configured(workbook: Path) -> None:
    """``ingest.watch_dir`` is optional, so an unset one is a message rather than a traceback."""
    result = runner.invoke(app, ["watch", str(workbook)])

    assert result.exit_code == 1
    output = flattened(result.output)
    assert "ingest.watch_dir" in output
    assert "--dir" in output


def test_watch_names_the_folder_that_is_not_there(workbook: Path, tmp_path: Path) -> None:
    argv = ["watch", str(workbook), "--dir", str(tmp_path / "not-there"), "--once"]

    result = runner.invoke(app, argv)

    assert result.exit_code == 1
    assert "does not exist" in flattened(result.output)


def test_watch_once_receives_what_is_already_sitting_in_the_folder(
    workbook: Path, inbox: Path
) -> None:
    """The sweep with no filesystem watcher at all: what a scheduled task runs (PLAN 2.8)."""
    (inbox / "exposures.xlsx").write_bytes(b"a,b\n1,2\n")

    result = runner.invoke(
        app, ["watch", str(workbook), "--dir", str(inbox), "--once", "--settle", "0"]
    )

    assert result.exit_code == 0, result.output
    output = flattened(result.output)
    assert "1 hand-in(s) received" in output
    assert "watched" in output and "exposures.xlsx" in output, "the audit line is the point"
    copies = list(workbook.parent.rglob("*exposures.xlsx"))
    assert len(copies) == 1, "the hand-in must be copied into the managed store"
    assert "handins" in copies[0].parts


def test_watch_once_twice_over_does_not_receive_the_same_hand_in_again(
    workbook: Path, inbox: Path
) -> None:
    """Idempotence is what makes an hourly scheduled sweep safe to leave running."""
    (inbox / "exposures.xlsx").write_bytes(b"a,b\n1,2\n")
    argv = ["watch", str(workbook), "--dir", str(inbox), "--once", "--settle", "0"]

    first = runner.invoke(app, argv)
    second = runner.invoke(app, argv)

    assert "1 hand-in(s) received" in flattened(first.output)
    assert "0 hand-in(s) received" in flattened(second.output)


def test_watch_takes_the_folder_and_the_pattern_from_the_project_config(
    workbook: Path, inbox: Path
) -> None:
    """A watched folder is a property of the process, so it belongs in kedge.toml, not a flag."""
    (workbook.parent / "kedge.toml").write_text(
        f'[ingest]\nwatch_dir = "{inbox.as_posix()}"\nwatch_glob = "*.csv"\n', encoding="utf-8"
    )
    (inbox / "exposures.csv").write_bytes(b"a,b\n1,2\n")
    (inbox / "ignored.xlsx").write_bytes(b"a,b\n3,4\n")

    result = runner.invoke(app, ["watch", str(workbook), "--once", "--settle", "0"])

    assert result.exit_code == 0, result.output
    output = flattened(result.output)
    assert "exposures.csv" in output
    assert "ignored.xlsx" not in output


def test_a_relative_watch_dir_is_relative_to_the_workbook_not_to_the_shell(
    workbook: Path, inbox: Path, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A scheduled sweep starts in whatever directory the scheduler chose, and must still work.

    ``ingest.watch_dir = "inbox"`` is exactly what the "set it in a kedge.toml beside the
    workbook" message invites, and resolving it against the process working directory made the
    setting work from the project folder and quietly find nothing from anywhere else.
    """
    beside_the_workbook = workbook.parent / "inbox"
    beside_the_workbook.mkdir()
    (beside_the_workbook / "exposures.xlsx").write_bytes(b"a,b\n1,2\n")
    (workbook.parent / "kedge.toml").write_text('[ingest]\nwatch_dir = "inbox"\n', encoding="utf-8")
    # `inbox` is a decoy of the same name in the directory the command is run from, and empty.
    monkeypatch.chdir(tmp_path)

    result = runner.invoke(app, ["watch", str(workbook), "--once", "--settle", "0"])

    assert result.exit_code == 0, result.output
    assert "1 hand-in(s) received" in flattened(result.output)
    assert not list(inbox.iterdir()), "the folder beside the shell must not have been touched"


def test_a_relative_dir_flag_is_relative_to_the_shell(
    workbook: Path, inbox: Path, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The other half of the rule: a flag means what it means where it was typed."""
    (inbox / "exposures.xlsx").write_bytes(b"a,b\n1,2\n")
    monkeypatch.chdir(tmp_path)

    result = runner.invoke(
        app, ["watch", str(workbook), "--dir", "inbox", "--once", "--settle", "0"]
    )

    assert result.exit_code == 0, result.output
    assert "1 hand-in(s) received" in flattened(result.output)


def test_watch_hands_the_configured_store_and_pattern_to_the_watcher(
    workbook: Path, inbox: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The event-driven path is a thin wrapper, and it has to be given the same settings.

    Including the two booleans: with both defaulting to True in ``watch()`` itself, dropping
    them here would leave the config quietly ignored on this path and every test still green.
    """
    (workbook.parent / "kedge.toml").write_text(
        "[ingest]\ncopy_on_select = false\ndedupe_by_hash = false\n", encoding="utf-8"
    )
    captured: dict[str, object] = {}

    def _start(directory: Path, _on_handin: object, **kwargs: object) -> object:
        captured["directory"] = directory
        captured.update(kwargs)
        return _StubFolder()

    monkeypatch.setattr(WATCH_MODULE, "watch", _start)
    runner.invoke(app, ["watch", str(workbook), "--dir", str(inbox), "--contract", "exposures"])

    assert captured["directory"] == inbox
    assert captured["glob"] == "*.xlsx"
    assert captured["contract"] == "exposures"
    assert Path(str(captured["store_dir"])).name == "handins"
    assert captured["copy_on_select"] is False
    assert captured["dedupe"] is False


def test_watch_does_not_announce_a_folder_it_could_not_start_watching(
    workbook: Path, tmp_path: Path
) -> None:
    """The reassuring line comes after the watcher is up, so it is never contradicted below it."""
    result = runner.invoke(app, ["watch", str(workbook), "--dir", str(tmp_path / "not-there")])

    assert result.exit_code == 1
    output = flattened(result.output)
    assert "does not exist" in output
    assert "watching" not in output, "it never started, so it must not say that it did"


def test_watch_stops_the_watcher_when_the_user_interrupts(
    workbook: Path, inbox: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Ctrl-C is how a watch is meant to end, so it exits 0 and leaves no observer thread."""
    folder = _StubFolder(interrupt=True)
    monkeypatch.setattr(WATCH_MODULE, "watch", lambda *_a, **_k: folder)

    result = runner.invoke(app, ["watch", str(workbook), "--dir", str(inbox)])

    assert result.exit_code == 0
    assert folder.stopped, "the observer thread must not outlive the command"
    assert "0 hand-in(s) received" in flattened(result.output)


# ── the real command line ────────────────────────────────────────────────────────────────────
#
# Everything above goes through CliRunner, which hands click a list of arguments. The console
# script does not: click reads `sys.argv` itself, and on Windows it rewrites what it finds there
# before any kedge code sees it. That is a whole class of bug no CliRunner test can reach, so
# these three drive the real argv instead.


def test_a_glob_pattern_reaches_the_command_exactly_as_it_was_typed(
    workbook: Path, inbox: Path, capsys: pytest.CaptureFixture[str], monkeypatch: pytest.MonkeyPatch
) -> None:
    """Regression: click globbed `--glob` against the current directory and swallowed the result.

    ``kedge watch book.xlsx --glob "*.xlsx"`` is normally run from the folder the workbook is in
    -- which by definition contains a match -- so click's Windows argv expansion turned the
    pattern into ``process.xlsx``, the sweep found no file of that name in the inbox, and the
    command reported nothing received and exited 0.
    """
    (inbox / "exposures.xlsx").write_bytes(b"a,b\n1,2\n")
    assert list(Path.cwd().glob("*.xlsx")) == [workbook], "the cwd must hold the tempting match"
    monkeypatch.setattr(sys, "excepthook", sys.excepthook)  # typer replaces it; put it back after
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "kedge",
            "watch",
            str(workbook),
            "--dir",
            str(inbox),
            "--once",
            "--settle",
            "0",
            "--glob",
            "*.xlsx",
        ],
    )

    with pytest.raises(SystemExit) as exit_info:
        app()

    assert exit_info.value.code == 0
    assert "1 hand-in(s) received" in flattened(capsys.readouterr().out)
    stored = [path for path in workbook.parent.rglob("*exposures.xlsx") if "handins" in path.parts]
    assert len(stored) == 1, "the hand-in the pattern matched must be in the managed store"


def test_the_command_line_expands_a_home_directory_but_not_a_wildcard(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Turning click's expansion off must not cost the one part of it that was worth having.

    No Windows shell expands ``~`` for a native program, so kedge does it; both shells expand
    their own variables and neither expands wildcards, so kedge does neither.
    """
    monkeypatch.setattr(
        sys, "argv", ["kedge", "watch", "~/book.xlsx", "--glob", "*.xlsx", "--contract", "~$a.xlsx"]
    )

    expanded = cli._console_args()

    assert expanded[1] == str(Path.home() / "book.xlsx")
    assert expanded[3] == "*.xlsx"
    assert expanded[5] == "~$a.xlsx", "an Excel lock-file pattern is not a home directory"


def test_the_entry_point_uses_that_expansion_and_not_clicks(
    capsys: pytest.CaptureFixture[str], monkeypatch: pytest.MonkeyPatch
) -> None:
    """And the app really does read its argv through it, or `~` would stop working on Windows.

    Nothing in the home directory is touched: the file does not exist, and the point is the
    path kedge says it could not find.
    """
    monkeypatch.setattr(sys, "excepthook", sys.excepthook)
    monkeypatch.setattr(sys, "argv", ["kedge", "inspect", "~/no-such-workbook.xlsx"])

    with pytest.raises(SystemExit) as exit_info:
        app()

    assert exit_info.value.code == 1
    reported = flattened(capsys.readouterr().err)
    assert "no such workbook" in reported
    assert "~" not in reported, "the home directory was not expanded on the way in"


def test_reconcile_exit_code_folds_on_status_not_on_the_passed_list() -> None:
    """Regression: `report.passed` is the list of passing regions, not a verdict.

    A report with one passing and one failing region has a truthy `.passed`, so branching on it
    exited 0 on a failed reconciliation -- a silent green light on exactly the case the command
    exists to catch. `.status` folds the regions and is falsy for anything but a clean pass.
    """

    @dataclass
    class _Report:
        status: ReconciliationStatus
        passed: list[str]

    mixed = _Report(ReconciliationStatus.FAILED, ["one region did pass"])
    clean = _Report(ReconciliationStatus.PASSED, ["everything passed"])
    vacuous = _Report(ReconciliationStatus.NOT_RECONCILED, [])

    assert bool(mixed.passed), "the bug only bites when .passed is truthy but the report failed"
    assert cli._reconcile_exit_code(mixed) == 1
    assert cli._reconcile_exit_code(clean) == 0
    assert cli._reconcile_exit_code(vacuous) == 1
