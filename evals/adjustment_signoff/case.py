"""Grading the adjustment eval: what to feed the runbook, and what to check when it stops.

Every grader here is keyed to an id in ``expected.yaml``, and
``tests/unit/test_evals_harness.py`` asserts the two sets are identical in both directions. That
link is the whole reason a rubric written in prose is worth having: the prose says what is being
checked and why, the code says how, and neither can drift without the other going red.

**The hand-ins are derived from the workbook, never committed beside it.** ``Pre-Adjustment`` and
``Post-Adjustment`` are already the two grids a user would bring back, so the harness reads them
out and writes them to a temporary directory as CSV. Committing copies would mean two sources of
truth for the same 120 rows and one of them going stale.

**Deterministic graders compare at a tolerance, never with ``==``.** polars' vectorised execution
of the rounding chain lands a few parts in 1e11 away from the scalar path, which is far inside
the penny that matters and far outside float equality. A grader written with ``==`` would fail a
correct notebook, and the obvious fix -- loosening the expected value -- would be the wrong one.
:data:`PENNY` is the threshold, and it is half a penny because that is the smallest difference
anybody would act on.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
import logging
import re
import tempfile
from dataclasses import dataclass, field
from functools import cache
from pathlib import Path
from typing import TYPE_CHECKING, Any

from harness.model import ItemResult, Outcome
from openpyxl import load_workbook

if TYPE_CHECKING:
    from collections.abc import Callable, Mapping

    from harness.drive import NotebookRun

logger = logging.getLogger(__name__)

CASE_DIR = Path(__file__).resolve().parent
WORKBOOK = CASE_DIR / "q2_accrual_adjustment.xlsx"
REFERENCE_NOTEBOOK = CASE_DIR / "notebook.py"
RUBRIC = CASE_DIR / "expected.yaml"

DATA_HEADER_ROW = 18
"""Where the extract tabs' header sits. Rows 3-16 hold the query, which a result grid would not."""

DATA_ROWS = 120

PENNY = 0.005
"""Half a penny: the tolerance every money comparison here uses. See the module docstring."""

BLANK_TRADE = "ACC-00001"
"""The trade with no accrual. Excel predicts 0.00 from the blank; the warehouse keeps NULL."""

APOSTROPHE_TRADE = "ACC-00041"
"""``O'Brien & Partners``. The workbook's own concatenated statement for this row is invalid."""


@cache
def _sheet_names() -> frozenset[str]:
    """Every sheet in the workbook, case-folded, for checking that a citation points somewhere.

    Cached because it is asked once per grading run and opening the workbook is the expensive
    part of this module. Case-folded because a plan writes ``Sign-off`` or ``sign-off`` as it
    pleases and neither is wrong.
    """
    workbook = load_workbook(WORKBOOK, read_only=True)
    try:
        return frozenset(name.strip().casefold() for name in workbook.sheetnames)
    finally:
        workbook.close()


# =============================================================================
# WHAT THE RUNBOOK IS FED
# =============================================================================


def grid_csv(sheet_name: str) -> str:
    """One extract tab as the CSV a user would bring back from the query client.

    Only the data region: the query text above it is workbook furniture, not something a result
    grid carries. Dates are rendered ISO because that is what every client emits and what the
    paste path would receive.
    """
    workbook = load_workbook(WORKBOOK, data_only=True)
    sheet = workbook[sheet_name]
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\n")
    for row in sheet.iter_rows(
        min_row=DATA_HEADER_ROW, max_row=DATA_HEADER_ROW + DATA_ROWS, max_col=7
    ):
        writer.writerow([_cell_text(cell.value) for cell in row])
    return buffer.getvalue()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value)


def excel_style_paste(sheet_name: str) -> str:
    """One extract tab as Excel puts it on the clipboard: displayed text, tab-delimited.

    The difference from :func:`grid_csv` is the whole point. A file written by a query tool
    carries values; a copy out of a spreadsheet carries what the cell *looked like*, so a column
    formatted ``#,##0.00`` arrives as ``364,422.95`` and a date as ``30/06/2026``. That is the
    paste a user actually makes, and it is the one that broke.
    """
    rows = list(csv.reader(io.StringIO(grid_csv(sheet_name))))
    lines: list[str] = []
    for index, row in enumerate(rows):
        if index == 0:
            lines.append("\t".join(row))
            continue
        rendered = list(row)
        if rendered[4]:
            rendered[4] = f"{float(rendered[4]):,.2f}"
        if rendered[5]:
            rendered[5] = dt.date.fromisoformat(rendered[5]).strftime("%d/%m/%Y")
        lines.append("\t".join(rendered))
    return "\n".join(lines)


def write_handins(directory: Path) -> tuple[Path, Path]:
    """Write the two grids the runbook will ask for, and return their paths."""
    directory.mkdir(parents=True, exist_ok=True)
    pre = directory / "pre_adjustment.csv"
    post = directory / "post_adjustment.csv"
    pre.write_text(grid_csv("Pre-Adjustment"), encoding="utf-8")
    post.write_text(grid_csv("Post-Adjustment"), encoding="utf-8")
    return pre, post


def script_for(pre: Path, post: Path) -> dict[str, Any]:
    """The human's part, played the same way every run.

    Keyed by the variable each widget is assigned to. A notebook that names its widgets
    differently will not be driven by this -- which is a real limitation of grading a runbook
    headlessly, and is reported as unused inputs rather than silently scored as a stop.
    """
    return {
        "period_end": dt.date(2026, 6, 30),
        "ledger": "STATUTORY",
        "pre_adjustment_pick": (pre,),
        "approve_adjustment_decision": "approve",
        "approve_adjustment_note": (
            "Uplift agreed at the June finance committee; scope checked against the statement."
        ),
        "update_statement_ran": True,
        "update_statement_ran_note": "76 rows affected, ticket FC-2291.",
        "post_adjustment_pick": (post,),
    }


# =============================================================================
# CONTEXT
# =============================================================================


@dataclass(frozen=True)
class Context:
    """Everything a grader is allowed to look at."""

    run: NotebookRun
    facts: Mapping[str, Any]
    notebook: Path
    unused_inputs: tuple[str, ...] = ()
    plan: Any = None
    notes: list[str] = field(default_factory=list)

    @property
    def defs(self) -> dict[str, Any]:
        return self.run.definitions

    def need(self, *names: str) -> ItemResult | None:
        """A skip naming what is missing, or ``None`` when everything is there.

        A notebook that stopped early has not *failed* the items about cells it never reached --
        it failed the item about stopping. Marking the rest as failures would report one problem
        five times and bury it.
        """
        absent = [name for name in names if name not in self.defs]
        if not absent:
            return None
        return _skip(
            f"the notebook defines no {', '.join(absent)}. "
            + (
                f"It {self.run.summary_line()}."
                if not self.run.completed
                else "It ran to completion, so this is a naming difference rather than a stop."
            )
        )


def _pass(detail: str = "") -> ItemResult:
    return ItemResult(id="", outcome=Outcome.PASS, detail=detail)


def _fail(detail: str) -> ItemResult:
    return ItemResult(id="", outcome=Outcome.FAIL, detail=detail)


def _skip(detail: str) -> ItemResult:
    return ItemResult(id="", outcome=Outcome.SKIP, detail=detail)


def _close(actual: float, expected: float) -> bool:
    return abs(actual - expected) < PENNY


# =============================================================================
# TIER 1 - DETERMINISTIC
# =============================================================================


def ran_to_completion(ctx: Context) -> ItemResult:
    """Nothing else is meaningful if the runbook never reached the end."""
    if ctx.unused_inputs:
        return _fail(
            f"the script named widgets the notebook does not have: "
            f"{', '.join(ctx.unused_inputs)}. Either the notebook names them differently or a "
            f"step is missing; nothing below was driven as intended."
        )
    if not ctx.run.completed:
        return _fail(ctx.run.summary_line())
    return _pass(f"{len(ctx.run.cells_run)} cells")


def totals_to_the_penny(ctx: Context) -> ItemResult:
    missing = ctx.need("adjust_totals")
    if missing:
        return missing
    totals = ctx.defs["adjust_totals"]
    try:
        got = {name: float(totals[name][0]) for name in ("before", "uplift", "after")}
    except (KeyError, IndexError, TypeError) as error:
        return _fail(f"adjust_totals is not a frame with before/uplift/after columns: {error}")

    expected = {
        "before": float(ctx.facts["accrual_total_before"]),
        "uplift": float(ctx.facts["uplift_total"]),
        "after": float(ctx.facts["accrual_total_after"]),
    }
    wrong = {
        name: (got[name], value) for name, value in expected.items() if not _close(got[name], value)
    }
    if wrong:
        lines = [
            f"{name}: expected {want:,.2f}, got {have:,.2f} (out by {have - want:+,.2f})"
            for name, (have, want) in wrong.items()
        ]
        return _fail(
            "\n".join(lines)
            + "\nFour of the 76 rows need Excel's 15-significant-digit collapse before "
            "rounding. A bare .round(2) is a penny out on each."
        )
    return _pass(f"before {got['before']:,.2f}, uplift {got['uplift']:,.2f}")


def reconciles_against_the_workbook(ctx: Context) -> ItemResult:
    """Every region the notebook claims must pass, and none may fail.

    Deliberately not "the overall status is PASSED". ``Adjustment!G`` holds the statements the
    workbook built by concatenation, one of which is invalid SQL; a good conversion does not
    reproduce it, so the report's overall status is honestly ``not_reconciled``. Requiring PASSED
    would fail the right answer, and quietly dropping the unmapped region from the denominator
    would be the false pass this project exists not to produce.
    """
    missing = ctx.need("reconciliation")
    if missing:
        return missing
    check = ctx.defs["reconciliation"]
    # The notebook's value is a TranslationCheck now: reconciliation is an acceptance test on
    # the conversion, not a live comparison on every run. `.report` is the live one when there
    # was one, which at conversion time -- which is what this grader drives -- there always is.
    report = getattr(check, "report", check)
    if report is None:
        return _fail(
            "nothing was compared against the workbook on the very first run of a freshly "
            "converted notebook. That run *is* the acceptance test; there is nothing to cite yet."
        )
    regions = list(getattr(report, "regions", []))
    if not regions:
        return _fail("the reconciliation report has no regions, so nothing was compared")

    failed = [region for region in regions if region.status.value == "failed"]
    passed = [region for region in regions if region.status.value == "passed"]
    if failed:
        lines = [f"{region.reference}: {region.status.value}" for region in failed]
        return _fail("regions disagreeing with the workbook's cached values:\n" + "\n".join(lines))
    if not passed:
        return _fail(
            "no region was compared at all. The workbook carries a full cache, so this means "
            "the notebook mapped nothing -- not that there was no baseline."
        )
    unreconciled = [region for region in regions if region.status.value == "not_reconciled"]
    detail = f"{len(passed)} region(s) passed, none failed"
    if unreconciled:
        detail += "; unmapped: " + ", ".join(region.reference for region in unreconciled)
    return _pass(detail)


def generated_sql_is_valid(ctx: Context) -> ItemResult:
    """Execute the statements the notebook generated and check they do what it predicted.

    Graded by running them, not by reading them. The row this catches is ``ACC-00041``, whose
    entity is ``O'Brien & Partners``: the workbook's own statement for it does not parse, and a
    conversion that reproduced the concatenation faithfully would reproduce that. A string
    comparison would have to know what correct escaping looks like; an engine already does.

    Compared against the notebook's own prediction rather than against ``Post-Adjustment``,
    because those are two different claims. This one is "the SQL says what the notebook worked
    out"; whether the warehouse then agreed is
    :func:`verification_finds_exactly_one_break`.
    """
    try:
        import duckdb
    except ImportError:
        return _skip(
            "duckdb is not installed, so the generated SQL was not executed. "
            "Install it with `uv sync --group evals`."
        )

    missing = ctx.need("update_statement", "adjust")
    if missing:
        return missing

    statements = str(ctx.defs["update_statement"])
    if not statements.strip():
        return _fail("the notebook generated no statements")
    predicted = ctx.defs["adjust"]
    predicted = predicted.collect() if hasattr(predicted, "collect") else predicted

    connection = duckdb.connect()
    try:
        connection.execute(
            "CREATE TABLE fin_accruals (trade_id VARCHAR, accrual_gbp DOUBLE, "
            "adjustment_note VARCHAR)"
        )
        rows = [
            (row["trade_id"], row["accrual_gbp_before"], None)
            for row in predicted.iter_rows(named=True)
        ]
        connection.executemany("INSERT INTO fin_accruals VALUES (?, ?, ?)", rows)
        # The notebook writes `fin.accruals`; DuckDB has no such schema here, and creating one
        # would be pretending the eval knows the warehouse's layout. Rewriting the qualifier is
        # the smallest possible change and touches nothing the grader is about.
        executed = 0
        for statement in statements.splitlines():
            text = statement.strip()
            if not text or text.startswith("--"):
                continue
            connection.execute(text.replace("fin.accruals", "fin_accruals"))
            executed += 1
        result = connection.execute(
            "SELECT trade_id, accrual_gbp, adjustment_note FROM fin_accruals ORDER BY trade_id"
        ).fetchall()
    except Exception as error:
        return _fail(
            f"the generated SQL did not execute: {type(error).__name__}: {error}\n"
            f"The row to look at is {APOSTROPHE_TRADE} -- its entity name contains an "
            f"apostrophe, and `&` in Excel escapes nothing."
        )
    finally:
        connection.close()

    applied = {trade: (value, note) for trade, value, note in result}
    wrong = [
        f"{row['trade_id']}: notebook computed {row['accrual_gbp_after']:,.2f}, "
        f"its own SQL produced {applied.get(row['trade_id'], (None, None))[0]}"
        for row in predicted.iter_rows(named=True)
        if not _close(
            applied.get(row["trade_id"], (float("nan"), None))[0] or 0.0, row["accrual_gbp_after"]
        )
    ]
    if wrong:
        return _fail(
            f"{len(wrong)} statement(s) did not apply the value the notebook computed:\n"
            + "\n".join(wrong[:5])
        )
    if APOSTROPHE_TRADE not in applied:
        return _fail(
            f"no statement reached {APOSTROPHE_TRADE}, the row whose entity name carries an "
            f"apostrophe. That is the row this check exists for."
        )
    note = applied[APOSTROPHE_TRADE][1] or ""
    if "O'Brien" not in note:
        return _fail(
            f"{APOSTROPHE_TRADE} was updated but its note came back as {note!r}, so the "
            f"apostrophe was mangled rather than escaped."
        )
    return _pass(f"{executed} statements executed, including {APOSTROPHE_TRADE}")


def null_is_not_zero(ctx: Context) -> ItemResult:
    """The workbook says 0.00 for the blank accrual; the notebook must say so too.

    Reproducing Excel is the whole job, and this is the row where doing so feels wrong. Excel
    read the empty cell as zero inside ROUND, so the workbook's cache holds 0.00 and any
    notebook that "improves" on it by predicting NULL stops reconciling. The place to notice the
    problem is the verification against the re-extract, not here.
    """
    missing = ctx.need("adjust")
    if missing:
        return missing
    frame = ctx.defs["adjust"]
    frame = frame.collect() if hasattr(frame, "collect") else frame
    row = next(
        (item for item in frame.iter_rows(named=True) if item["trade_id"] == BLANK_TRADE), None
    )
    if row is None:
        return _fail(
            f"{BLANK_TRADE} is not in the adjusted set. It has no accrual, but the UPDATE's "
            f"WHERE clause still reaches it -- dropping it silently loses a row the statement "
            f"touches."
        )
    after = row.get("accrual_gbp_after")
    if after is None:
        return _fail(
            f"{BLANK_TRADE} came out as null. Excel reads the blank cell as zero inside ROUND, "
            f"so the workbook's cached value is 0.00 and this will not reconcile. Route the "
            f"column through .xl.empty_as_zero() rather than letting polars propagate the null."
        )
    if not _close(float(after), 0.0):
        return _fail(f"{BLANK_TRADE} came out as {after}, where the workbook holds 0.00")
    return _pass(f"{BLANK_TRADE} predicted at 0.00, matching the workbook")


def verification_finds_exactly_one_break(ctx: Context) -> ItemResult:
    """The dangerous direction is a pass. This is the row a total-level check never finds."""
    missing = ctx.need("verification")
    if missing:
        return missing
    report = ctx.defs["verification"]
    if getattr(report, "ok", False):
        return _fail(
            "the verification passed. It should not: the warehouse left "
            f"{BLANK_TRADE} NULL where the notebook predicted 0.00. A pass here is a signed "
            "claim that production holds what the notebook says it holds."
        )
    if report.status.value == "not_reconciled":
        return _fail(
            f"nothing was verified: {report.reason}. Not verifying is not the same as "
            f"verifying and finding a break, and only one of them is the right answer here."
        )
    breaks = list(getattr(report, "breaks", ()))
    if report.break_count != 1:
        return _fail(
            f"expected exactly 1 break, got {report.break_count} "
            f"({report.missing_count} missing, {report.unexpected_count} unexpected). "
            f"First few: {[item.key for item in breaks[:3]]}"
        )
    if breaks and breaks[0].key != (BLANK_TRADE,):
        return _fail(f"the break is on {breaks[0].key}, expected ({BLANK_TRADE!r},)")
    return _pass(f"one break, on {BLANK_TRADE}, as designed")


def cancelled_rows_excluded(ctx: Context) -> ItemResult:
    missing = ctx.need("adjust")
    if missing:
        return missing
    frame = ctx.defs["adjust"]
    frame = frame.collect() if hasattr(frame, "collect") else frame
    expected = int(ctx.facts["rows_the_update_reaches"])
    if frame.height != expected:
        return _fail(
            f"{frame.height} rows adjusted, expected {expected}. Four trades inside the "
            f"in-scope entities are CANCELLED, and the UPDATE's WHERE clause excludes them."
        )
    return _pass(f"{expected} rows")


def no_pandas(ctx: Context) -> ItemResult:
    """The validation gate over every cell of the notebook as written."""
    from kedge.agent.validate import validate_cell
    from kedge.notebook.codegen import read_notebook

    document = read_notebook(ctx.notebook)
    offenders: list[str] = []
    for index, cell in enumerate(document.cells):
        report = validate_cell(cell.code, cell=cell.name or str(index))
        offenders.extend(
            f"{cell.name or index}: {violation.message}"
            for violation in report.violations
            if not violation.message.startswith("no name registry")
        )
    if offenders:
        return _fail("\n".join(offenders[:8]))
    return _pass(f"{len(document.cells)} cells")


# =============================================================================
# THE RUNBOOK AS A USER MEETS IT
# =============================================================================
#
# Everything above grades the *result* of a completed run. These grade the experience of
# working through one, and every one of them exists because a real user hit the problem in app
# mode while the deterministic tier was reporting 38/38. A conversion can be numerically perfect
# and still be a page that stalls with no explanation.

STEPS_BEFORE_ANY_INPUT = ("extract_query", "pre_adjustment_input")
"""What a user should see on opening a fresh notebook: the first statement, and where to reply."""

MUST_BE_HIDDEN_AT_THE_START = (
    "post_adjustment_input",
    "post_adjustment_paste",
    "update_statement",
    "update_statement_ran",
)
"""What must not be on screen before it is that step's turn.

``post_adjustment_input`` is the one this list was written for. A selector that constructs
widgets and reads nothing has no dataflow edge, so marimo cannot gate it -- it renders from the
moment the notebook opens, and invites a re-extract taken before the update has been run. There
is no way to detect that afterwards, which is what makes it worth a grader of its own.
"""


def _drive(ctx: Context, inputs: dict[str, Any], root: Path) -> Any:
    """Drive the notebook again, in a workspace of this grader's own."""
    from harness.drive import run_notebook, workspace_overrides

    return run_notebook(
        ctx.notebook,
        inputs=inputs,
        overrides=workspace_overrides(root, WORKBOOK),
    )


def progressive_disclosure(ctx: Context) -> ItemResult:
    """Nothing is on screen before its turn, and the first step is."""
    from harness.drive import visible_cells, workspace_overrides

    with tempfile.TemporaryDirectory(prefix="kedge-eval-visible-") as workspace:
        root = Path(workspace)
        seen = set(
            visible_cells(
                ctx.notebook,
                inputs={},
                overrides=workspace_overrides(root, WORKBOOK),
            )
        )

    early = [name for name in MUST_BE_HIDDEN_AT_THE_START if name in seen]
    if early:
        return _fail(
            f"visible on a fresh open, before their step: {', '.join(early)}.\n"
            f"A cell that only constructs widgets reads nothing, so marimo has no edge to hide "
            f"it on. Have it read the token of the step before it."
        )
    missing = [name for name in STEPS_BEFORE_ANY_INPUT if name not in seen]
    if missing:
        return _fail(
            f"not visible on a fresh open: {', '.join(missing)}. A runbook that shows the user "
            f"nothing to do has stalled, whatever it is waiting for."
        )
    return _pass(f"{len(seen)} cell(s) shown; the rest wait their turn")


def a_blocked_step_says_which_step_it_is(ctx: Context) -> ItemResult:
    """In app mode a stopped cell is the only thing left on the page. It has to explain itself.

    Everything below a stop disappears, with nothing to distinguish "waiting for you", "broken"
    and "finished". The stop message is the entire user interface at that moment, so it has to
    carry the position as well as the instruction.
    """
    with tempfile.TemporaryDirectory(prefix="kedge-eval-blocked-") as workspace:
        run = _drive(ctx, {"period_end": dt.date(2026, 6, 30)}, Path(workspace))

    if run.completed:
        return _fail("the notebook ran to completion with no hand-in supplied, so nothing blocked")
    message = run.stopped_because or ""
    if not message.strip():
        return _fail(
            f"the notebook stopped at {run.stopped_at!r} with an empty message. In app mode "
            f"that is a blank page."
        )
    if not re.search(r"step\s+\d+\s+of\s+\d+", message, re.IGNORECASE):
        return _fail(
            f"the blocking message does not say which step it is:\n  {message[:200]}\n"
            f"Without it the user cannot tell a page that is waiting from one that has died."
        )
    if not re.search(r"\b(run|drop|select|paste|type|choose|tick|enter|fill)\b", message, re.I):
        return _fail(
            f"the blocking message explains itself but never says what to do:\n"
            f"  {message[:200]}\n"
            f"A user reading that is stuck. Lead with the instruction; put the reason after it."
        )
    return _pass(message.split(".")[0])


def a_declared_exception_does_not_read_as_a_defect(ctx: Context) -> ItemResult:
    """A region the notebook deliberately does not reproduce must not send anyone bug-hunting.

    ``Adjustment!G`` is the workbook's own concatenated UPDATE statements, one of which is
    invalid SQL. A good conversion renders them properly instead, so there is nothing to
    reconcile -- and the default message for an unmapped region ("the notebook produced no
    values for this region, check that the cell ran and that the variable names match") is then
    actively misleading.

    It is still not a pass. Nothing here can make an unchecked region one, and this insists on
    that as firmly as it insists on the wording.
    """
    missing = ctx.need("reconciliation")
    if missing:
        return missing
    check = ctx.defs["reconciliation"]
    report = getattr(check, "report", check)
    if report is None:
        return _skip("this run cited a recorded acceptance, so there are no live regions")
    region = next(
        (item for item in getattr(report, "regions", []) if "G17" in (item.reference or "")),
        None,
    )
    if region is None:
        return _skip("the report has no region for Adjustment!G, so there is nothing to judge")
    if region.status.value == "passed":
        return _fail(
            "Adjustment!G is reported as passed. Nothing was compared against it -- an "
            "unchecked region is never a pass, however deliberate the omission."
        )
    if getattr(region.reason, "value", None) != "not_reproduced":
        return _fail(
            f"Adjustment!G is reported as {getattr(region.reason, 'value', None)!r}, which "
            f"reads as a cell that failed to run. Declare it with `not_reproduced=` so the "
            f"panel says it was a decision, and why."
        )
    if "kedge.sql" not in region.detail and "invalid" not in region.detail.lower():
        return _fail("the region is declared, but the reason given explains nothing")
    headline = report.headline()
    if headline.startswith("NOT RECONCILED"):
        return _fail(
            f"the panel still leads with NOT RECONCILED:\n  {headline[:160]}\n"
            f"Every run of a correct notebook would show that for ever, and a signal that is "
            f"permanently amber is one people stop reading."
        )
    return _pass(headline.split(".")[0][:90])


def the_period_is_an_input(ctx: Context) -> ItemResult:
    """The extract SQL carries the period the user chose, not one typed into the query.

    In the workbook the period end was typed into a cell *and then typed again* into the WHERE
    clause of both statements -- right three times and wrong the fourth. A conversion that
    hardcodes it has kept the defect.
    """
    missing = ctx.need("extract_query")
    if missing:
        return missing

    chosen = dt.date(2026, 3, 31)
    with tempfile.TemporaryDirectory(prefix="kedge-eval-period-") as workspace:
        root = Path(workspace)
        pre, post = write_handins(root / "handins-in")
        script = {**script_for(pre, post), "period_end": chosen}
        run = _drive(ctx, script, root)

    statement = str(run.definitions.get("extract_query", ""))
    if not statement:
        return _skip("the notebook defines no extract_query under a changed period")
    if chosen.isoformat() not in statement:
        return _fail(
            f"the period was set to {chosen}, and the extract query does not mention it:\n"
            f"  {statement.splitlines()[-1][:160]}\n"
            f"The period has to reach the statement, or the user edits SQL by hand every quarter."
        )
    if str(ctx.facts["period_end"]) in statement:
        return _fail(
            f"the extract query still carries {ctx.facts['period_end']} after the period was "
            f"changed to {chosen}. It is hardcoded."
        )
    return _pass(f"the query follows the input ({chosen})")


def a_paste_out_of_excel_works(ctx: Context) -> ItemResult:
    """The exact thing that broke in app mode: Excel copies displayed text, not values.

    A column formatted ``#,##0.00`` arrives as ``364,422.95`` -- comma included -- and polars
    types it ``String``. The failure then surfaces several operations later as ``arithmetic on
    dtypes str and dyn float is not allowed``, from inside a query plan, in a cell app mode
    renders as nothing at all.

    Graded through the paste box rather than a file, because a file written by a query tool is
    the case that already worked.
    """
    grid = excel_style_paste("Pre-Adjustment")
    with tempfile.TemporaryDirectory(prefix="kedge-eval-paste-") as workspace:
        root = Path(workspace)
        _pre, post = write_handins(root / "handins-in")
        script = {
            **script_for(post, post),
            "pre_adjustment_pick": (),
            "pre_adjustment_paste": grid,
        }
        run = _drive(ctx, script, root)

    if not run.completed:
        detail = run.summary_line()
        if run.error is not None:
            detail = f"{detail}\n  {type(run.error).__name__}: {str(run.error)[:220]}"
        return _fail(
            f"a grid pasted the way Excel puts it on the clipboard did not get through: "
            f"{detail}\nRead the hand-in through kedge.ingest.read_data, which types text "
            f"columns and reports that it did."
        )
    totals = run.definitions.get("adjust_totals")
    if totals is None:
        return _fail("the paste was accepted but nothing was computed from it")
    uplift = float(totals["uplift"][0])
    if not _close(uplift, float(ctx.facts["uplift_total"])):
        return _fail(
            f"the pasted grid produced an uplift of {uplift:,.2f}, not "
            f"{float(ctx.facts['uplift_total']):,.2f}. The thousands separators were probably "
            f"parsed as something other than a number."
        )
    return _pass(f"tab-delimited, thousands separators, uplift {uplift:,.2f}")


def the_run_resumes_after_the_kernel_dies(ctx: Context) -> ItemResult:
    """Close the notebook halfway through, come back, and carry on.

    marimo's state lives in the kernel: reopen and every paste box is empty and every dropdown
    is unset. Everything the user already did has to come back off disk, or a two-day process is
    one nobody can leave.

    Driven twice against the same run directory. The second pass supplies *nothing* -- no paste,
    no file, no decision -- and must still reach the end.
    """
    with tempfile.TemporaryDirectory(prefix="kedge-eval-resume-") as workspace:
        root = Path(workspace)
        pre, post = write_handins(root / "handins-in")
        first = _drive(ctx, script_for(pre, post), root)
        if not first.completed:
            return _skip(
                f"the first pass did not complete, so resuming cannot be graded: {first.summary_line()}"
            )
        # A new kernel: no widget values at all, only what is on disk.
        second = _drive(ctx, {"period_end": dt.date(2026, 6, 30), "ledger": "STATUTORY"}, root)

    if not second.completed:
        return _fail(
            f"reopening the notebook did not pick the run up: {second.summary_line()}\n"
            f"The hand-ins are in the managed store and the decisions are in the run file; "
            f"read them rather than asking again."
        )
    totals = second.definitions.get("adjust_totals")
    if totals is None or not _close(float(totals["uplift"][0]), float(ctx.facts["uplift_total"])):
        return _fail("the resumed run reached the end but not with the same numbers")
    return _pass("resumed with no inputs re-supplied, same totals")


def starting_fresh_keeps_the_old_run(ctx: Context) -> ItemResult:
    """Starting again must not erase what was signed off last time.

    A run record is evidence. A tool where "start again" quietly deletes it is one nobody should
    put a control through.
    """
    import kedge.runs

    with tempfile.TemporaryDirectory(prefix="kedge-eval-fresh-") as workspace:
        root = Path(workspace)
        pre, post = write_handins(root / "handins-in")
        first = _drive(ctx, script_for(pre, post), root)
        if not first.completed:
            return _skip("the first pass did not complete, so starting fresh cannot be graded")
        original = str(first.definitions.get("KEDGE_RUN_ID", ""))

        second = _drive(
            ctx,
            {**script_for(pre, post), "kedge_run_mode": "start a new run"},
            root,
        )
        store = kedge.runs.RunStore(root / "runs")
        ids = store.run_ids()

    restarted = str(second.definitions.get("KEDGE_RUN_ID", ""))
    if restarted == original:
        return _fail(
            "choosing 'start a new run' carried on with the same run. Starting fresh has to "
            "start something."
        )
    if original not in ids:
        return _fail(
            f"the previous run {original} is no longer on disk. Starting fresh must write a new "
            f"record beside the old one, never over it."
        )
    return _pass(f"{len(ids)} runs kept: {', '.join(ids)}")


def reconciliation_does_not_rot_on_a_later_period(ctx: Context) -> ItemResult:
    """Run the notebook for the next quarter. Reconciliation must not go red.

    This is the defect a user found by asking a simple question of the panel: why is a notebook
    I will run every month still comparing itself to an Excel file from Q2? Reconciliation asks
    whether the *translation* is faithful -- a question about the conversion, answered once,
    against the data the spreadsheet holds. It is not a check on this month's numbers, and
    running it as one reports FAILED on every run after the first, on runs where nothing is
    wrong. Within a few months it is pointing at a spreadsheet nobody has opened since the
    process changed.

    Driven twice against one workspace: once as the conversion (which is the acceptance test),
    then again on different data for a different period.
    """
    with tempfile.TemporaryDirectory(prefix="kedge-eval-rot-") as workspace:
        root = Path(workspace)
        pre, post = write_handins(root / "handins-in")
        first = _drive(ctx, script_for(pre, post), root)
        if not first.completed:
            return _skip(f"the conversion run did not complete: {first.summary_line()}")
        accepted = first.definitions.get("reconciliation")
        if accepted is None or not getattr(accepted, "translation_accepted", False):
            return _fail(
                "the first run did not accept the translation, so there is nothing for a later "
                "run to cite. That first run *is* the acceptance test."
            )
        # Next quarter: same process, different data.
        later = _drive(ctx, {**script_for(post, post), "period_end": dt.date(2026, 9, 30)}, root)

    if not later.completed:
        return _fail(f"the later run did not complete: {later.summary_line()}")
    check = later.definitions.get("reconciliation")
    if check is None:
        return _fail("the later run produced no translation check at all")

    status = getattr(getattr(check, "status", None), "value", None)
    if status == "failed":
        return _fail(
            "reconciliation reported FAILED on a run for a different period. The workbook is not "
            "that run's baseline -- the numbers are supposed to differ. Record the acceptance "
            "once and cite it, rather than comparing every run against a spreadsheet that does "
            "not describe its data."
        )
    if getattr(check, "applies", True):
        return _fail(
            "the later run compared itself against the workbook again. Nothing about a Q2 "
            "spreadsheet is a baseline for a Q3 run."
        )
    if not getattr(check, "translation_accepted", False):
        return _fail("the later run cites no acceptance, so it claims nothing about anything")
    cited = check.summary_line()
    if "reconciled against" not in cited:
        return _fail(f"the citation does not say what was reconciled or when:\n  {cited[:160]}")
    if not getattr(check, "watching_this_run", ()):
        return _fail(
            "the panel cites an old acceptance and does not say what is checking today's "
            "numbers, which reads as 'nothing is being checked'."
        )
    return _pass(cited[:100])


def the_notebook_says_why_this_process_exists(ctx: Context) -> ItemResult:
    """A conversion that keeps the numbers and drops the reasons has lost the important half.

    The workbook opens with a Sign-off tab: Purpose, Background, Scope, Known issues -- prose
    somebody wrote precisely so the next person would not have to guess. The analyser recovers
    every word of it, with the cells it came from. A notebook that reproduces the arithmetic
    perfectly and says nothing about what it is for is one that, eight months on, nobody can
    safely change.

    Graded on the notebook's *rendered* output rather than on a variable, because the question
    is whether a reader is told, not whether a field is populated somewhere.
    """
    missing = ctx.need("kedge_briefing")
    if missing:
        return missing
    briefing = "\n".join(ctx.run.panels[:4])
    if not briefing.strip():
        return _fail("the notebook renders no briefing at all")

    wanted = {
        "what it is for": ("uplift", "accrual"),
        "why it exists": ("reforecast", "rate card", "committee"),
        "what to watch": ("cancelled", "no accrual", "null"),
    }
    absent = [
        label
        for label, needles in wanted.items()
        if not any(needle.lower() in briefing.lower() for needle in needles)
    ]
    if absent:
        return _fail(
            f"the briefing does not cover: {', '.join(absent)}. All of it is in the workbook's "
            f"Sign-off tab and reaches the plan through the analyser's notes."
        )
    if "Sign-off!" not in briefing:
        return _fail(
            "the briefing states what the process is for but never says where that came from. "
            "Invented background in a finance notebook is worse than none -- it is confident, "
            "plausible, and unattributable. Cite the sheet and cells."
        )
    return _pass("purpose, background and known issues, cited to the Sign-off tab")


def unanswered_questions_are_not_buried(ctx: Context) -> ItemResult:
    """A caveat nobody opens is a caveat nobody has.

    The plan raised two questions the conversion could not settle -- three entities or four, and
    what a blank accrual should mean. Both qualify every number the notebook produces, every
    time it runs. Tucking them inside a collapsed section, or into a YAML file beside the
    notebook, is the same as not raising them.
    """
    if ctx.plan is None:
        return _skip("no plan supplied, so there are no open questions to look for")
    unanswered = [item for item in ctx.plan.open_questions if not item.answer]
    if not unanswered:
        return _skip("the plan has no unanswered questions")

    missing = ctx.need("kedge_briefing")
    if missing:
        return missing
    shown = "\n".join(ctx.run.panels[:4]).lower()
    unmentioned = [
        item.question
        for item in unanswered
        if not any(word in shown for word in item.question.lower().split()[:6])
    ]
    if unmentioned:
        return _fail(
            f"{len(unmentioned)} unanswered question(s) from the plan are nowhere in what the "
            f"notebook shows. They qualify every number below them:\n  - "
            + "\n  - ".join(question[:110] for question in unmentioned)
        )
    return _pass(f"{len(unanswered)} unanswered question(s) surfaced")


DETERMINISTIC: dict[str, Callable[[Context], ItemResult]] = {
    "ran_to_completion": ran_to_completion,
    "totals_to_the_penny": totals_to_the_penny,
    "reconciles_against_the_workbook": reconciles_against_the_workbook,
    "generated_sql_is_valid": generated_sql_is_valid,
    "null_is_not_zero": null_is_not_zero,
    "verification_finds_exactly_one_break": verification_finds_exactly_one_break,
    "cancelled_rows_excluded": cancelled_rows_excluded,
    "no_pandas": no_pandas,
    "progressive_disclosure": progressive_disclosure,
    "a_blocked_step_says_which_step_it_is": a_blocked_step_says_which_step_it_is,
    "the_period_is_an_input": the_period_is_an_input,
    "a_paste_out_of_excel_works": a_paste_out_of_excel_works,
    "the_run_resumes_after_the_kernel_dies": the_run_resumes_after_the_kernel_dies,
    "starting_fresh_keeps_the_old_run": starting_fresh_keeps_the_old_run,
    "the_notebook_says_why_this_process_exists": the_notebook_says_why_this_process_exists,
    "unanswered_questions_are_not_buried": unanswered_questions_are_not_buried,
    "reconciliation_does_not_rot_on_a_later_period": (
        reconciliation_does_not_rot_on_a_later_period
    ),
    "a_declared_exception_does_not_read_as_a_defect": (
        a_declared_exception_does_not_read_as_a_defect
    ),
}


# =============================================================================
# TIER 2 - STRUCTURAL
# =============================================================================
#
# Every one of these grades a ProcessPlan. Without one they all skip -- which is the honest
# outcome, and why the report always prints its denominator.


def _no_plan() -> ItemResult:
    return _skip("no plan supplied; pass --plan to grade the structural tier")


def hands_over_rather_than_pretends(ctx: Context) -> ItemResult:
    if ctx.plan is None:
        return _no_plan()
    handoffs = [stage for stage in ctx.plan.stages if stage.is_handoff]
    if len(handoffs) < 2:
        return _fail(
            f"{len(handoffs)} handoff stage(s). The process has two statements a person has to "
            f"run -- the extract and the update -- and neither is something kedge can execute."
        )
    return _pass(", ".join(stage.id for stage in handoffs))


def _declared_handins(plan: Any) -> dict[str, str]:
    """Stage id to the label of the hand-in it declares, for every stage that declares one.

    Mirrors :func:`kedge.notebook.scaffold._named_handin`'s *reading* of a plan -- a
    ``{origin: handin, ref: ...}`` source -- and nothing about whether the scaffolder acts on it.
    That second question is :func:`_emitted_handins`, and the whole point of
    :func:`takes_two_handins` is that the two answers can differ.
    """
    from kedge.plan.model import SourceOrigin

    declared: dict[str, str] = {}
    for stage in plan.stages:
        for source in stage.sources:
            if source.origin is SourceOrigin.HANDIN and source.ref and stage.id not in declared:
                declared[stage.id] = source.ref
    return declared


def _scaffold(plan: Any) -> tuple[list[Any], str]:
    """The cells kedge would build from this plan, or why it could not build any.

    Every item that grades what the *notebook* will do runs the real scaffolder rather than
    re-deriving its rules, because re-deriving them is how this tier came to be green on a plan
    whose second hand-in had nowhere to arrive: ``build_cells`` returns early for a ``checkpoint``
    stage before it ever looks for a hand-in source. A predicate copied out of the scaffolder is a
    second copy of a rule, and a second copy rots.

    Returns:
        The cells, and ``""``; or an empty list and why there are none. A plan that will not
        scaffold at all is a finding of its own and belongs to whichever item is about that, so it
        comes back as a reason rather than as a traceback or as a silent zero.
    """
    from kedge.notebook.scaffold import build_cells

    try:
        return build_cells(plan, allow_unapproved=True), ""
    except Exception as error:
        # Broad on purpose: *every* way a plan can fail to scaffold is somebody else's finding.
        # Turning one of them into a red hand-in item would report the wrong defect, and letting
        # it propagate would take the whole tier down with it.
        logger.warning("the plan would not scaffold, so nothing about its cells can be counted")
        return [], f"{type(error).__name__}: {error}"


def _emitted_handins(cells: list[Any]) -> tuple[set[str], bool]:
    """The stage ids the scaffolder emits hand-in cells for, and whether the head one is emitted."""
    from kedge.notebook.scaffold import HEAD_CELL_NAMES

    staged = {cell.stage_id for cell in cells if cell.role == "handin" and cell.stage_id}
    head = any(cell.role == "handin" and cell.name in HEAD_CELL_NAMES for cell in cells)
    return staged, head


def takes_two_handins(ctx: Context) -> ItemResult:
    """Graded on the cells the scaffolder emits, never on the sources the plan declares.

    The weak version of this item asked only that *some* stage declare a hand-in with a ``ref``,
    and a real model-written plan passed it while the notebook it scaffolded had exactly one place
    to put a grid. Both its load steps carried a hand-in; one of them was typed ``checkpoint``, and
    ``build_cells`` returns early for a checkpoint before ``_named_handin`` is consulted. An item
    that is green on a plan whose re-extract has nowhere to arrive is worse than no item.
    """
    if ctx.plan is None:
        return _no_plan()
    declared = _declared_handins(ctx.plan)
    if not declared:
        return _fail(
            "no stage declares a hand-in of its own. The re-extract cannot be the notebook's "
            "head hand-in: it does not exist when the notebook is opened."
        )

    cells, unscaffoldable = _scaffold(ctx.plan)
    if unscaffoldable:
        return _skip(
            f"the plan declares hand-ins on {', '.join(sorted(declared))} but would not "
            f"scaffold, so what the notebook asks for could not be counted -- {unscaffoldable}"
        )
    staged, head = _emitted_handins(cells)

    lost = sorted(set(declared) - staged)
    if lost:
        detail = ", ".join(f"{stage} ({declared[stage]!r})" for stage in lost)
        return _fail(
            f"the plan declares a hand-in on {detail} and the scaffolder emits no cell for it, "
            f"so that grid has nowhere to arrive. `build_cells` returns early for a `checkpoint` "
            f"stage before it looks for a hand-in source: the source is read, shown on the "
            f"approval card, and then dropped. Declaring the input is not the same as the "
            f"notebook asking for it."
        )
    # Grids, not cells. Two stages naming the *same* ref scaffold two selectors and ask for one
    # thing twice, which is one grid short of a process that extracts, updates, and re-extracts.
    refs = {declared[stage].strip().casefold() for stage in staged if declared.get(stage)}
    grids = len(refs) + (1 if head else 0)
    if grids < 2:
        duplicated = len(staged) + (1 if head else 0) > grids
        because = (
            " -- two stages declare the same hand-in, so the notebook asks for one grid twice"
            if duplicated
            else ""
        )
        return _fail(
            f"the notebook will ask for {grids} distinct grid, and this process brings back two: "
            f"the extract, then the re-extract that proves the update took{because}. Emitted "
            f"for: {', '.join(sorted(staged)) or 'no stage'}."
        )

    # Mid-process-ness. A hand-in on a stage that waits for nothing is another opening input, and
    # the re-extract is precisely the one that cannot be: it does not exist when the notebook is
    # opened, so its cell has to sit below the cell that told the user how to produce it.
    mid_process = sorted(
        stage.id for stage in ctx.plan.stages if stage.id in staged and stage.depends_on
    )
    if not mid_process:
        return _fail(
            f"every hand-in the notebook asks for is an opening input: none of "
            f"{', '.join(sorted(staged))} waits on an earlier stage. The re-extract arrives "
            f"mid-process by definition -- it does not exist until the update has been run."
        )

    named = ", ".join(f"{stage} ({declared.get(stage, 'unnamed')})" for stage in sorted(staged))
    return _pass(
        f"{grids} distinct grid(s) from {len(staged)} staged hand-in(s)"
        + (" plus the notebook's own head hand-in" if head else "")
        + f"; arriving mid-process: {', '.join(mid_process)}. {named}"
    )


def generates_the_update_from_the_frame(ctx: Context) -> ItemResult:
    if ctx.plan is None:
        return _no_plan()
    generated = [
        stage
        for stage in ctx.plan.stages
        if stage.handoff is not None and stage.handoff.is_generated
    ]
    if not generated:
        return _fail(
            "no hand-off is generated from an upstream frame. Adjustment!G is one UPDATE per "
            "row built by concatenation; the equivalent is a template rendered over the "
            "computed frame."
        )
    return _pass(", ".join(stage.id for stage in generated))


def does_not_drop_the_sql_column(ctx: Context) -> ItemResult:
    """The sharpest item in the rubric: a plausible, well-reasoned, catastrophic edit."""
    if ctx.plan is None:
        return _no_plan()
    dropped = [
        item.range
        for item in ctx.plan.dropped
        if "G" in item.range.upper() and "ADJUSTMENT" in item.range.upper()
    ]
    if dropped:
        return _fail(
            f"the plan drops {', '.join(dropped)}. The analyser reports that region as dead "
            f"because nothing in the workbook reads it -- its consumer is a person with a "
            f"clipboard. Dropping it deletes the step that changes the data."
        )
    return _pass("Adjustment!G is kept")


def raises_the_memo_discrepancy(ctx: Context) -> ItemResult:
    if ctx.plan is None:
        return _no_plan()
    haystack = " ".join(
        [
            *(question.question for question in ctx.plan.open_questions),
            *(assumption for stage in ctx.plan.stages for assumption in stage.assumptions),
            *(stage.notes or "" for stage in ctx.plan.stages),
            *(
                stage.checkpoint.question
                for stage in ctx.plan.stages
                if stage.checkpoint is not None
            ),
        ]
    ).lower()
    if "three" in haystack or "four" in haystack or "sign-off" in haystack:
        return _pass("the disagreement is named somewhere a reviewer will see it")
    return _fail(
        "nothing in the plan mentions that the Sign-off tab says three entities where the "
        "scope and the statement both say four. Adopting either number silently is the failure."
    )


def _upstream_closure(plan: Any, start: str) -> set[str]:
    """Every stage ``start`` runs after, transitively, following ``depends_on``.

    Bounded by the visited set, so a plan whose graph has a cycle in it -- which the schema does
    not forbid -- terminates rather than hanging a sweep.
    """
    by_id = {stage.id: stage for stage in plan.stages}
    seen: set[str] = set()
    queue = list(by_id[start].depends_on) if start in by_id else []
    while queue:
        current = queue.pop()
        if current in seen:
            continue
        seen.add(current)
        upstream = by_id.get(current)
        if upstream is not None:
            queue.extend(upstream.depends_on)
    return seen


def _handoff_text(handoff: Any) -> str:
    """The text a hand-off actually hands over, fixed or templated."""
    return (handoff.statement or handoff.template or "") if handoff is not None else ""


def _writes(handoff: Any) -> bool:
    """Whether this hand-off's text would change data, read off the text and nothing else.

    Deliberately not :attr:`~kedge.plan.model.Handoff.statement_writes`, and the two differences
    are both defects this tier was found to have.

    It does not consult ``mutates``. ``statement_writes`` does not either, but
    :attr:`~kedge.plan.model.Handoff.needs_confirmation` -- which this used to call -- is
    ``mutates or statement_writes``, so a plan that declared the *read-only* extract query
    ``mutates: true`` was marked down for wanting a checkpoint in front of a ``SELECT``. Erring
    towards a tick-box is a shape ``expected.yaml``'s sibling item explicitly accepts and the
    propose prompt nudges towards; failing a plan for caution, in a message asserting the
    ``SELECT`` writes, is telling a reviewer something false.

    And it does not consult ``medium``. ``statement_writes`` returns ``False`` for anything that
    is not ``sql`` without reading a character of it, which is right in the product -- ``text`` is
    a filename to request or a colleague to ask, and no verb at the front of it means anything.
    Here it is a one-field escape: ``UPDATE fin.accruals SET ...`` under ``medium: text`` with
    ``mutates: false`` passed every item in this tier while the approval card rendered "changes
    nothing" over a production write. The medium is another claim; the text is the fact.
    """
    from kedge.sql import changes_data

    return changes_data(_handoff_text(handoff))


def _mutating_handoffs(plan: Any) -> list[Any]:
    """Hand-off stages whose own text would change data. See :func:`_writes` for what "own" costs.

    ``effective_handoff()`` is not used: it synthesises a ``-- TODO(kedge)`` placeholder for a
    stage that declares no hand-off at all, and a placeholder is not a statement anybody was
    handed. A ``kind: handoff`` stage with ``handoff: null`` is a plan that named the shape and
    not the substance, and it must not be able to satisfy an item about statements by supplying
    none -- see :func:`mutates_agrees_with_the_statement`, which is the item that names it.
    """
    return [stage for stage in plan.stages if stage.handoff is not None and _writes(stage.handoff)]


def has_a_checkpoint_before_the_update(ctx: Context) -> ItemResult:
    """Position, not presence -- and reachability, not a single edge.

    Two ways the presence version of this was wrong, and a real model-written plan exercised
    both. A checkpoint sitting *after* the update satisfies "the plan has a checkpoint", and a
    checkpoint gating the read-only extract satisfies "some hand-off names a checkpoint in
    ``depends_on``" while the UPDATE has nothing in front of it. Neither is a recorded decision
    before a production write. So the question is asked of the mutating hand-off specifically,
    and answered over the whole upstream closure rather than one edge: a checkpoint two stages
    up is still a decision the user made before being handed the statement, and refusing it
    would mark a correct plan down for its decomposition.

    Which hand-off is "the update" comes from :func:`_writes` -- ``kedge.sql.changes_data`` over
    the statement -- and from nothing else. It used to come from
    :attr:`~kedge.plan.model.Handoff.needs_confirmation`, which is ``mutates or statement_writes``,
    so a plan declaring the read-only extract query ``mutates: true`` was failed here for wanting
    a checkpoint in front of a ``SELECT``, in a message asserting the ``SELECT`` writes.
    """
    if ctx.plan is None:
        return _no_plan()
    checkpoints = {stage.id for stage in ctx.plan.stages if stage.is_checkpoint}
    if not checkpoints:
        return _fail("the plan has no checkpoint at all")

    mutating = _mutating_handoffs(ctx.plan)
    if not mutating:
        return _fail(
            "no hand-off carries a statement that writes, so there is nothing for a checkpoint "
            "to gate. The process applies an UPDATE to fin.accruals; a plan that never hands it "
            "over has lost the step that changes the data. Judged by kedge.sql.changes_data over "
            "the statement itself, so a plan cannot escape this item by under-declaring `mutates` "
            "-- nor be marked down for over-declaring it, which is the safe direction."
        )

    upstream = {stage.id: _upstream_closure(ctx.plan, stage.id) for stage in mutating}
    ungated = [stage for stage in mutating if not checkpoints & upstream[stage.id]]
    if ungated:
        return _fail(
            f"{', '.join(stage.id for stage in ungated)} hands over a statement that writes with "
            f"no checkpoint anywhere upstream of it. The plan's checkpoints are "
            f"{', '.join(sorted(checkpoints))}, and being in the plan is not being in front of "
            f"the UPDATE: a decision recorded afterwards is a decision about something that has "
            f"already happened. What does run first: "
            f"{', '.join(sorted(upstream[ungated[0].id])) or 'nothing at all'}."
        )
    return _pass(
        ", ".join(
            f"{stage.id} after {', '.join(sorted(checkpoints & upstream[stage.id]))}"
            for stage in mutating
        )
    )


def the_re_extract_waits_for_the_update(ctx: Context) -> ItemResult:
    """The one defect in this rubric's history that cannot be detected after the fact.

    Every other structural item asks *whether* the notebook does something. This asks **where**,
    and it exists because a plan that gets every one of them right can still scaffold a notebook
    that puts the re-extract box on screen the moment it opens. One edge does it: point
    ``post_adjustment`` at ``pre_adjustment`` instead of at ``update_statement`` and the selector
    is emitted seven cells above the UPDATE with no gate token in it, so marimo has no dataflow
    edge to hide it on. Graded against the committed reference bodies, that plan scored exactly
    what the correct one scored.

    What it costs is not a mark. A re-extract taken *before* the statement ran looks exactly like
    one taken after, and there is no way to tell afterwards -- which is why
    :attr:`~kedge.plan.model.Handoff.needs_confirmation` was changed to err towards a tick-box in
    the first place. A runbook that invites that is worse than one that stops, because it produces
    a verification that passes.

    Two things are asked of the notebook, and the second is the one a reader would forget. At
    least one hand-in must be **emitted below** the mutating hand-off's cells; and it must
    **read that hand-off's confirmation token**, because reading the token is the only thing that
    creates the edge. A cell that constructs widgets and references no upstream name has no edges
    at all, and marimo renders it immediately however far down the file it sits.
    """
    if ctx.plan is None:
        return _no_plan()
    mutating = _mutating_handoffs(ctx.plan)
    if not mutating:
        return _skip(
            "the plan hands over no statement that writes, so there is no update for a "
            "re-extract to wait for. `hands_over_rather_than_pretends` is the item about that."
        )

    cells, unscaffoldable = _scaffold(ctx.plan)
    if unscaffoldable:
        return _skip(
            f"the plan would not scaffold, so cell order could not be read -- {unscaffoldable}"
        )

    numbered = list(enumerate(cells))
    selectors = [
        (index, cell)
        for index, cell in numbered
        if cell.role == "handin" and cell.stage_id and cell.name.endswith("_input")
    ]
    last_cell_of = {
        stage.id: max((index for index, cell in numbered if cell.stage_id == stage.id), default=-1)
        for stage in mutating
    }
    if not selectors:
        return _skip(
            "the notebook asks for no hand-in of its own, so there is no re-extract to place. "
            "`takes_two_handins` is the item about that."
        )

    for stage in mutating:
        last = last_cell_of[stage.id]
        gates = {
            cell.name
            for cell in cells
            if cell.stage_id == stage.id and cell.name.endswith("_confirmed")
        }
        below = [(index, cell) for index, cell in selectors if index > last]
        if not below:
            where = ", ".join(f"{cell.name} at {index}" for index, cell in selectors)
            return _fail(
                f"every hand-in the notebook asks for is emitted above {stage.id}, whose "
                f"statement writes: {where}, and {stage.id}'s last cell is at {last}. The "
                f"re-extract box is therefore on screen before the UPDATE it is meant to follow, "
                f"and a grid pasted into it looks exactly like one taken afterwards. Nothing can "
                f"detect that later -- the verification passes either way."
            )
        if not gates:
            return _fail(
                f"{stage.id} writes and emits no confirmation cell, so there is no token for a "
                f"later hand-in to gate on. Nothing downstream can know the statement was run."
            )
        gated = [(index, cell) for index, cell in below if any(gate in cell.code for gate in gates)]
        if not gated:
            names = ", ".join(cell.name for _, cell in below)
            return _fail(
                f"{names} is emitted below {stage.id} and reads none of its confirmation tokens "
                f"({', '.join(sorted(gates))}), so marimo has no dataflow edge to hide it on. A "
                f"cell that builds `mo.ui` elements and references no upstream name renders from "
                f"the moment the notebook opens, wherever in the file it sits -- position in the "
                f"file is not the gate, reading the token is."
            )

    placed = ", ".join(
        f"{cell.name} at {index} (below {stage.id} at {last_cell_of[stage.id]})"
        for stage in mutating
        for index, cell in selectors
        if index > last_cell_of[stage.id]
    )
    return _pass(f"gated below {', '.join(stage.id for stage in mutating)}: {placed}")


def mutates_agrees_with_the_statement(ctx: Context) -> ItemResult:
    """``mutates`` is a claim; the statement is the fact, and the plan is where they must agree.

    The notebook now errs the safe way on its own -- :attr:`~kedge.plan.model.Handoff.
    needs_confirmation` is ``mutates or statement_writes``, so an ``UPDATE`` declared read-only
    still gets a tick-box -- which is exactly why this belongs in the rubric rather than being
    left to the scaffolder to compensate for. A contradiction the product silently survives is
    one nobody has to fix, and the approval card still renders the flag: a reviewer reading
    "changes nothing" over an ``UPDATE fin.accruals`` is being told something false at the moment
    they are deciding.
    """
    if ctx.plan is None:
        return _no_plan()
    claiming = [stage for stage in ctx.plan.stages if stage.is_handoff or stage.handoff is not None]
    if not claiming:
        return _skip(
            "the plan hands nothing over, so there is no statement for `mutates` to agree or "
            "disagree with. `hands_over_rather_than_pretends` is the item about that."
        )

    # A hand-off with no text of its own scaffolds to `effective_handoff()`'s `-- TODO(kedge)`
    # placeholder, which `changes_data` reads as read-only -- so a stage that supplies no
    # statement used to score full marks here for having nothing to contradict. `mutates` on a
    # hand-off with no statement is a claim about text that is not there.
    empty = [stage for stage in claiming if not _handoff_text(stage.handoff)]
    if empty:
        return _fail(
            f"{', '.join(stage.id for stage in empty)} hands something over and supplies no "
            f"statement, so `mutates` is a claim about text that is not there. The scaffolder "
            f"emits its `-- TODO(kedge)` placeholder in place of the statement, which reads as "
            f"read-only whatever the step was meant to do."
        )

    contradicting = [
        stage for stage in claiming if _writes(stage.handoff) and not stage.handoff.mutates
    ]
    if contradicting:
        mediums = ", ".join(
            f"{stage.id} (medium: {stage.handoff.medium.value})" for stage in contradicting
        )
        return _fail(
            f"{mediums} declares `mutates: false` over a statement kedge.sql.changes_data reads "
            f"as a write. The flag is what the approval card shows a reviewer, and it is telling "
            f"them the opposite of what the text does. Note the medium is not a defence: "
            f"`Handoff.statement_writes` stops reading at anything but `sql`, so retyping an "
            f"UPDATE as `text` silences the product's own check and changes nothing about what "
            f"the user is being asked to run."
        )

    writes = [stage.id for stage in claiming if _writes(stage.handoff)]
    over_declared = [
        stage.id for stage in claiming if stage.handoff.mutates and not _writes(stage.handoff)
    ]
    detail = f"{len(claiming)} hand-off(s) agree with their own statements"
    detail += f"; writing: {', '.join(writes)}" if writes else "; none of them writes"
    if over_declared:
        detail += (
            f"; {', '.join(over_declared)} declares `mutates: true` over a statement that only "
            f"reads, which costs a tick-box and is the safe direction"
        )
    return _pass(detail)


def the_briefing_survives_the_workbook(ctx: Context) -> ItemResult:
    """The one part of a workbook nobody can reconstruct, and the plan is where it is kept or lost.

    ``Sign-off`` carries Purpose, Background, Scope and Known issues, and the analyser extracts
    all eight notes with the sheet and cells each came from. A plan that leaves ``briefing`` null
    scaffolds a notebook whose first cell tells its reader that *"the workbook this was converted
    from carried no description of what the process is for"* -- which is false, confidently false,
    and in the one register this project exists to protect.

    The asymmetry is what makes this worth a rubric item. ``Briefing`` refuses prose with no
    citations, so an *invented* briefing cannot be written down; nothing anywhere notices one that
    never arrived. Sources are graded as well as prose because a briefing with neither is the
    honest answer for some workbooks and is the wrong answer for this one.
    """
    if ctx.plan is None:
        return _no_plan()
    briefing = getattr(ctx.plan, "briefing", None)
    if briefing is None or briefing.is_empty:
        return _fail(
            "the plan carries no briefing. The workbook's Sign-off tab holds Purpose, "
            "Background, Scope and Known issues, and the analyser recovers every one of them "
            "with the cells it came from -- so the notebook this scaffolds opens by telling its "
            "reader the workbook explained nothing, which is not true."
        )
    if not briefing.sources:
        return _fail(
            "the briefing states a purpose or a background and cites nothing. Invented "
            "background in a finance notebook is worse than none: it is confident, plausible, "
            "and the next reader cannot tell it from the real thing."
        )
    if not (briefing.purpose and briefing.background):
        missing = ", ".join(
            name for name in ("purpose", "background") if not getattr(briefing, name)
        )
        return _fail(
            f"the briefing has no {missing}. The background is the half nobody can reconstruct "
            f"from the code -- why a flat 4.5% uplift, why statutory ledger only -- and it is "
            f"written out in full on the Sign-off tab."
        )
    # A citation has to point somewhere. `Briefing` only asks that `sources` be non-empty, so
    # `sources=["nowhere in particular"]` satisfies the schema and satisfies a grader that counts
    # them -- which is the same asymmetry one level down: an unattributable briefing is refused,
    # an unfalsifiable attribution is not. At least one source must name a sheet this workbook
    # actually has. The rest may be anything: a procedure document or a person who said so is a
    # perfectly good source and is not checkable from here.
    sheets = _sheet_names()
    located = [
        source
        for source in briefing.sources
        if source.split("!")[0].strip().casefold() in sheets and "!" in source
    ]
    if not located:
        return _fail(
            f"none of the briefing's {len(briefing.sources)} source(s) names a sheet this "
            f"workbook has ({', '.join(sorted(sheets))}): {'; '.join(briefing.sources)}. A "
            f"citation the next reader cannot follow is not attribution -- it is the appearance "
            f"of it, and this workbook's Purpose, Background, Scope and Known issues are all "
            f"sitting on one tab with cell references the analyser already recovered."
        )
    return _pass(
        f"purpose, background and {len(briefing.watch_for)} thing(s) to watch for, "
        f"{len(located)} of {len(briefing.sources)} source(s) locatable in the workbook: "
        f"{', '.join(located[:3])}" + (" and more" if len(located) > 3 else "")
    )


def does_not_trust_the_impact_summary(ctx: Context) -> ItemResult:
    """Graded on the notebook, not the plan: the figures are either used or they are not."""
    missing = ctx.need("signoff")
    if missing:
        return missing
    stale_movement = float(ctx.facts["stale_movement"])
    stale_rows = int(ctx.facts["stale_rows_adjusted"])
    signoff = ctx.defs["signoff"]
    if not isinstance(signoff, dict):
        return _skip("the notebook's `signoff` is not a mapping, so its figures cannot be read")
    if _close(float(signoff.get("movement", 0.0)), stale_movement):
        return _fail(
            f"the notebook reports a movement of {stale_movement:,.2f}, which is the Sign-off "
            f"tab's own figure. It is stale -- it predates two trades joining the scope."
        )
    if int(signoff.get("rows_adjusted", 0)) == stale_rows:
        return _fail(
            f"the notebook reports {stale_rows} rows adjusted, which is the memo's number. "
            f"The statement reaches {ctx.facts['rows_the_update_reaches']}."
        )
    return _pass("the impact figures are recomputed, not carried across")


def consults_the_knowledge_pack(ctx: Context) -> ItemResult:
    return _skip(
        "no knowledge pack describes fin.accruals; context/databases/example.yaml is a "
        "different schema. Add one to grade this."
    )


STRUCTURAL: dict[str, Callable[[Context], ItemResult]] = {
    "hands_over_rather_than_pretends": hands_over_rather_than_pretends,
    "takes_two_handins": takes_two_handins,
    "generates_the_update_from_the_frame": generates_the_update_from_the_frame,
    "does_not_drop_the_sql_column": does_not_drop_the_sql_column,
    "raises_the_memo_discrepancy": raises_the_memo_discrepancy,
    "does_not_trust_the_impact_summary": does_not_trust_the_impact_summary,
    "has_a_checkpoint_before_the_update": has_a_checkpoint_before_the_update,
    "the_re_extract_waits_for_the_update": the_re_extract_waits_for_the_update,
    "mutates_agrees_with_the_statement": mutates_agrees_with_the_statement,
    "the_briefing_survives_the_workbook": the_briefing_survives_the_workbook,
    "consults_the_knowledge_pack": consults_the_knowledge_pack,
}
