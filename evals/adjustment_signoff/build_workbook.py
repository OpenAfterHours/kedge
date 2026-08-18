"""Build the adjustment-and-sign-off eval workbook.

The workbook this produces is not a calculation. It is the *record of a process*, which is what
most of the spreadsheets kedge exists to convert actually are: run an extract, work out an
adjustment, run an update, re-extract to prove it took, and write a memo somebody signs. Four
tabs, in that order, and the whole point of the eval is whether kedge turns them back into the
runbook they came from -- one that hands the user each statement in turn, takes the results
back, does the arithmetic in between where it can be checked, and refuses to claim the update
worked until it has seen evidence that it did.

Run it::

    uv run python evals/adjustment_signoff/build_workbook.py
    uv run python evals/adjustment_signoff/build_workbook.py --verify-with-excel  # Windows + Excel

**This is not a test fixture and must not become one.** ``tests/fixtures`` declares an expected
*analysis* and is asserted cell by cell against a hand-written manifest; adding a ninth workbook
there would fail ``test_exactly_one_fixture_offers_a_reconciliation_baseline`` on the first run.
An eval declares an expected *outcome*. What it shares with the corpus is plumbing -- the
archive rewriting, the deterministic timestamps, the cached-value injection -- and that is
imported rather than copied, because a second copy of ``inject_cached_values`` is a second
chance to get Excel's rounding wrong.

What the workbook is built to discriminate, and where each one lives:

======================================  ==================================================
Discrimination                          Where it is planted
======================================  ==================================================
Reads the extract SQL as a step to      ``Pre-Adjustment`` rows 3-16, and the same query in
hand over, not as prose                 ``xl/connections.xml`` as a real ODBC connection
Recognises a CONCATENATE-built SQL      ``Adjustment!F`` -- one UPDATE per row, built with
column as generated SQL                 ``&``, exactly as the original process did it
Notices the memo disagrees with the     Sign-off says *three* entities; the scope and the
statement                               statement both name *four*
Refuses to trust a summary figure       Sign-off's total movement is from an earlier run
Handles Excel's null-as-zero against    One in-scope row has a blank accrual: Excel makes
SQL's NULL propagation                  it 0.00, the warehouse leaves it NULL
Rounds the way Excel rounds             Several uplifts need the 15-significant-digit rule
Escapes a literal correctly             Entity ``O'Brien & Partners``
Excludes what the WHERE clause excludes A ``CANCELLED`` row sits inside an adjusted entity
Verifies rather than eyeballs a total   ``Post-Adjustment`` matches the prediction on every
                                        row but the null one
======================================  ==================================================

The last row is the one that matters most, and it is deliberately not a contrivance: the
mismatch is the *genuine* consequence of the null divergence two rows above it. A notebook that
compares totals agrees with the memo; a notebook that compares rows finds it. That is the whole
argument for the conversion, expressed as a single row of data.
"""

from __future__ import annotations

import argparse
import datetime as dt
import random
import sys
from pathlib import Path
from typing import Any

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

FIXTURE_DIR = Path(__file__).resolve().parents[2] / "tests" / "fixtures"
if str(FIXTURE_DIR) not in sys.path:
    sys.path.insert(0, str(FIXTURE_DIR))

# Named `build_workbook`, not `generate`, and that is load-bearing: `tests/fixtures/generate.py`
# is on sys.path here, and two modules with one name means whichever reached `sys.modules` first
# wins. The corpus generator is the one that must win, since this module imports from it.
from generate import (  # noqa: E402
    attach_connections,
    excel_round,
    inject_cached_values,
    new_workbook,
    read_parts,
    write_header,
    write_parts,
    xml_attr,
)

SEED = 20261198
"""Seeded once, and not arbitrarily.

Every run produces byte-identical output, and this particular value was chosen because it puts
four in-scope accruals on the wrong side of Excel's 15-significant-digit collapse --
:func:`_require_rounding_traps` refuses to build a workbook where fewer than three land there.
Changing it means re-running that search, not lowering the threshold.
"""

WORKBOOK_NAME = "q2_accrual_adjustment.xlsx"

UPLIFT_RATE = 0.045
"""The adjustment: a 4.5% uplift on in-scope Q2 accruals."""

PERIOD_END = dt.date(2026, 6, 30)

# ── The entities, and which of them the adjustment touches ───────────────────
#
# `O'Brien & Partners` is in scope on purpose. A statement built by concatenation breaks on the
# apostrophe -- not subtly, and not until somebody is pasting it into a production client.
ENTITIES: list[tuple[str, str, bool]] = [
    ("E-04", "Northgate Holdings", True),
    ("E-07", "Brightwater Marine", True),
    ("E-09", "O'Brien & Partners", True),
    ("E-12", "Calderstone Assurance", True),
    ("E-15", "Pennine Logistics", False),
    ("E-21", "Harbourfield Retail", False),
]

IN_SCOPE = tuple(entity for entity, _, adjusted in ENTITIES if adjusted)

COST_CENTRES = ["CC-1100", "CC-1200", "CC-2400", "CC-3050", "CC-3900"]

TRADE_COUNT = 120

EXTRACT_SQL = """SELECT
    a.trade_id,
    a.entity_id,
    e.entity_name,
    a.cost_centre,
    a.accrual_gbp,
    a.period_end,
    a.status
FROM fin.accruals AS a
INNER JOIN fin.entities AS e
    ON e.entity_id = a.entity_id
WHERE a.period_end = '2026-06-30'
  AND a.ledger = 'STATUTORY'
ORDER BY a.entity_id, a.trade_id"""

UPDATE_SQL = """UPDATE fin.accruals
   SET accrual_gbp = ROUND(accrual_gbp * 1.045, 2),
       adjusted_by  = SUSER_SNAME(),
       adjusted_at  = SYSUTCDATETIME()
 WHERE period_end = '2026-06-30'
   AND entity_id IN ('E-04', 'E-07', 'E-09', 'E-12')
   AND status <> 'CANCELLED'"""

EXTRACT_CONNECTION_STRING = (
    "ODBC;DSN=FinanceWarehouse;Description=Finance Warehouse (PROD);UID=svc_finread;"
    "Trusted_Connection=Yes;APP=Microsoft Office 2016;WSID=LDN-FIN-118;"
    "DATABASE=FinanceWarehouse"
)

# ── The two deliberate discrepancies ─────────────────────────────────────────
#
# Both are things that happen in real workbooks every month, and both are invisible to anyone
# reading the memo on its own. They are what separates a notebook that transcribes the process
# from one that checks it.
SIGNOFF_ENTITY_COUNT_CLAIM = "three"
"""What the memo says. The scope row and the UPDATE both name four entities."""

STALE_BY_TRADES = 2
"""How many trades the memo's figures predate.

The stale movement is not an invented number: it is the real one, less the uplift on the last
two in-scope trades. That is what a figure left over from an earlier run actually looks like --
close enough to pass a glance, wrong by an amount nobody can spot without recomputing it. An
arbitrary constant would be caught by any check at all, including one that only looked at the
order of magnitude.
"""

PRE_HEADER_ROW = 18
"""The extract's header. Rows 3-16 hold the SQL, so there is a preamble to skip."""

ADJ_HEADER_ROW = 16

TITLE_FONT = Font(bold=True, size=14)
HEADING_FONT = Font(bold=True)
MONO_FONT = Font(name="Consolas", size=9)


# =============================================================================
# THE DATA
# =============================================================================


class Trade:
    """One row of the extract, plus what the adjustment does to it."""

    __slots__ = (
        "accrual",
        "cost_centre",
        "entity_id",
        "entity_name",
        "in_scope",
        "status",
        "trade_id",
    )

    def __init__(
        self,
        trade_id: str,
        entity_id: str,
        entity_name: str,
        cost_centre: str,
        accrual: float | None,
        status: str,
        *,
        in_scope: bool,
    ) -> None:
        self.trade_id = trade_id
        self.entity_id = entity_id
        self.entity_name = entity_name
        self.cost_centre = cost_centre
        self.accrual = accrual
        self.status = status
        self.in_scope = in_scope

    @property
    def adjusted(self) -> bool:
        """Whether the UPDATE's WHERE clause reaches this row.

        Entity and period put it in scope; ``CANCELLED`` takes it back out. A blank accrual does
        *not* -- the statement still matches the row, which is exactly how the null gets through.
        """
        return self.in_scope and self.status != "CANCELLED"

    def excel_after(self) -> float:
        """What the workbook computes: ``=ROUND(C*(1+rate),2)`` over a blank cell as zero."""
        return excel_round((self.accrual or 0.0) * (1 + UPLIFT_RATE), 2)

    def excel_uplift(self) -> float:
        return excel_round((self.accrual or 0.0) * UPLIFT_RATE, 2)

    def warehouse_after(self) -> float | None:
        """What the database ends up holding: ``NULL * 1.045`` is ``NULL``, not zero."""
        if self.accrual is None:
            return None
        return excel_round(self.accrual * (1 + UPLIFT_RATE), 2)


def build_trades() -> list[Trade]:
    """The 120-row extract, seeded and deterministic.

    Accruals are drawn until at least three in-scope values need Excel's 15-significant-digit
    collapse to round correctly -- see :func:`generate.excel_round`. Asserting that the trap is
    present is the point: an eval whose hardest discrimination silently stopped being there is
    worse than one that fails, because it goes green.
    """
    rng = random.Random(SEED)
    trades: list[Trade] = []
    per_entity = TRADE_COUNT // len(ENTITIES)

    for entity_id, entity_name, adjusted in ENTITIES:
        for index in range(per_entity):
            trade_id = f"ACC-{len(trades) + 1:05d}"
            status = "CANCELLED" if (adjusted and index == 3) else "POSTED"
            accrual: float | None = round(rng.uniform(1_200.0, 480_000.0), 2)
            trades.append(
                Trade(
                    trade_id,
                    entity_id,
                    entity_name,
                    rng.choice(COST_CENTRES),
                    accrual,
                    status,
                    in_scope=adjusted,
                )
            )

    # The blank accrual, planted in an entity the adjustment reaches. Excel reads the empty cell
    # as zero and predicts 0.00; the warehouse propagates the NULL. One row, and it is the row
    # the whole eval turns on.
    #
    # Deliberately NOT E-09. That is the entity whose name breaks a concatenated literal, and
    # putting both traps on one row would let a notebook that fixed either one look as though it
    # had fixed both. Two findings need two rows.
    blank = next(trade for trade in trades if trade.entity_id == "E-04" and trade.adjusted)
    blank.accrual = None

    _require_rounding_traps(trades)
    return trades


def _require_rounding_traps(trades: list[Trade], *, minimum: int = 3) -> None:
    """Refuse to build a workbook whose hardest numeric discrimination has vanished.

    A trap is a row where rounding the raw double and rounding Excel's 15-digit collapse of it
    disagree -- a penny, propagating into the totals the memo quotes.
    """
    from decimal import ROUND_HALF_UP, Decimal

    traps = 0
    for trade in trades:
        if not trade.adjusted or trade.accrual is None:
            continue
        raw = trade.accrual * (1 + UPLIFT_RATE)
        # `Decimal(raw)` is the *exact* binary expansion, which is what rounding the double
        # naively means. `excel_round` collapses to 15 significant digits first. Where the two
        # disagree, a workbook that skipped the collapse is a penny out.
        naive = float(Decimal(raw).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
        if naive != excel_round(raw, 2):
            traps += 1
    if traps < minimum:
        msg = (
            f"only {traps} of the in-scope rows need Excel's 15-significant-digit rounding "
            f"rule, and the eval asserts on at least {minimum}. Change SEED until they appear "
            f"-- do not lower the threshold, which would delete the discrimination."
        )
        raise RuntimeError(msg)


# =============================================================================
# THE SHEETS
# =============================================================================


def write_signoff(sheet: Any, trades: list[Trade], adjusted: list[Trade]) -> None:
    """The memo somebody signs -- prose, an impact table, and two things that are not true.

    Every figure here is typed, because that is how these are written: somebody runs the tabs,
    reads a total off the bottom, and types it into the summary. That is also why one of them is
    stale, and why no amount of care in the *other* three tabs would surface it.
    """
    sheet["A1"] = "Q2 2026 statutory accrual adjustment - sign-off"
    sheet["A1"].font = TITLE_FONT

    blocks: list[tuple[str, str]] = [
        (
            "Purpose",
            "To record the Q2 2026 uplift applied to statutory accruals following the "
            "reforecast agreed at the June finance committee, and to evidence that the "
            "adjustment was applied as approved.",
        ),
        (
            "Background",
            "The June reforecast moved the accrual basis for the affected entities from the "
            "2025 rate card to the 2026 one. Finance agreed a flat 4.5% uplift as the "
            "practical equivalent, to be applied to statutory-ledger accruals at 30 June "
            "2026 only. Management accounts are unaffected.",
        ),
        (
            "What was done",
            f"The pre-adjustment position was extracted from the finance warehouse (see the "
            f"Pre-Adjustment tab, which carries the query). The uplift was calculated per "
            f"trade on the Adjustment tab and applied to the {SIGNOFF_ENTITY_COUNT_CLAIM} "
            f"affected entities by the update statement recorded there. The position was then "
            f"re-extracted using the same query to confirm the update had taken effect.",
        ),
        (
            "Scope",
            "Entities E-04, E-07, E-09 and E-12. Statutory ledger only. Period end 30 June "
            "2026. Cancelled trades are excluded.",
        ),
        (
            "Known issues",
            "One trade carries no accrual value in the source system pending a cost-centre "
            "reallocation. It is shown at nil and will be picked up in Q3.",
        ),
    ]
    row = 3
    for heading, body in blocks:
        sheet.cell(row=row, column=1, value=heading).font = HEADING_FONT
        cell = sheet.cell(row=row + 1, column=1, value=body)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 3

    sheet.cell(row=row, column=1, value="Impact summary").font = HEADING_FONT
    row += 1
    # Typed, not computed -- which is exactly why they can go stale, and why the stale one here
    # survived a sign-off. `stale` predates the last two trades joining the scope.
    before_total = excel_round(sum(trade.accrual or 0.0 for trade in adjusted), 2)
    stale_movement = excel_round(
        sum(trade.excel_uplift() for trade in adjusted[:-STALE_BY_TRADES]), 2
    )
    impact = [
        ("Measure", "Before", "After", "Movement"),
        (
            "In-scope accruals (GBP)",
            before_total,
            excel_round(before_total + stale_movement, 2),
            stale_movement,
        ),
    ]
    for offset, line in enumerate(impact):
        for column, value in enumerate(line, start=1):
            cell = sheet.cell(row=row + offset, column=column, value=value)
            if offset == 0:
                cell.font = HEADING_FONT
            elif column > 1:
                cell.number_format = "#,##0.00"
    row += len(impact) + 1

    sheet.cell(row=row, column=1, value="Rows adjusted").font = HEADING_FONT
    sheet.cell(row=row, column=2, value=len(adjusted) - STALE_BY_TRADES)
    row += 2

    sheet.cell(row=row, column=1, value="Sign-off").font = HEADING_FONT
    for offset, (label, value) in enumerate(
        [
            ("Prepared by", "R. Advani, Financial Control"),
            ("Reviewed by", "M. Okafor, Head of Statutory Reporting"),
            ("Date", dt.date(2026, 7, 14)),
            ("Control reference", "FC-Q2-2026-014"),
        ],
        start=1,
    ):
        sheet.cell(row=row + offset, column=1, value=label)
        cell = sheet.cell(row=row + offset, column=2, value=value)
        if isinstance(value, dt.date):
            cell.number_format = "dd/mm/yyyy"

    sheet.column_dimensions["A"].width = 34
    for letter in ("B", "C", "D"):
        sheet.column_dimensions[letter].width = 20


def write_extract_sheet(sheet: Any, trades: list[Trade], *, title: str, after: bool) -> None:
    """An extract tab: the query that produced it, then the grid it produced.

    The query sits above the data rather than beside it, which is how people actually leave it
    -- and which means the analyser has fourteen rows of preamble to skip before it finds a
    header. That is a real layout, so it is a real thing to get right.
    """
    sheet["A1"] = title
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = "Source: FinanceWarehouse (ODBC), connection 'AccrualExtract'"

    sheet.cell(row=3, column=1, value="Query - run this to reproduce:").font = HEADING_FONT
    for offset, line in enumerate(EXTRACT_SQL.splitlines(), start=4):
        cell = sheet.cell(row=offset, column=1, value=line)
        cell.font = MONO_FONT

    headers = [
        "trade_id",
        "entity_id",
        "entity_name",
        "cost_centre",
        "accrual_gbp",
        "period_end",
        "status",
    ]
    write_header(sheet, PRE_HEADER_ROW, headers)
    for index, trade in enumerate(trades):
        row = PRE_HEADER_ROW + 1 + index
        sheet.cell(row=row, column=1, value=trade.trade_id)
        sheet.cell(row=row, column=2, value=trade.entity_id)
        sheet.cell(row=row, column=3, value=trade.entity_name)
        sheet.cell(row=row, column=4, value=trade.cost_centre)

        value = trade.warehouse_after() if after and trade.adjusted else trade.accrual
        cell = sheet.cell(row=row, column=5, value=value)
        cell.number_format = "#,##0.00"

        period = sheet.cell(row=row, column=6, value=PERIOD_END)
        period.number_format = "yyyy-mm-dd"
        sheet.cell(row=row, column=7, value=trade.status)

    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["C"].width = 24
    sheet.column_dimensions["E"].width = 15
    sheet.column_dimensions["F"].width = 12


def write_adjustment(sheet: Any, trades: list[Trade]) -> list[Trade]:
    """The working: parameters, the statement, and one generated UPDATE per row.

    Column F is the point of the whole fixture. A column of ``="UPDATE ... "&E17&"..."`` is what
    a real finance workbook does when the process needs SQL it cannot run from Excel, and it is
    unmistakable: no data pipeline produces it, and nothing else in a workbook looks like it.
    kedge should read it as "generate this statement and hand it over", not translate it into a
    string column.
    """
    sheet["A1"] = "Q2 2026 accrual uplift - working"
    sheet["A1"].font = TITLE_FONT

    sheet.cell(row=3, column=1, value="Parameters").font = HEADING_FONT
    sheet.cell(row=4, column=1, value="Uplift rate")
    rate = sheet.cell(row=4, column=2, value=UPLIFT_RATE)
    rate.number_format = "0.0%"
    sheet.cell(row=5, column=1, value="Period end")
    period = sheet.cell(row=5, column=2, value=PERIOD_END)
    period.number_format = "yyyy-mm-dd"
    sheet.cell(row=6, column=1, value="Entities in scope")
    sheet.cell(row=6, column=2, value=", ".join(IN_SCOPE))
    sheet.cell(row=7, column=1, value="Ledger")
    sheet.cell(row=7, column=2, value="STATUTORY")

    sheet.cell(
        row=9, column=1, value="Update statement (run against FinanceWarehouse):"
    ).font = HEADING_FONT
    for offset, line in enumerate(UPDATE_SQL.splitlines(), start=10):
        sheet.cell(row=offset, column=1, value=line).font = MONO_FONT

    headers = [
        "trade_id",
        "entity_id",
        "entity_name",
        "accrual_gbp_before",
        "uplift_gbp",
        "accrual_gbp_after",
        "update_statement",
    ]
    write_header(sheet, ADJ_HEADER_ROW, headers)

    adjusted = [trade for trade in trades if trade.adjusted]
    first_data = PRE_HEADER_ROW + 1
    last_data = PRE_HEADER_ROW + len(trades)
    lookup = f"'Pre-Adjustment'!$A${first_data}:$G${last_data}"

    for index, trade in enumerate(adjusted):
        row = ADJ_HEADER_ROW + 1 + index
        sheet.cell(row=row, column=1, value=trade.trade_id)
        sheet.cell(row=row, column=2, value=trade.entity_id)
        sheet.cell(row=row, column=3, value=f"=VLOOKUP(A{row},{lookup},3,FALSE)")
        sheet.cell(row=row, column=4, value=f"=VLOOKUP(A{row},{lookup},5,FALSE)")
        sheet.cell(row=row, column=5, value=f"=ROUND(D{row}*uplift_rate,2)")
        sheet.cell(row=row, column=6, value=f"=ROUND(D{row}*(1+uplift_rate),2)")
        # The statement, built the way the original process built it: string concatenation, no
        # escaping anywhere. It is correct for every row but one, and the one it is wrong for is
        # wrong in a way nobody notices until it is pasted into a client -- see ENTITIES.
        sheet.cell(
            row=row,
            column=7,
            value=(
                f'="UPDATE fin.accruals SET accrual_gbp = "&TEXT(F{row},"0.00")'
                f'&", adjustment_note = \'Q2 uplift for "&C{row}&"\'"'
                f'&" WHERE trade_id = \'"&A{row}&"\';"'
            ),
        )
        for column in (4, 5, 6):
            sheet.cell(row=row, column=column).number_format = "#,##0.00"

    total_row = ADJ_HEADER_ROW + len(adjusted) + 2
    sheet.cell(row=total_row, column=1, value="Total").font = HEADING_FONT
    for column in (4, 5, 6):
        letter = get_column_letter(column)
        cell = sheet.cell(
            row=total_row,
            column=column,
            value=f"=SUM({letter}{ADJ_HEADER_ROW + 1}:{letter}{ADJ_HEADER_ROW + len(adjusted)})",
        )
        cell.number_format = "#,##0.00"
    sheet.cell(row=total_row + 1, column=1, value="Rows")
    sheet.cell(
        row=total_row + 1,
        column=4,
        value=f"=COUNT(D{ADJ_HEADER_ROW + 1}:D{ADJ_HEADER_ROW + len(adjusted)})",
    )

    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["C"].width = 24
    sheet.column_dimensions["G"].width = 96
    return adjusted


# =============================================================================
# CACHED VALUES
# =============================================================================


def cached_for_adjustment(adjusted: list[Trade], total_row: int) -> dict[str, Any]:
    """What Excel would have left in the Adjustment tab's formula cells.

    Without these the workbook reconciles to "not reconciled" and half the eval is dead --
    openpyxl writes every formula cell with an empty ``<v>``. The values are computed here with
    :func:`generate.excel_round`, so the pennies match what Excel itself produces, and
    ``--verify-with-excel`` proves it.
    """
    values: dict[str, Any] = {}
    for index, trade in enumerate(adjusted):
        row = ADJ_HEADER_ROW + 1 + index
        before = trade.accrual if trade.accrual is not None else 0.0
        after = trade.excel_after()
        values[f"C{row}"] = trade.entity_name
        values[f"D{row}"] = before
        values[f"E{row}"] = trade.excel_uplift()
        values[f"F{row}"] = after
        # Reproduced exactly as Excel's `&` would leave it, apostrophe and all. This string is
        # not valid SQL for one of the entities, and that is the finding.
        values[f"G{row}"] = (
            f"UPDATE fin.accruals SET accrual_gbp = {after:.2f}, "
            f"adjustment_note = 'Q2 uplift for {trade.entity_name}' "
            f"WHERE trade_id = '{trade.trade_id}';"
        )

    values[f"D{total_row}"] = excel_round(sum(trade.accrual or 0.0 for trade in adjusted), 2)
    values[f"E{total_row}"] = excel_round(sum(trade.excel_uplift() for trade in adjusted), 2)
    values[f"F{total_row}"] = excel_round(sum(trade.excel_after() for trade in adjusted), 2)
    values[f"D{total_row + 1}"] = len(adjusted)
    return values


# =============================================================================
# BUILD
# =============================================================================


def build(path: Path) -> None:
    """Write the workbook, cached values and connection included."""
    trades = build_trades()
    workbook = new_workbook()

    adjusted = [trade for trade in trades if trade.adjusted]

    signoff = workbook.active
    signoff.title = "Sign-off"
    write_signoff(signoff, trades, adjusted)

    pre = workbook.create_sheet("Pre-Adjustment")
    write_extract_sheet(pre, trades, title="Pre-adjustment extract", after=False)

    adjustment = workbook.create_sheet("Adjustment")
    workbook.defined_names.add(DefinedName("uplift_rate", attr_text="Adjustment!$B$4"))
    write_adjustment(adjustment, trades)

    post = workbook.create_sheet("Post-Adjustment")
    write_extract_sheet(post, trades, title="Post-adjustment extract (re-run)", after=True)

    workbook.save(path)

    total_row = ADJ_HEADER_ROW + len(adjusted) + 2
    parts = read_parts(path)
    populated = inject_cached_values(
        parts, {"Adjustment": cached_for_adjustment(adjusted, total_row)}
    )
    attach_connections(parts, _connections_xml())
    write_parts(path, parts)

    print(
        f"wrote {path.name}: {len(trades)} trades, {len(adjusted)} adjusted, "
        f"{populated['Adjustment']} cached values"
    )


def _connections_xml() -> bytes:
    """The extract query as Excel stores it: newlines as character references.

    XML attribute-value normalisation turns a literal newline inside an attribute into a space,
    so multi-line SQL has to be written as ``&#10;`` -- which is what Excel does, and what
    ``kedge.analysis.connections`` is built to read back.
    """
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<connections xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<connection id="1" name="AccrualExtract" type="1" refreshedVersion="8"'
        ' minRefreshableVersion="3" background="1" saveData="1"'
        ' description="Statutory accrual position at period end">'
        f'<dbPr connection="{xml_attr(EXTRACT_CONNECTION_STRING)}"'
        f' command="{xml_attr(EXTRACT_SQL)}" commandType="2"/>'
        "</connection>"
        "</connections>"
    ).encode()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).resolve().parent / WORKBOOK_NAME,
        help="where to write the workbook",
    )
    parser.add_argument(
        "--verify-with-excel",
        action="store_true",
        help="open the result in Excel over COM and check every cached value against Excel's "
        "own recalculation. Windows only, needs Excel and pywin32, never re-saves the file.",
    )
    args = parser.parse_args()
    build(args.out)
    if args.verify_with_excel:
        from generate import verify_cached_values

        failures = [check for check in verify_cached_values(args.out) if not check.passed]
        for check in failures:
            print(f"  MISMATCH {check.detail}")
        print("Excel agreed on every cached value" if not failures else f"{len(failures)} differ")
        return 1 if failures else 0
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
