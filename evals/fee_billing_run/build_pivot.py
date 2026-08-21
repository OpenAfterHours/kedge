"""Finish the fee-billing-run workbook with its ``Summary`` pivot, by driving Excel.

openpyxl reads a pivot table in full -- source range, cache fields, row and column fields, the
data field and its aggregation -- and cannot author one. Nothing it could write would satisfy
Excel, because a pivot is five parts that have to agree: the cache definition, the cache records,
the table definition, the relationship between them and the rendered grid on the sheet. So the
only way to get a genuine pivot into the eval workbook is to ask Excel for one, which is the same
bargain ``tests/fixtures/generate.py --verify-with-excel`` already strikes -- where ground truth
about Excel is needed, drive Excel rather than reason about it.

Deliberately not part of :func:`build_workbook.build`. That function is pure Python,
deterministic to the byte, and CI runs it on Ubuntu; this needs Windows, Excel and pywin32, and
produces a file Excel wrote rather than one this repository wrote. It is opt-in::

    uv run --with pywin32 python evals/fee_billing_run/build_workbook.py --with-pivot
    uv run --with pywin32 python evals/fee_billing_run/build_pivot.py     # on an existing file

**Opening this workbook in Excel destroys the eval, and the fix is one line.** Measured, not
assumed: opened with Excel's default automatic calculation, a pivot added and the file saved, 34
of its 5,253 cached values move -- and five of them are the ones the eval is built around.
``Allocation`` is planted *stale*, holding the fees as they stood before the three overrides were
agreed (discrimination 9, and the only place in the eval where the cached values are deliberately
wrong). Excel recalculates it into agreement with ``Working``, ``Recon`` follows it, and the
discrimination is gone from a file that still opens, still analyses and still passes every
assertion that does not look at those five cells. ``fullCalcOnLoad="0"`` does not stop it: Excel
compares the stored ``calcId`` against its own engine version and recalculates anyway.

Forcing ``Application.Calculation`` to manual *before* the workbook is opened moves that figure
to zero of 5,253. It has to be before, because a workbook opened while the application is on
automatic has already recalculated by the time the property is settable -- and the property is
application-level, which Excel refuses to set with no workbook open at all, hence the empty
scratch workbook opened first and closed last. :func:`add_pivot` re-reads every cached value
afterwards and refuses to leave a damaged workbook behind, so the guard cannot rot silently.

Excel also rewrites ``<calcPr>`` on the way out (``calcId="0" calcCompleted="0"``), which is a
claim about the file this repository has not made. The element ``build`` wrote is put back
verbatim, and the archive is re-zipped through :func:`generate.write_parts` so the timestamps
stay pinned. The result is still not byte-reproducible -- Excel stamps its own revision ids into
the parts it writes -- which is the other reason this path is opt-in.
"""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

FIXTURE_DIR = Path(__file__).resolve().parents[2] / "tests" / "fixtures"
if str(FIXTURE_DIR) not in sys.path:
    sys.path.insert(0, str(FIXTURE_DIR))

from generate import read_parts, write_parts  # noqa: E402

WORKBOOK_NAME = "m11_management_fee_run.xlsx"

SOURCE_SHEET = "Allocation"
TARGET_SHEET = "Summary"
PIVOT_NAME = "FeeSummary"
ANCHOR = "A3"
ROW_FIELDS = ("legal_entity", "cost_centre")
COLUMN_FIELD = "period"
DATA_FIELD = "fee_gbp"
DATA_CAPTION = "Sum of fee_gbp"
"""What the pivot is, in one place, because the Excel side and the read-back both need it.

No page field. The proposal's probe had one and the extractor of section 7.1 will want to see one
eventually, but ``Allocation`` carries five columns and none of them is a natural filter --
``period`` is already the column axis and ``client_code`` is the grain. A page field over the
grain would be decoration, and decoration in an eval is something a grader has to explain away.
"""

# Excel's own enumeration members. Spelled out because a late-bound COM client has no constants
# module to import them from, and because `win32com.client.constants` is populated only after a
# makepy pass this script deliberately does not require.
XL_DATABASE = 1
XL_ROW_FIELD = 1
XL_COLUMN_FIELD = 2
XL_DATA_FIELD = 4
XL_SUM = -4157
XL_CALCULATION_MANUAL = -4135
XL_CALCULATION_AUTOMATIC = -4105

_CALC_PR = re.compile(r"<calcPr\b[^>]*/>")


@dataclass(frozen=True, slots=True)
class PivotBuilt:
    """What the COM pass produced, read back off the saved file rather than claimed.

    Every field here comes from openpyxl re-parsing the archive, so a report built from it is
    evidence that the parts agree rather than a transcript of what was asked for.
    """

    location: str
    source_sheet: str
    source_ref: str
    cache_fields: tuple[str, ...]
    row_fields: tuple[str, ...]
    column_fields: tuple[str, ...]
    page_fields: tuple[str, ...]
    data_fields: tuple[str, ...]
    records: int
    cached_values_checked: int

    def render(self) -> str:
        """The verification, as the lines somebody reading the build output needs."""
        return "\n".join(
            (
                f"pivot            {PIVOT_NAME!r} at {TARGET_SHEET}!{self.location}",
                f"source           sheet={self.source_sheet!r} ref={self.source_ref!r}",
                f"cache fields     {list(self.cache_fields)}",
                f"rowFields        {list(self.row_fields)}",
                f"colFields        {list(self.column_fields)}",
                f"pageFields       {list(self.page_fields)}",
                f"dataFields       {list(self.data_fields)}",
                f"cache records    {self.records}",
                f"cached values    {self.cached_values_checked} checked, none moved",
            )
        )


# =============================================================================
# THE CACHED VALUES THE PIVOT MUST NOT DISTURB
# =============================================================================


def cached_snapshot(path: Path) -> dict[str, Any]:
    """Every cached value in the workbook, keyed ``sheet!ref``.

    ``Summary`` is excluded because that is the tab gaining the pivot's rendered grid, and a
    sheet that is meant to change is not evidence about a sheet that is not.
    """
    workbook = load_workbook(path, data_only=True)
    try:
        return {
            f"{sheet.title}!{cell.coordinate}": cell.value
            for sheet in workbook.worksheets
            if sheet.title != TARGET_SHEET
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        }
    finally:
        workbook.close()


def moved_values(before: dict[str, Any], after: dict[str, Any]) -> list[str]:
    """Cached values that are not what they were, rendered for a human.

    Floats are compared with a tolerance because a value that survives is still a value Excel has
    round-tripped through its own decimal formatting; a value that has been *recalculated* moves
    by pounds, not by ulps.
    """
    changed: list[str] = []
    for key, value in before.items():
        found = after.get(key, "<gone>")
        if isinstance(value, float) and isinstance(found, float):
            if abs(value - found) > 1e-6:
                changed.append(f"{key}: {value!r} -> {found!r}")
        elif value != found:
            changed.append(f"{key}: {value!r} -> {found!r}")
    return changed


# =============================================================================
# DRIVING EXCEL
# =============================================================================


def _excel() -> Any:
    """Start Excel, importing pywin32 lazily so this module imports anywhere.

    The eval suite imports every generator it can find, on every platform CI runs, and a
    top-level ``import win32com`` would make this module unimportable on all of them.
    """
    try:
        import win32com.client
    except ImportError as error:
        msg = (
            "building the Summary pivot needs pywin32 and a local Excel installation. Run it as "
            "`uv run --with pywin32 python evals/fee_billing_run/build_pivot.py`."
        )
        raise RuntimeError(msg) from error
    return win32com.client.Dispatch("Excel.Application")


def _drop_existing(sheet: Any) -> None:
    """Remove a previous ``FeeSummary`` so a second run replaces it rather than failing.

    Counted downwards because clearing a table removes it from the collection, and Excel
    renumbers what is left. ``TableRange2`` is the whole table including the page-field rows;
    ``TableRange1`` omits them, and clearing that leaves a header nothing owns. Excel drops the
    orphaned cache when the file is saved.
    """
    for index in range(sheet.PivotTables().Count, 0, -1):
        table = sheet.PivotTables(index)
        if table.Name == PIVOT_NAME:
            table.TableRange2.Clear()


def _write_pivot(book: Any) -> None:
    """Create the pivot over ``Allocation``'s used range and lay its fields out."""
    source = book.Worksheets(SOURCE_SHEET)
    target = book.Worksheets(TARGET_SHEET)
    _drop_existing(target)

    # `Address` is a property to a late-bound client, not a method: calling it raises
    # `'str' object is not callable`. The default it returns is already absolute and local, so
    # the sheet name is prefixed here.
    ref = f"'{source.Name}'!{source.UsedRange.Address}"
    cache = book.PivotCaches().Create(SourceType=XL_DATABASE, SourceData=ref)
    table = cache.CreatePivotTable(TableDestination=target.Range(ANCHOR), TableName=PIVOT_NAME)

    for position, name in enumerate(ROW_FIELDS, start=1):
        field = table.PivotFields(name)
        field.Orientation = XL_ROW_FIELD
        field.Position = position
    table.PivotFields(COLUMN_FIELD).Orientation = XL_COLUMN_FIELD

    data = table.PivotFields(DATA_FIELD)
    data.Orientation = XL_DATA_FIELD
    data.Function = XL_SUM
    data.Caption = DATA_CAPTION


def _drive_excel(path: Path) -> None:
    """Open the workbook with calculation frozen, add the pivot, save, and always quit.

    The scratch workbook is not decoration. ``Application.Calculation`` cannot be set with no
    workbook open, and it has to be manual *before* the eval workbook loads -- see this module's
    docstring for what a single automatic load costs.
    """
    excel = _excel()
    excel.Visible = False
    excel.DisplayAlerts = False
    scratch = None
    try:
        scratch = excel.Workbooks.Add()
        excel.Calculation = XL_CALCULATION_MANUAL
        excel.CalculateBeforeSave = False
        book = excel.Workbooks.Open(str(path.resolve()), UpdateLinks=0)
        try:
            _write_pivot(book)
            book.Save()
        finally:
            book.Close(SaveChanges=False)
    finally:
        try:
            if scratch is not None:
                excel.Calculation = XL_CALCULATION_AUTOMATIC
                scratch.Close(SaveChanges=False)
        finally:
            excel.Quit()


# =============================================================================
# READING IT BACK
# =============================================================================


def read_back(path: Path, cached_values_checked: int = 0) -> PivotBuilt:
    """Re-parse the saved archive and report what the pivot actually is.

    ``Worksheet._pivots`` is private and is the only accessor openpyxl offers; everything below
    it is public API.
    """
    workbook = load_workbook(path)
    try:
        pivots = workbook[TARGET_SHEET]._pivots
        found = [pivot for pivot in pivots if pivot.name == PIVOT_NAME]
        if len(found) != 1:
            names = [pivot.name for pivot in pivots]
            msg = (
                f"expected exactly one pivot named {PIVOT_NAME!r} on {TARGET_SHEET}, found {names}"
            )
            raise RuntimeError(msg)
        pivot = found[0]
        cache = pivot.cache
        source = cache.cacheSource.worksheetSource
        names = [field.name for field in cache.cacheFields]

        def named(index: int) -> str:
            """A field index as its cache-field name. -2 is Excel's marker for the data axis."""
            return names[index] if 0 <= index < len(names) else "(values)"

        return PivotBuilt(
            location=str(pivot.location.ref),
            source_sheet=str(source.sheet),
            source_ref=str(source.ref),
            cache_fields=tuple(names),
            row_fields=tuple(named(field.x) for field in pivot.rowFields),
            column_fields=tuple(named(field.x) for field in pivot.colFields),
            page_fields=tuple(named(field.fld) for field in pivot.pageFields),
            data_fields=tuple(
                f"{field.name} = {field.subtotal}({named(field.fld)})" for field in pivot.dataFields
            ),
            records=len(cache.records.r) if cache.records is not None else 0,
            cached_values_checked=cached_values_checked,
        )
    finally:
        workbook.close()


# =============================================================================
# BUILD
# =============================================================================


def add_pivot(path: Path) -> PivotBuilt:
    """Add the ``Summary`` pivot to an existing workbook, in place.

    Idempotent: an existing ``FeeSummary`` is replaced rather than duplicated. Refuses to return
    a workbook whose other cached values have moved, because the one thing this step must not do
    is quietly recalculate the staleness out of ``Allocation``.
    """
    if not path.is_file():
        msg = f"{path} does not exist; build the workbook before adding the pivot"
        raise FileNotFoundError(msg)

    before = cached_snapshot(path)
    calc_pr = _CALC_PR.search(read_parts(path)["xl/workbook.xml"].decode("utf-8"))
    if calc_pr is None:
        msg = f"{path.name} has no <calcPr> element, so its calculation settings cannot be restored"
        raise RuntimeError(msg)

    _drive_excel(path)

    parts = read_parts(path)
    workbook_xml = parts["xl/workbook.xml"].decode("utf-8")
    parts["xl/workbook.xml"] = _CALC_PR.sub(calc_pr.group(0), workbook_xml, count=1).encode("utf-8")
    write_parts(path, parts)

    moved = moved_values(before, cached_snapshot(path))
    if moved:
        head = "\n  ".join(moved[:10])
        msg = (
            f"Excel moved {len(moved)} of {len(before)} cached values while adding the pivot, so "
            f"the planted staleness in {SOURCE_SHEET} can no longer be trusted:\n  {head}"
        )
        raise RuntimeError(msg)

    return read_back(path, cached_values_checked=len(before))


def main() -> int:
    path = Path(__file__).resolve().parent / WORKBOOK_NAME
    print(add_pivot(path).render())
    return 0


if __name__ == "__main__":
    sys.exit(main())
