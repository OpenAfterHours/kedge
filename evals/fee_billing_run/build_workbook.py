"""Build the management-fee billing-run eval workbook.

The workbook this produces records a monthly process: work out what each client owes, apply the
floors, caps and discounts that were agreed, allocate the result to the entity that carries it,
and post the invoices. Ten tabs, laid out in the order somebody added them over four years --
which is deliberately not dependency order. See ``evals/proposals/fee_billing_run.md``.

What makes it an eval rather than a fixture is that the difficulty is *placed*. Ten
discriminations are planted, each so it can fail on its own; the proposal names them and this
module is where each one physically lives.

Two structural rules govern the whole file, and breaking either quietly destroys the eval.

**Complexity comes from variety, never from size.** R1C1 normalisation collapses a formula
filled down a column and across thirty columns to one logical operation, so a big uniform grid
is a *simple* workbook however many cells it has. ``WORKING_COLUMNS`` is therefore a list of
distinct shapes, and it is the complexity of the workbook almost by itself.

**The opposite error is just as easy.** A reference whose target moves per row -- a fill-down of
``=Positions!$A4``, ``=Positions!$A6``, ``=Positions!$A8`` -- normalises to a *different* R1C1
string on every row, so it never compresses and one column becomes eighty-four operations. The
first build of this workbook came in at 208 operations and complexity 0.855 for exactly that
reason. Both offenders are now written the way the process would really do it, and
``--calibrate`` is what caught it.

Columns are addressed **by name**, never by letter. Every formula template interpolates
``{some_key}`` and gets that column's letter, so inserting a column cannot silently repoint an
arithmetic chain at its neighbour. The same keys drive the Python model in ``compute``, which is
what makes the cached values a parallel implementation of the sheet rather than a copy of it.

Run it::

    uv run python evals/fee_billing_run/build_workbook.py          # rebuild
    uv run python evals/fee_billing_run/build_workbook.py --calibrate   # measure, write nothing

**Building and measuring are different operations and this file keeps them apart.**
``--calibrate`` and ``--verify-with-excel`` read the workbook that is there; they never rebuild
it, because the committed file is the artifact the eval grades and a rebuilt one is a different
input. See :func:`main`.

:func:`build` is pure Python and byte-reproducible, which is what lets CI regenerate the workbook
on any platform and compare. The ``Summary`` pivot is the one thing it cannot write -- openpyxl
reads a pivot table and cannot author one -- so that is a separate opt-in pass over the finished
file, needing Windows, Excel and pywin32::

    uv run --with pywin32 python evals/fee_billing_run/build_workbook.py --with-pivot --calibrate

A rebuild over a workbook that already carries a pivot is refused unless it will put the pivot
back, or ``--force`` says the loss is intended.

See ``build_pivot.py``, and note the warning it carries: Excel recalculates on open, which
silently repairs the deliberately stale ``Allocation`` figures that discrimination 9 is made of.
"""

from __future__ import annotations

import argparse
import datetime as dt
import logging
import random
import shutil
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

FIXTURE_DIR = Path(__file__).resolve().parents[2] / "tests" / "fixtures"
if str(FIXTURE_DIR) not in sys.path:
    sys.path.insert(0, str(FIXTURE_DIR))

# Imported rather than copied, for the reason CLAUDE.md gives about `kedge.xl`: a second copy of
# a rounding or normalisation routine is a second chance to get it subtly wrong. `excel_round`
# carries the 15-significant-digit collapse, and `normalise` has to run *after* the save --
# openpyxl assigns `properties.modified = datetime.now()` inside `save_workbook`, so a timestamp
# set on the workbook object is discarded and the file is reproducible only within one second.
from generate import (  # noqa: E402
    attach_connections,
    excel_round,
    inject_cached_values,
    normalise,
    read_parts,
    set_full_calc_on_load,
    write_parts,
    xml_attr,
)

logger = logging.getLogger(__name__)

WORKBOOK_NAME = "m11_management_fee_run.xlsx"

# The structural band the proposal commits to, asserted by `--calibrate` and by
# `tests/unit/test_evals_fee_billing_run.py`. A band rather than a floor because this eval can
# rot in both directions: too simple and it stresses nothing the corpus does not already stress,
# too complex and it triages to `stop`, at which point it cannot be converted and so cannot be
# graded on the quality of its conversion.
TARGET_OPERATIONS = (45, 60)
TARGET_PATTERNS = 12
TARGET_COMPLEXITY = (0.68, 0.75)
FORBIDDEN_VERDICT = "stop"

SEED = 20261103
"""The one source of randomness, named so that anything checking the workbook's figures can
reproduce them without copying a literal out of :func:`build`. ``expected.yaml``'s ``facts`` are
checked against ``compute(build_clients(random.Random(SEED)))``, which is only the same run as
the committed file for as long as this is the seed the build uses."""

CLIENTS = 84
ENTITIES = ("LuxCo", "UKCo", "DubCo")
MANDATES = ("  balanced ", "equity", " fixed income")

PERIOD_START = dt.date(2026, 11, 1)
PERIOD_END = dt.date(2026, 11, 30)
EXCEL_EPOCH = dt.date(1899, 12, 30)
"""Excel's day zero. 1900 is treated as a leap year, which is why this is the 30th."""

MIN_FEE = 750.0
MAX_FEE = 250_000.0
DISCOUNT = 0.025
OLD_FLAT_BPS = 26.0
OLD_ADJUSTMENT = 0.97

BANDS: tuple[tuple[float, float, float], ...] = (
    (0.0, 5_000_000.0, 35.0),
    (5_000_000.0, 20_000_000.0, 27.5),
    (20_000_000.0, 50_000_000.0, 22.0),
    (50_000_000.0, 100_000_000.0, 18.5),
)
"""Tiered rate card, matched with an approximate ``VLOOKUP`` on the band floor.

Discrimination 1, and the highest-value translation risk in the eval: the polars equivalent is
``join_asof(strategy="backward")`` on a sorted frame, the boundary is ``>=`` on the floor, and
an off-by-one moves an entire tier of clients onto the wrong rate.
"""

NEGOTIATED_STEP = 5
"""Every fifth client has a negotiated rate that overrides the band.

Not every client, deliberately: if all of them had one the banded lookup would be dead code and
discrimination 1 would evaporate.
"""

SMALL_ACCOUNTS: dict[int, float] = {3: 120_000.0, 27: 185_000.0, 71: 96_500.0}
"""Client index to AUM, for the handful of accounts small enough that the minimum fee binds.

Without them the floor column is dead: an average AUM of tens of millions never produces a fee
anywhere near 750, so ``MAX(prorated, minimum)`` is the identity on every row and a conversion
that drops it entirely still reconciles. ``require_planted_discriminations`` refuses to build a
workbook where that is true, and it caught this on the first run.
"""

ONBOARDED_MID_PERIOD: dict[int, int] = {12: 12, 33: 20, 58: 6}
"""Client index to day of November they were onboarded. Everyone else bills a full month.

Discrimination 3. Without these the pro-rating columns compute ``fee * 30 / 30`` for every row
and a conversion that drops them entirely still reconciles.
"""

OVERRIDES: tuple[tuple[str, float, str, dt.date], ...] = (
    (
        "00007",
        0.0,
        "Fee waived for Q4 pending the mandate transfer agreed with the client in September.",
        dt.date(2026, 9, 18),
    ),
    (
        "00041",
        18_500.0,
        "Capped at the 2025 level; the uplift was not communicated before the billing run.",
        dt.date(2026, 10, 2),
    ),
    (
        "00062",
        4_220.75,
        "Manual correction: two accounts were double-counted in the position extract.",
        dt.date(2026, 11, 3),
    ),
)
"""Three computed fees the billing manager typed over, each with a written reason.

Discrimination 4, and the sharpest judgement call in the eval. Three answers are wrong in
different directions: reproducing the numbers bakes last quarter's decision into every future
month, dropping them loses three real decisions, and reconciling against them reports a pass for
a figure nobody calculated. The reasons are prose because that is what makes them re-askable.
"""

CLIENT_NAMES: dict[int, str] = {40: "O'Hanlon & Reid Nominees"}
"""Client index to the legal name on the account, where it is not the generated one.

**A workbook in which no literal needs escaping cannot test non-negotiable 3.** For its first
year this one had none: ``Post`` inserted a five-digit code, a month string and a number, so
naive ``&`` concatenation produced valid SQL on every row and ``generated_sql_is_valid`` passed a
conversion that had never heard of :mod:`kedge.sql`. ``adjustment_signoff`` has ``O'Brien &
Partners`` and calls it the one place where matching Excel is *wrong*; this is the same plant at
this workbook's scale, and :func:`write_post` is what makes it bite -- the statement carries the
client's name, so the apostrophe reaches a quoted literal that ``&`` does not escape and the
workbook's own cached statement for ``00041`` is not valid SQL.

Index 40 is client ``00041``, which is also an :data:`OVERRIDES` row. That is deliberate: the
one statement whose fee was typed over by a person is the one whose text is broken, so a
conversion that reproduces the workbook faithfully gets both wrong in the same row.
"""

EXTRACT_SQL = """SELECT
       p.client_code,
       p.client_name,
       p.as_of_date,
       p.period_month,
       p.mandate,
       p.market_value_gbp
  FROM fin.positions p
  JOIN fin.mandates m
    ON m.client_code = p.client_code
 WHERE p.period_month IN ('2026-10', '2026-11')
   AND p.status <> 'CLOSED'
 ORDER BY p.client_code, p.as_of_date"""

EXTRACT_CONNECTION_STRING = (
    "ODBC;DSN=FinanceWarehouse;Description=Finance Warehouse (PROD);UID=svc_finread;"
    "Trusted_Connection=Yes;APP=Microsoft Office 2016;WSID=LDN-FIN-118;"
    "DATABASE=FinanceWarehouse"
)
"""The DSN the extract was refreshed through, as Excel records it in ``xl/connections.xml``.

The tab table in ``evals/proposals/fee_billing_run.md`` §3 promised this and the generator did
not write it, so ``analyse(...).connections`` came back empty and this workbook was *easier*
than ``q2_accrual_adjustment.xlsx`` on the one axis the whole case exists to make harder. A
connection part is what turns "the extract is a step to hand over" into a claim the analyser can
check twice -- once from the SQL a person left in ``Positions`` rows 3-13, and once from the
query Excel itself stored.
"""

SQL_FIRST_ROW = 3
POSITIONS_HEADER_ROW = SQL_FIRST_ROW + len(EXTRACT_SQL.splitlines()) + 1
POSITIONS_FIRST_ROW = POSITIONS_HEADER_ROW + 1
POSITIONS_LAST_ROW = POSITIONS_FIRST_ROW + CLIENTS * 2 - 1

FEE_HEADER_ROW = 4
FEE_BAND_FIRST_ROW = 5
FEE_BAND_LAST_ROW = FEE_BAND_FIRST_ROW + len(BANDS) - 1
FEE_NEGOTIATED_FIRST_ROW = 5
FEE_NEGOTIATED_LAST_ROW = FEE_NEGOTIATED_FIRST_ROW + (CLIENTS - 1) // NEGOTIATED_STEP

WORKING_HEADER_ROW = 18
WORKING_FIRST_ROW = WORKING_HEADER_ROW + 1
WORKING_LAST_ROW = WORKING_FIRST_ROW + CLIENTS - 1
WORKING_TOTAL_ROW = WORKING_LAST_ROW + 2
OLD_RATE_CELL = "B14"
OLD_FACTOR_CELL = "B15"
OLD_RATE_REF = "$B$14"
OLD_FACTOR_REF = "$B$15"
"""Fully absolute, and that is not a style preference.

``$B14`` is column-absolute but *row-relative*, so filled down it normalises to ``R[-5]C2``,
``R[-6]C2``, ``R[-7]C2`` -- a different R1C1 string on every row. Two columns written that way
added 168 single-cell operations and pushed the workbook to 216. A parameter reference is
absolute in both axes, which is also what makes it classify as ``parameter_ref`` rather than as
arithmetic."""


def serial(value: dt.date) -> int:
    """A date as Excel stores it: days since 1899-12-30."""
    return (value - EXCEL_EPOCH).days


# =============================================================================
# THE FORMULA SHAPES
# =============================================================================


@dataclass(frozen=True, slots=True)
class Column:
    """One column of ``Working``: what it is called, and how it is calculated.

    ``formula`` is a template. ``{row}`` and ``{prev}`` are the current and previous row
    numbers; every other placeholder is a *column key* and resolves to that column's letter, so
    a chain of arithmetic says what it reads rather than encoding a letter that a later
    insertion would silently invalidate. An empty template means the column holds data.

    ``key`` and ``header`` are deliberately different vocabularies. The key is the generator's,
    stable and readable; the header is the one a person typed into row 18, and it is prose. See
    the note above :data:`WORKING_COLUMNS`.
    """

    key: str
    header: str
    formula: str = ""


BAND_RANGE = f"'Fee Schedule'!$D${FEE_BAND_FIRST_ROW}:$F${FEE_BAND_LAST_ROW}"
NEGOTIATED_RANGE = f"'Fee Schedule'!$A${FEE_NEGOTIATED_FIRST_ROW}:$B${FEE_NEGOTIATED_LAST_ROW}"
POSITIONS_VALUE_RANGE = f"Positions!$F${POSITIONS_FIRST_ROW}:$F${POSITIONS_LAST_ROW}"
POSITIONS_CODE_RANGE = f"Positions!$A${POSITIONS_FIRST_ROW}:$A${POSITIONS_LAST_ROW}"
"""Ranges the formula templates splice in.

Precomputed rather than interpolated inline: a template's ``{key}`` placeholders have to survive
untouched until :func:`formula_for` resolves them, so an f-string would need every one of them
doubled. Concatenating a ready-made range reads better and cannot get that wrong.
"""


# A word on the headers, because the tidy ones were an artifact and the reason is worth keeping.
#
# They used to read `Avg AUM (GBP)`, `Net fee (GBP)`, `Tier bps` -- headers written by a
# programmer, which normalise straight onto the identifiers a conversion would pick. A workbook
# whose headers are *already* identifiers is not testing what a real conversion faces. Four years
# and three people produce `Nov-26 avg AUM`, `Fee @ tier`, `Adj'd fee (net)`, `# posns`,
# `Days b'ld`: abbreviations, apostrophes, ampersands, units in brackets, capitalisation that
# drifted when somebody added a column beside one that was already there.
#
# Two pairs **collide** under `reconcile.baseline._normalise_name`, and that is the point rather
# than an accident. `Tier (bps)` and `Tier bps` -- the rate the schedule gives and the rate
# actually applied, adjacent, named a year apart -- both normalise to `tier_bps`, as do the two
# legacy `2024 adj` columns. A naive normalisation cannot tell either pair apart, so at most one
# of each can resolve against a frame and the map has to be filled in by hand.
#
# Both pairs differ only in punctuation, which is deliberate and is not the same thing as the
# duplicate labels `findings.duplicate_headers` already catches: that one casefolds and compares,
# so it sees `2024 Adj` beside `2024 adj` and is blind to `2024 adj.` beside `2024 adj`.
# `_normalise_name` erases exactly the characters it compares on. The corpus already has a
# `duplicate_headers` fixture; what it does not have is a collision only the reconciliation map
# can feel.
#
# What this cost, measured with `infer_regions` and `compare.to_vector` semantics against a
# conversion carrying the identifiers the old headers normalised to: **41 of 45** regions
# resolved against the scaffolder's default before, **6** after -- and four of those six are the
# two collided pairs, so the six carry four distinct names between them and two of the four
# cannot say which region they mean. Four regions resolved under neither: the two whose fill
# starts a row below the header, so `_header_above` reads a formula cell rather than a label,
# and the two totals-row regions, which have no header at all.
#
# It is **not** an attempt to make proposal prediction P6 come true. P6 was measured against the
# old workbook and stands refuted there; this is the removal of an artifact that was quietly
# making the eval easier than the thing it is a model of, and the honest reading is that the old
# 91% was a measurement of the generator's own tidiness.
#
# `key` never changes with a header. It is what the formulas and :func:`compute` are written
# against, so a header rewrite cannot move a number.
WORKING_COLUMNS: tuple[Column, ...] = (
    # Data, not a formula, and that is load-bearing twice over. A person pastes the codes in --
    # and `=Positions!$A{n}` filled down would normalise to a different R1C1 string per row and
    # turn one column into eighty-four operations.
    Column("client", "Client"),
    Column(
        "client_name",
        "Name",
        "=INDEX('Entity Map'!$B:$B,MATCH({client}{row},'Entity Map'!$A:$A,0))",
    ),
    Column(
        "onboarded",
        "On b'd",
        "=INDEX('Entity Map'!$F:$F,MATCH({client}{row},'Entity Map'!$A:$A,0))",
    ),
    Column("period_start", "Bill from", "=MAX(DATE(2026,11,1),{onboarded}{row})"),
    Column("period", "Mth", '=TEXT({period_start}{row},"yyyy-mm")'),
    Column(
        "avg_aum",
        "Nov-26 avg AUM",
        "=AVERAGEIFS(Positions!$F:$F,Positions!$A:$A,{client}{row})",
    ),
    Column("prior_close", "Prior mth close", "={avg_aum}{prev}"),
    Column(
        "opening",
        "Opening bal.",
        "=IF({client}{row}={client}{prev},{prior_close}{row},{avg_aum}{row})",
    ),
    # Discrimination 1. Approximate match: the band whose floor is the largest not exceeding the
    # AUM. `join_asof`, not `join`. Its header collides with the one below under
    # `_normalise_name`, which is the collision worth having: these are the two columns a reader
    # most needs told apart.
    Column(
        "band_bps",
        "Tier (bps)",
        "=VLOOKUP({avg_aum}{row}," + BAND_RANGE + ",3,TRUE)",
    ),
    # Discrimination 8 is here rather than decorative: the negotiated rate arrives as *text*,
    # Excel coerces it in the arithmetic below, and polars will not.
    Column(
        "tier_bps",
        "Tier bps",
        "=IFERROR(VLOOKUP({client}{row}," + NEGOTIATED_RANGE + ",2,FALSE),{band_bps}{row})",
    ),
    Column(
        "legal_entity",
        "Entity",
        "=INDEX('Entity Map'!$C:$C,MATCH({client}{row},'Entity Map'!$A:$A,0))",
    ),
    Column(
        "cost_centre",
        "C/C",
        "=INDEX('Entity Map'!$D:$D,MATCH({client}{row},'Entity Map'!$A:$A,0))",
    ),
    Column("gross_fee", "Fee @ tier", "=ROUND({avg_aum}{row}*{tier_bps}{row}/10000,2)"),
    Column("days_in_month", "Days in mth", "=DAY(EOMONTH({period_start}{row},0))"),
    Column(
        "days_billed",
        "Days b'ld",
        '=DATEDIF({period_start}{row},EOMONTH({period_start}{row},0),"d")+1',
    ),
    Column(
        "prorated",
        "Fee (p/r)",
        "=ROUND({gross_fee}{row}*{days_billed}{row}/{days_in_month}{row},2)",
    ),
    Column("floored", "Fee after min", "=MAX({prorated}{row},'Fee Schedule'!$I$3)"),
    Column("capped", "Fee after max", "=MIN({floored}{row},'Fee Schedule'!$I$4)"),
    Column("discount", "Disc'nt", "=IFERROR({capped}{row}*'Fee Schedule'!$I$5,0)"),
    Column("net_fee", "Adj'd fee (net)", "=ROUND({capped}{row}-{discount}{row},2)"),
    # Discrimination 4: the workbook resolves the override silently, and a runbook must re-ask it.
    Column(
        "agreed_fee",
        "Fee agreed w/ client",
        "=IFERROR(VLOOKUP({client}{row},Overrides!$A:$C,3,FALSE),{net_fee}{row})",
    ),
    Column(
        "override_flag",
        "O/R?",
        '=IF({agreed_fee}{row}={net_fee}{row},"","OVERRIDE")',
    ),
    Column("invoice_key", "Inv. ref", '=LEFT({client}{row},3)&"-"&{period}{row}'),
    Column("positions_seen", "# posns", "=COUNTIFS(Positions!$A:$A,{client}{row})"),
    Column("running_total", "Running tot.", "=SUM(${agreed_fee}${first}:{agreed_fee}{row})"),
    Column(
        "mandate",
        "Mandate (clean)",
        "=UPPER(TRIM(INDEX('Entity Map'!$E:$E,MATCH({client}{row},'Entity Map'!$A:$A,0))))",
    ),
    Column(
        "mandate_aum",
        "AUM in mandate",
        "=SUMIFS(Positions!$F:$F,Positions!$E:$E,{mandate}{row})",
    ),
    Column("share", "% of mandate", "=IFERROR({avg_aum}{row}/{mandate_aum}{row},0)"),
    Column(
        "weighted_check",
        "AUM chk (Oct & Nov)",
        "=SUMPRODUCT(" + POSITIONS_VALUE_RANGE + ",--(" + POSITIONS_CODE_RANGE + "={client}{row}))",
    ),
    Column(
        "fee_band",
        "Band (A/B/C)",
        '=IF({agreed_fee}{row}>40000,"A",IF({agreed_fee}{row}>8000,"B","C"))',
    ),
    Column("accounts", "# a/cs", "=COUNTIF('Entity Map'!$A:$A,{client}{row})"),
    Column("latest_asof", "Data as at", "=MAX(Positions!$C:$C)"),
    Column("billing_year", "Yr", "=YEAR({period_start}{row})"),
    # The abandoned 2024 method. Real formulas, filled down like everything else, read by
    # nothing -- so they become `dead_region` findings, and they are the haystack that
    # discrimination 5 asks a conversion to find the `Post` column in. The second collision is
    # here: the raw variance and the factored one were labelled the same thing a month apart,
    # one of them with a full stop.
    Column("old_fee", "2024 fee", "=ROUND({avg_aum}{row}*" + OLD_RATE_REF + "/10000,2)"),
    Column("old_delta", "2024 adj", "={old_fee}{row}-{net_fee}{row}"),
    Column("old_adj", "2024 adj.", "={old_delta}{row}*" + OLD_FACTOR_REF),
    Column(
        "old_note",
        "2024 adj ref",
        '=CONCATENATE({client}{row}," / ",TEXT({old_adj}{row},"0.00"))',
    ),
)

LETTERS: dict[str, str] = {
    column.key: get_column_letter(index) for index, column in enumerate(WORKING_COLUMNS, start=1)
}

TOTALLED = ("gross_fee", "prorated", "net_fee", "agreed_fee")
"""The columns the workbook sums under the grid. A totals row is its own logical operation and
has no column name for `infer_regions` to match, which is why the reconciliation map cannot be
left at the scaffolder's default for it."""


def formula_for(column: Column, row: int) -> str:
    """Render one cell's formula, resolving column keys to letters."""
    return column.formula.format(
        row=row, prev=max(WORKING_FIRST_ROW, row - 1), first=WORKING_FIRST_ROW, **LETTERS
    )


# =============================================================================
# THE PYTHON MODEL
# =============================================================================
#
# A parallel implementation of the sheet, not a transcription of it. These are the values Excel
# would leave in the cells, and they are what gets injected as the cached values that make the
# workbook reconcilable at all. Written against the same column keys as the formulas above so
# the two can be read side by side.


@dataclass(frozen=True, slots=True)
class Client:
    """One client's inputs: what the extract returned and what the reference data says."""

    index: int
    code: str
    name: str
    entity: str
    cost_centre: str
    mandate_raw: str
    onboarded: dt.date
    october_aum: float
    november_aum: float

    @property
    def negotiated_bps(self) -> str | None:
        """The rate as it arrives from the onboarding team: text, because Excel copies what a
        cell *looks like* rather than what it holds."""
        if self.index % NEGOTIATED_STEP:
            return None
        return f"{20 + (self.index % 9) * 1.5:.1f}"


def build_clients(rng: random.Random) -> list[Client]:
    """The client list, seeded and deterministic."""
    clients: list[Client] = []
    for index in range(CLIENTS):
        day = ONBOARDED_MID_PERIOD.get(index)
        onboarded = dt.date(2026, 11, day) if day is not None else dt.date(2024, 3, 1 + index % 27)
        small = SMALL_ACCOUNTS.get(index)
        october = small if small is not None else round(rng.uniform(250_000, 90_000_000), 2)
        november = (
            round(small * 1.04, 2)
            if small is not None
            else round(rng.uniform(250_000, 90_000_000), 2)
        )
        clients.append(
            Client(
                index=index,
                code=f"{index + 1:05d}",
                name=CLIENT_NAMES.get(index, f"Client {index + 1:03d} Holdings"),
                entity=ENTITIES[index % 3],
                cost_centre=f"CC-{100 + (index % 6) * 10}",
                mandate_raw=MANDATES[index % 3],
                onboarded=onboarded,
                october_aum=october,
                november_aum=november,
            )
        )
    return clients


def band_bps_for(aum: float) -> float:
    """Excel's approximate ``VLOOKUP``: the last band whose floor does not exceed the value."""
    chosen = BANDS[0][2]
    for floor, _ceiling, bps in BANDS:
        if aum >= floor:
            chosen = bps
    return chosen


def compute(clients: list[Client]) -> list[dict[str, Any]]:
    """Every cell of ``Working``, in the order Excel would evaluate it.

    Money goes through :func:`generate.excel_round` rather than :func:`round`, because Excel
    collapses an operand to 15 significant digits before rounding half away from zero and the
    difference is a penny that then propagates into every total below it.
    """
    override_by_code = {code: agreed for code, agreed, _reason, _on in OVERRIDES}
    mandate_totals: dict[str, float] = {}
    for client in clients:
        key = client.mandate_raw.strip().upper()
        mandate_totals[key] = (
            mandate_totals.get(key, 0.0) + client.october_aum + client.november_aum
        )

    rows: list[dict[str, Any]] = []
    running = 0.0
    for position, client in enumerate(clients):
        avg_aum = (client.october_aum + client.november_aum) / 2
        prior_close = avg_aum if position == 0 else rows[position - 1]["avg_aum"]
        period_start = max(PERIOD_START, client.onboarded)
        band_bps = band_bps_for(avg_aum)
        negotiated = client.negotiated_bps
        # Excel's VLOOKUP returns the cell it found, and the negotiated column is *text* -- so
        # this cell holds the string '20.0', not the number 20.0. Verified by recalculating the
        # workbook in Excel: the seventeen negotiated rows come back as text and everything
        # downstream still works, because Excel coerces text in arithmetic. That coercion is
        # discrimination 8, and it has to be true of the cached values too or the baseline the
        # eval reconciles against is fiction.
        tier_bps_shown: str | float = negotiated if negotiated is not None else band_bps
        tier_bps = float(negotiated) if negotiated is not None else band_bps

        gross_fee = excel_round(avg_aum * tier_bps / 10000, 2)
        days_in_month = PERIOD_END.day
        days_billed = (PERIOD_END - period_start).days + 1
        prorated = excel_round(gross_fee * days_billed / days_in_month, 2)
        floored = max(prorated, MIN_FEE)
        capped = min(floored, MAX_FEE)
        discount = capped * DISCOUNT
        net_fee = excel_round(capped - discount, 2)
        agreed_fee = override_by_code.get(client.code, net_fee)
        running += agreed_fee

        mandate = client.mandate_raw.strip().upper()
        mandate_aum = mandate_totals[mandate]
        old_fee = excel_round(avg_aum * OLD_FLAT_BPS / 10000, 2)
        old_delta = old_fee - net_fee
        old_adj = old_delta * OLD_ADJUSTMENT

        rows.append(
            {
                "client": client.code,
                "client_name": client.name,
                "onboarded": serial(client.onboarded),
                "period_start": serial(period_start),
                "period": f"{period_start:%Y-%m}",
                "avg_aum": avg_aum,
                "prior_close": prior_close,
                "opening": prior_close if position == 0 else avg_aum,
                "band_bps": band_bps,
                "tier_bps": tier_bps_shown,
                "tier_bps_numeric": tier_bps,
                "legal_entity": client.entity,
                "cost_centre": client.cost_centre,
                "gross_fee": gross_fee,
                "days_in_month": days_in_month,
                "days_billed": days_billed,
                "prorated": prorated,
                "floored": floored,
                "capped": capped,
                "discount": discount,
                "net_fee": net_fee,
                "agreed_fee": agreed_fee,
                "override_flag": "" if agreed_fee == net_fee else "OVERRIDE",
                "invoice_key": f"{client.code[:3]}-{period_start:%Y-%m}",
                "positions_seen": 2,
                "running_total": excel_round(running, 2),
                "mandate": mandate,
                "mandate_aum": mandate_aum,
                "share": avg_aum / mandate_aum,
                "weighted_check": client.october_aum + client.november_aum,
                "fee_band": "A" if agreed_fee > 40000 else ("B" if agreed_fee > 8000 else "C"),
                "accounts": 1,
                "latest_asof": serial(dt.date(2026, 11, 28)),
                "billing_year": period_start.year,
                "old_fee": old_fee,
                "old_delta": old_delta,
                "old_adj": old_adj,
                # TEXT(x,"0.00") rounds the way Excel rounds, not the way format() does, so the
                # rounding goes through `excel_round` before it is formatted.
                "old_note": f"{client.code} / {excel_round(old_adj, 2):.2f}",
            }
        )
    return rows


def require_planted_discriminations(rows: list[dict[str, Any]]) -> None:
    """Refuse to build a workbook whose discriminations have gone.

    A seed change or a threshold tweak can remove one silently, leaving a file that opens fine
    and grades fine against everything except the thing it was built to measure.
    """
    if not any(row["days_billed"] < row["days_in_month"] for row in rows):
        msg = "no client is billed for a part month, so the pro-rating columns are a no-op"
        raise RuntimeError(msg)
    if len({row["band_bps"] for row in rows}) < 3:
        msg = "the banded lookup resolves to fewer than three bands, so it is barely a band"
        raise RuntimeError(msg)
    if not any(row["override_flag"] for row in rows):
        msg = "no override survived into the computed rows"
        raise RuntimeError(msg)
    if not any(row["floored"] > row["prorated"] for row in rows):
        msg = "the minimum fee never binds, so the floor column is dead"
        raise RuntimeError(msg)
    negotiated = sum(1 for row in rows if row["tier_bps_numeric"] != row["band_bps"])
    if not 5 <= negotiated <= len(rows) - 5:
        msg = f"{negotiated} negotiated rates: the band lookup and the override must both matter"
        raise RuntimeError(msg)
    if not any("'" in str(row["client_name"]) for row in rows):
        msg = (
            "no client name carries an apostrophe, so every literal the Post column concatenates "
            "is valid SQL and non-negotiable 3 is untested by this workbook"
        )
        raise RuntimeError(msg)


# =============================================================================
# SHEETS
# =============================================================================


def write_positions(sheet: Any, clients: list[Client]) -> None:
    """The warehouse extract: the query that produced it, then the grid it returned.

    The SQL sits in cells exactly as the original process kept it, so a conversion has something
    real to read when it decides this is a step to hand over rather than prose to summarise.

    Its header row stays machine-shaped where the hand-typed tabs do not: these are the columns
    ``EXTRACT_SQL`` selects, pasted in with the grid, and a person who renamed them would be
    breaking the paste next month.
    """
    sheet["A1"] = "Positions extract -- run this against the warehouse, paste below"
    sheet["A1"].font = Font(bold=True)
    for offset, line in enumerate(EXTRACT_SQL.splitlines()):
        cell = sheet.cell(SQL_FIRST_ROW + offset, 1, line)
        cell.font = Font(name="Consolas", size=9)
        cell.alignment = Alignment(horizontal="left")

    headers = (
        "client_code",
        "client_name",
        "as_of_date",
        "period_month",
        "mandate",
        "market_value_gbp",
    )
    for index, header in enumerate(headers, start=1):
        sheet.cell(POSITIONS_HEADER_ROW, index, header).font = Font(bold=True)

    row = POSITIONS_FIRST_ROW
    for client in clients:
        for month, day, value in (
            ("2026-10", dt.date(2026, 10, 28), client.october_aum),
            ("2026-11", dt.date(2026, 11, 28), client.november_aum),
        ):
            sheet.cell(row, 1, client.code)
            sheet.cell(row, 2, client.name)
            sheet.cell(row, 3, day)
            sheet.cell(row, 4, month)
            sheet.cell(row, 5, client.mandate_raw.strip())
            sheet.cell(row, 6, value)
            row += 1


def write_fee_schedule(sheet: Any, clients: list[Client]) -> None:
    """The file the onboarding team emailed, pasted in.

    Two preamble rows and a header on row 4 -- the shape a pasted extract actually has. Two
    tables side by side: the banded rate card that most clients are on, and the negotiated rates
    that override it for a minority. Both have to be there, or one of them is dead code.

    The negotiated rate is written as **text**, which is discrimination 8 and not decoration:
    Excel copies what a cell looks like rather than what it holds, coerces it back to a number
    in the arithmetic downstream, and polars does not.

    Machine-shaped headers, like ``Positions`` and for the same reason: this grid is pasted from
    somebody else's file every time the schedule changes, so its columns are named by whoever
    exports it rather than by whoever pastes it.
    """
    sheet["A1"] = "Fee schedule -- from Client Onboarding, effective 2026-11-01"
    sheet["A1"].font = Font(bold=True)
    sheet["A2"] = "Do not edit. Superseded schedules are kept on the shared drive."

    for column, header in (
        (1, "client_code"),
        (2, "negotiated_bps"),
        (4, "band_floor"),
        (5, "band_ceiling"),
        (6, "band_bps"),
    ):
        sheet.cell(FEE_HEADER_ROW, column, header).font = Font(bold=True)

    for offset, (floor, ceiling, bps) in enumerate(BANDS):
        row = FEE_BAND_FIRST_ROW + offset
        sheet.cell(row, 4, floor)
        sheet.cell(row, 5, ceiling)
        sheet.cell(row, 6, bps)

    row = FEE_NEGOTIATED_FIRST_ROW
    for client in clients:
        negotiated = client.negotiated_bps
        if negotiated is None:
            continue
        sheet.cell(row, 1, client.code)
        sheet.cell(row, 2, negotiated)
        row += 1

    for label_row, label, value in (
        (3, "Minimum fee", MIN_FEE),
        (4, "Maximum fee", MAX_FEE),
        (5, "Standard discount", DISCOUNT),
    ):
        sheet.cell(label_row, 8, label)
        sheet.cell(label_row, 9, value)


def write_entity_map(sheet: Any, clients: list[Client]) -> None:
    """Client to legal entity to cost centre, plus when each client came on.

    The eighth tab in the file and a dependency of the fourth, which is discrimination 10: tab
    order is the order somebody added things, never dependency order.

    A CRM export, so machine-shaped headers again -- and ``case.py`` locates its grid by looking
    for ``client_code``, which is what an export is for.
    """
    headers = ("client_code", "client_name", "legal_entity", "cost_centre", "mandate", "onboarded")
    for index, header in enumerate(headers, start=1):
        sheet.cell(1, index, header).font = Font(bold=True)
    for offset, client in enumerate(clients):
        row = 2 + offset
        sheet.cell(row, 1, client.code)
        sheet.cell(row, 2, client.name)
        sheet.cell(row, 3, client.entity)
        sheet.cell(row, 4, client.cost_centre)
        sheet.cell(row, 5, client.mandate_raw)
        sheet.cell(row, 6, client.onboarded)


def write_working(sheet: Any, clients: list[Client]) -> None:
    """The wide one: parameters, then a column per formula shape, filled down.

    This is where the workbook's complexity lives. See ``WORKING_COLUMNS``.
    """
    sheet["A1"] = "Management fee working -- November 2026"
    sheet["A1"].font = Font(bold=True, size=12)
    sheet["A3"] = "Period"
    sheet["B3"] = "2026-11"
    sheet["A4"] = "Prepared by"
    sheet["B4"] = "Billing Operations"
    sheet["A14"] = "2024 flat rate (bps)"
    sheet[OLD_RATE_CELL] = OLD_FLAT_BPS
    sheet["A15"] = "2024 adjustment factor"
    sheet[OLD_FACTOR_CELL] = OLD_ADJUSTMENT

    for index, column in enumerate(WORKING_COLUMNS, start=1):
        sheet.cell(WORKING_HEADER_ROW, index, column.header).font = Font(bold=True)

    for offset, client in enumerate(clients):
        row = WORKING_FIRST_ROW + offset
        for index, column in enumerate(WORKING_COLUMNS, start=1):
            if not column.formula:
                sheet.cell(row, index, client.code)
                continue
            sheet.cell(row, index, formula_for(column, row))

    sheet.cell(WORKING_TOTAL_ROW, 1, "TOTAL").font = Font(bold=True)
    for key in TOTALLED:
        letter = LETTERS[key]
        sheet.cell(
            WORKING_TOTAL_ROW,
            get_column_letter_index(letter),
            f"=SUM({letter}{WORKING_FIRST_ROW}:{letter}{WORKING_LAST_ROW})",
        )


def get_column_letter_index(letter: str) -> int:
    """Column letter to 1-based index. openpyxl has the inverse; this direction is trivial."""
    index = 0
    for character in letter:
        index = index * 26 + (ord(character) - ord("A") + 1)
    return index


def write_overrides(sheet: Any) -> None:
    """Three computed fees the billing manager typed over, each with a reason.

    Headers as the billing manager typed them, for the reason recorded above
    :data:`WORKING_COLUMNS`. This tab is nobody's extract: it is a person keeping a note.
    """
    headers = ("Client", "Calc'd fee", "Agreed fee (GBP)", "Reason", "Agreed on", "Agreed by")
    for index, header in enumerate(headers, start=1):
        sheet.cell(1, index, header).font = Font(bold=True)
    net = LETTERS["net_fee"]
    client = LETTERS["client"]
    for offset, (code, agreed, reason, agreed_on) in enumerate(OVERRIDES):
        row = 2 + offset
        sheet.cell(row, 1, code)
        sheet.cell(
            row,
            2,
            f"=INDEX(Working!${net}:${net},MATCH($A{row},Working!${client}:${client},0))",
        )
        sheet.cell(row, 3, agreed)
        sheet.cell(row, 4, reason)
        sheet.cell(row, 5, agreed_on)
        sheet.cell(row, 6, "Billing Operations")


def allocation_rows(clients: list[Client]) -> list[tuple[int, Client | None, str]]:
    """The Allocation grid's layout: client rows in entity blocks, each block subtotalled.

    Returns ``(row, client, entity)`` with ``client`` None on a subtotal row. Shared by the
    writer and the cached-value model so the two cannot drift.
    """
    layout: list[tuple[int, Client | None, str]] = []
    row = 2
    for entity in ENTITIES:
        for client in clients:
            if client.entity != entity:
                continue
            layout.append((row, client, entity))
            row += 1
        layout.append((row, None, entity))
        row += 1
    return layout


def write_allocation(sheet: Any, clients: list[Client]) -> None:
    """Fees aggregated to invoice grain, with a subtotal row per entity *inside* the grid.

    Discrimination 6. An embedded subtotal is not data, and a conversion that reads the grid flat
    double-counts every entity. ``SUBTOTAL`` rather than ``SUM`` because that is what the toolbar
    button produces, and because it excludes itself.

    The headers are the ledger's: this grid is shaped for ``fin.fee_invoice`` and the ``Post``
    statements name the same three columns. They are also the ``Summary`` pivot's field names,
    since a pivot takes its fields from its source header row -- so ``build_pivot.ROW_FIELDS``
    is these cells, and renaming them here renames them there.
    """
    headers = ("client_code", "legal_entity", "cost_centre", "period", "fee_gbp")
    for index, header in enumerate(headers, start=1):
        sheet.cell(1, index, header).font = Font(bold=True)

    agreed = LETTERS["agreed_fee"]
    client_letter = LETTERS["client"]
    block_start = 2
    for row, client, entity in allocation_rows(clients):
        if client is None:
            sheet.cell(row, 2, f"{entity} subtotal").font = Font(bold=True)
            sheet.cell(row, 5, f"=SUBTOTAL(9,E{block_start}:E{row - 1})")
            block_start = row + 1
            continue
        sheet.cell(row, 1, client.code)
        sheet.cell(row, 2, entity)
        sheet.cell(row, 3, client.cost_centre)
        sheet.cell(row, 4, "2026-11")
        # Aggregated by key rather than by position: `=Working!$S{n}` would point at a different
        # row from each Allocation row and so refuse to compress, and a lookup is what an
        # allocation tab does anyway.
        sheet.cell(
            row,
            5,
            f"=SUMIFS(Working!${agreed}:${agreed},"
            f"Working!${client_letter}:${client_letter},$A{row})",
        )


def write_post(sheet: Any, clients: list[Client]) -> None:
    """One INSERT per invoice, built by concatenation. The manual carry.

    Nothing in the workbook reads this column, so it is a ``dead_region`` with fan-out zero --
    which is exactly what sorts it into the tail of the planner's context. It is also the step
    that posts the fees, so dropping it deletes the point of the process.

    **The statement carries the client's name, and that is what makes ``&`` wrong here rather
    than merely unwise.** See :data:`CLIENT_NAMES`: one of the eighty-four names contains an
    apostrophe, so this formula renders ``'O'Hanlon & Reid Nominees'`` -- seven apostrophes in
    the statement where a valid one has six, and a syntax error at the moment somebody pastes it
    into a production client. ``&`` quotes nothing and escapes nothing; :func:`kedge.sql.literal`
    doubles the apostrophe (``'O''Hanlon & Reid Nominees'``), which is the whole difference, so a
    conversion has to render the literal rather than reproduce the workbook's text.
    """
    sheet["A1"] = "Run these against the ledger once the fee run is approved."
    sheet["A1"].font = Font(bold=True)
    sheet["A3"] = "SQL to run"
    sheet["A3"].font = Font(bold=True)
    client = LETTERS["client"]
    period = LETTERS["period"]
    agreed = LETTERS["agreed_fee"]
    for offset in range(len(clients)):
        working_row = WORKING_FIRST_ROW + offset
        # The name is read off the CRM export by position, not out of `Working!B`, which holds
        # the same string one lookup away. That is deliberate. Nothing in this workbook reads
        # `Working!B`, so it is one of the nineteen `dead_region` findings that discrimination 5
        # hides the manual carry among -- and pointing this column at it would take the haystack
        # down to eighteen and the complexity with it. `Entity Map` is already read (by
        # `Working!B` itself), so reaching past the display column costs the measured structure
        # nothing. The offset is constant, the way `Working!${client}` already is, so the column
        # still normalises to one R1C1 string and stays one logical operation.
        sheet.cell(
            4 + offset,
            1,
            '="INSERT INTO fin.fee_invoice '
            "(client_code, client_name, period_month, fee_gbp) VALUES ('\""
            f"&Working!${client}{working_row}"
            "&\"', '\""
            f"&'Entity Map'!$B{2 + offset}"
            "&\"', '\""
            f"&Working!${period}{working_row}"
            '&"\', "'
            f"&Working!${agreed}{working_row}"
            '&");"',
        )


def write_recon(sheet: Any) -> None:
    """Last month beside this month, with a variance column and typed commentary.

    Hand-built, so hand-typed headers -- and the two month columns are dated rather than named,
    which is how a tab that is copied forward every month actually reads.
    """
    headers = (
        "Entity",
        "Oct-26 (GBP)",
        "Nov-26 (GBP)",
        "Var (GBP)",
        "Commentary",
        "Var %",
        "Flag",
    )
    for index, header in enumerate(headers, start=1):
        sheet.cell(1, index, header).font = Font(bold=True)
    priors = (412_880.44, 388_100.02, 205_990.18)
    for offset, (entity, prior) in enumerate(zip(ENTITIES, priors, strict=True)):
        row = 2 + offset
        sheet.cell(row, 1, entity)
        sheet.cell(row, 2, prior)
        sheet.cell(row, 3, f"=SUMIF(Allocation!$B:$B,$A{row},Allocation!$E:$E)")
        sheet.cell(row, 4, f"=C{row}-B{row}")
        sheet.cell(row, 6, f"=IFERROR(D{row}/B{row},0)")
        sheet.cell(row, 7, f'=IF(ABS(F{row})>0.05,"REVIEW","")')
    sheet["E2"] = "Transfer in from the Geneva book; expected."
    sheet["E3"] = "Two mandates closed in October."
    sheet["E4"] = "Unexplained. Raised with the desk 2026-11-04."


def write_signoff(sheet: Any) -> None:
    """Purpose, background, known issues. The part nobody can reconstruct.

    ``WorkbookAnalysis.notes`` picks this up and the plan's ``Briefing`` is where it goes, prose
    citing its sources. Written rather than generated, because invented background in a finance
    notebook is confident, plausible and unattributable.

    The "known issues" entry is load-bearing for discrimination 9: the staleness of Allocation is
    recorded *here and nowhere else*, because no analyser can detect it from the cells.
    """
    sheet["A1"] = "Management fee billing run -- sign-off"
    sheet["A1"].font = Font(bold=True, size=12)

    entries = (
        (
            "Purpose",
            "Calculate the management fee owed by each client for the month, apply the agreed "
            "floors, caps and discounts, allocate the result to the legal entity and cost "
            "centre that will carry it, and post the invoices to the ledger.",
        ),
        (
            "Background",
            "The tiered schedule replaced a flat 26bps rate in January 2024. The old columns "
            "are still in the Working tab because the 2024 comparatives were needed for the "
            "audit; nothing reads them now.",
        ),
        (
            "Known issues",
            "The Allocation tab is left on manual calculation because recalculating it over a "
            "full extract takes several minutes, so its figures are from before the overrides "
            "were agreed and do not match the Working tab. Recalculate before relying on them. "
            "Three clients have agreed fees that override the calculation -- see the Overrides "
            "tab, and confirm each one still applies before posting.",
        ),
        (
            "Cadence",
            "Monthly, on the third working day, after the position extract has been signed off "
            "by Middle Office.",
        ),
    )
    row = 3
    for heading, text in entries:
        sheet.cell(row, 1, heading).font = Font(bold=True)
        cell = sheet.cell(row + 1, 1, text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 3


def write_summary(sheet: Any) -> None:
    """The tab the manager reads: a title, and the pivot that Excel puts underneath it.

    Discrimination 7. The pivot itself cannot be written here -- openpyxl reads one and cannot
    author one -- so ``build_pivot.add_pivot`` finishes this tab by driving Excel, behind
    ``--with-pivot``. Without that pass the tab is a title and nothing else, which is honest:
    :func:`build` is what CI can reproduce, and a pivot is not.

    The title occupies row 1 and the pivot is anchored at ``A3``, so the two do not collide.
    """
    sheet["A1"] = "Fee summary by legal entity and cost centre -- November 2026"
    sheet["A1"].font = Font(bold=True, size=12)


# =============================================================================
# CACHED VALUES
# =============================================================================


def _as_excel_text(value: float) -> str:
    """A number as Excel's ``&`` operator renders it into a string.

    Excel's general number format drops a trailing ``.0``: concatenating the fee 18500 yields
    ``18500`` where Python's f-string yields ``18500.0``. The difference is invisible until the
    generated statement is compared against the workbook's own cached text, and then it is every
    row whose fee happens to be integral. Found by recalculating this workbook in Excel and
    diffing all 5,254 cached values, which is the only way anybody was going to notice.
    """
    return repr(int(value)) if float(value).is_integer() else repr(value)


def cached_values(clients: list[Client], rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """What Excel would have left in every formula cell.

    Without these the workbook has no reconciliation baseline at all: triage reports a blocker,
    ``infer_regions`` proposes nothing, and every region degrades to "not reconciled" -- which
    makes the whole reconciliation half of the rubric ungradeable.

    ``Allocation`` is the deliberate exception, and it is discrimination 9. Its values are real
    but **stale**: they are what the tab held before the three overrides were agreed, because the
    sheet is left on manual calculation. A conversion that reconciles against them and "fixes"
    itself until they match has adopted numbers Excel itself would disown; the correct answer is
    to declare the region not reproduced, with the reason, and report CHECKED WITH EXCEPTIONS.
    """
    working: dict[str, Any] = {}
    for offset, row in enumerate(rows):
        excel_row = WORKING_FIRST_ROW + offset
        for column in WORKING_COLUMNS:
            if not column.formula:
                continue
            working[f"{LETTERS[column.key]}{excel_row}"] = row[column.key]
    for key in TOTALLED:
        working[f"{LETTERS[key]}{WORKING_TOTAL_ROW}"] = excel_round(
            sum(row[key] for row in rows), 2
        )

    by_code = {row["client"]: row for row in rows}
    overrides = {
        f"B{2 + offset}": by_code[code]["net_fee"]
        for offset, (code, _agreed, _reason, _on) in enumerate(OVERRIDES)
    }

    # The stale allocation: pre-override fees, which is what "left on manual calculation" means.
    allocation: dict[str, Any] = {}
    block_total = 0.0
    for row_number, client, _entity in allocation_rows(clients):
        if client is None:
            allocation[f"E{row_number}"] = excel_round(block_total, 2)
            block_total = 0.0
            continue
        stale = by_code[client.code]["net_fee"]
        allocation[f"E{row_number}"] = stale
        block_total += stale

    recon: dict[str, Any] = {}
    priors = (412_880.44, 388_100.02, 205_990.18)
    for offset, (entity, prior) in enumerate(zip(ENTITIES, priors, strict=True)):
        row_number = 2 + offset
        november = excel_round(
            sum(
                by_code[client.code]["net_fee"]
                for _r, client, ent in allocation_rows(clients)
                if client is not None and ent == entity
            ),
            2,
        )
        variance = november - prior
        recon[f"C{row_number}"] = november
        recon[f"D{row_number}"] = variance
        recon[f"F{row_number}"] = variance / prior
        recon[f"G{row_number}"] = "REVIEW" if abs(variance / prior) > 0.05 else ""

    # Excel's own text, apostrophe and all. For `00041` this is not valid SQL and the rot guard
    # asserts that it is not: reproducing it faithfully reproduces a bug, and a conversion that
    # renders the literal through `kedge.sql` has to be distinguishable from one that copied.
    post = {
        f"A{4 + offset}": (
            "INSERT INTO fin.fee_invoice "
            "(client_code, client_name, period_month, fee_gbp) VALUES "
            f"('{row['client']}', '{row['client_name']}', '{row['period']}', "
            f"{_as_excel_text(row['agreed_fee'])});"
        )
        for offset, row in enumerate(rows)
    }

    return {
        "Working": working,
        "Overrides": overrides,
        "Allocation": allocation,
        "Recon": recon,
        "Post": post,
    }


# =============================================================================
# BUILD
# =============================================================================


def build(path: Path) -> None:
    """Write the workbook, cached values included. Deterministic: nothing reads the clock."""
    rng = random.Random(SEED)
    clients = build_clients(rng)
    rows = compute(clients)
    require_planted_discriminations(rows)

    wb = Workbook()
    # Tab order is the order somebody added things over four years, which is discrimination 10.
    positions = wb.active
    positions.title = "Positions"
    write_positions(positions, clients)
    write_fee_schedule(wb.create_sheet("Fee Schedule"), clients)
    write_working(wb.create_sheet("Working"), clients)
    write_overrides(wb.create_sheet("Overrides"))
    write_allocation(wb.create_sheet("Allocation"), clients)
    write_summary(wb.create_sheet("Summary"))
    write_post(wb.create_sheet("Post"), clients)
    write_entity_map(wb.create_sheet("Entity Map"), clients)
    write_recon(wb.create_sheet("Recon"))
    write_signoff(wb.create_sheet("Sign-off"))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

    parts = read_parts(path)
    populated = inject_cached_values(parts, cached_values(clients, rows))
    # False, so Excel trusts what is cached rather than rebuilding it on open. That is what makes
    # the stale Allocation figures survive to be found.
    set_full_calc_on_load(parts, False)
    attach_connections(parts, _connections_xml())
    write_parts(path, parts)
    normalise(path)

    total = sum(populated.values())
    if total < 3_000:
        msg = f"only {total} cached values landed; the workbook will not reconcile"
        raise RuntimeError(msg)


def _connections_xml() -> bytes:
    """The positions query as Excel stores it: newlines as character references.

    XML attribute-value normalisation turns a literal newline inside an attribute into a space,
    so multi-line SQL has to be written as ``&#10;`` -- which is what Excel does, and what
    :mod:`kedge.analysis.connections` is built to read back. ``xml_attr`` does that encoding;
    open-coding the escape here is how the ``Positions`` query would come back as one long line.

    ``refreshOnLoad`` is deliberately absent. This part exists to be *read*, and a workbook that
    asks Excel to go to the warehouse the moment it opens would make ``--verify-with-excel`` a
    network operation against a DSN nobody has.
    """
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<connections xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<connection id="1" name="PositionsExtract" type="1" refreshedVersion="8"'
        ' minRefreshableVersion="3" background="1" saveData="1"'
        ' description="Client positions at day-end, current month and prior">'
        f'<dbPr connection="{xml_attr(EXTRACT_CONNECTION_STRING)}"'
        f' command="{xml_attr(EXTRACT_SQL)}" commandType="2"/>'
        "</connection>"
        "</connections>"
    ).encode()


def restore_connection(path: Path) -> str:
    """Put ``xl/connections.xml`` back the way :func:`build` wrote it, after Excel re-saved it.

    **Driving Excel found a third thing, and this is it.** Excel rewrites the connection part on
    every save, and two things do not survive. The newlines inside ``command`` come back as
    ``_x000a_`` -- Excel's escaped-character convention, *not* the ``&#10;`` character reference
    the part is authored with and not what ``adjustment_signoff``'s ``_connections_xml`` used to
    say Excel does -- and ``commandType="2"`` is dropped outright, because ``2`` is the
    attribute's schema default.

    ``kedge.analysis.connections`` decoded neither, so the extract query read back as one
    unbroken line with ``_x000a_`` littered through it and no command type at all -- which is
    what ``build_proposal_context`` would have put in front of the planner as the query to hand
    over. This eval was the first thing in the repository to reach that: every other connection
    part here is hand-authored and has never been near Excel. **Both are fixed in the reader
    now**, which decodes the general ``_xHHHH_`` escape and reads an absent ``commandType`` as
    its default, and the verbatim bytes Excel wrote are pinned in
    ``tests/unit/test_connections.py`` so the fix regresses on every platform CI runs rather
    than only where Excel is installed.

    So this function is no longer load-bearing, and it stays for the reason ``build_pivot``
    restores ``<calcPr>``: the committed artifact should be a file this repository authored, not
    one carrying claims Excel made on its way past. Keeping it also keeps the part identical
    whichever build path last ran -- ``--with-pivot`` on Windows, or the pure-Python ``build``
    CI reproduces -- and the ``&#10;`` form is the readable one for anyone who opens the part.
    Nothing depends on it any more: drop it the day this eval wants a workbook that is Excel's
    artifact end to end, and the reader will take it.
    """
    parts = read_parts(path)
    if "xl/connections.xml" in parts:
        # The bytes only. Excel's save already carries the content-type override and the
        # workbook relationship, and `attach_connections` would append a second one.
        parts["xl/connections.xml"] = _connections_xml()
    else:
        attach_connections(parts, _connections_xml())
    write_parts(path, parts)
    return (
        "connection       restored to the authored form; Excel re-encodes the query's "
        "newlines as _x000a_ on save and drops commandType"
    )


def carries_a_pivot(path: Path) -> bool:
    """Whether the file on disk already holds a pivot table part.

    :func:`build` cannot author one -- openpyxl reads a pivot table and will not write one -- so
    a pure-Python rebuild over the committed workbook *deletes* the ``Summary`` tab's pivot and
    with it discrimination 7. That used to happen silently, and to happen on the two flags whose
    whole job is to measure the artifact rather than change it. See :func:`main`.
    """
    return any(name.startswith("xl/pivotTables/") for name in read_parts(path))


# =============================================================================
# THE EXCEL ORACLE
# =============================================================================


STALE_BY_DESIGN: tuple[str, ...] = (
    "Allocation",
    "Recon",
)
"""The only sheets whose cached values Excel is allowed to disagree with.

``Allocation`` holds the figures from before the three overrides were agreed, because the tab is
left on manual calculation -- discrimination 9. ``Recon`` reads ``Allocation``, so the staleness
propagates there and it would be wrong if it did not. Everywhere else a disagreement means the
Python model in :func:`compute` has drifted from what Excel actually does, and the baseline the
eval reconciles against is fiction.
"""


def verify_with_excel(path: Path) -> int:
    """Recalculate a copy in Excel and report every cached value that moves.

    The cached values are a *parallel implementation* of the sheet, so nothing but Excel can say
    whether they are right. Running this found two real defects that no amount of reading the
    code would have: ``VLOOKUP`` over a text-formatted column returns **text**, so seventeen
    ``tier_bps`` cells held a number where Excel holds ``'20.0'``; and Excel's ``&`` renders an
    integral number without its trailing ``.0``, so every generated statement whose fee happened
    to be a round number disagreed with the workbook's own text.

    Windows only, and opt-in: it needs Excel and pywin32. It works on a **copy**, because
    recalculating the committed workbook is precisely what would destroy the deliberate
    staleness it exists to protect.
    """
    try:
        import win32com.client
    except ImportError:
        logger.warning("pywin32 is not installed, so the Excel oracle cannot run")
        return 0

    with tempfile.TemporaryDirectory() as directory:
        scratch = Path(directory) / path.name
        shutil.copy2(path, scratch)
        before = _snapshot(scratch)

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        book = None
        try:
            book = excel.Workbooks.Open(str(scratch))
            excel.CalculateFullRebuild()
            book.Save()
        finally:
            if book is not None:
                book.Close(SaveChanges=False)
            excel.Quit()

        after = _snapshot(scratch)

    moved = [
        (sheet, ref, old, after.get((sheet, ref)))
        for (sheet, ref), old in before.items()
        if not _agrees(old, after.get((sheet, ref)))
    ]
    unexpected = [item for item in moved if item[0] not in STALE_BY_DESIGN]

    print(f"cells compared            {len(before):>6}")
    print(f"moved on recalculation    {len(moved):>6}")
    print(f"  of those, by design     {len(moved) - len(unexpected):>6}   {STALE_BY_DESIGN}")
    print(f"  unexplained             {len(unexpected):>6}")
    for sheet, ref, old, new in unexpected[:20]:
        print(f"    {sheet}!{ref}: {old!r} -> {new!r}")

    if unexpected:
        print()
        print("FAIL  the Python model disagrees with Excel outside the stale sheets")
        return 1
    if not moved:
        print()
        print("FAIL  nothing moved at all, so the deliberate staleness has gone")
        return 1
    print()
    print("OK    every disagreement is the planted staleness and nothing else")
    return 0


def _snapshot(path: Path) -> dict[tuple[str, str], Any]:
    """Every non-empty cell value in the workbook, keyed by sheet and reference."""
    book = load_workbook(path, data_only=True)
    return {
        (name, cell.coordinate): cell.value
        for name in book.sheetnames
        for row in book[name].iter_rows()
        for cell in row
        if cell.value is not None
    }


def _agrees(old: Any, new: Any) -> bool:
    """Whether two cached values match, at the half-penny that matters rather than exactly."""
    if isinstance(old, float) and isinstance(new, float):
        return abs(old - new) < 0.005
    return bool(old == new)


def calibrate(path: Path) -> int:
    """Measure the built workbook against the structural band, and say which way it missed."""
    from kedge.analysis.analyse import analyse
    from kedge.analysis.workbook import open_workbook
    from kedge.plan.propose import build_proposal_context
    from kedge.plan.triage import complexity, triage
    from kedge.reconcile.baseline import infer_regions

    analysis = analyse(path)
    result = triage(analysis)
    score = complexity(analysis)
    patterns = sorted({op.excel_pattern.value for op in analysis.operations})
    cross_sheet = sum(1 for op in analysis.operations if any(r.sheet for r in op.references))
    dead = [f for f in analysis.findings if f.kind.value == "dead_region"]
    reconcilable = [op for op in analysis.operations if op.cached_values_present]
    context = build_proposal_context(analysis, result)
    ranked = sorted(analysis.operations, key=lambda op: (-op.downstream_ref_count, op.id))
    carry = next((op for op in analysis.operations if op.sheet == "Post"), None)
    with open_workbook(path) as handle:
        regions = infer_regions(handle, analysis)

    print(f"workbook            {path.name}  ({path.stat().st_size:,} bytes)")
    print(f"operations          {len(analysis.operations):>6}   target {TARGET_OPERATIONS}")
    print(f"distinct patterns   {len(patterns):>6}   target >= {TARGET_PATTERNS}")
    print(f"complexity          {score:>6.3f}   target {TARGET_COMPLEXITY}")
    print(f"verdict             {result.verdict.value:>6}   must not be {FORBIDDEN_VERDICT!r}")
    print(f"convertible         {result.convertible:>6.2f}")
    print(f"reconcilable ops    {len(reconcilable):>6}   of {len(analysis.operations)}")
    print(f"inferred regions    {len(regions):>6}")
    print(f"sheets              {len(analysis.sheets):>6}")
    print(f"cross-sheet ops     {cross_sheet:>6}")
    print(f"findings            {len(analysis.findings):>6}")
    print(f"dead regions        {len(dead):>6}")
    print(f"dependency edges    {len(analysis.graph.edges):>6}")
    print(f"column profiles     {len(analysis.profiles):>6}")
    print(f"patterns            {patterns}")
    if carry is not None:
        position = [op.id for op in ranked].index(carry.id) + 1
        print(
            f"manual carry        ranks {position} of {len(ranked)}  (fan-out {carry.downstream_ref_count})"
        )
    else:
        print("manual carry        NOT FOUND -- the Post column did not become an operation")
    print(f"digest truncation   {[k for k in context if k.endswith('_omitted')] or 'none'}")
    for blocker in result.blockers:
        print(f"  blocker: {blocker.render()}")

    problems: list[str] = []
    if not TARGET_OPERATIONS[0] <= len(analysis.operations) <= TARGET_OPERATIONS[1]:
        problems.append(f"operations {len(analysis.operations)} outside {TARGET_OPERATIONS}")
    if len(patterns) < TARGET_PATTERNS:
        problems.append(f"only {len(patterns)} distinct patterns, want >= {TARGET_PATTERNS}")
    if not TARGET_COMPLEXITY[0] <= score <= TARGET_COMPLEXITY[1]:
        problems.append(f"complexity {score:.3f} outside {TARGET_COMPLEXITY}")
    if result.verdict.value == FORBIDDEN_VERDICT:
        problems.append("triage says stop, so the workbook cannot be converted or graded")
    if not regions:
        problems.append("no reconcilable regions, so the reconciliation rubric is ungradeable")

    print()
    if problems:
        for problem in problems:
            print(f"MISS  {problem}")
        return 1
    print("IN BAND")
    return 0


def main() -> int:
    """Build, or measure, but never both by accident.

    **Measuring the workbook must not damage it, and both flags used to.** ``--calibrate`` and
    ``--verify-with-excel`` each began by calling :func:`build`, which rewrites the file without
    the ``Summary`` pivot -- so asking either question destroyed discrimination 7, silently, and
    the answer they printed was about a workbook that no longer matched the committed one. That
    is the sharper half: the file on disk *is* the artifact the eval grades, it carries a part no
    pure-Python build can reproduce, and a rebuilt workbook is a different input. Measuring
    something else and calling it the measurement is worse than refusing to measure.

    So the two questions are separated. A measuring run reads the file as it stands and writes
    nothing. A building run refuses to overwrite a workbook that already carries a pivot unless
    it is going to put the pivot back (``--with-pivot``) or the caller says ``--force`` -- which
    is the escape hatch for regenerating on a machine with no Excel, where the loss is real,
    intended and now stated out loud rather than discovered a week later.
    """
    parser = argparse.ArgumentParser(description="Build the fee-billing-run eval workbook.")
    parser.add_argument(
        "--calibrate",
        action="store_true",
        help="Measure the workbook on disk against the structural band. Writes nothing.",
    )
    parser.add_argument(
        "--verify-with-excel",
        action="store_true",
        help="Recalculate a copy of the workbook on disk in Excel and check the cached values. "
        "Windows, needs pywin32. Writes nothing.",
    )
    parser.add_argument(
        "--with-pivot",
        action="store_true",
        help=(
            "Rebuild, then finish the Summary tab by driving Excel over COM. Needs Windows, "
            "Excel and pywin32, and produces a file this repository cannot reproduce byte for "
            "byte."
        ),
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Rebuild even though it drops the Summary pivot the file already carries.",
    )
    args = parser.parse_args()

    path = Path(__file__).resolve().parent / WORKBOOK_NAME
    measuring = args.calibrate or args.verify_with_excel
    building = args.with_pivot or not measuring

    if building:
        if path.exists() and carries_a_pivot(path) and not (args.with_pivot or args.force):
            print(f"REFUSED  {path.name} already carries a Summary pivot, and a pure-Python")
            print("         rebuild would delete it -- openpyxl reads a pivot table and cannot")
            print("         author one. Discrimination 7 is that pivot.")
            print("         Rebuild it with `--with-pivot` (Windows, Excel, pywin32), or pass")
            print("         `--force` to rebuild without it and restore the file from git after.")
            return 2
        build(path)
        print(f"wrote {path}")
        if args.with_pivot:
            # Imported here rather than at the top so that `build` -- the pure-Python path CI
            # runs -- does not depend on a module whose whole purpose is to start Excel.
            from build_pivot import add_pivot

            print()
            print(add_pivot(path).render())
            print(restore_connection(path))
    elif not path.exists():
        print(f"MISSING  {path.name} is not there to measure. Build it first.")
        return 2
    else:
        print(f"measuring {path.name} as it stands; nothing was rebuilt")

    status = 0
    if args.calibrate:
        print()
        status |= calibrate(path)
    if args.verify_with_excel:
        print()
        status |= verify_with_excel(path)
    return status


if __name__ == "__main__":
    sys.exit(main())
