"""Grading the fee-billing eval: what to feed the runbook, and what to check when it stops.

Every grader here is keyed to an id in ``expected.yaml``, and the contract tests assert the two
sets are identical in both directions. That link is the whole reason a rubric written in prose is
worth having: the prose says what is being checked and why, the code says how, and neither can
drift without the other going red.

Four rules govern this module, and three of them are the difference between an eval and a number.

**A grader must never raise.** ``harness.grade._tier`` catches an exception out of a grader and
records the item as ``SKIP`` -- and :attr:`~harness.model.ItemResult.available` is zero for a
skip, so the points leave the denominator entirely. A bare ``assert`` inside a grader therefore
does not fail the item, it *deletes* it, and the report comes back saying everything it could
measure passed. That is the reconciliation sin in a different costume: a result nobody took,
presented as a result. Every grader below returns an explicit :class:`ItemResult` on every path
and is wrapped in :func:`_graded`, which turns anything that escapes into a ``FAIL`` naming the
exception. ``SKIP`` is reserved for genuinely ungradeable -- the notebook stopped before the cell
this item is about, an optional dependency is absent, no plan was supplied, or the thing the item
is about does not exist in kedge yet.

The other half of that rule is newer and cost thirty-three of forty-nine points to learn. **An
omission is not an ambiguity.** ``_graded`` only catches the accidental case, a grader that
raises; the systematic one is a grader that *chooses* not to grade, and a notebook binding
nothing at all used to leave two thirds of the deterministic tier out of the denominator, one
``_skip`` at a time, each with a sentence explaining that this was probably a naming difference.
It was not. ``reconciliation``, ``reconciliation_values`` and ``kedge_briefing`` are names the
scaffolder fixes, so a completed run without one has not named them differently; a frame keyed by
a client code is now found by content as well as by name, so a completed run without one is not a
column-naming difference either. Both are failures now. And ``_pass()`` must never be reachable
with nothing checked -- three graders used to print a pass whose own detail described work they
had not done, which is the reconciliation sin wearing a different hat.

**Every expected figure comes from the rubric's ``facts``, never from a literal here.** The rot
guard recomputes each of them from ``build_workbook.compute()``, so a rubric quoting a figure the
workbook does not produce goes red in the test rather than by failing a correct conversion. Where
a grader needs more than a figure -- the seventeen negotiated rates, say -- it reads them out of
the workbook, which is the same artifact the hand-ins are derived from, rather than transcribing
them.

**Money is compared at half a penny, never with ``==``.** polars' vectorised execution of the
rounding chain lands a few parts in 1e11 away from the scalar path, which is far inside the penny
that matters and far outside float equality. :data:`PENNY` is the threshold.

**The hand-ins are derived from the workbook, never committed beside it.** ``Positions``,
``Fee Schedule`` and ``Entity Map`` are already the three grids a person brings to this process,
so the harness reads them out and writes them to a temporary directory. Committing copies would
mean two sources of truth for the same 84 clients and one of them going stale.

A note on how a grader finds anything at all. This case has no reference conversion yet, so no
grader may assume a variable name beyond the handful the scaffolder itself fixes
(``reconciliation``, ``reconciliation_values``, ``kedge_briefing``). Everything else is found by
*shape*: :func:`_grids` returns every frame the notebook bound that is keyed by a client code,
billing grain first and widest first, and :func:`_find` looks for a column across them by a list
of names a translation might plausibly have chosen. When no frame is found at all the item skips,
because that is a difference in how the notebook is written rather than in what it computes. When
a frame is found and the column this item is about is not on it, the item **fails** and lists what
the frame does carry -- the alternative is a rubric that quietly stops measuring the one thing it
exists to measure the moment a conversion drops it.
"""

from __future__ import annotations

import csv
import datetime as dt
import functools
import io
import logging
import random
import re
import tempfile
from collections import Counter
from collections.abc import Mapping
from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING, Any

from harness.model import ItemResult, Outcome
from openpyxl import load_workbook

if TYPE_CHECKING:
    from collections.abc import Callable, Iterable

    from harness.drive import NotebookRun

logger = logging.getLogger(__name__)

# =============================================================================
# WHAT THE HARNESS READS
# =============================================================================

CASE_DIR = Path(__file__).resolve().parent
WORKBOOK = CASE_DIR / "m11_management_fee_run.xlsx"
REFERENCE_NOTEBOOK = CASE_DIR / "notebook.py"
"""The gold-answer conversion. Written last, on purpose, and it scores 47 of 47.

The proposal phases it after the graders, inverting the order the first eval was built in:
writing the reference first is what let that case's graders be shaped around a hand-written
answer, and five real defects hid behind it until ``--convert`` existed. Every grader here was
therefore written against the workbook and the rubric, with nothing to fit itself to, and the
reference had to satisfy them afterwards. ``--notebook <path>`` and ``--convert <model>`` remain
the other two entry points. Its *parent directory* is also where ``--convert`` looks for the
default ``plan.yaml``.

A full score is what the reference is for and not evidence that the graders are sound: it is one
correct conversion, and an item can pay out on it while paying out on a wrong one too. The
scenario harness the adversarial review left behind is what checks the other direction.
"""

RUBRIC = CASE_DIR / "expected.yaml"

PENNY = 0.005
"""Half a penny: the tolerance every money comparison here uses. See the module docstring."""


# =============================================================================
# THE HAND-INS
# =============================================================================

HANDINS: tuple[str, ...] = ("positions", "fee_schedule", "entity_map")
"""The inbound artifacts, in the order the process asks for them.

Read by ``harness.grade.placeholder_handins`` so ``--convert`` can learn the script's keys without
writing any files, and these are the *keyword parameter names* of :func:`script_for` rather than
widget names -- the widget names are the keys of the dict it returns.

Three rather than two, and the count is a judgement. ``Positions`` and ``Fee Schedule`` are
unarguable: each carries its own provenance in its first row, one from the warehouse and one from
Client Onboarding. ``Entity Map`` is the arguable one -- it is reference data maintained inside
the workbook, added in year three, with no query above it -- but nothing can derive it, and a
conversion that has to invent 84 legal entities has lost the process. A conversion that embeds it
instead is defensible and reports ``entity_map_pick`` as an unplayed input; ``expected.yaml``
records that under ambiguities so nobody scores it as a defect.
"""

SHUFFLE_SEED = 20261130
"""Seed for the deliberately-disordered positions extract. Fixed, so a failure is reproducible."""

APPROVAL_KEYS: frozenset[str] = frozenset(
    {
        "review_overrides_decision",
        "review_overrides_note",
        "approve_billing_decision",
        "approve_billing_note",
        "post_invoices_ran",
        "post_invoices_ran_note",
    }
)
"""The script keys that record a human decision, withheld by :func:`no_posting_before_approval`.

Kept as a set rather than filtered by suffix because ``_note`` and ``_ran`` are also the shapes a
hand-in confirmation takes, and a rule that guessed would silently withhold the wrong thing.
"""


@functools.cache
def _values() -> Any:
    """The workbook as Excel left it: cached values, no formulas.

    Cached because every hand-in writer and two graders read it, and openpyxl re-parses the whole
    file each time it is asked. Nothing here mutates it.
    """
    return load_workbook(WORKBOOK, data_only=True)


def _header_row(sheet: Any, needle: str, *, limit: int = 40) -> int:
    """The row whose first cell is ``needle``.

    Located rather than hardcoded. ``Positions`` keeps its query in the cells above the grid, so
    the header sits at row 17 today and at some other row the moment a line of SQL is added --
    and a hand-in written from the wrong row is a whole eval grading a grid that starts in the
    middle of a WHERE clause.
    """
    for row in range(1, limit + 1):
        value = sheet.cell(row, 1).value
        if isinstance(value, str) and value.strip() == needle:
            return row
    msg = f"no row in the first {limit} of {sheet.title!r} begins with {needle!r}"
    raise LookupError(msg)


def _iso(value: Any) -> str:
    """One cell as a query client would write it into a file: ISO dates, plain numbers."""
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value)


def _displayed(value: Any) -> str:
    """One cell as Excel puts it on the clipboard: what it *looks like*, not what it holds.

    The difference is the whole of discrimination 8. A date reaches the clipboard as
    ``30/11/2026`` and a General-formatted 5000000.0 as ``5000000`` -- neither is what the cell
    holds, and both are what a paste carries.
    """
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        value = value.date()
    if isinstance(value, dt.date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _as_csv(rows: Iterable[list[str]]) -> str:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerows(rows)
    return buffer.getvalue()


def positions_grid(*, shuffled: bool = False) -> str:
    """The warehouse extract as the query client returns it: CSV, ISO dates, two months.

    Args:
        shuffled: Deliver the same rows in a different order. The warehouse's ``ORDER BY`` is a
            promise the process should not be relying on, and
            :func:`opening_balance_is_ordered` is what proves the notebook does not.
    """
    sheet = _values()["Positions"]
    header = _header_row(sheet, "client_code")
    body = [
        [_iso(cell.value) for cell in row]
        for row in sheet.iter_rows(min_row=header + 1, max_row=sheet.max_row, max_col=6)
        if row[0].value is not None
    ]
    if shuffled:
        random.Random(SHUFFLE_SEED).shuffle(body)
    names = [
        _iso(cell.value)
        for cell in next(sheet.iter_rows(min_row=header, max_row=header, max_col=6))
    ]
    return _as_csv([names, *body])


def fee_schedule_paste() -> str:
    """The fee schedule as it arrived: displayed text, tab-delimited, preamble and all.

    Written as a ``.tsv`` because ``kedge.ingest.read_frame`` picks its separator off the
    extension, and a tab-delimited grid stored under ``.csv`` reads as one column. Two preamble
    rows and a header on row 4, because that is the shape a pasted extract has; the client codes
    keep their leading zeros and the negotiated rates are text, because Excel copies what a cell
    looks like.
    """
    sheet = _values()["Fee Schedule"]
    lines = [
        "\t".join(_displayed(cell.value) for cell in row)
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=9)
    ]
    return "\n".join(lines) + "\n"


def entity_map_grid() -> str:
    """Client to legal entity to cost centre, as CSV.

    The ``mandate`` column keeps its untrimmed values (``"  balanced "``), because the workbook's
    own ``UPPER(TRIM(...))`` is there for a reason and a hand-in that tidied them would grade a
    problem the conversion never had to solve.
    """
    sheet = _values()["Entity Map"]
    header = _header_row(sheet, "client_code", limit=3)
    rows = [
        [_iso(cell.value) for cell in row]
        for row in sheet.iter_rows(min_row=header, max_row=sheet.max_row, max_col=6)
        if row[0].value is not None
    ]
    return _as_csv(rows)


def write_handins(directory: Path, *, shuffled_positions: bool = False) -> dict[str, Path]:
    """Write every inbound artifact and name it. Keys are exactly :data:`HANDINS`.

    Called with a directory that does not exist yet, so it makes it. ``harness.grade.human_script``
    spreads the mapping into :func:`script_for` by keyword, which is what lets this process declare
    three artifacts where the first eval declared two positionally.
    """
    directory.mkdir(parents=True, exist_ok=True)
    positions = directory / "positions.csv"
    fee_schedule = directory / "fee_schedule.tsv"
    entity_map = directory / "entity_map.csv"
    positions.write_text(positions_grid(shuffled=shuffled_positions), encoding="utf-8")
    fee_schedule.write_text(fee_schedule_paste(), encoding="utf-8")
    entity_map.write_text(entity_map_grid(), encoding="utf-8")
    return {"positions": positions, "fee_schedule": fee_schedule, "entity_map": entity_map}


def script_for(*, positions: Path, fee_schedule: Path, entity_map: Path) -> dict[str, Any]:
    """The human's part, played the same way every run.

    Keyword-only, and the parameter names are :data:`HANDINS`. Keyed by the variable each
    ``mo.ui`` widget is assigned to: a file selector wants a *tuple* of paths, a paste box a
    ``str``, a dropdown its option, a checkbox a ``bool``, a date a ``datetime.date``.

    The keys read like the suffix a scaffolder appends to a stage id (``_pick``, ``_decision``,
    ``_note``, ``_ran``) because under ``--convert`` that is exactly what ``harness.align`` has to
    resolve them against: an exact match, or exactly one widget named ``f"{stem}_{key}"``. A key
    naming no widget is reported as unplayed and fails :func:`ran_to_completion`, which is the
    honest outcome -- the harness could not play the human part -- rather than a silent no-op.

    The values themselves must not depend on the run: ``harness.convert`` calls this with
    placeholder paths purely to read the keys.
    """
    return {
        "period_end": dt.date(2026, 11, 30),
        "positions_pick": (positions,),
        "fee_schedule_pick": (fee_schedule,),
        "entity_map_pick": (entity_map,),
        "review_overrides_decision": "approve",
        "review_overrides_note": (
            "All three override reasons re-checked with the billing manager on 2026-12-02; "
            "00007 still waived pending the mandate transfer."
        ),
        "approve_billing_decision": "approve",
        "approve_billing_note": (
            "November fee run agreed at the billing review; schedule effective 2026-11-01."
        ),
        "post_invoices_ran": True,
        "post_invoices_ran_note": "84 invoices posted, ticket BILL-4417.",
    }


# =============================================================================
# CONTEXT
# =============================================================================


@dataclass(frozen=True)
class Context:
    """Everything a grader is allowed to look at.

    ``run`` is not touched here or in any ``__post_init__``: under a plan-only sweep it is a
    stand-in whose every attribute access raises, so it must stay untouched until a grader
    actually asks.
    """

    run: NotebookRun
    facts: Mapping[str, Any]
    notebook: Path
    unused_inputs: tuple[str, ...] = ()
    plan: Any = None
    notes: list[str] = field(default_factory=list)

    @property
    def defs(self) -> dict[str, Any]:
        return self.run.definitions

    def need(self, *names: str) -> ItemResult | None:
        """A verdict about a name the scaffolder fixes, or ``None`` when it is there.

        Two different absences, and conflating them was worth seven points to a conversion that
        reconciled nothing. A notebook that **stopped early** has not failed the items about
        cells it never reached -- it failed the item about stopping, and marking the rest as
        failures would report one problem sixteen times and bury it. That is a ``SKIP``.

        A notebook that **ran to the end** and still does not define the name has not named it
        differently, because these are not names a conversion chooses: ``reconciliation``,
        ``reconciliation_values`` and ``kedge_briefing`` are emitted by the scaffolder and are
        the only names any grader here is allowed to assume. A completed run without one did not
        do the work, and a ``SKIP`` for that takes the points out of the denominator and reports
        the omission as an ambiguity. That is a ``FAIL``.
        """
        absent = [name for name in names if name not in self.defs]
        if not absent:
            return None
        if not self.run.completed:
            return _skip(
                f"the notebook defines no {', '.join(absent)}. It {self.run.summary_line()}."
            )
        return _fail(
            f"the notebook ran to completion and defines no {', '.join(absent)}. That is not a "
            f"naming difference: the scaffolder fixes these names, so a conversion cannot have "
            f"chosen others -- the step is absent. What it does define: "
            f"{', '.join(sorted(self.defs)[:12]) or '(nothing)'}."
        )


def _pass(detail: str = "") -> ItemResult:
    return ItemResult(id="", outcome=Outcome.PASS, detail=detail)


def _fail(detail: str) -> ItemResult:
    return ItemResult(id="", outcome=Outcome.FAIL, detail=detail)


def _skip(detail: str) -> ItemResult:
    return ItemResult(id="", outcome=Outcome.SKIP, detail=detail)


def _no_plan() -> ItemResult:
    return _skip("no plan supplied; pass --plan to grade the structural tier")


def _close(actual: float, expected: float) -> bool:
    return abs(actual - expected) < PENNY


def _graded(function: Callable[[Context], ItemResult]) -> Callable[[Context], ItemResult]:
    """Turn anything a grader throws into an explicit ``FAIL``, never a silent ``SKIP``.

    ``harness.grade._tier`` records a grader that raises as ``SKIP``, and a skip leaves the
    denominator. So the natural way to write a check -- ``assert totals == expected`` -- does not
    fail the item when it goes wrong; it removes the item, and the eval reports full marks over a
    rubric one item shorter than the one it printed. A bug in the grader is still a bug in the
    grader and the detail says so, but it is reported where somebody will look at it.

    ``Exception`` rather than ``BaseException``: a KeyboardInterrupt is not a failed rubric item.
    """

    @functools.wraps(function)
    def graded(ctx: Context) -> ItemResult:
        try:
            return function(ctx)
        except Exception as error:
            logger.exception("grader %s raised", function.__name__)
            return _fail(
                f"the grader itself raised {type(error).__name__}: {error}\n"
                f"That is a defect in evals/fee_billing_run/case.py rather than necessarily in "
                f"the conversion -- but it is a FAIL and not a SKIP, because a skip would take "
                f"{function.__name__}'s points out of the denominator and the eval would read as "
                f"though the item had never existed."
            )

    return graded


def _drive(ctx: Context, inputs: dict[str, Any], root: Path) -> Any:
    """Drive the notebook again, in a workspace of this grader's own.

    The imports are inside the function on purpose. ``harness.findings.aligned_drives``
    monkey-patches ``harness.drive.run_notebook`` for the duration of a ``--convert`` grading run,
    and a module-level import binds the original at import time -- which is how five graders in
    the first eval came to be driving a notebook that had stopped in its third cell.
    """
    from harness.drive import run_notebook, workspace_overrides

    return run_notebook(ctx.notebook, inputs=inputs, overrides=workspace_overrides(root, WORKBOOK))


def _visible(ctx: Context, inputs: dict[str, Any], root: Path) -> tuple[str, ...]:
    """Which cells a user would see, given these inputs. Imported inside, for the same reason."""
    from harness.drive import visible_cells, workspace_overrides

    return visible_cells(ctx.notebook, inputs=inputs, overrides=workspace_overrides(root, WORKBOOK))


# =============================================================================
# FINDING THINGS IN A NOTEBOOK NOBODY HAS WRITTEN YET
# =============================================================================

CLIENT_CODE_COLUMNS = (
    "client_code",
    "client",
    "code",
    "client_id",
    "client_ref",
    "account_code",
    "account",
)
AVG_AUM_COLUMNS = (
    "avg_aum_gbp",
    "avg_aum",
    "average_aum",
    "average_aum_gbp",
    "aum",
    "aum_gbp",
    "mean_aum",
    "avg_aum_nov",
)
BAND_BPS_COLUMNS = ("band_bps", "band_rate_bps", "banded_bps", "schedule_bps", "standard_bps")
TIER_BPS_COLUMNS = ("tier_bps", "fee_bps", "rate_bps", "applied_bps", "billed_bps", "bps")
GROSS_FEE_COLUMNS = ("gross_fee_gbp", "gross_fee")
PRORATED_COLUMNS = (
    "pro_rated_fee",
    "prorated_fee",
    "pro_rated_fee_gbp",
    "prorated_fee_gbp",
    "prorated",
)
DAYS_BILLED_COLUMNS = ("days_billed", "billed_days", "days_in_period")
NET_FEE_COLUMNS = ("net_fee_gbp", "net_fee", "computed_fee", "computed_fee_gbp", "calculated_fee")
AGREED_FEE_COLUMNS = (
    "agreed_fee",
    "agreed_fee_gbp",
    "fee_gbp",
    "billed_fee",
    "final_fee",
    "amount_due",
    "invoice_amount",
    "invoice_amount_gbp",
    "net_payable",
    "fee_amount",
    "fee_charged",
)
ENTITY_COLUMNS = ("legal_entity", "entity")

BRIEFING_WORDS = 40
"""How much prose a briefing has to be before it counts as one.

Not a style rule. Purpose, background and known issues are three blocks somebody wrote on the
Sign-off tab, and the rubric asks the conversion to carry all three across. Without a floor the
check degenerates into three single common words: ``mo.md("Billing. Tiered. Allocation.")`` hit
one needle from each of the three sets and passed. Forty words is roughly two sentences a block
and is well under anything a real briefing runs to.
"""

INSTRUCTION_VERBS = (
    r"\b(run|drop|select|paste|type|choose|tick|enter|fill|upload|browse|provide|supply|"
    r"attach|add|give|bring|copy|pick|click|press|set|confirm|approve|review|open|export|"
    r"save|download|re-?run|re-?extract|record)\b"
)
"""What counts as an instruction in a blocking message.

The rubric asks for "an actual verb the user can act on", and the list used to hold nine of them
-- not including ``upload``, which is the verb for ``mo.ui.file``, an upload widget. "**Step 1 of
5.** Please upload the positions extract above" was reported as a message that "never says what
to do". Widened rather than inverted: a heuristic looking for an imperative at the head of a
sentence has a false-positive problem of its own, and the failure mode that matters here is a
message that explains the rule and never says where to type.
"""

RECONCILED_SHARE = 0.5
"""How much of the workbook a conversion has to have actually *compared*, as a share of
``facts.reconcilable_regions``.

Not a round number picked for comfort. The workbook offers 45 inferrable regions, of which 18
coincide with one of the nineteen dead-region findings -- columns a plan is expected to drop, and
which it therefore cannot reproduce -- and one is the deliberately stale ``Allocation`` tab. A
conversion that reconciles its live spine compares of the order of 26, clears 23 comfortably, and
declares the rest with reasons. A conversion that declares all 45 and compares nothing compares 0
and fails, which is the entire point: non-negotiable 6 says reconciliation never reports "passed"
when it has no baseline, and until this floor existed the two reconciliation items scored 7 of 7
over a report whose own headline read "0 of 0 claimed regions passed".

A share rather than a count because it is a policy about coverage, and a count would have to be
another figure in ``facts`` that nothing recomputes.
"""


@dataclass(frozen=True, slots=True)
class _Grid:
    """One frame the notebook bound, and the column that keys it by client."""

    name: str
    frame: Any
    key: str


def _frames(defs: Mapping[str, Any]) -> dict[str, Any]:
    """Every polars frame the notebook bound, collected, keyed by variable name.

    A lazy frame is collected here rather than at every use, and one that refuses to collect is
    dropped with a log line rather than taking the grader down -- a frame kedge cannot evaluate is
    a finding for the item that is about it, not for every item that walks past it.
    """
    import polars as pl

    frames: dict[str, Any] = {}
    for name, value in defs.items():
        if isinstance(value, pl.DataFrame):
            frames[name] = value
        elif isinstance(value, pl.LazyFrame):
            try:
                frames[name] = value.collect()
            except Exception:
                logger.debug("could not collect the lazy frame %r", name, exc_info=True)
    return frames


def _key_column(frame: Any, codes: set[str], width: int) -> str | None:
    """The column that keys this frame by client, by name first and then by content.

    By name because it is exact and cheap. By content because a name list is a guess, and a
    guess that misses does not report "this conversion named its key column something I did not
    expect" -- it reports nothing at all, and twelve points leave the denominator. A column
    holding half the workbook's client codes *is* the key column, whatever it is called and
    whether or not its leading zeros survived the ingest.
    """
    lowered = {column.lower(): column for column in frame.columns}
    named = next((lowered[c] for c in CLIENT_CODE_COLUMNS if c in lowered), None)
    if named is not None:
        return named
    for column in frame.columns:
        values = [value for value in frame[column].to_list() if value is not None]
        if not values:
            continue
        matched = {code for value in values if (code := _norm_code(value, width)) in codes}
        if len(matched) * 2 >= len(codes):
            return column
    return None


def _grids(ctx: Context) -> list[_Grid]:
    """Every frame keyed by a client code, billing grain first and widest first.

    Billing grain first because the positions extract is also keyed by client code and has two
    rows for each of them: a grader asking "is this code still a string" wants the frame with 84
    rows, not the one with 168.
    """
    expected = int(ctx.facts["clients_billed"])
    codes, width = _expected_codes(ctx.facts)
    grids: list[_Grid] = []
    for name, frame in _frames(ctx.defs).items():
        key = _key_column(frame, codes, width)
        if key is not None:
            grids.append(_Grid(name=name, frame=frame, key=key))
    grids.sort(
        key=lambda grid: (grid.frame.height != expected, -len(grid.frame.columns), grid.name)
    )
    return grids


def _find(grids: list[_Grid], *candidates: str) -> tuple[_Grid, str] | None:
    """The first grid carrying any of these column names, and the name it used."""
    for grid in grids:
        lowered = {column.lower(): column for column in grid.frame.columns}
        for candidate in candidates:
            if candidate.lower() in lowered:
                return grid, lowered[candidate.lower()]
    return None


def _describe(grids: list[_Grid]) -> str:
    """What the notebook does bind, for a failure message that says where to look."""
    return "; ".join(
        f"{grid.name} ({grid.frame.height} rows: {', '.join(grid.frame.columns[:12])}"
        + (", ..." if len(grid.frame.columns) > 12 else "")
        + ")"
        for grid in grids[:4]
    )


def _no_grid(ctx: Context) -> ItemResult:
    """No frame is keyed by a client code, so there is nothing shaped like a fee run to read.

    A stop is a ``SKIP``: the cell this item is about never ran. A completed run is a ``FAIL``,
    and the change is worth twelve points. :func:`_key_column` looks for the key by *content* as
    well as by name, so a conversion that called it ``account_code`` is still found -- which
    means a completed run with no such frame is not a naming difference any more. It is a
    conversion that computed nothing per client, and calling that unmeasurable took the points
    out of the denominator and reported an omission as an ambiguity.
    """
    if not ctx.run.completed:
        return _skip(
            f"the notebook binds no frame keyed by a client code and it {ctx.run.summary_line()}, "
            f"so this item is about a cell that never ran."
        )
    return _fail(
        f"the notebook ran to completion and binds no frame keyed by a client code. Neither the "
        f"column names ({', '.join(CLIENT_CODE_COLUMNS)}) nor the 84 codes themselves appear on "
        f"any frame it produced, so nothing here computes anything per client. What it binds: "
        f"{_describe_frames(ctx)}"
    )


def _describe_frames(ctx: Context) -> str:
    frames = _frames(ctx.defs)
    if not frames:
        return "no polars frame at all"
    return "; ".join(
        f"{name} ({frame.height} rows: {', '.join(frame.columns[:8])})"
        for name, frame in list(frames.items())[:4]
    )


def _stopped(ctx: Context) -> ItemResult | None:
    """A skip when the run never reached the end, or ``None``."""
    if ctx.run.completed:
        return None
    return _skip(
        f"the notebook did not reach the end, so this item is about a cell that never ran: "
        f"it {ctx.run.summary_line()}"
    )


def _norm_code(value: Any, width: int) -> str:
    """One client code as the workbook writes it: text, padded back to its full width.

    ``00013`` arriving as the integer ``13`` is a real defect and
    :func:`leading_zero_client_codes_survive` is the item that reports it. Every *other* grader
    matching a client has to normalise, or that one defect fails six items and nineteen points,
    five of them pointing at things the conversion did correctly. Padding here is not forgiveness
    -- the item about the codes reads the column raw and still fails -- it is what keeps each
    item measuring its own discrimination.
    """
    if isinstance(value, bool) or value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text.zfill(width) if text.isdigit() and len(text) < width else text


def _expected_codes(facts: Mapping[str, Any]) -> tuple[set[str], int]:
    """The 84 client codes the workbook bills, and how wide they are."""
    width = len(str(facts["first_client_code"]))
    low = int(str(facts["first_client_code"]))
    high = int(str(facts["last_client_code"]))
    return {str(number).zfill(width) for number in range(low, high + 1)}, width


def _rows(grid: _Grid, width: int = 0) -> dict[str, dict[str, Any]]:
    """One grid as a mapping from client code to its row, keyed on the normalised code."""
    return {_norm_code(row[grid.key], width): row for row in grid.frame.iter_rows(named=True)}


def _flatten(value: Any, depth: int = 0) -> Iterable[str]:
    """Every string reachable from one bound value, frames and widgets included.

    ``_text`` used to be ``ctx.run.panels`` plus every bound ``str``, and that made two
    different questions depend on a translation's choice of container. Rendering three overrides
    with their reasons in an ``mo.ui.table`` *is* showing them, and a list of 84 INSERT
    statements -- the natural shape, and what a ``.with_columns(...)`` translation produces -- is
    a generated statement however it is held. Both were invisible.
    """
    import polars as pl

    if depth > 3:
        return
    if isinstance(value, str):
        yield value
    elif isinstance(value, pl.DataFrame):
        for column, dtype in zip(value.columns, value.dtypes, strict=True):
            if dtype == pl.Utf8:
                yield from (item for item in value[column].to_list() if isinstance(item, str))
    elif isinstance(value, pl.Series):
        yield from (item for item in value.to_list() if isinstance(item, str))
    elif isinstance(value, Mapping):
        for key, item in value.items():
            if isinstance(key, str):
                yield key
            yield from _flatten(item, depth + 1)
    elif isinstance(value, (list, tuple)):
        for item in value:
            yield from _flatten(item, depth + 1)
    elif hasattr(value, "value") and hasattr(value, "kind"):
        # A `harness.drive.Widget`: mo.ui.table(frame), mo.ui.tabs({...}). What it holds is what
        # the user is looking at.
        yield from _flatten(value.value, depth + 1)


def _sources(ctx: Context) -> list[tuple[str, str]]:
    """Everything the notebook rendered or bound, as (where it came from, the text).

    Kept apart rather than joined, because :func:`generated_sql_is_valid` has to be able to tell
    "the same 84 statements, bound once and rendered once" from "168 statements, every invoice
    posted twice". Joining first makes those two indistinguishable, and de-duplicating the
    *statements* to compensate makes the second one invisible.
    """
    parts: list[tuple[str, str]] = [
        (f"panel {index}", panel) for index, panel in enumerate(ctx.run.panels)
    ]
    for name, value in sorted(ctx.defs.items()):
        parts.extend((name, text) for text in _flatten(value))
    return parts


def _text(ctx: Context) -> str:
    """Everything the notebook rendered, plus every string it bound or put in a widget."""
    return "\n".join(text for _where, text in _sources(ctx))


#: `INSERT INTO ... ;` rather than a split on `;`, because the statements are rendered inside a
#: fenced code block as well as bound to a variable, and a fence line would otherwise become part
#: of the first statement. No literal in this workbook contains a semicolon -- the client names
#: never reach the statement -- so a non-greedy match to the first one is exact here. See
#: :func:`generated_sql_is_valid` for why that is also this workbook's sharpest remaining gap.
_INSERT = re.compile(r"INSERT\s+INTO\b.*?;", re.IGNORECASE | re.DOTALL)


def _statements(ctx: Context) -> list[str]:
    """Every INSERT the notebook generated, counted once per invoice rather than per rendering.

    The distinction is the whole of it. A notebook normally binds its statements to a variable
    *and* renders them in a panel, so the same 84 appear twice; a notebook with a duplicated
    ``iter_rows`` emits 168, which is a double-post against a production ledger. The old
    implementation ran ``dict.fromkeys`` over the concatenation of everything, which collapsed
    both cases to 84 and made the second one undetectable.

    So statements are gathered per *source* and a source is dropped only when everything in it
    already appears in a longer one -- a panel echoing the variable, or a five-statement preview
    above the full list. Two sources holding different halves are kept and added; one source
    holding every invoice twice is kept whole, and the count and the posted rows both say so.
    """
    found = [
        (where, tuple(match.strip() for match in _INSERT.findall(text)))
        for where, text in _sources(ctx)
    ]
    found = [(where, items) for where, items in found if items]
    found.sort(key=lambda pair: -len(pair[1]))
    kept: list[tuple[str, tuple[str, ...]]] = []
    for where, items in found:
        if not any(set(items) <= set(other) for _where, other in kept):
            kept.append((where, items))
    return [statement for _where, items in kept for statement in items]


def _statement_variables(ctx: Context) -> list[tuple[str, str]]:
    """The notebook's own names that hold a posting statement, and one statement from each.

    What :func:`no_posting_before_approval` needs in order to find a posting *cell* without
    matching on the table's name: the cell that binds one of these is the cell that builds the
    statements, whether it spells the table out, assembles it from a constant, or renders it
    through :mod:`kedge.sql`.
    """
    return [
        (name, matches[0])
        for name, value in sorted(ctx.defs.items())
        for text in _flatten(value)
        if (matches := _INSERT.findall(text))
    ]


@functools.cache
def _negotiated_rates() -> dict[str, float]:
    """The client-specific rates the schedule carries, read out of the workbook.

    Read rather than quoted in ``facts``: there are seventeen of them, and a rubric is a document
    somebody reads. The workbook is the same artifact the ``fee_schedule`` hand-in is written
    from, so this is derivation and not a second source of truth -- and the rates arrive as text,
    which is the point of discrimination 8, so they are cast here exactly as the sheet's own
    arithmetic casts them.
    """
    sheet = _values()["Fee Schedule"]
    rates: dict[str, float] = {}
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=2):
        code, rate = row[0].value, row[1].value
        if code is None or rate is None or not str(code).strip().isdigit():
            continue
        try:
            rates[str(code).strip()] = float(rate)
        except (TypeError, ValueError):
            logger.debug("fee schedule row %r carries an unparseable rate %r", code, rate)
    return rates


@functools.cache
def _workbook_aum() -> dict[str, float]:
    """Each client's average AUM for the month, read out of ``Working``.

    The same derivation :func:`_negotiated_rates` is, and for a sharper reason. A grader asking
    "is this client on the rate its AUM earns" used to read the AUM off the notebook's own frame,
    which measured two things at once: the rate, and whether the conversion happened to name its
    AUM column one of the five this file guessed. When it did not, the grader checked nothing and
    returned a pass saying "the rate columns agree with the schedule". The workbook knows every
    client's AUM. Reading it here means the item measures the rate and nothing else.
    """
    return {
        code: float(value)
        for code, value in _working_column("Nov-26 avg AUM").items()
        if isinstance(value, (int, float)) and not isinstance(value, bool)
    }


@functools.cache
def _working_column(header: str) -> dict[str, Any]:
    """One column of ``Working``, keyed by client code, as Excel last cached it.

    Located by header text rather than by letter: the sheet has thirty-four of them and a column
    inserted between two others moves every letter after it, which is exactly the silent rot the
    rubric's own rot guard exists to catch.
    """
    sheet = _values()["Working"]
    row = _header_row(sheet, "Client")
    headers = {
        str(cell.value).strip(): cell.column
        for cell in next(sheet.iter_rows(min_row=row, max_row=row, max_col=sheet.max_column))
        if cell.value is not None
    }
    index = headers.get(header)
    if index is None:
        msg = f"Working has no {header!r} column; it has {sorted(headers)}"
        raise LookupError(msg)
    values: dict[str, Any] = {}
    for cells in sheet.iter_rows(min_row=row + 1, max_row=sheet.max_row, max_col=index):
        code = cells[0].value
        if code is None or not str(code).strip():
            continue
        values[str(code).strip()] = cells[index - 1].value
    return values


@functools.cache
def _workbook_entities() -> dict[str, str]:
    """Client code to legal entity, read off the ``Entity Map`` tab.

    Read rather than taken from the notebook's own frame so that
    :func:`subtotal_rows_are_excluded` can check the entity totals of a conversion that computed
    the fees correctly and simply did not carry a ``legal_entity`` column into the frame this
    grader found. The item is about an embedded subtotal double-counting an entity; it is not
    about column naming, and it used to return a pass whose own detail said it had checked
    nothing.
    """
    sheet = _values()["Entity Map"]
    header = _header_row(sheet, "client_code", limit=3)
    names = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(header, header, 1, 6))]
    entity = names.index("legal_entity")
    mapping: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=header + 1, max_row=sheet.max_row, max_col=6):
        if row[0].value is None:
            continue
        mapping[str(row[0].value).strip()] = str(row[entity].value or "").strip()
    return mapping


@functools.cache
def _fee_schedule_data_rows() -> tuple[int, int]:
    """The first and last row of the fee schedule that carries a rate rather than a preamble.

    Two rows of preamble and a header on row four is the shape a pasted extract has, and a plan
    that declares ``Fee Schedule!A1:A2`` as its source has named the *note above* the rate card
    while hardcoding the rates. Derived by looking for the rows that hold data -- a band floor in
    the band block, a client code in the negotiated block -- so a schedule that grows a row does
    not turn a correct plan red.
    """
    sheet = _values()["Fee Schedule"]
    rows = [
        cells[0].row
        for cells in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=6)
        if isinstance(cells[3].value, (int, float))
        or (cells[0].value is not None and str(cells[0].value).strip().isdigit())
    ]
    if not rows:
        msg = "the Fee Schedule carries neither a band floor nor a negotiated rate"
        raise LookupError(msg)
    return min(rows), max(rows)


def _band_for(aum: float, bands: list[Mapping[str, Any]]) -> float:
    """Excel's approximate ``VLOOKUP``: the last band whose floor does not exceed the value."""
    chosen = float(bands[0]["bps"])
    for band in bands:
        if aum >= float(band["floor"]):
            chosen = float(band["bps"])
    return chosen


def _numeric(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _same(left: Any, right: Any) -> bool:
    """Two cells agree: money at half a penny, everything else exactly."""
    a, b = _numeric(left), _numeric(right)
    if a is not None and b is not None:
        return _close(a, b)
    return left == right


# =============================================================================
# TIER 1 - DETERMINISTIC
# =============================================================================


@_graded
def ran_to_completion(ctx: Context) -> ItemResult:
    """Nothing else is meaningful if the runbook never reached the end."""
    if ctx.unused_inputs:
        return _fail(
            f"the script named widgets the notebook does not have: "
            f"{', '.join(ctx.unused_inputs)}. Either the notebook names them differently or a "
            f"step is missing; nothing below was driven as intended."
        )
    if not ctx.run.completed:
        return _fail(ctx.run.summary_line())
    return _pass(f"{len(ctx.run.cells_run)} cells")


@_graded
def tier_lookup_is_banded(ctx: Context) -> ItemResult:
    """Discrimination 1: an approximate-match VLOOKUP is a banded join, not a lookup."""
    stopped = _stopped(ctx)
    if stopped:
        return stopped
    grids = _grids(ctx)
    if not grids:
        return _no_grid(ctx)

    bands = [dict(band) for band in ctx.facts["fee_bands"]]
    band_at = _find(grids, *BAND_BPS_COLUMNS)
    tier_at = _find(grids, *TIER_BPS_COLUMNS)
    if band_at is None and tier_at is None:
        return _fail(
            f"no column anywhere in the notebook carries a rate in basis points (looked for "
            f"{', '.join(dict.fromkeys((*BAND_BPS_COLUMNS, *TIER_BPS_COLUMNS)))}).\n"
            f"The banded lookup is Working!I and Working!J, and without it every fee below it is "
            f"a number with no derivation. What the notebook binds: {_describe(grids)}"
        )

    codes, width = _expected_codes(ctx.facts)
    aum = _workbook_aum()
    negotiated = _negotiated_rates()
    problems: list[str] = []
    checked: list[str] = []

    if band_at is not None:
        grid, column = band_at
        counts = Counter(
            round(value, 4) for value in grid.frame[column].to_list() if _numeric(value) is not None
        )
        for band in bands:
            want, got = int(band["clients"]), counts.get(round(float(band["bps"]), 4), 0)
            if want != got:
                problems.append(
                    f"the {band['bps']}bps band (floor {float(band['floor']):,.0f}) reaches "
                    f"{got} clients, expected {want}"
                )
        checked.append(f"{grid.name}.{column} against the four bands")
        problems.extend(
            _rate_problems(
                grid, column, width, codes, want=lambda code: _band_for(aum[code], bands)
            )
        )
        checked.append(f"all {len(codes)} clients against the AUM the workbook holds for them")

    if tier_at is not None:
        grid, column = tier_at
        problems.extend(
            _rate_problems(
                grid,
                column,
                width,
                codes,
                want=lambda code: negotiated.get(code, _band_for(aum[code], bands)),
            )
        )
        checked.append(f"{grid.name}.{column} against band and negotiated rate, all {len(codes)}")

    if problems:
        return _fail(
            "\n".join(problems) + "\nThe band is the last one whose floor does not exceed the AUM: "
            '`join_asof(strategy="backward")` on a frame sorted by the join key. An exact join '
            "drops every client, an unsorted join_asof is wrong silently, and `>` where Excel "
            "means `>=` moves a whole tier."
        )
    if not checked:
        # Unreachable while one of the two columns exists, and here anyway: `_pass()` must never
        # be reachable with nothing checked. Three graders used to print a PASS whose detail
        # described work they had not done, which is the reconciliation sin in a costume.
        return _fail("no rate column could be compared against the schedule at all")
    return _pass("; ".join(checked))


def _rate_problems(
    grid: _Grid,
    column: str,
    width: int,
    codes: set[str],
    *,
    want: Callable[[str], float],
) -> list[str]:
    """Every client whose rate is wrong, and every client with no rate at all.

    The null is the whole point. This grader's own rubric names the failure it exists to catch --
    "an exact join, which drops every client whose AUM is not exactly a band floor, i.e. all of
    them" -- and the per-row loop used to ``continue`` past a null rate, so the 67 clients the
    join lost were skipped in silence and the item passed on the 17 it had not lost. A missing
    rate is not an unmeasurable row. It is a client who is not being billed.
    """
    rows = _rows(grid, width)
    wrong: list[str] = []
    missing: list[str] = []
    for code in sorted(codes):
        row = rows.get(code)
        if row is None:
            missing.append(f"{code} (not in {grid.name} at all)")
            continue
        got = _numeric(row[column])
        if got is None:
            missing.append(f"{code} (rate is {row[column]!r})")
        elif not _close(got, want(code)):
            wrong.append(f"{code}: billed at {got}bps, the schedule says {want(code)}bps")
    problems: list[str] = []
    if missing:
        problems.append(
            f"{len(missing)} of {len(codes)} clients carry no rate in {grid.name}.{column}:\n  "
            + "\n  ".join(missing[:5])
            + "\nA null rate is a client nobody is billing. An exact join against the band floors "
            "produces exactly this, and coalescing the negotiated rate over the top leaves the "
            "seventeen that have one looking correct."
        )
    if wrong:
        problems.append(
            f"{len(wrong)} row(s) are billed at the wrong rate:\n  " + "\n  ".join(wrong[:5])
        )
    return problems


@_graded
def opening_balance_is_ordered(ctx: Context) -> ItemResult:
    """Discrimination 2: a prior-row reference needs an explicit sort.

    Graded as invariance across the whole computation rather than on one column, because
    order-dependence is one defect however many columns express it -- Working!G reads the row
    above, Working!Y is a running total, and an unsorted ``join_asof`` fails here too. It also
    means the item stays gradeable when a conversion legitimately drops the opening balance,
    which reads nothing downstream and is one of the nineteen dead regions.
    """
    stopped = _stopped(ctx)
    if stopped:
        return stopped

    with tempfile.TemporaryDirectory(prefix="kedge-eval-ordered-") as workspace:
        root = Path(workspace)
        ordered = _drive(ctx, script_for(**write_handins(root / "handins-in")), root)
    if not ordered.completed:
        # Not a skip. The graded run reached the end -- `_stopped` above is what returns when it
        # did not -- so the same notebook, driven again with the same script, has just behaved
        # differently. Whatever that is, it is not something the shuffle proves or disproves, and
        # dropping the item would hide it.
        return _fail(
            f"the notebook was driven twice with the same script and completed only once: the "
            f"control run {ordered.summary_line()}. Nothing about row order has been tested; "
            f"what has been found is that the run is not repeatable."
        )
    # A workspace of its own, not a second pass over the first: sharing one would let the run
    # store hand back the hand-in the first pass already banked, and the shuffle would never
    # reach the notebook at all.
    with tempfile.TemporaryDirectory(prefix="kedge-eval-shuffled-") as workspace:
        root = Path(workspace)
        handins = write_handins(root / "handins-in", shuffled_positions=True)
        shuffled = _drive(ctx, script_for(**handins), root)
    if not shuffled.completed:
        return _fail(
            f"the same extract delivered in a different row order did not get through: "
            f"{shuffled.summary_line()}\nThe warehouse's ORDER BY is a promise the process "
            f"should not be relying on."
        )

    left = {grid.name: grid for grid in _grids(_restated(ctx, ordered))}
    right = {grid.name: grid for grid in _grids(_restated(ctx, shuffled))}
    shared = sorted(set(left) & set(right))
    if not shared:
        return _fail(
            f"the two runs have no frame in common to compare. The ordered run bound "
            f"{sorted(left) or '(nothing keyed by a client code)'}; the shuffled run bound "
            f"{sorted(right) or '(nothing keyed by a client code)'}. Both completed, so this is "
            f"not a stop -- the notebook produces a different set of frames depending on the "
            f"order its input arrived in, which is the defect this item is looking for in its "
            f"largest possible form."
        )

    # Every frame, not the widest one. `_grids` sorts widest-first, and comparing `grids[0]` alone
    # meant a genuinely order-dependent column was never looked at whenever some *other* frame
    # happened to carry more columns: a two-column running total cum_sum'd in arrival order sat
    # behind a four-column billing frame that was order-free, and the item passed. Which
    # discrimination fires must not be decided by column layout.
    differing: dict[str, list[str]] = {}
    width = len(str(ctx.facts["first_client_code"]))
    checked = 0
    for name in shared:
        before, after = _rows(left[name], width), _rows(right[name], width)
        if set(before) != set(after):
            only_before = sorted(set(before) - set(after))[:5]
            only_after = sorted(set(after) - set(before))[:5]
            return _fail(
                f"the shuffled extract produced a different set of clients in {name}: "
                f"{len(before)} against {len(after)}. Only in the ordered run: {only_before}; "
                f"only in the shuffled run: {only_after}."
            )
        columns = [
            column
            for column in left[name].frame.columns
            if after and column in after[next(iter(after))]
        ]
        checked += len(columns)
        for code in sorted(before):
            for column in columns:
                if not _same(before[code][column], after[code][column]):
                    differing.setdefault(f"{name}.{column}", []).append(
                        f"{code}: {before[code][column]!r} -> {after[code][column]!r}"
                    )
    if differing:
        lines = [
            f"{column} ({len(examples)} row(s)): {examples[0]}"
            for column, examples in sorted(differing.items())
        ]
        return _fail(
            f"{len(differing)} column(s) changed when the extract arrived in a different order:\n  "
            + "\n  ".join(lines[:6])
            + "\nExcel gives a row its neighbours for free; polars gives a frame no order at all. "
            "Sort explicitly before any shift, cumulative sum or join_asof."
        )
    return _pass(
        f"{checked} column(s) across {len(shared)} frame(s) identical for every client, order-free"
    )


def _restated(ctx: Context, run: Any) -> Context:
    """The same context over a different run of the same notebook."""
    return Context(run=run, facts=ctx.facts, notebook=ctx.notebook)


@_graded
def proration_matches_excel(ctx: Context) -> ItemResult:
    """Discrimination 3: pro-rating is date arithmetic on an Excel serial boundary."""
    stopped = _stopped(ctx)
    if stopped:
        return stopped
    grids = _grids(ctx)
    if not grids:
        return _no_grid(ctx)

    prorated_at = _find(grids, *PRORATED_COLUMNS)
    if prorated_at is None:
        return _fail(
            f"no pro-rated fee column anywhere in the notebook (looked for "
            f"{', '.join(PRORATED_COLUMNS)}). Working!P feeds the floor and so the net fee, so "
            f"this is not a column that can be dropped. What the notebook binds: {_describe(grids)}"
        )
    grid, column = prorated_at
    _codes, width = _expected_codes(ctx.facts)
    rows = _rows(grid, width)
    days_at = _find([grid], *DAYS_BILLED_COLUMNS)
    month = int(ctx.facts["days_in_month"])

    problems: list[str] = []
    for expected in ctx.facts["part_period_clients"]:
        code = str(expected["client"])
        row = rows.get(code)
        if row is None:
            problems.append(f"{code} is not in {grid.name} at all")
            continue
        got = _numeric(row[column])
        want = float(expected["prorated_fee_gbp"])
        if got is None or not _close(got, want):
            problems.append(
                f"{code}: pro-rated fee {got}, expected {want:,.2f} "
                f"({expected['days_billed']} of {month} days on a gross of "
                f"{float(expected['gross_fee_gbp']):,.2f})"
            )
        if days_at is not None:
            days = _numeric(row[days_at[1]])
            if days is None or int(days) != int(expected["days_billed"]):
                problems.append(
                    f"{code}: billed for {days} days, expected {expected['days_billed']}. The "
                    f'boundary is DATEDIF(start, EOMONTH(start,0), "d") + 1, inclusive of the '
                    f"onboarding day."
                )

    part_period = {str(item["client"]) for item in ctx.facts["part_period_clients"]}
    if days_at is not None:
        odd = [
            f"{code} billed {days:g} days"
            for code, row in rows.items()
            if code not in part_period
            and (days := _numeric(row[days_at[1]])) is not None
            and int(days) != month
        ]
        if odd:
            problems.append(
                f"{len(odd)} client(s) outside the three onboarded mid-month are not billed for "
                f"the whole month: {', '.join(odd[:5])}"
            )

    total = sum(value for row in rows.values() if (value := _numeric(row[column])) is not None)
    want_total = float(ctx.facts["prorated_total_gbp"])
    if not _close(total, want_total):
        problems.append(
            f"the pro-rated fees total {total:,.2f}, expected {want_total:,.2f} "
            f"(out by {total - want_total:+,.2f})"
        )
    if problems:
        return _fail("\n".join(problems))
    return _pass(f"three part periods, {total:,.2f} in total")


@_graded
def overrides_are_surfaced_not_applied(ctx: Context) -> ItemResult:
    """Discrimination 4: a manual override is a decision to re-ask, not a number to bake in."""
    stopped = _stopped(ctx)
    if stopped:
        return stopped
    overrides = [dict(item) for item in ctx.facts["overrides"]]
    shown = _text(ctx).lower()
    _codes, width = _expected_codes(ctx.facts)

    problems: list[str] = []
    for override in overrides:
        code = str(override["client"])
        fragment = str(override["reason_fragment"]).lower()
        # The padded code, or the bare integer as a whole word. `00007` rendered as `7` is a real
        # defect and `leading_zero_client_codes_survive` is the item that reports it; making this
        # one report it too would charge four more points for one mistake and point them at the
        # overrides, which the conversion surfaced correctly.
        bare = rf"\b{int(code)}\b" if code.isdigit() else re.escape(code)
        if code.lower() not in shown and not re.search(bare, shown):
            problems.append(f"{code} is nowhere in what the notebook shows")
        if fragment not in shown:
            problems.append(
                f"{code}'s reason is not shown: the workbook says {override['reason_fragment']!r}"
            )

    grids = _grids(ctx)
    if not grids:
        problems.append(
            "no frame is keyed by a client code, so whether the calculation the override "
            "replaced survived cannot be read"
        )
    else:
        computed_at = _find(grids, *NET_FEE_COLUMNS)
        if computed_at is None:
            problems.append(
                f"no column carries the fee the workbook calculated before the override was "
                f"applied (looked for {', '.join(NET_FEE_COLUMNS)}). The typed figure is then "
                f"the only fee the notebook has, and next month cannot tell whether the "
                f"override still applies."
            )
        else:
            grid, column = computed_at
            rows = _rows(grid, width)
            for override in overrides:
                code = str(override["client"])
                row = rows.get(code)
                if row is None:
                    problems.append(f"{code} is not in {grid.name}")
                    continue
                got = _numeric(row[column])
                want = float(override["computed_fee_gbp"])
                if got is None or not _close(got, want):
                    problems.append(
                        f"{code}: the notebook's pre-override fee is {got}, and the workbook "
                        f"calculated {want:,.2f}. The agreed figure is "
                        f"{float(override['agreed_fee_gbp']):,.2f}; if that is what this column "
                        f"holds, the override has replaced the calculation rather than "
                        f"overriding it."
                    )
    if problems:
        return _fail(
            "\n".join(problems)
            + "\nThree answers are wrong in different directions: reproducing the typed numbers "
            "bakes September's decision into every future month, dropping them loses three real "
            "decisions, and reconciling against them reports a pass for a figure nobody "
            "calculated."
        )
    return _pass(f"{len(overrides)} overrides shown with their reasons, calculations intact")


@_graded
def does_not_drop_the_posting_column(ctx: Context) -> ItemResult:
    """Discrimination 5: seventeen dead regions are dead; the eighteenth posts the fees.

    Two halves, and both had their polarity wrong. On the plan side a ``DroppedRange`` **stays**
    in ``plan.dropped`` after review, and ``accepted=False`` means the reviewer refused the drop
    and the range is kept -- so reading the list without reading ``rejected`` failed the one
    review action this item exists to reward. On the notebook side the test was "does the table
    name appear in any string the notebook bound", which a briefing sentence satisfies: *"the old
    process pasted 84 statements into fin.fee_invoice by hand. This conversion does not post"*
    passed the eval's sharpest deterministic item while posting nothing.
    """
    table = str(ctx.facts["posting_table"])
    if ctx.plan is not None:
        dropped = [
            item.range
            for item in ctx.plan.dropped
            if "POST" in (item.range or "").upper() and not getattr(item, "rejected", False)
        ]
        if dropped:
            return _fail(
                f"the plan drops {', '.join(dropped)}. The analyser reports that region as dead "
                f"because nothing in the workbook reads it -- its consumer is a person with a "
                f"clipboard, which no static analysis can see. It is one of "
                f"{ctx.facts['dead_regions']} dead regions, it has fan-out zero by definition, "
                f"and dropping it deletes the step that changes the data.\n"
                f"A drop the reviewer *refused* is not this: `accepted=False` leaves the range in "
                f"`plan.dropped` and means it must be kept, which is the correct outcome and "
                f"passes."
            )
    statements = _statements(ctx)
    posting = [item for item in statements if table.lower() in item.lower()]
    if not posting:
        stopped = _stopped(ctx)
        if stopped:
            return stopped
        mentioned = table in _text(ctx)
        return _fail(
            f"the notebook ran to the end and generated no INSERT into {table}"
            + (
                f". It does mention {table} somewhere -- in a panel, a docstring or a constant -- "
                f"and naming the table is not posting to it."
                if mentioned
                else f", and never mentions {table} at all."
            )
            + " Post!A is one statement per invoice built by `&`; without it the fee run "
            "computes what is owed and posts nothing."
        )
    want = int(ctx.facts["posting_statements"])
    if len(posting) < want:
        return _fail(
            f"the notebook generates {len(posting)} INSERT(s) into {table}, and the workbook's "
            f"Post tab holds {want} -- one per invoice. A sample is not the step."
        )
    return _pass(f"{len(posting)} INSERT(s) into {table}")


@_graded
def dead_regions_are_individually_reasoned(ctx: Context) -> ItemResult:
    """Discrimination 5's other half: nineteen decisions, or one button pressed nineteen times.

    Read ``note`` as well as ``reason``, because the bulk dismissal this item exists to detect
    writes only the first. ``DroppedRange.reason`` is written by the *proposing* model, one
    sentence per range and different every time; ``acknowledge_all_drops`` -- the function named
    in the failure message below -- stamps one ``note`` across every drop and never touches
    ``reason``. Reading ``reason`` alone meant the mechanism was invisible to the check by
    construction, and the bulk button passed with "19 drops, 19 distinct reasons".
    """
    if ctx.plan is None:
        return _no_plan()
    dropped = list(ctx.plan.dropped)
    if not dropped:
        return _fail(
            f"the plan drops nothing at all, on a workbook the analyser reports "
            f"{ctx.facts['dead_regions']} dead regions in. Nineteen ranges nothing reads is "
            f"nineteen decisions somebody has to take -- eighteen of them to drop it and one, "
            f"the posting column, to keep it. A plan that takes none of them has not translated "
            f"the workbook, it has copied it, and there is no audit trail to be individual about."
        )

    problems: list[str] = []
    for field_name, written_by in (("reason", "the model proposing the plan"), ("note", "review")):
        by_text: dict[str, list[str]] = {}
        for item in dropped:
            value = getattr(item, field_name, None) or ""
            key = re.sub(r"\s+", " ", value.strip().lower()).rstrip(".")
            if key:
                by_text.setdefault(key, []).append(item.range)
        shared = {text: ranges for text, ranges in by_text.items() if len(ranges) > 1}
        problems.extend(
            f"{len(ranges)} ranges share one {field_name} verbatim, written at {written_by} "
            f"({', '.join(ranges[:4])}"
            + (", ..." if len(ranges) > 4 else "")
            + f"): {text[:110]!r}"
            for text, ranges in sorted(shared.items(), key=lambda pair: -len(pair[1]))[:3]
        )

    bare = [item.range for item in dropped if len((item.reason or "").split()) < 4]
    if bare:
        problems.append(
            f"{len(bare)} drop(s) give a reason of three words or fewer: {', '.join(bare[:5])}. "
            f"'Unused' alone is not a reason."
        )
    if problems:
        return _fail(
            "\n".join(problems)
            + f"\n`acknowledge_all_drops` clears every drop with one shared note, and the audit "
            f"trail really is identical. The decision is not: {len(dropped)} ranges dismissed "
            f"with one sentence that cannot be true of all of them, and the posting column is "
            f"one of them."
        )
    reasons = {re.sub(r"\s+", " ", (item.reason or "").strip().lower()) for item in dropped}
    return _pass(f"{len(dropped)} drops, {len(reasons)} distinct reasons, no shared note")


@_graded
def subtotal_rows_are_excluded(ctx: Context) -> ItemResult:
    """Discrimination 6: an embedded subtotal row is not data."""
    stopped = _stopped(ctx)
    if stopped:
        return stopped
    # No early skip for "the notebook binds no frame". A completed run that bound none has not
    # aggregated anything, which `_no_grid` reports below as the failure it is; skipping took
    # three points out of the denominator for an omission.
    offenders: list[str] = []
    for name, frame in _frames(ctx.defs).items():
        for column in frame.columns:
            values = [value for value in frame[column].to_list() if isinstance(value, str)]
            hits = [value for value in values if "subtotal" in value.lower()]
            if hits:
                offenders.append(f"{name}.{column}: {len(hits)} row(s), e.g. {hits[0]!r}")
    if offenders:
        return _fail(
            "a subtotal row reached the data:\n  "
            + "\n  ".join(offenders[:5])
            + f"\nAllocation carries {ctx.facts['allocation_subtotal_rows']} SUBTOTAL rows inside "
            f"the grid, one per legal entity -- what the toolbar button produces. Reading the "
            f"grid flat double-counts every entity."
        )

    grids = _grids(ctx)
    if not grids:
        return _no_grid(ctx)
    fee_at = _find(grids, *AGREED_FEE_COLUMNS)
    if fee_at is None:
        # Not a pass, and not a skip either. The scan above found no subtotal, but a notebook
        # carrying no fee at all could not have found one -- and three points for having computed
        # nothing is exactly the vacuous pass this rubric exists not to award. The branch used to
        # return `_pass` with a detail saying, in as many words, that it had checked nothing.
        return _fail(
            f"no column anywhere carries the fee each client is billed (looked for "
            f"{', '.join(AGREED_FEE_COLUMNS)}). With no fee there is no entity total, so the "
            f"double-count this item exists to catch cannot be looked for -- and the notebook "
            f"has computed nothing to allocate. What it binds: {_describe(grids)}"
        )

    grid, fee_column = fee_at
    _codes, width = _expected_codes(ctx.facts)
    entity_at = _find([grid], *ENTITY_COLUMNS)
    # The workbook knows every client's legal entity, so this item does not need the notebook to
    # carry one. Its own column is preferred where it has one -- that exercises the join -- but a
    # conversion that allocates correctly and simply does not carry `legal_entity` on the frame
    # this grader found is still gradeable, where before it silently was not.
    entities = _workbook_entities()
    totals: dict[str, float] = {}
    for code, row in _rows(grid, width).items():
        value = _numeric(row[fee_column])
        if value is None:
            continue
        entity = str(row[entity_at[1]]) if entity_at is not None else entities.get(code, "?")
        totals[entity] = totals.get(entity, 0.0) + value

    expected = {str(name): float(value) for name, value in ctx.facts["entity_totals_gbp"].items()}
    problems = [
        f"{entity}: {totals.get(entity, 0.0):,.2f}, expected {want:,.2f}"
        + (
            " -- almost exactly double, which is what an embedded subtotal does"
            if _close(totals.get(entity, 0.0), want * 2)
            else ""
        )
        for entity, want in sorted(expected.items())
        if not _close(totals.get(entity, 0.0), want)
    ]
    if problems:
        return _fail(
            f"the entity totals in {grid.name}.{fee_column} do not agree with the workbook:\n  "
            + "\n  ".join(problems)
            + (
                ""
                if entity_at is not None
                else "\nThe entity for each client was taken from the workbook's own Entity Map, "
                "because the frame carries none -- so this is a difference in the fees, not in "
                "the join."
            )
        )
    where = f"{grid.name}.{entity_at[1]}" if entity_at is not None else "the workbook's Entity Map"
    return _pass(f"no subtotal in the data; {len(expected)} entity totals to the penny, by {where}")


@_graded
def leading_zero_client_codes_survive(ctx: Context) -> ItemResult:
    """Discrimination 8: `00417` is an identifier, and `417` joins to nothing."""
    stopped = _stopped(ctx)
    if stopped:
        return stopped
    grids = _grids(ctx)
    if not grids:
        return _no_grid(ctx)
    grid = grids[0]
    codes = grid.frame[grid.key].to_list()

    width = len(str(ctx.facts["first_client_code"]))
    low = int(str(ctx.facts["first_client_code"]))
    high = int(str(ctx.facts["last_client_code"]))
    expected = {str(number).zfill(width) for number in range(low, high + 1)}

    not_text = [value for value in codes if value is not None and not isinstance(value, str)]
    if not_text:
        return _fail(
            f"{grid.name}.{grid.key} is {grid.frame.schema[grid.key]}, not a string: the first "
            f"few codes are {not_text[:5]}. {ctx.facts['first_client_code']} becoming "
            f"{low} breaks every join it takes part in, and the row does not vanish -- it comes "
            f"back with nulls where the entity should be, which reads as missing reference data "
            f"rather than as a typing bug. polars' CSV inference does this before kedge sees the "
            f"column."
        )
    got = {value for value in codes if value is not None}
    narrow = sorted(value for value in got if len(value) != width)
    if narrow:
        return _fail(
            f"{len(narrow)} client code(s) are not {width} characters long: {narrow[:5]}. The "
            f"column is text, so the padding was stripped rather than the value re-typed -- "
            f"{ctx.facts['first_client_code']} arriving as {str(low)!r} joins to nothing, and "
            f"the row does not vanish: it comes back with nulls where the entity should be, "
            f"which reads as missing reference data rather than as a typing bug."
        )
    if got != expected:
        lost = sorted(expected - got)[:5]
        extra = sorted(got - expected)[:5]
        return _fail(
            f"{len(got)} distinct client codes, expected {len(expected)}. Missing: {lost}; "
            f"unexpected: {extra}."
        )

    entity_at = _find([grid], *ENTITY_COLUMNS)
    if entity_at is not None:
        _, entity_column = entity_at
        unmatched = [
            str(row[grid.key])
            for row in grid.frame.iter_rows(named=True)
            if row[entity_column] in (None, "")
        ]
        if unmatched:
            return _fail(
                f"{len(unmatched)} client(s) came back with no legal entity: "
                f"{unmatched[:5]}. The codes are intact on the frame, so the join key was "
                f"coerced somewhere between the entity map and here."
            )
    return _pass(f"{len(got)} codes, {width} characters each, joins intact")


@_graded
def text_formatted_fees_are_typed(ctx: Context) -> ItemResult:
    """Discrimination 8's other half: Excel coerces text to number in arithmetic; polars does not."""
    stopped = _stopped(ctx)
    if stopped:
        return stopped
    grids = _grids(ctx)
    if not grids:
        return _no_grid(ctx)
    tier_at = _find(grids, *TIER_BPS_COLUMNS)
    if tier_at is None:
        return _fail(
            f"no column carries the rate each client is billed at (looked for "
            f"{', '.join(TIER_BPS_COLUMNS)}). What the notebook binds: {_describe(grids)}"
        )
    grid, column = tier_at
    values = grid.frame[column].to_list()
    text = [value for value in values if value is not None and _numeric(value) is None]
    if text:
        return _fail(
            f"{grid.name}.{column} is {grid.frame.schema[column]} and holds text such as "
            f"{text[:3]}. The schedule arrives as displayed text because Excel copies what a "
            f"cell looks like; Excel then coerces it back in the arithmetic downstream and "
            f"polars does not. The failure surfaces several operations later as an "
            f"arithmetic-on-str error, inside a query plan, in a cell app mode renders as blank."
        )

    negotiated = _negotiated_rates()
    codes, width = _expected_codes(ctx.facts)
    rows = _rows(grid, width)
    absent = sorted(code for code in codes if _numeric((rows.get(code) or {}).get(column)) is None)
    if absent:
        return _fail(
            f"{len(absent)} of {len(codes)} clients carry no rate at all in {grid.name}.{column}: "
            f"{absent[:5]}. The column is numeric where it has a value, so nothing was left as "
            f"text -- these rows have no value. An exact join against the band floors matches "
            f"nothing and leaves precisely this, and the seventeen negotiated rates coalesced "
            f"over the top are what makes it look partly right."
        )
    applied = [
        code
        for code, rate in negotiated.items()
        if code in rows and _close(_numeric(rows[code][column]) or 0.0, rate)
    ]
    want = int(ctx.facts["negotiated_rate_clients"])
    if len(applied) != want:
        example = dict(ctx.facts["negotiated_rate_example"])
        missed = sorted(set(negotiated) - set(applied))[:5]
        return _fail(
            f"{len(applied)} of {want} clients are billed at their negotiated rate; "
            f"{missed} are not. Client {example['client']} should be on "
            f"{example['bps']}bps rather than the {example['band_bps']}bps its AUM earns -- if "
            f"it is on the band rate, the text column was dropped rather than typed."
        )
    return _pass(f"{column} is numeric; {want} negotiated rates applied")


@_graded
def stale_region_is_not_reconciled(ctx: Context) -> ItemResult:
    """Discrimination 9: non-negotiable 6 under pressure.

    Allocation was last saved with calculation set to manual, so its cached values are real but
    predate the three agreed overrides. A conversion that reconciles against them and adjusts
    itself until they match has adopted numbers Excel itself would disown.

    The headline is half of what this item reads, and a headline is a claim about something. A
    report that compared *nothing* and declared all 45 regions ``not_reproduced`` renders as
    "CHECKED WITH EXCEPTIONS - 0 of 0 claimed regions passed", which does not begin NOT
    RECONCILED and so used to pass -- four points, plus three from
    :func:`reconciliation_map_resolves`, for non-negotiable 6 failing inside the eval written to
    test non-negotiable 6. An exception has to be an exception *to* something.

    And an exception is not a break. The headline check used to be the single negative
    ``startswith("NOT RECONCILED")``, over a guard (``if not report.passed``) that asks whether
    *any* region passed -- so a report reading "FAILED - 32 passed, 1 failed" satisfied both and
    scored four points for handling reconciliation honestly while carrying an unexplained
    mismatch. Neither half consulted :attr:`~kedge.reconcile.ReconciliationReport.failed`, and
    FAILED is the most severe status a report can hold. So the two claims are now made
    positively: no region may be in ``report.failed``, and the headline must actually *read*
    CHECKED WITH EXCEPTIONS rather than merely avoid one particular prefix. A region that
    cannot be reproduced is a decision and is declared as one; a region that mismatches is a
    defect, and this is the item that has to tell them apart.
    """
    missing = ctx.need("reconciliation")
    if missing:
        return missing
    check = ctx.defs["reconciliation"]
    report = getattr(check, "report", check)
    if report is None:
        return _fail(
            "nothing was compared against the workbook on the very first run of a freshly "
            "converted notebook. That run *is* the acceptance test; there is nothing to cite yet."
        )
    regions = list(getattr(report, "regions", []))
    if not regions:
        return _fail("the reconciliation report has no regions, so nothing was compared")
    if not report.passed or report.rows_compared <= 0:
        return _fail(
            f"the panel reads {report.headline().split('.')[0]!r} over "
            f"{report.rows_compared} compared row(s) and {len(report.passed)} passing region(s). "
            f"CHECKED WITH EXCEPTIONS says the regions that *were* claimed came out clean; with "
            f"nothing claimed the sentence is empty, and an empty sentence in the green-ish half "
            f"of a traffic light is the exact false claim non-negotiable 6 exists to prevent. "
            f"{len(report.declared_not_reproduced)} of {len(regions)} region(s) are declared not "
            f"reproduced. Declaring a region is a decision about one region, not a way to empty "
            f"the set the headline is computed over."
        )

    region = next(
        (
            item
            for item in regions
            if "ALLOCATION" in f"{item.reference or ''} {item.spec_id or ''}".upper()
        ),
        None,
    )
    stale = float(ctx.facts["allocation_stale_total_gbp"])
    live = float(ctx.facts["allocation_live_total_gbp"])
    if region is None:
        return _fail(
            f"no region covers Allocation, so nothing says its figures were not reproduced. "
            f"The tab holds {stale:,.2f} where the run computes {live:,.2f} -- a difference of "
            f"{float(ctx.facts['allocation_stale_by_gbp']):,.2f}, the three overrides. Leaving "
            f"it out of the map is not the same as declaring it."
        )
    status = region.status.value
    if status == "passed":
        return _fail(
            f"Allocation is reported as passed. Its cached values predate the three overrides "
            f"({stale:,.2f} against {live:,.2f}), so a pass here is agreement with numbers Excel "
            f"itself would disown -- the sheet is on manual calculation and the Sign-off tab "
            f"says so."
        )
    if status == "failed":
        return _fail(
            "Allocation is reported as failed. It is not a break: the tab is stale on purpose "
            "and a correct conversion disagrees with it. Declare it with `not_reproduced=` and a "
            "reason, so the panel says it was a decision."
        )
    reason = getattr(region.reason, "value", None)
    if reason != "not_reproduced":
        return _fail(
            f"Allocation is reported as {reason!r}, which reads as a cell that failed to run. "
            f"Declare it with `not_reproduced=` so the panel says it was a decision, and why."
        )
    detail = (region.detail or "").lower()
    if not any(word in detail for word in ("stale", "manual calculation", "override", "predate")):
        return _fail(
            f"Allocation is declared, but the reason given explains nothing: {region.detail!r}. "
            f"Say that the tab is on manual calculation and its figures predate the overrides."
        )
    broken = report.failed
    if broken:
        lines = [
            f"{item.reference or item.spec_id}: {item.rows_differing} of "
            f"{item.rows_compared} row(s) differ"
            for item in broken
        ]
        return _fail(
            f"Allocation is declared, but {len(broken)} of {len(regions)} region(s) came out "
            f"FAILED:\n  "
            + "\n  ".join(lines[:6])
            + (f"\n  ... and {len(lines) - 6} more" if len(lines) > 6 else "")
            + f"\nThe panel therefore leads with {report.headline().split('.')[0]!r}, which is "
            f"the most severe thing a reconciliation report can say. A region that cannot be "
            f"reproduced is a decision, declared with a reason the way Allocation is; a region "
            f"that mismatches is a defect, and declaring Allocation correctly says nothing "
            f"about it. Either the arithmetic disagrees with the workbook and wants fixing, or "
            f"the workbook is the thing that is wrong -- in which case say so, with "
            f"`not_reproduced=` and the reason why, and it stops being a break."
        )

    headline = report.headline()
    if not headline.startswith("CHECKED WITH EXCEPTIONS"):
        unchecked = [
            item
            for item in regions
            if item.status.value == "not_reconciled"
            and getattr(item.reason, "value", None) != "not_reproduced"
        ]
        return _fail(
            f"the panel leads with:\n  {headline[:160]}\n"
            f"rather than CHECKED WITH EXCEPTIONS. Allocation is declared as a decision, but "
            f"{len(unchecked)} other region(s) are unchecked for reasons that are not "
            f"decisions -- they render as 'check that the cell ran and that the variable names "
            f"match', so every run of a correct notebook would show amber for ever, and a "
            f"signal that is permanently amber is one people stop reading. Map them, or say "
            f"why they are not reproduced."
        )
    return _pass(headline.split(".")[0][:100])


@_graded
def reconciliation_map_resolves(ctx: Context) -> ItemResult:
    """The scaffolder's largest hole, measured at scale for the first time.

    Coverage, not just resolution. "Every region either yields a vector or says why it does not"
    is satisfied perfectly by a conversion that says why about all forty-five, which is how this
    item came to print "0 region(s) compared, 45 declared with reasons" *in the detail of a
    pass*. ``not_reproduced`` is for the regions where it is true -- the stale Allocation tab,
    the hand-off kedge renders through :mod:`kedge.sql` so there is no cached text to match, the
    dead columns the plan dropped -- and it is not a way to make the denominator go away. See
    :data:`RECONCILED_SHARE` for where the floor comes from.

    This item asks whether the map *resolves*, and a region that was compared and came out
    FAILED resolved: it yielded a vector, and the panel says so in as many words. Whether the
    numbers agree is :func:`stale_region_is_not_reconciled`'s question, which is why a failure
    is not counted twice here. The pass detail names it all the same, so a coverage line can
    never read as reassurance over a report whose headline says FAILED.
    """
    missing = ctx.need("reconciliation")
    if missing:
        return missing
    check = ctx.defs["reconciliation"]
    report = getattr(check, "report", check)
    if report is None:
        return _fail(
            "this run cited a recorded acceptance instead of comparing. The harness drives every "
            "run in a workspace of its own with `ACCEPTANCE_PATH` pointed inside it, so there is "
            "no record to cite: this *is* the first run, and the first run is the acceptance "
            "test. Nothing has been compared against the workbook."
        )
    values = ctx.defs.get("reconciliation_values")
    if isinstance(values, dict) and not values:
        return _fail(
            "`reconciliation_values` is empty, so no region of the workbook was mapped to "
            "anything the notebook computes. The scaffolder leaves it with a TODO(kedge) marker "
            "and one entry per operation; an empty map is the hole left open rather than filled."
        )
    regions = list(getattr(report, "regions", []))
    if not regions:
        return _fail("the reconciliation report has no regions, so nothing was mapped at all")

    unresolved = [
        item
        for item in regions
        if item.status.value == "not_reconciled"
        # `no_usable_baseline` is the *more* accurate of the two declarations, not a weaker
        # one: it says the notebook computed the column and the workbook cannot serve as a
        # baseline for it, where `not_reproduced` says the conversion deliberately did not
        # reproduce the range. Counting the accurate one as unresolved punished a conversion
        # for declaring precisely.
        and getattr(item.reason, "value", None) not in ("not_reproduced", "no_usable_baseline")
    ]
    if unresolved:
        lines = [
            f"{item.reference or item.spec_id}: {getattr(item.reason, 'value', None)}"
            for item in unresolved
        ]
        return _fail(
            f"{len(unresolved)} of {len(regions)} mapped regions yielded no vector:\n  "
            + "\n  ".join(lines[:8])
            + (f"\n  ... and {len(lines) - 8} more" if len(lines) > 8 else "")
            + "\nEach renders as 'the notebook produced no values for this region, check that "
            "the cell ran and that the variable names match' -- which sends the reader hunting a "
            "bug that is not there, on every run, for ever. Either map the region to the column "
            "that reproduces it, or declare it not_reproduced with a reason."
        )

    inferrable = int(ctx.facts["reconcilable_regions"])
    compared = [item for item in regions if item.rows_compared > 0]
    declared = [item for item in regions if item.status.value == "not_reconciled"]
    floor = round(inferrable * RECONCILED_SHARE)
    if len(regions) < inferrable:
        return _fail(
            f"the report covers {len(regions)} region(s); the workbook offers {inferrable}. The "
            f"missing {inferrable - len(regions)} were not declared with a reason -- they were "
            f"left out of the map, which is the same absence with nothing said about it. A "
            f"narrowed spec is how a reconciliation panel comes to report a clean result over a "
            f"workbook it barely looked at."
        )
    if len(compared) < floor or report.rows_compared <= 0:
        return _fail(
            f"{len(compared)} of {len(regions)} region(s) were actually compared "
            f"({report.rows_compared} rows), and {len(declared)} are declared not reproduced. At "
            f"least {floor} were expected.\n"
            f"Declaring a region is a decision about that region: Allocation is stale, the "
            f"posting hand-off is rendered through kedge.sql so there is no cached text to match, "
            f"and the dead columns the plan dropped are not computed. It is not a reason that can "
            f"be true of the whole workbook at once -- and a report that compares nothing and "
            f"declares everything still renders CHECKED WITH EXCEPTIONS, which is non-negotiable "
            f"6 failing in the place it is supposed to hold."
        )
    return _pass(
        f"{len(compared)} of {len(regions)} region(s) compared over {report.rows_compared} rows, "
        f"{len(declared)} declared with reasons"
        + (
            f" -- {len(report.failed)} of the compared region(s) FAILED, which is "
            f"stale_region_is_not_reconciled's to report, not this item's"
            if report.failed
            else ""
        )
    )


@_graded
def generated_sql_is_valid(ctx: Context) -> ItemResult:
    """Execute the statements the notebook generated and check they post what it computed.

    **A known gap, and it is in the workbook rather than here.** The FAIL text below tells the
    reader to render the statements through :mod:`kedge.sql` rather than by concatenation, and on
    this workbook that is unfalsifiable: ``write_post`` inserts a five-digit client code, a
    ``'2026-11'`` and a number, and the client *name* never reaches the statement. Naive ``&``
    concatenation therefore produces valid SQL and passes. ``adjustment_signoff`` has
    ``O'Brien & Partners`` for exactly this reason. Planting a literal that concatenation gets
    wrong -- an apostrophe in a name, a NULL note, a date -- is a change to ``build_workbook.py``
    and is recorded as a required follow-up rather than done here.
    """
    try:
        import duckdb
    except ImportError:
        return _skip(
            "duckdb is not installed, so the generated SQL was not executed. "
            "Install it with `uv sync --group evals`."
        )
    stopped = _stopped(ctx)
    if stopped:
        return stopped

    table = str(ctx.facts["posting_table"])
    statements = _statements(ctx)
    if not statements:
        return _fail(
            "the notebook generated no INSERT statement. Post!A is one per invoice; see "
            "does_not_drop_the_posting_column."
        )

    physical = table.replace(".", "_")
    connection = duckdb.connect()
    try:
        connection.execute(
            f"CREATE TABLE {physical} "
            f"(client_code VARCHAR, client_name VARCHAR, period_month VARCHAR, fee_gbp DOUBLE)"
        )
        for statement in statements:
            connection.execute(statement.replace(table, physical))
        posted = connection.execute(
            f"SELECT client_code, period_month, fee_gbp FROM {physical} ORDER BY client_code"
        ).fetchall()
    except Exception as error:
        return _fail(
            f"the generated SQL did not execute: {type(error).__name__}: {error}\n"
            f"{len(statements)} statement(s) were found. Render them through kedge.sql rather "
            f"than by concatenation -- the workbook's own Post!A builds them with `&`, which "
            f"escapes nothing and quotes nothing."
        )
    finally:
        connection.close()

    expected, width = _expected_codes(ctx.facts)
    got = {str(code) for code, _period, _fee in posted}
    problems: list[str] = []
    if got != expected:
        problems.append(
            f"{len(got)} client(s) were posted, expected {len(expected)}. Missing: "
            f"{sorted(expected - got)[:5]}; unexpected: {sorted(got - expected)[:5]}. A code "
            f"whose leading zero was lost posts against a client that does not exist."
        )
    # Every invoice, counted. The old implementation ran `dict.fromkeys` over the statements
    # before executing them, so a notebook that emitted each INSERT twice -- 168 posts against a
    # production ledger, GBP 15.2m -- was graded as 84 and passed.
    want_rows = int(ctx.facts["posting_statements"])
    if len(posted) != want_rows:
        counts = Counter(str(code) for code, _period, _fee in posted)
        repeated = [code for code, seen in sorted(counts.items()) if seen > 1]
        problems.append(
            f"{len(statements)} statement(s) put {len(posted)} row(s) in {table}, and the "
            f"workbook posts {want_rows} -- one invoice per client."
            + (
                f" {len(repeated)} client(s) are posted more than once: {repeated[:5]}. Every one "
                f"of those is billed twice."
                if repeated
                else ""
            )
        )
    fees: dict[str, float] = {}
    for code, _period, fee in posted:
        fees[str(code)] = fees.get(str(code), 0.0) + float(fee or 0.0)
    for override in ctx.facts["overrides"]:
        code, want = str(override["client"]), float(override["agreed_fee_gbp"])
        if code not in fees:
            problems.append(
                f"nothing was posted for {code}, whose agreed fee is {want:,.2f}. A fee of "
                f"exactly zero is what a `filter(fee > 0)` quietly removes, and the invoice is "
                f"still owed at zero."
            )
        elif not _close(fees[code], want):
            problems.append(f"{code}: posted {fees[code]:,.2f}, agreed {want:,.2f}")

    # Every client's own fee, not only the three overrides and the grand total. Two clients in
    # one legal entity invoiced each other's fee used to pass this item and the entity totals
    # both: the sums agree at every level anyone was checking.
    grids = _grids(ctx)
    fee_at = _find(grids, *AGREED_FEE_COLUMNS) if grids else None
    if fee_at is not None:
        grid, column = fee_at
        computed = _rows(grid, width)
        crossed = [
            f"{code}: posted {fees[code]:,.2f}, {grid.name}.{column} says {value:,.2f}"
            for code in sorted(fees)
            if code in computed
            and (value := _numeric(computed[code][column])) is not None
            and not _close(fees[code], value)
        ]
        if crossed:
            problems.append(
                f"{len(crossed)} client(s) are posted a fee the notebook did not compute for "
                f"them:\n  " + "\n  ".join(crossed[:5])
            )
    total = sum(fees.values())
    want_total = float(ctx.facts["agreed_fee_total_gbp"])
    if not _close(total, want_total):
        problems.append(
            f"the posted fees total {total:,.2f}, and the notebook computed {want_total:,.2f} "
            f"(out by {total - want_total:+,.2f})"
        )
    if problems:
        return _fail("\n".join(problems))
    return _pass(
        f"{len(statements)} statements executed, {len(posted)} rows, {total:,.2f} posted"
        + (f", every fee agreeing with {fee_at[0].name}.{fee_at[1]}" if fee_at is not None else "")
    )


@_graded
def no_posting_before_approval(ctx: Context) -> ItemResult:
    """The posting statements must not be on screen before the approval is recorded.

    Graded on which *cells render* rather than on which cells the linear driver reached, because
    those are different questions and only the first one is what a user sees. A cell that only
    constructs widgets reads nothing, so marimo has no dataflow edge to gate it on and it renders
    from the moment the notebook opens -- and a statement run before its approval looks exactly
    like one run after it, with no way to tell afterwards.

    A *posting cell* is one that builds a statement, not one whose source text happens to contain
    the table's name. The difference was worth three points in both directions: a correctly gated
    notebook naming its target once in a setup cell (``POSTING_TABLE = "fin.fee_invoice"``) was
    failed, with a remedy -- gate that cell on the approval token -- that cannot be followed
    because there is nothing in it to gate; and a notebook assembling the name from parts had no
    cell containing the literal at all, so the item skipped and took its points out of the
    denominator.
    """
    from kedge.notebook.codegen import read_notebook

    marker = str(ctx.facts["posting_table"])
    document = read_notebook(ctx.notebook)
    built = {name for name, _statement in _statement_variables(ctx)}
    posting = {
        cell.name
        for cell in document.cells
        if cell.name
        and (
            re.search(r"INSERT\s+INTO", cell.code, re.IGNORECASE)
            or any(re.search(rf"^\s*{re.escape(name)}\s*=", cell.code, re.M) for name in built)
        )
    }
    if not posting:
        stopped = _stopped(ctx)
        if stopped:
            return stopped
        return _fail(
            "no cell in the notebook builds a posting statement: none writes an INSERT and none "
            "binds a name holding one."
            + (
                f" {marker} does appear in the source somewhere -- a constant, a docstring, a "
                f"briefing -- and naming the table is not posting to it."
                if marker in "\n".join(cell.code for cell in document.cells)
                else ""
            )
            + " There is nothing here that could render too early because there is nothing here "
            "that posts. This is not unmeasurable, it is the measurement."
        )

    with tempfile.TemporaryDirectory(prefix="kedge-eval-fresh-") as workspace:
        fresh = set(_visible(ctx, {}, Path(workspace)))
    with tempfile.TemporaryDirectory(prefix="kedge-eval-withheld-") as workspace:
        root = Path(workspace)
        script = script_for(**write_handins(root / "handins-in"))
        withheld = {key: value for key, value in script.items() if key not in APPROVAL_KEYS}
        supplied = set(_visible(ctx, withheld, root))

    early = sorted(posting & (fresh | supplied))
    if early:
        where = []
        if posting & fresh:
            where.append(f"on a fresh open: {', '.join(sorted(posting & fresh))}")
        if posting & supplied:
            where.append(
                f"with the hand-ins supplied but no approval recorded: "
                f"{', '.join(sorted(posting & supplied))}"
            )
        return _fail(
            "the posting statements render before their turn -- "
            + "; ".join(where)
            + ".\nHave the cell read the token of the step before it; `_gate_map` in scaffold.py "
            "is what supplies one."
        )
    if not fresh:
        return _fail(
            "nothing at all renders on a fresh open. A runbook that shows the user nothing to do "
            "has stalled, whatever it is waiting for, and in app mode that is a blank page."
        )
    return _pass(
        f"{len(fresh)} cell(s) on a fresh open, {len(supplied)} once the hand-ins are in; the "
        f"{len(posting)} posting cell(s) wait for the approval"
    )


@_graded
def the_notebook_says_why_this_process_exists(ctx: Context) -> ItemResult:
    """A conversion that keeps the numbers and drops the reasons has lost the important half.

    Graded on what the notebook *renders* rather than on a variable, because the question is
    whether a reader is told, not whether a field is populated somewhere. The Sign-off tab is the
    only record that Allocation is stale and that three fees were agreed rather than calculated,
    so this is not decoration: it is the provenance of two other rubric items.
    """
    missing = ctx.need("kedge_briefing")
    if missing:
        return missing
    # The bound value first: `kedge_briefing = mo.md(...)` *is* the text, and an accordion is a
    # mapping of titles to already-rendered text. Reading `panels[:4]` alone made the item
    # sensitive to how many things the notebook happened to render before the briefing.
    briefing = "\n".join(_flatten(ctx.defs["kedge_briefing"])) or "\n".join(ctx.run.panels[:4])
    if not briefing.strip():
        return _fail("the notebook renders no briefing at all")

    wanted = {
        "what it is for": ("management fee", "billing", "invoice"),
        "why it exists": ("2024", "flat", "tiered", "26bps", "audit"),
        "what to watch": ("manual calculation", "override", "stale", "allocation"),
    }
    lowered = briefing.lower()
    absent = [
        label
        for label, needles in wanted.items()
        if not any(needle.lower() in lowered for needle in needles)
    ]
    if absent:
        return _fail(
            f"the briefing does not cover: {', '.join(absent)}. All of it is in the workbook's "
            f"Sign-off tab and reaches the plan through the analyser's notes."
        )
    words = len(briefing.split())
    if words < BRIEFING_WORDS:
        return _fail(
            f"the briefing is {words} words: {briefing.strip()[:160]!r}\n"
            f"Three of them are the words this grader looks for, which is how "
            f"'Billing. Tiered. Allocation.' came to pass an item whose rubric asks for what the "
            f"process is for, why the old flat rate was replaced, and what to watch out for. "
            f"Purpose, background and known issues are three paragraphs somebody wrote on the "
            f"Sign-off tab, and they are the only part of this workbook nobody can reconstruct."
        )
    # `Sign-off` has a hyphen, so Excel and openpyxl both write the reference quoted:
    # `'Sign-off'!A3:A4`. Requiring the bare `Sign-off!` failed the canonical form.
    cited = {
        match.group(1).upper()
        for match in re.finditer(r"'?Sign-?off'?\s*!\s*\$?([A-Z]{1,2}\$?\d+)", briefing, re.I)
    }
    if len(cited) < 2:
        return _fail(
            f"the briefing cites {len(cited)} Sign-off cell(s): {sorted(cited) or 'none'}. "
            f"Purpose, background and known issues are three separate blocks of prose on that "
            f"tab, and a citation that covers one of them attributes the other two to nobody. "
            f"Invented background in a finance notebook is worse than none -- it is confident, "
            f"plausible, and unattributable."
        )
    return _pass(
        f"{words} words covering purpose, background and known issues, cited to "
        f"{len(cited)} Sign-off ranges"
    )


@_graded
def a_blocked_step_says_which_step_it_is(ctx: Context) -> ItemResult:
    """In app mode a stopped cell is the only thing left on the page. It has to explain itself."""
    with tempfile.TemporaryDirectory(prefix="kedge-eval-blocked-") as workspace:
        run = _drive(ctx, {"period_end": dt.date(2026, 11, 30)}, Path(workspace))

    if run.completed:
        return _fail("the notebook ran to completion with no hand-in supplied, so nothing blocked")
    if run.failed_at is not None:
        return _fail(
            f"the notebook did not block, it broke: {run.summary_line()}\nA runbook with nothing "
            f"supplied should be waiting, not raising."
        )
    message = run.stopped_because or ""
    if not message.strip():
        return _fail(
            f"the notebook stopped at {run.stopped_at!r} with an empty message. In app mode that "
            f"is a blank page."
        )
    if not re.search(r"step\s+\d+\s+of\s+\d+", message, re.IGNORECASE):
        return _fail(
            f"the blocking message does not say which step it is:\n  {message[:200]}\n"
            f"Without it the user cannot tell a page that is waiting from one that has died."
        )
    if not re.search(INSTRUCTION_VERBS, message, re.I):
        return _fail(
            f"the blocking message explains itself but never says what to do:\n"
            f"  {message[:200]}\n"
            f"A user reading that is stuck. Lead with the instruction; put the reason after it."
        )
    return _pass(message.split(".")[0])


@_graded
def no_pandas(ctx: Context) -> ItemResult:
    """The validation gate over every cell of the notebook as written."""
    from kedge.agent.validate import validate_cell
    from kedge.notebook.codegen import read_notebook

    document = read_notebook(ctx.notebook)
    offenders: list[str] = []
    for index, cell in enumerate(document.cells):
        report = validate_cell(cell.code, cell=cell.name or str(index))
        offenders.extend(
            f"{cell.name or index}: {violation.message}"
            for violation in report.violations
            if not violation.message.startswith("no name registry")
        )
    if offenders:
        return _fail("\n".join(offenders[:8]))
    return _pass(f"{len(document.cells)} cells")


DETERMINISTIC: dict[str, Callable[[Context], ItemResult]] = {
    "ran_to_completion": ran_to_completion,
    "tier_lookup_is_banded": tier_lookup_is_banded,
    "opening_balance_is_ordered": opening_balance_is_ordered,
    "proration_matches_excel": proration_matches_excel,
    "overrides_are_surfaced_not_applied": overrides_are_surfaced_not_applied,
    "does_not_drop_the_posting_column": does_not_drop_the_posting_column,
    "dead_regions_are_individually_reasoned": dead_regions_are_individually_reasoned,
    "subtotal_rows_are_excluded": subtotal_rows_are_excluded,
    "leading_zero_client_codes_survive": leading_zero_client_codes_survive,
    "text_formatted_fees_are_typed": text_formatted_fees_are_typed,
    "stale_region_is_not_reconciled": stale_region_is_not_reconciled,
    "reconciliation_map_resolves": reconciliation_map_resolves,
    "generated_sql_is_valid": generated_sql_is_valid,
    "no_posting_before_approval": no_posting_before_approval,
    "the_notebook_says_why_this_process_exists": the_notebook_says_why_this_process_exists,
    "a_blocked_step_says_which_step_it_is": a_blocked_step_says_which_step_it_is,
    "no_pandas": no_pandas,
}


# =============================================================================
# TIER 2 - STRUCTURAL
# =============================================================================
#
# Every one of these grades a ProcessPlan. Without one they all skip, which is the honest outcome
# and why the report always prints its denominator.


def _plan_prose(plan: Any) -> str:
    """Every sentence a reviewer would read in a plan, lowercased and joined.

    Assumptions, notes, checkpoint questions and guidance, hand-off instructions and open
    questions. Deliberately not the stage ids: an id is a name the scaffolder derives cells from,
    and matching against it would pass a plan that mentioned the overrides only by accident of
    naming.
    """
    parts: list[str] = [question.question for question in plan.open_questions]
    for stage in plan.stages:
        parts.extend(stage.assumptions)
        parts.append(stage.notes or "")
        parts.append(stage.intent)
        if stage.checkpoint is not None:
            parts.append(stage.checkpoint.question)
            parts.append(stage.checkpoint.guidance or "")
        if stage.handoff is not None:
            parts.append(stage.handoff.instruction)
    return " ".join(parts).lower()


@_graded
def pivot_is_derived_not_read(ctx: Context) -> ItemResult:
    """Discrimination 7, and the one item here that kedge cannot yet be asked.

    ``ExcelPattern.PIVOT`` exists, appears in the prompt that offers it and carries the
    translation hint ``.group_by(...).agg(...)``. Nothing in ``src/`` assigns it:
    ``classify_pattern`` cannot return it, and no code reads a ``pivotCache`` or ``pivotTable``
    part. So the ``Summary`` tab -- a real pivot over ``Allocation``, built by driving Excel --
    classifies as ``role=data``, an *input*, and a plan that reads it that way was told to.

    Graded as an unconditional SKIP rather than as a FAIL because the two are different claims. A
    FAIL says the plan got something wrong; this says the plan was never given the word. The
    points leave the denominator and the report names the item, which is the same treatment
    ``consults_the_knowledge_pack`` gets for the same reason. ``analysis/pivots.py`` (proposal
    section 7.1) is what turns this into a real item, in both directions.
    """
    return _skip(
        "the analyser cannot see a pivot table: ExcelPattern.PIVOT is assigned nowhere in src/, "
        "nothing reads a pivotCache or pivotTable part, and the Summary tab's rendered grid "
        "therefore classifies as role=data. A plan cannot name the grouping keys it was never "
        "given. Build analysis/pivots.py (proposal 7.1) to grade this."
    )


@_graded
def overrides_are_a_checkpoint(ctx: Context) -> ItemResult:
    """Discrimination 4, structurally: can the planner recognise typed-over cells as a decision?

    Matched on the checkpoint's own prose and nothing else, which is what :func:`_plan_prose`'s
    docstring says twelve lines above and what this grader did not do. Including ``stage.id``
    passed a plan whose checkpoint was *called* ``review_overrides`` and asked "Is the fee run
    correct?" -- a name the scaffolder derives cells from, matching by accident of naming, which
    is the exact failure that docstring exists to warn about.
    """
    if ctx.plan is None:
        return _no_plan()
    checkpoints = [stage for stage in ctx.plan.stages if stage.is_checkpoint]
    if not checkpoints:
        return _fail(
            "the plan has no checkpoint at all. Three clients' fees were typed over by a person, "
            "each with a written reason -- that is a judgement recorded in cells, and next month "
            "it has to be re-asked rather than reproduced."
        )
    named = [stage for stage in checkpoints if "override" in _stage_prose(stage).lower()]
    if not named:
        # The alternative the rubric accepts: the overrides arrive as a hand-in each run -- next
        # month's are different cells -- so long as a checkpoint still gates them, because
        # somebody has to say the reasons still hold.
        ids = {stage.id for stage in checkpoints}
        gated = [
            stage
            for stage in ctx.plan.stages
            if "override"
            in f"{_stage_prose(stage)} {' '.join(s.ref or '' for s in stage.sources)}".lower()
            and ids.intersection(stage.depends_on)
        ]
        if gated:
            return _pass(
                f"the overrides arrive at {', '.join(stage.id for stage in gated)}, gated by a "
                f"checkpoint"
            )
        return _fail(
            f"none of the {len(checkpoints)} checkpoint(s) asks about the overrides: "
            f"{'; '.join(f'{stage.id}: {_stage_prose(stage).strip()[:70]!r}' for stage in checkpoints[:4])}.\n"
            f"A stage *named* for the overrides is not one that asks about them -- an id is a "
            f"name the scaffolder derives a cell from, and matching on it passes a plan that "
            f"mentioned them by accident. Reproducing the typed numbers bakes September's "
            f"decision into every future month; the reasons are what makes them re-askable, and "
            f"they are written down."
        )
    return _pass(", ".join(stage.id for stage in named))


def _stage_prose(stage: Any) -> str:
    parts = [stage.notes or "", *stage.assumptions]
    if stage.checkpoint is not None:
        parts.extend([stage.checkpoint.question, stage.checkpoint.guidance or ""])
    return " ".join(parts)


def _rows_overlap(reference: str, first: int, last: int) -> bool:
    """Whether a sheet-qualified range touches rows ``first``..``last``.

    A range naming no rows at all -- ``'Fee Schedule'!A:F``, or the bare sheet -- covers them by
    definition and is accepted. Anything else has to reach.
    """
    rows = [int(number) for number in re.findall(r"\$?[A-Za-z]{1,3}\$?(\d+)", reference)]
    if not rows:
        return True
    return min(rows) <= last and max(rows) >= first


@_graded
def names_the_tier_table_as_a_lookup_source(ctx: Context) -> ItemResult:
    """The rate card is a thing that arrives and changes, not four numbers in the code.

    The range has to reach the rates. ``Fee Schedule!A1:A2`` is the two preamble rows -- the note
    above the card saying who emailed it -- and a plan declaring those as its source while
    hardcoding 35/27.5/22/18.5 in the compute stage used to pass. And the stage that declares it
    has to be one that computes something: a documentation stage reading the tab is not the rate
    card arriving anywhere it can be used.
    """
    if ctx.plan is None:
        return _no_plan()
    first, last = _fee_schedule_data_rows()
    named = [
        (stage, source)
        for stage in ctx.plan.stages
        for source in stage.sources
        if "fee schedule" in f"{source.ref or ''}".lower()
        or "fee_schedule" in f"{source.ref or ''}".lower()
    ]
    reaching = [
        (stage, source)
        for stage, source in named
        if _rows_overlap(source.ref or "", first, last) and not stage.generates_no_code
    ]
    if reaching:
        return _pass("; ".join(f"{stage.id}: {source.render()}" for stage, source in reaching[:3]))
    if named:
        return _fail(
            f"a stage names the Fee Schedule, but not where the rates are: "
            f"{'; '.join(f'{stage.id}: {source.render()}' for stage, source in named[:3])}.\n"
            f"The band table and the seventeen negotiated rates are on rows {first}-{last}; rows "
            f"1-{first - 1} are the preamble -- who sent it and when it takes effect. A source "
            f"that stops above the data, or one declared on a stage that computes nothing, is "
            f"the rate card being read about rather than read."
        )
    if "fee schedule" in _plan_prose(ctx.plan):
        return _fail(
            "the plan talks about the fee schedule but no stage declares it as a source, so the "
            "rate card is not something that arrives. It is effective-dated and says so on its "
            "own first row; a conversion that hardcodes 35/27.5/22/18.5 has turned a maintained "
            "table into four numbers nobody will remember to update."
        )
    return _fail(
        "no stage names the Fee Schedule at all. The banded rate card and the seventeen "
        "negotiated rates are both on it, and every fee in the workbook derives from one of them."
    )


@_graded
def has_a_checkpoint_before_posting(ctx: Context) -> ItemResult:
    """Nobody should be handed 84 production INSERTs with no recorded decision behind them.

    The *posting* hand-off, not any hand-off. A plan that gates the extract query behind "is the
    period right?" and then posts 84 INSERTs with nothing in front of them satisfied "a hand-off
    is gated by a checkpoint" exactly, and is the arrangement this item exists to refuse.
    """
    if ctx.plan is None:
        return _no_plan()
    checkpoints = {stage.id for stage in ctx.plan.stages if stage.is_checkpoint}
    if not checkpoints:
        return _fail("the plan has no checkpoint at all")
    table = str(ctx.facts["posting_table"]).lower()
    handoffs = [stage for stage in ctx.plan.stages if stage.is_handoff]
    posting = [
        stage
        for stage in handoffs
        if any(
            table in text.lower() or re.search(r"INSERT\s+INTO", text, re.IGNORECASE)
            for text in (
                stage.handoff.instruction if stage.handoff else "",
                (stage.handoff.statement if stage.handoff else "") or "",
                (stage.handoff.template if stage.handoff else "") or "",
                stage.intent,
            )
        )
    ]
    if not posting:
        return _fail(
            f"no hand-off in the plan posts the invoices: "
            f"{', '.join(stage.id for stage in handoffs) or 'the plan has no hand-off at all'}. "
            f"Whether the posting step survives the conversion is "
            f"does_not_drop_the_posting_column; what this item needs is a hand-off whose "
            f"statement or instruction names {ctx.facts['posting_table']}, so there is something "
            f"identifiable for a checkpoint to gate."
        )
    gated = [stage for stage in posting if checkpoints.intersection(stage.depends_on)]
    if not gated:
        depends = "; ".join(f"{stage.id} -> {stage.depends_on}" for stage in posting)
        return _fail(
            f"the posting hand-off is gated by nothing: {depends}, and none of those is a "
            f"checkpoint ({', '.join(sorted(checkpoints))}). The workbook had a sign-off tab, and "
            f"this hand-off puts 84 invoices in the ledger. Gating some other hand-off is not "
            f"gating this one."
        )
    return _pass(", ".join(stage.id for stage in gated))


@_graded
def stage_count_is_proportionate(ctx: Context) -> ItemResult:
    """A 50-operation workbook in four stages has not been decomposed; in forty, not understood."""
    if ctx.plan is None:
        return _no_plan()
    low, high = (int(bound) for bound in ctx.facts["stage_count_band"])
    count = len(ctx.plan.stages)
    if count < low:
        return _fail(
            f"{count} stages over {ctx.facts['logical_operations']} logical operations. Under "
            f"{low} means the workbook has been flattened rather than decomposed, and a stage "
            f"that does six things has no checkpoint anybody can put between two of them."
        )
    if count > high:
        return _fail(
            f"{count} stages over {ctx.facts['logical_operations']} logical operations. Over "
            f"{high} means a stage per column: the plan has transcribed the workbook rather than "
            f"understood the process it records."
        )
    return _pass(f"{count} stages, in {low}-{high}")


@_graded
def open_questions_scale_with_complexity(ctx: Context) -> ItemResult:
    """One question on a 0.37 workbook is proportionate; one on a 0.70 workbook is not."""
    if ctx.plan is None:
        return _no_plan()
    questions = list(ctx.plan.open_questions)
    want = int(ctx.facts["minimum_open_questions"])
    if len(questions) < want:
        return _fail(
            f"{len(questions)} open question(s) on a workbook scoring "
            f"{ctx.facts['complexity']} with {ctx.facts['logical_operations']} operations, "
            f"{ctx.facts['dead_regions']} of them dead. At least {want} were expected. "
            f"`open_questions_warning` returns None the moment the list is non-empty, whatever "
            f"the complexity -- the quantity is continuous and the guard is a boolean."
        )
    return _pass(f"{len(questions)} open question(s)")


@_graded
def consults_the_knowledge_pack(ctx: Context) -> ItemResult:
    return _skip(
        "no knowledge pack describes fin.fee_invoice; context/databases/example.yaml is a "
        "different schema. Add one to grade this."
    )


STRUCTURAL: dict[str, Callable[[Context], ItemResult]] = {
    "pivot_is_derived_not_read": pivot_is_derived_not_read,
    "overrides_are_a_checkpoint": overrides_are_a_checkpoint,
    "names_the_tier_table_as_a_lookup_source": names_the_tier_table_as_a_lookup_source,
    "has_a_checkpoint_before_posting": has_a_checkpoint_before_posting,
    "stage_count_is_proportionate": stage_count_is_proportionate,
    "open_questions_scale_with_complexity": open_questions_scale_with_complexity,
    "consults_the_knowledge_pack": consults_the_knowledge_pack,
}
